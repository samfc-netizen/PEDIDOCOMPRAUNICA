import re
import math
import difflib
import csv
import json
import base64
import hashlib
import unicodedata
import urllib.request
import urllib.error
import urllib.parse
from email.mime.text import MIMEText
from email.utils import formataddr
from io import BytesIO, StringIO
from datetime import datetime, date


# Dependências da multipage de Previsão Financeira
import io
import zipfile
import xml.etree.ElementTree as ET
from typing import Dict, Iterable, List, Optional, Tuple

import pdfplumber

try:
    import fitz  # PyMuPDF - leitura de PDF muito mais rápida quando instalado
except Exception:
    fitz = None

try:
    from PIL import Image
except Exception:
    Image = None

try:
    import pytesseract
except Exception:
    pytesseract = None

import pandas as pd
import streamlit as st

try:
    from openpyxl import Workbook
except Exception:
    Workbook = None

try:
    from google.oauth2.credentials import Credentials
    from google.oauth2 import service_account
    from google.auth.transport.requests import Request
    from googleapiclient.discovery import build
    from googleapiclient.http import MediaIoBaseUpload
except Exception:
    Credentials = None
    service_account = None
    Request = None
    build = None
    MediaIoBaseUpload = None

st.set_page_config(page_title="Análise de Giro e Pedido de Compra", layout="wide", page_icon="📊")

# =========================================================
# CONFIGURAÇÕES DO NEGÓCIO
# =========================================================

LOJAS_MAP = {
    "004": "ADE",
    "006": "GAMA",
    "009": "ÚNICA",
    "012": "SOFNORTE",
    "013": "CEILÂNDIA",
    "014": "SIA",
    "015": "UNAÍ",
    "016": "AG LINDAS",
    "022": "GUARÁ",
    "024": "LUZIÂNIA",
}

CODIGOS_LOJAS = ["004", "006", "012", "013", "014", "015", "016", "022", "024"]
CODIGO_UNICA = "009"
MESES_PADRAO = ["01/2026", "02/2026", "03/2026", "04/2026"]
MESES = MESES_PADRAO.copy()

GOOGLE_DRIVE_ROOT_FOLDER_ID = "1PqWXzphyeU_Q5Wc7UDYeZUzFA4kKZL-H"
GOOGLE_SUBPASTA_PEDIDOS = "Pedidos Editaveis"
GOOGLE_SUBPASTA_FINAIS = "Arquivos Finais"
GOOGLE_PLANILHA_CONTROLE = "Controle de Pedidos - Dauto"
GOOGLE_PLANILHA_CADASTRO = "Cadastro de Produtos - Dauto"
GOOGLE_PLANILHA_CADASTRO_ID = "1iu9dbvhQCqdfWTrRL2_HMnAkQQnbK_cqN_k-9pZpkBw"
GOOGLE_PLANILHA_PRECOS_BRASILUX_ID = "1OkQ-QamdcrCIg_LQUuIwdJrqrEAAoa5arLzkZD6rvi4"
GOOGLE_PLANILHA_PRECOS_BRASILUX_LINK = "https://docs.google.com/spreadsheets/d/1OkQ-QamdcrCIg_LQUuIwdJrqrEAAoa5arLzkZD6rvi4/edit"
GOOGLE_MODELO_PEDIDO_ID = "1iu9dbvhQCqdfWTrRL2_HMnAkQQnbK_cqN_k-9pZpkBw"
GOOGLE_PASTA_APROVACAO_ID = "1ZlTC720fGMHk6cqApXtjeNPBe84Vgx_b"
APPS_SCRIPT_WEB_APP_URL_PADRAO = "https://script.google.com/macros/s/AKfycbwihVeGlWG3-SpecqZhHR2TsFVYmYZEGlvyrImpLX9eliv-fl7CqUwWCWBnmnlFIBww/exec"
GOOGLE_PASTA_APROVADOS_ID = "1Ez4LgDFh964iF-MjUl1KFjvGGQMKtRz_"
GOOGLE_APROVADORES_EMAILS = [
    "samuel@dautotintas.com.br",
    "victor@dautotintas.com.br",
    "compras@unicaatacadista.com.br",
]
GOOGLE_PEDIDOS_COLUNAS = [
    "id_pedido", "nome_pedido", "fornecedor", "status", "valor",
    "criado_em", "criado_por", "aprovado_em", "aprovado_por",
    "link_pedido", "spreadsheet_id", "link_autcom", "link_fornecedor", "observacao",
]
GOOGLE_ACOMPANHAMENTO_COLUNAS = [
    "data", "mes", "fornecedor", "nome_pedido", "valor",
    "status", "link_pedido", "link_autcom", "link_fornecedor",
]

# =========================================================
# FUNÇÕES AUXILIARES
# =========================================================

def br_to_float(value):
    if value is None:
        return 0.0
    value = str(value).strip()
    if value == "" or value.lower() in ["nan", "none", "-"]:
        return 0.0
    value = value.replace("R$", "").replace(" ", "")
    value = value.replace(".", "").replace(",", ".")
    try:
        return float(value)
    except Exception:
        return 0.0


def numero_planilha_para_float(value):
    """
    Converte números vindos de Excel/Google Sheets sem perder milhares.

    Correções principais:
    - 1,000  -> 1000   quando a vírgula vier como separador de milhar do Sheets
    - 12,000 -> 12000
    - 1.000  -> 1000   quando o ponto vier como separador de milhar
    - 28,12  -> 28.12  quando a vírgula vier como decimal brasileiro
    - 28.12  -> 28.12  quando o ponto vier como decimal
    - 1.234,56 -> 1234.56
    - 1,234.56 -> 1234.56
    """
    if value is None:
        return 0.0

    if isinstance(value, (int, float)) and not isinstance(value, bool):
        try:
            if pd.isna(value):
                return 0.0
            return float(value)
        except Exception:
            pass

    txt = str(value).strip()
    if txt == "" or txt.lower() in ["nan", "none", "-"]:
        return 0.0

    txt = txt.replace("R$", "").replace(" ", "").replace(" ", "")

    # Remove sinais/artefatos comuns sem afetar números válidos.
    txt = txt.replace("+", "")

    # Quando tem vírgula e ponto, o último separador define o decimal.
    # BR: 1.234,56 | US/Sheets: 1,234.56
    if "," in txt and "." in txt:
        if txt.rfind(",") > txt.rfind("."):
            txt = txt.replace(".", "").replace(",", ".")
        else:
            txt = txt.replace(",", "")

    elif "," in txt:
        # Google Sheets/CSV em padrão americano: 1,000 / 12,000 / 123,456
        # Aqui a vírgula é milhar, não decimal.
        if re.fullmatch(r"-?\d{1,3}(?:,\d{3})+", txt):
            txt = txt.replace(",", "")
        else:
            # Decimal brasileiro simples: 28,12 / 0,5
            txt = txt.replace(",", ".")

    elif "." in txt:
        # Ponto como milhar: 1.000 / 12.000 / 123.456
        if re.fullmatch(r"-?\d{1,3}(?:\.\d{3})+", txt):
            txt = txt.replace(".", "")
        else:
            # Mantém ponto como decimal: 28.12 / 25.940000000000001
            pass

    try:
        return float(txt)
    except Exception:
        return 0.0


def corrigir_mojibake_texto(texto):
    texto = str(texto or "")
    if not any(marca in texto for marca in ["Ã", "Â", "â", "ðŸ"]):
        return texto
    try:
        corrigido = texto.encode("cp1252", errors="strict").decode("utf-8", errors="strict")
        if corrigido:
            return corrigido
    except Exception:
        return texto
    return texto


def remover_acentos(texto):
    texto = unicodedata.normalize("NFD", str(texto or ""))
    return "".join(ch for ch in texto if unicodedata.category(ch) != "Mn")


def normalizar_coluna(nome):
    nome = corrigir_mojibake_texto(nome)
    nome = str(nome).strip().upper().replace("\ufeff", "")
    nome = remover_acentos(nome)
    nome = re.sub(r"[^A-Z0-9]+", " ", nome)
    nome = re.sub(r"\s+", " ", nome).strip()
    return nome


def normalizar_texto_simples(value):
    texto = str(value or "").strip().lower()
    texto = unicodedata.normalize("NFD", texto)
    texto = "".join(ch for ch in texto if unicodedata.category(ch) != "Mn")
    texto = re.sub(r"\s+", " ", texto)
    return texto


def parse_data_br(value):
    if value is None:
        return pd.NaT
    value = str(value).strip()
    if not value or value.lower() in ["nan", "none", "-"]:
        return pd.NaT
    match = re.search(r"(\d{1,2}/\d{1,2}/\d{2,4})", value)
    if not match:
        return pd.NaT
    data_txt = match.group(1)
    for fmt in ["%d/%m/%Y", "%d/%m/%y"]:
        try:
            return datetime.strptime(data_txt, fmt).date()
        except Exception:
            continue
    return pd.NaT


def format_data_br(value):
    if pd.isna(value) or value is None or value == "":
        return ""
    if isinstance(value, pd.Timestamp):
        return value.strftime("%d/%m/%Y")
    if isinstance(value, (datetime, date)):
        return value.strftime("%d/%m/%Y")
    parsed = parse_data_br(value)
    if pd.isna(parsed):
        return ""
    return parsed.strftime("%d/%m/%Y")


def format_num_br(value, casas=1):
    try:
        value = round(float(value), casas)
        texto = f"{value:,.{casas}f}".replace(",", "X").replace(".", ",").replace("X", ".")
        if "," in texto:
            texto = texto.rstrip("0").rstrip(",")
        return texto
    except Exception:
        return value


def format_int_br(value):
    try:
        return f"{int(round(float(value))):,}".replace(",", ".")
    except Exception:
        return value


def format_moeda_br(value):
    try:
        value = float(value)
        return "R$ " + f"{value:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
    except Exception:
        return "R$ 0,00"


PDF_MAX_PAGINAS_PADRAO = 1000


def _pdf_bytes(uploaded_file):
    """Transforma UploadedFile/bytes em bytes e reposiciona o arquivo quando possível."""
    if uploaded_file is None:
        return b""
    if isinstance(uploaded_file, (bytes, bytearray)):
        return bytes(uploaded_file)
    try:
        return uploaded_file.getvalue()
    except Exception:
        try:
            uploaded_file.seek(0)
            return uploaded_file.read()
        except Exception:
            return b""


@st.cache_data(show_spinner=False, ttl=3600, max_entries=8)
def extract_text_from_pdf_cached(pdf_bytes, max_pages=PDF_MAX_PAGINAS_PADRAO):
    """
    Extrai texto de PDF com cache para evitar que o Streamlit releia o mesmo arquivo
    a cada interação. Prioriza PyMuPDF quando disponível, que costuma ser muito mais
    rápido que pdfplumber para PDFs grandes.
    """
    pdf_bytes = bytes(pdf_bytes or b"")
    if not pdf_bytes:
        return ""

    textos = []

    # Caminho rápido: PyMuPDF. Se não estiver instalado, cai para pdfplumber.
    if fitz is not None:
        try:
            with fitz.open(stream=pdf_bytes, filetype="pdf") as doc:
                total = min(len(doc), int(max_pages or len(doc)))
                for i in range(total):
                    page_text = doc[i].get_text("text", sort=True) or ""
                    if page_text.strip():
                        textos.append(page_text)
            return "\n".join(textos)
        except Exception:
            textos = []

    # Fallback: pdfplumber, página a página.
    with pdfplumber.open(BytesIO(pdf_bytes)) as pdf:
        total = min(len(pdf.pages), int(max_pages or len(pdf.pages)))
        for i in range(total):
            page_text = pdf.pages[i].extract_text(x_tolerance=1, y_tolerance=3) or ""
            if page_text.strip():
                textos.append(page_text)

    return "\n".join(textos)


def extract_text_from_pdf(uploaded_file, max_pages=PDF_MAX_PAGINAS_PADRAO):
    return extract_text_from_pdf_cached(_pdf_bytes(uploaded_file), max_pages=max_pages)


@st.cache_data(show_spinner=False, ttl=3600, max_entries=8)
def extract_text_from_pdf_pdfplumber_cached(pdf_bytes, max_pages=PDF_MAX_PAGINAS_PADRAO):
    """
    Extrai texto preservando melhor a ordem visual do relatório.
    Uso recomendado para o PDF de Giro de Estoque, porque o PyMuPDF pode
    reorganizar as linhas e quebrar o parser por empresa/produto.
    """
    pdf_bytes = bytes(pdf_bytes or b"")
    if not pdf_bytes:
        return ""

    textos = []
    with pdfplumber.open(BytesIO(pdf_bytes)) as pdf:
        total = min(len(pdf.pages), int(max_pages or len(pdf.pages)))
        for i in range(total):
            page = pdf.pages[i]
            page_text = page.extract_text(x_tolerance=1, y_tolerance=3) or ""
            if page_text.strip():
                textos.append(page_text)
    return "\n".join(textos)


def extract_text_from_pdf_pdfplumber(uploaded_file, max_pages=PDF_MAX_PAGINAS_PADRAO):
    return extract_text_from_pdf_pdfplumber_cached(_pdf_bytes(uploaded_file), max_pages=max_pages)


def aviso_pymupdf_ausente_para_giro(uploaded_file):
    try:
        tamanho_mb = len(_pdf_bytes(uploaded_file)) / (1024 * 1024)
    except Exception:
        tamanho_mb = 0

    if fitz is None and tamanho_mb >= 3:
        st.warning(
            "Este PDF de Giro tem muitas paginas e o PyMuPDF nao esta ativo neste ambiente. "
            "A leitura ainda funciona pelo fallback, mas pode demorar bastante. "
            "Para leitura rapida, instale/atualize as dependencias com: pip install -r requirements.txt"
        )


def diagnosticar_pdf_giro(texto):
    """Retorna uma mensagem amigável quando o arquivo no campo Giro não é o relatório esperado."""
    txt = str(texto or "")
    up = txt.upper()
    if not txt.strip():
        return (
            "O PDF enviado no campo Giro de Estoque não retornou texto. "
            "Provavelmente ele é escaneado/imagem ou está protegido. Gere/exporte o relatório em PDF textual pelo sistema."
        )
    if "ABERTO" in up and ("PEDIDO" in up or "PEDIDOS" in up or "QTDE" in up):
        return (
            "O arquivo enviado no campo Giro de Estoque parece ser o relatório de Pedidos em Aberto. "
            "Na primeira caixa envie o PDF de Giro de Estoque; na segunda caixa envie o PDF de Pedidos em Aberto."
        )
    if "GIRO" not in up and "EMPRESA" not in up:
        return (
            "O arquivo enviado no campo Giro de Estoque não parece ter o layout do relatório de Giro. "
            "Confira se você selecionou o PDF correto."
        )
    return (
        "Não consegui extrair os dados do Giro de Estoque. "
        "Confira se o PDF é o relatório de Giro de Estoque no layout padrão, com EMPRESA, código do item, meses, estoque e preço."
    )


def aviso_pdf_grande(uploaded_file, limite_mb=25):
    try:
        tamanho_mb = len(_pdf_bytes(uploaded_file)) / (1024 * 1024)
        if tamanho_mb >= limite_mb:
            st.warning(
                f"PDF com {tamanho_mb:.1f} MB. A leitura pode demorar. "
                "O app agora usa cache e leitura otimizada para evitar travamentos."
            )
    except Exception:
        pass


@st.cache_data(show_spinner=False, ttl=3600, max_entries=8)
def extract_pdf_linhas_e_tabelas_cached(pdf_bytes, max_pages=PDF_MAX_PAGINAS_PADRAO):
    """Extrai linhas de texto e linhas de tabelas com cache para PDFs de fornecedor."""
    pdf_bytes = bytes(pdf_bytes or b"")
    linhas_texto = []
    linhas_tabela = []
    if not pdf_bytes:
        return linhas_texto, linhas_tabela

    with pdfplumber.open(BytesIO(pdf_bytes)) as pdf:
        total = min(len(pdf.pages), int(max_pages or len(pdf.pages)))
        for i in range(total):
            page = pdf.pages[i]
            for tabela in (page.extract_tables() or []):
                for linha in tabela:
                    celulas = [str(c or "").strip() for c in (linha or [])]
                    if any(celulas):
                        linhas_tabela.append(celulas)
                        linhas_texto.append(" | ".join(celulas))

            page_text = page.extract_text(x_tolerance=1, y_tolerance=3) or ""
            for linha in page_text.splitlines():
                linha = linha.strip()
                if linha:
                    linhas_texto.append(linha)

    return linhas_texto, linhas_tabela


def extract_pdf_linhas_e_tabelas(uploaded_file, max_pages=PDF_MAX_PAGINAS_PADRAO):
    return extract_pdf_linhas_e_tabelas_cached(_pdf_bytes(uploaded_file), max_pages=max_pages)


MESES_ABREV_PT = {
    "01": "Jan", "02": "Fev", "03": "Mar", "04": "Abr", "05": "Mai", "06": "Jun",
    "07": "Jul", "08": "Ago", "09": "Set", "10": "Out", "11": "Nov", "12": "Dez",
}


def extrair_meses_giro_pdf(text):
    """
    Puxa os meses diretamente do PDF de Giro.
    Prioriza a ordem real das colunas da tabela, porque alguns relatórios exibem
    "REFERENTE AOS MESES" em ordem decrescente, mas as colunas de giro aparecem em
    ordem crescente no cabeçalho.
    """
    text = str(text or "")

    # Ordem usada para mapear os valores de cada linha do produto.
    # Ex.: o PDF pode dizer "REFERENTE AOS MESES: 06, 05, 04, 03", mas a tabela
    # estar visualmente como "03/2026 04/2026 05/2026 06/2026".
    for raw_line in text.splitlines():
        line = raw_line.strip()
        line_upper = line.upper()
        if "COD" not in line_upper or "MEDIA" not in line_upper:
            continue

        meses_linha = re.findall(r"\b\d{2}/\d{4}\b", line)
        if len(meses_linha) >= 2:
            return meses_linha

    for raw_line in text.splitlines():
        line = raw_line.strip()
        meses_linha = re.findall(r"\b\d{2}/\d{4}\b", line)
        if len(meses_linha) >= 2 and ("ESTOQUE" in line.upper() or "MEDIA" in line.upper() or "MÉDIA" in line.upper()):
            return meses_linha

    padrao_ref = re.search(r"REFERENTE\s+AOS\s+MESES\s*:\s*([^\n]+)", text, flags=re.IGNORECASE)
    if padrao_ref:
        meses = re.findall(r"\b\d{2}/\d{4}\b", padrao_ref.group(1))
        if meses:
            return sorted(meses, key=lambda mes: (int(mes.split("/")[1]), int(mes.split("/")[0])))

    meses = []
    for mes in re.findall(r"\b\d{2}/\d{4}\b", text):
        if mes not in meses:
            meses.append(mes)
    return meses[:4] if meses else MESES_PADRAO.copy()


def label_mes_giro(mes):
    try:
        mm, yyyy = str(mes).split("/")
        return f"{MESES_ABREV_PT.get(mm.zfill(2), mm)}/{str(yyyy)[-2:]}"
    except Exception:
        return str(mes)


def col_giro(prefixo, mes):
    return f"{prefixo} {label_mes_giro(mes)}"


def mes_atual_referencia():
    hoje = date.today()
    return f"{hoje.month:02d}/{hoje.year}"


def colunas_pedido_compras(meses_ref=None):
    """
    Ordem oficial da tela Pedido de Compra e do download Pedido Editável.
    Mantém Código Fábrica e Embalagem no final, conforme solicitado.
    """
    meses_ref = meses_ref or MESES
    return [
        "codigo",
        "descricao",
        *[col_giro("Giro Geral", mes) for mes in meses_ref],
        "Média Giro Geral",
        "Estoque Lojas",
        "Estoque Única",
        "Estoque Geral",
        "Saldo em Trânsito/ABERTO",
        "Estoque Final",
        "Estoque Alvo",
        "Sugestão Sistema",
        "Sugestão arredondada",
        "Preço Última Compra",
        "Data Última Compra",
        "PEDIDO Final",
        "Origem Sugestão",
        "Valor Final do Pedido",
        "Embalagem",
        "Código Fábrica",
    ]

# =========================================================
# LEITURA DO PDF DE GIRO DE ESTOQUE
# =========================================================

def _token_numero_giro(valor):
    """
    Identifica números válidos nas colunas numéricas do relatório de giro.
    Evita tratar pedaços da descrição/código de fábrica como giro.
    """
    txt = str(valor or "").strip()
    if txt == "":
        return False
    return bool(re.fullmatch(r"-?\d+(?:\.\d{3})*,\d{1,4}|-?\d+,\d{1,4}|-?\d+", txt))


def _encontrar_unidade_giro(partes, qtd_meses):
    """
    No PDF de giro, algumas descrições contêm a palavra/tipo 'UN' antes da unidade real.
    Exemplo real:
    85582 ... 0.5L UN 1263 T UN 0,00 0,00 0,00 ...
    A unidade correta é o último UN antes da sequência dos meses.

    Esta função procura a unidade que vem imediatamente antes de uma sequência numérica
    com a quantidade de meses do cabeçalho. Assim o sistema não puxa '1263' da descrição
    como se fosse giro de abril.
    """
    candidatos = []

    for i, token in enumerate(partes):
        token_upper = str(token).strip().upper()
        unidade_explicita = token_upper in ["UN", "UND", "UNID", "UNIDADE"]

        # Alguns PDFs colam a unidade no fim do texto, ex.: "...-STUN 2,45 1,00..."
        unidade_colada = (
            not unidade_explicita
            and len(token_upper) > 2
            and token_upper.endswith("UN")
            and not _token_numero_giro(token_upper)
        )

        if not unidade_explicita and not unidade_colada:
            continue

        proximos = partes[i + 1:i + 1 + qtd_meses]
        if len(proximos) < qtd_meses:
            continue

        if all(_token_numero_giro(v) for v in proximos):
            candidatos.append((i, unidade_colada))

    if not candidatos:
        return None, False

    # Usa o último candidato válido, porque a descrição pode conter "UN" antes da unidade real.
    return candidatos[-1]


def parse_linha_giro(line, meses_ref=None):
    """
    Layout esperado:
    COD DESCRICAO DO ITEM UN 04/2026 05/2026 06/2026 MEDIA PREVI.30 ESTOQUE
    SUGESTAO PR.ULT.COMP DT.ULT.COMP PR.VENDA % LUCRO

    Correção importante:
    - Não usa mais o primeiro token "UN" encontrado.
    - Localiza a unidade pela sequência numérica dos meses logo depois dela.
    - Isso evita erro em itens cuja descrição contém "UN" ou códigos numéricos antes
      da unidade real, como o item 85582.
    """
    if not re.match(r"^\d{5}\s+", str(line).strip()):
        return None

    partes = str(line).strip().split()
    codigo = partes[0].zfill(5)

    meses_ref = meses_ref or MESES_PADRAO
    qtd_meses = len(meses_ref)

    un_index, unidade_colada = _encontrar_unidade_giro(partes, qtd_meses)
    if un_index is None:
        return None

    antes_un = partes[1:un_index]
    depois_un = partes[un_index + 1:]

    # Se a unidade veio colada no final do último token da descrição, remove só o "UN".
    if unidade_colada:
        token_sem_un = str(partes[un_index])[:-2].strip()
        if token_sem_un:
            antes_un = partes[1:un_index] + [token_sem_un]

    if len(depois_un) < qtd_meses + 4:
        return None

    # Garantia adicional: os meses precisam ser exatamente a primeira sequência após a unidade.
    if not all(_token_numero_giro(v) for v in depois_un[:qtd_meses]):
        return None

    codigo_fabrica_extraido = ""
    descricao_tokens = list(antes_un)

    if descricao_tokens:
        ultimo_token = descricao_tokens[-1]
        if re.fullmatch(r"\d{5,}", ultimo_token):
            codigo_fabrica_extraido = ultimo_token[-6:] if len(ultimo_token) > 7 else ultimo_token
            descricao_tokens = descricao_tokens[:-1]
        else:
            match_fabrica = re.search(r"(\d{5,})$", ultimo_token)
            if match_fabrica:
                codigo_raw = match_fabrica.group(1)
                codigo_fabrica_extraido = codigo_raw[-6:] if len(codigo_raw) > 7 else codigo_raw
                descricao_tokens[-1] = ultimo_token[:match_fabrica.start(1)].rstrip("- ")

    data_idx = None
    for i, token in enumerate(depois_un):
        if re.fullmatch(r"\d{1,2}/\d{1,2}/\d{2,4}", str(token).strip()):
            data_idx = i
            break

    dt_ult_compra = pd.NaT
    dt_ult_compra_txt = ""
    pr_ult_compra = 0.0

    if data_idx is not None:
        dt_ult_compra_txt = depois_un[data_idx]
        dt_ult_compra = parse_data_br(dt_ult_compra_txt)
        if data_idx - 1 >= 0:
            pr_ult_compra = br_to_float(depois_un[data_idx - 1])
    else:
        # Após os meses: MEDIA, PREVI.30, ESTOQUE, SUGESTAO, PR.ULT.COMP...
        idx_preco = qtd_meses + 4
        pr_ult_compra = br_to_float(depois_un[idx_preco]) if len(depois_un) > idx_preco else 0.0

    idx_estoque = qtd_meses + 2

    return {
        "codigo": codigo,
        "descricao": " ".join(descricao_tokens).strip(),
        "codigo_fabrica": codigo_fabrica_extraido,
        **{mes: br_to_float(depois_un[i]) for i, mes in enumerate(meses_ref)},
        "estoque": br_to_float(depois_un[idx_estoque]) if len(depois_un) > idx_estoque else 0,
        "pr_ult_compra": pr_ult_compra,
        "dt_ult_compra": dt_ult_compra,
        "dt_ult_compra_txt": dt_ult_compra_txt,
        "codigo_empresa": None,
        "loja": None,
    }


def parse_giro_estoque(text, meses_ref=None):
    meses_ref = meses_ref or extrair_meses_giro_pdf(text)
    registros = []
    empresa_atual = None

    for raw_line in text.splitlines():
        line = raw_line.strip()

        empresa_match = re.search(r"EMPRESA\s*:\s*(\d{3})\s*-", line)
        if empresa_match:
            empresa_atual = empresa_match.group(1)
            continue

        if not empresa_atual or empresa_atual not in LOJAS_MAP:
            continue

        produto = parse_linha_giro(line, meses_ref=meses_ref)
        if produto:
            produto["codigo_empresa"] = empresa_atual
            produto["loja"] = LOJAS_MAP[empresa_atual]
            registros.append(produto)

    return pd.DataFrame(registros)


@st.cache_data(show_spinner=False, ttl=3600, max_entries=4)
def parse_giro_estoque_pdf_cached(pdf_bytes, max_pages=PDF_MAX_PAGINAS_PADRAO):
    """
    Le o PDF de Giro ja retornando o DataFrame.

    O relatorio pode ter centenas de paginas. PyMuPDF costuma ler esse PDF em
    segundos; pdfplumber fica como fallback para ambientes sem a dependencia rapida.
    """
    pdf_bytes = bytes(pdf_bytes or b"")
    if not pdf_bytes:
        return pd.DataFrame(), MESES_PADRAO.copy(), "", "vazio"

    textos = []

    if fitz is not None:
        try:
            with fitz.open(stream=pdf_bytes, filetype="pdf") as doc:
                total = min(len(doc), int(max_pages or len(doc)))
                for i in range(total):
                    page_text = doc[i].get_text("text", sort=True) or ""
                    if page_text.strip():
                        textos.append(page_text)

            texto = "\n".join(textos)
            meses_ref = extrair_meses_giro_pdf(texto)
            df = parse_giro_estoque(texto, meses_ref)
            if df is not None and not df.empty:
                return df, meses_ref, texto[:4000], "pymupdf"
        except Exception:
            textos = []

    with pdfplumber.open(BytesIO(pdf_bytes)) as pdf:
        total = min(len(pdf.pages), int(max_pages or len(pdf.pages)))
        for i in range(total):
            page_text = pdf.pages[i].extract_text(x_tolerance=1, y_tolerance=3) or ""
            if page_text.strip():
                textos.append(page_text)

    texto = "\n".join(textos)
    meses_ref = extrair_meses_giro_pdf(texto)
    df = parse_giro_estoque(texto, meses_ref)
    return df, meses_ref, texto[:4000], "pdfplumber"


def parse_giro_estoque_pdf(uploaded_file, max_pages=PDF_MAX_PAGINAS_PADRAO):
    return parse_giro_estoque_pdf_cached(_pdf_bytes(uploaded_file), max_pages=max_pages)

# =========================================================
# LEITURA DO PDF DE PEDIDOS EM ABERTO / SALDO EM TRÂNSITO
# =========================================================

_NUMERO_BR_RE = re.compile(r"^-?\d{1,3}(?:\.\d{3})*,\d+$|^-?\d+,\d+$|^-?\d+$")


def _eh_numero_br(valor):
    return bool(_NUMERO_BR_RE.match(str(valor).strip()))


def extrair_descricao_pedido_aberto_tokens(tokens, un_index):
    if not tokens or un_index is None or un_index <= 0:
        return ""

    descricao_tokens = []
    primeiro = str(tokens[0]).strip()
    match_primeiro = re.match(r"^\d{5}[-\s]*(.*)$", primeiro)
    if match_primeiro and match_primeiro.group(1).strip():
        descricao_tokens.append(match_primeiro.group(1).strip(" -"))

    descricao_tokens.extend(str(t).strip() for t in tokens[1:un_index] if str(t).strip())
    return " ".join(descricao_tokens).strip()


def encontrar_indice_aberto_no_cabecalho(text):
    """
    No relatório de Pedidos de Compra, após a unidade UN, a sequência numérica é:
    QTDE, TOT.LIT, TOT.KIL, PES.ITE, BAIXADO, ABERTO, VR.UNIT, TOT.IPI, ALQ.IPI, TOT.SUB, TOTAL.

    Portanto, ABERTO é sempre o 6º número depois do UN, índice 5.
    """
    for raw_line in str(text or "").splitlines():
        line = str(raw_line or "").strip().upper()
        if " UN " not in f" {line} " or "ABERTO" not in line or "BAIXADO" not in line:
            continue

        tokens = line.split()
        try:
            un_index = tokens.index("UN")
        except ValueError:
            continue

        campos = []
        i = un_index + 1
        while i < len(tokens):
            token = tokens[i].strip()
            if token == "PES." and i + 1 < len(tokens) and tokens[i + 1].strip() == "ITE":
                campos.append("PES.ITE")
                i += 2
                continue
            campos.append(token)
            i += 1

        if "ABERTO" in campos:
            return campos.index("ABERTO")

    return 5


def parse_linha_pedido_aberto(line, indice_aberto=None):
    """
    Lê uma linha do PDF de pedidos em aberto.

    Regra corrigida:
    - Não usa a posição do cabeçalho inteiro, porque a descrição varia.
    - Localiza a unidade UN.
    - Depois da UN, considera apenas números.
    - Puxa a coluna ABERTO pelo índice fixo 5:
      QTDE=0, TOT.LIT=1, TOT.KIL=2, PES.ITE=3, BAIXADO=4, ABERTO=5, VR.UNIT=6.
    """
    line = line.strip()
    match = re.match(r"^(\d{5})[-\s]", line)
    if not match:
        return None

    codigo = match.group(1).zfill(5)
    partes = line.split()

    un_index = None
    for i, token in enumerate(partes):
        if token.upper() in ["UN", "UND", "UNID", "UNIDADE"]:
            un_index = i
            break

    if un_index is None:
        return None

    descricao = extrair_descricao_pedido_aberto_tokens(partes, un_index)
    valores_numericos = [p for p in partes[un_index + 1:] if _eh_numero_br(p)]
    idx_aberto = 5 if indice_aberto is None else int(indice_aberto)

    if len(valores_numericos) > idx_aberto:
        return {"codigo": codigo, "descricao": descricao, "Saldo em Trânsito/ABERTO": br_to_float(valores_numericos[idx_aberto])}

    return {"codigo": codigo, "descricao": descricao, "Saldo em Trânsito/ABERTO": 0.0}


def parse_pedidos_compra_aberto(text):
    registros = []
    indice_aberto = encontrar_indice_aberto_no_cabecalho(text)

    for raw_line in text.splitlines():
        produto = parse_linha_pedido_aberto(raw_line, indice_aberto=indice_aberto)
        if produto:
            registros.append(produto)

    if not registros:
        return pd.DataFrame(columns=["codigo", "descricao", "Saldo em Trânsito/ABERTO"])

    return pd.DataFrame(registros).groupby("codigo", as_index=False).agg({
        "descricao": "first",
        "Saldo em Trânsito/ABERTO": "sum",
    })


def _parse_pedidos_compra_aberto_pdf_stream(uploaded_file):
    """
    Lê o PDF de Pedidos em Aberto / Saldo em Trânsito.

    Correção principal:
    - Para este relatório, a coluna ABERTO é sempre o 6º número após a unidade UN.
      Sequência depois do UN:
      QTDE=0, TOT.LIT=1, TOT.KIL=2, PES.ITE=3, BAIXADO=4, ABERTO=5, VR.UNIT=6.
    - Por isso, a leitura por texto é usada primeiro. Ela é mais segura do que coordenadas,
      porque algumas linhas do PDF vêm com campos colados, ex.: ST000001-WANDA.
    - A leitura por coordenadas fica apenas como fallback.
    """

    # 1) Caminho principal: texto linha a linha, usando a posição fixa do ABERTO após UN.
    try:
        uploaded_file.seek(0)
    except Exception:
        pass

    try:
        texto = extract_text_from_pdf_pdfplumber(uploaded_file)
        df_texto = parse_pedidos_compra_aberto(texto)
        if df_texto is not None and not df_texto.empty:
            df_texto["codigo"] = df_texto["codigo"].astype(str).str.extract(r"(\d+)", expand=False).fillna("").str.zfill(5)
            df_texto["Saldo em Trânsito/ABERTO"] = pd.to_numeric(df_texto["Saldo em Trânsito/ABERTO"], errors="coerce").fillna(0)
            df_texto = df_texto.groupby("codigo", as_index=False).agg({
                "descricao": "first",
                "Saldo em Trânsito/ABERTO": "sum",
            })
            if float(df_texto["Saldo em Trânsito/ABERTO"].sum()) > 0:
                return df_texto
    except Exception:
        pass

    # 2) Fallback: tenta por coordenadas caso o texto não tenha retornado itens.
    registros = []

    try:
        uploaded_file.seek(0)
    except Exception:
        pass

    try:
        with pdfplumber.open(uploaded_file) as pdf:
            aberto_x = None

            for page in pdf.pages:
                words = page.extract_words(x_tolerance=1, y_tolerance=3, keep_blank_chars=False) or []
                if not words:
                    continue

                linhas = {}
                for w in words:
                    top_key = round(float(w.get("top", 0)) / 3) * 3
                    linhas.setdefault(top_key, []).append(w)

                for _, linha_words in sorted(linhas.items()):
                    linha_words = sorted(linha_words, key=lambda w: float(w.get("x0", 0)))
                    textos = [str(w.get("text", "")).strip() for w in linha_words]
                    textos_upper = [t.upper() for t in textos]

                    if (
                        "QTDE" in textos_upper
                        and "BAIXADO" in textos_upper
                        and "ABERTO" in textos_upper
                        and "VR.UNIT" in textos_upper
                    ):
                        idx = textos_upper.index("ABERTO")
                        w = linha_words[idx]
                        aberto_x = (float(w["x0"]) + float(w["x1"])) / 2
                        continue

                    if aberto_x is None or not textos:
                        continue

                    match_codigo = re.match(r"^(\d{5})(?:[-\s]|$)", textos[0])
                    if not match_codigo:
                        continue

                    codigo = match_codigo.group(1).zfill(5)

                    un_index = None
                    for i, token in enumerate(textos):
                        if token.upper() in ["UN", "UND", "UNID", "UNIDADE"]:
                            un_index = i
                            break

                    descricao = extrair_descricao_pedido_aberto_tokens(textos, un_index)

                    # Mesmo no fallback por coordenadas, se houver UN, prioriza a sequência numérica.
                    if un_index is not None:
                        valores_numericos = [p for p in textos[un_index + 1:] if _eh_numero_br(p)]
                        if len(valores_numericos) > 5:
                            registros.append({
                                "codigo": codigo,
                                "descricao": descricao,
                                "Saldo em Trânsito/ABERTO": br_to_float(valores_numericos[5]),
                            })
                            continue

                    candidatos = []
                    for w in linha_words[1:]:
                        txt = str(w.get("text", "")).strip()
                        if not _eh_numero_br(txt):
                            continue
                        cx = (float(w["x0"]) + float(w["x1"])) / 2
                        distancia = abs(cx - aberto_x)
                        candidatos.append((distancia, txt))

                    candidatos = [c for c in candidatos if c[0] <= 22]
                    if candidatos:
                        candidatos.sort(key=lambda x: x[0])
                        registros.append({
                            "codigo": codigo,
                            "descricao": descricao,
                            "Saldo em Trânsito/ABERTO": br_to_float(candidatos[0][1]),
                        })

    except Exception:
        registros = []

    if registros:
        return pd.DataFrame(registros).groupby("codigo", as_index=False).agg({
            "descricao": "first",
            "Saldo em Trânsito/ABERTO": "sum",
        })

    return pd.DataFrame(columns=["codigo", "descricao", "Saldo em Trânsito/ABERTO"])


@st.cache_data(show_spinner=False, ttl=3600, max_entries=8)
def parse_pedidos_compra_aberto_pdf_cached(pdf_bytes):
    pdf_bytes = bytes(pdf_bytes or b"")
    if not pdf_bytes:
        return pd.DataFrame(columns=["codigo", "descricao", "Saldo em Trânsito/ABERTO"])
    return _parse_pedidos_compra_aberto_pdf_stream(BytesIO(pdf_bytes))


def parse_pedidos_compra_aberto_pdf(uploaded_file):
    return parse_pedidos_compra_aberto_pdf_cached(_pdf_bytes(uploaded_file))


# =========================================================
# LEITURA DO CSV DE CADASTRO DE PRODUTOS
# =========================================================

@st.cache_data(show_spinner="Lendo cadastro CSV...")
def ler_cadastro_produtos_csv(uploaded_file):
    if uploaded_file is None:
        return pd.DataFrame()

    def mapear_colunas(colunas):
        colunas_norm = {normalizar_coluna(c): c for c in colunas}

        candidatos_codigo = [
            "CÓDIGO", "CODIGO", "CÓD.ITEM", "COD.ITEM", "CÓD ITEM", "COD ITEM",
            "CÓDIGO ITEM", "CODIGO ITEM",
        ]
        candidatos_descricao = [
            "DESCRIÇÃO DO ITEM", "DESCRICAO DO ITEM", "DESCRIÇÃO", "DESCRICAO",
            "DESC ITEM", "DESCRIÇÃO ITEM", "DESCRICAO ITEM",
        ]
        candidatos_fabrica = [
            "CÓD. FÁBRICA", "COD. FÁBRICA", "CÓD. FABRICA", "COD. FABRICA",
            "CÓDIGO DE FÁBRICA", "CODIGO DE FABRICA", "NOVO CÓDIGO DE FÁBRICA",
            "NOVO CODIGO DE FABRICA", "COD FABRICA", "CÓD FABRICA",
            "CÓDIGO FÁBRICA", "CODIGO FABRICA", "CODIGO_FABRICA", "codigo_fabrica",
        ]

        def encontrar(candidatos):
            for candidato in candidatos:
                candidato_norm = normalizar_coluna(candidato)
                if candidato_norm in colunas_norm:
                    return colunas_norm[candidato_norm]
            return None

        candidatos_embalagem = [
            "EMBALAGEM", "EMB", "QTD EMBALAGEM", "QUANTIDADE EMBALAGEM",
            "QTDE EMBALAGEM", "QTD. EMBALAGEM", "MULTIPLO", "MÚLTIPLO",
        ]

        return {
            "codigo": encontrar(candidatos_codigo),
            "descricao": encontrar(candidatos_descricao),
            "codigo_fabrica": encontrar(candidatos_fabrica),
            "embalagem": encontrar(candidatos_embalagem),
        }

    tentativas = [
        {"sep": ";", "encoding": "utf-8-sig"},
        {"sep": ";", "encoding": "latin1"},
        {"sep": ",", "encoding": "utf-8-sig"},
        {"sep": ",", "encoding": "latin1"},
        {"sep": "\t", "encoding": "utf-8-sig"},
        {"sep": "\t", "encoding": "latin1"},
    ]

    df = None
    colunas_mapeadas = None
    ultimo_erro = None

    for tentativa in tentativas:
        try:
            uploaded_file.seek(0)
            temp = pd.read_csv(
                uploaded_file,
                sep=tentativa["sep"],
                encoding=tentativa["encoding"],
                dtype=str,
                engine="python",
                on_bad_lines="skip",
            )
            temp.columns = [str(c).strip() for c in temp.columns]
            mapeadas = mapear_colunas(temp.columns)
            if (
                mapeadas["codigo"]
                and mapeadas["descricao"]
                and mapeadas["codigo_fabrica"]
            ):
                df = temp
                colunas_mapeadas = mapeadas
                break
        except Exception as e:
            ultimo_erro = str(e)
            continue

    if df is None or colunas_mapeadas is None:
        st.error("Não consegui ler o CSV de cadastro.")
        st.caption(
            "O CSV pode ter um destes padrões de colunas: CÓDIGO, DESCRIÇÃO DO ITEM, CÓD. FABRICA "
            "ou Cód.Item, Descrição, Novo Código de fábrica."
        )
        if ultimo_erro:
            st.caption(f"Último erro identificado: {ultimo_erro}")
        return pd.DataFrame()

    colunas_selecionadas = [
        colunas_mapeadas["codigo"],
        colunas_mapeadas["descricao"],
        colunas_mapeadas["codigo_fabrica"],
    ]
    novos_nomes = ["codigo", "descricao_cadastro", "codigo_fabrica_cadastro"]

    if colunas_mapeadas.get("embalagem"):
        colunas_selecionadas.append(colunas_mapeadas["embalagem"])
        novos_nomes.append("embalagem")

    cadastro = df[colunas_selecionadas].copy()
    cadastro.columns = novos_nomes

    cadastro["codigo"] = cadastro["codigo"].astype(str).str.extract(r"(\d+)")[0].str.zfill(5)
    cadastro["descricao_cadastro"] = cadastro["descricao_cadastro"].astype(str).str.strip()
    cadastro["codigo_fabrica_cadastro"] = cadastro["codigo_fabrica_cadastro"].astype(str).str.strip()

    if "embalagem" in cadastro.columns:
        cadastro["embalagem"] = cadastro["embalagem"].apply(br_to_float)
        cadastro["embalagem"] = pd.to_numeric(cadastro["embalagem"], errors="coerce").fillna(0).round(0).astype(int)
    else:
        cadastro["embalagem"] = 0

    cadastro = cadastro.dropna(subset=["codigo"])
    cadastro = cadastro[cadastro["codigo"].str.lower() != "nan"]
    cadastro = cadastro.drop_duplicates(subset=["codigo"], keep="first")
    return cadastro


def normalizar_cadastro_produtos_df(df):
    if df is None or df.empty:
        return pd.DataFrame()

    df = df.copy()
    df.columns = [str(c).strip() for c in df.columns]

    def mapear_colunas_existentes(colunas):
        colunas_norm_local = {normalizar_coluna(c): c for c in colunas}

        def encontrar_local(candidatos):
            for candidato in candidatos:
                candidato_norm = normalizar_coluna(candidato)
                if candidato_norm in colunas_norm_local:
                    return colunas_norm_local[candidato_norm]
            return None

        col_codigo_local = encontrar_local([
            "CODIGO", "CÓDIGO", "COD.ITEM", "CÓD.ITEM", "COD ITEM", "CÓD ITEM",
            "CODIGO ITEM", "CÓDIGO ITEM", "codigo",
        ])
        col_descricao_local = encontrar_local([
            "DESCRICAO DO ITEM", "DESCRIÇÃO DO ITEM", "DESCRICAO", "DESCRIÇÃO",
            "DESC ITEM", "DESCRICAO ITEM", "DESCRIÇÃO ITEM", "descricao",
        ])
        col_fabrica_local = encontrar_local([
            "COD. FÁBRICA", "CÓD. FÁBRICA", "COD. FABRICA", "CÓD. FABRICA",
            "CODIGO DE FABRICA", "CÓDIGO DE FÁBRICA", "NOVO CODIGO DE FABRICA",
            "NOVO CÓDIGO DE FÁBRICA", "COD FABRICA", "CÓD FABRICA",
            "CODIGO FABRICA", "CÓDIGO FÁBRICA", "CODIGO_FABRICA", "codigo_fabrica",
        ])
        col_embalagem_local = encontrar_local([
            "EMBALAGEM", "EMB", "QTD EMBALAGEM", "QUANTIDADE EMBALAGEM",
            "QTDE EMBALAGEM", "QTD. EMBALAGEM", "MULTIPLO", "MÚLTIPLO", "embalagem",
        ])
        return col_codigo_local, col_descricao_local, col_fabrica_local, col_embalagem_local

    col_codigo, col_descricao, col_fabrica, col_embalagem = mapear_colunas_existentes(df.columns)

    if not col_codigo or not col_descricao or not col_fabrica:
        for idx in range(min(8, len(df))):
            possivel_cabecalho = [str(v).strip() for v in df.iloc[idx].tolist()]
            teste_codigo, teste_descricao, teste_fabrica, teste_embalagem = mapear_colunas_existentes(possivel_cabecalho)
            if teste_codigo and teste_descricao and teste_fabrica:
                df = df.iloc[idx + 1:].copy()
                df.columns = possivel_cabecalho
                df = df.reset_index(drop=True)
                col_codigo, col_descricao, col_fabrica, col_embalagem = mapear_colunas_existentes(df.columns)
                break

    colunas_norm = {normalizar_coluna(c): c for c in df.columns}

    def encontrar(candidatos):
        for candidato in candidatos:
            candidato_norm = normalizar_coluna(candidato)
            if candidato_norm in colunas_norm:
                return colunas_norm[candidato_norm]
        return None

    if not col_codigo:
        col_codigo = encontrar([
            "CODIGO", "CÓDIGO", "COD.ITEM", "CÓD.ITEM", "COD ITEM", "CÓD ITEM",
            "CODIGO ITEM", "CÓDIGO ITEM", "codigo",
        ])
    if not col_descricao:
        col_descricao = encontrar([
            "DESCRICAO DO ITEM", "DESCRIÇÃO DO ITEM", "DESCRICAO", "DESCRIÇÃO",
            "DESC ITEM", "DESCRICAO ITEM", "DESCRIÇÃO ITEM", "descricao",
        ])
    if not col_fabrica:
        col_fabrica = encontrar([
            "COD. FÁBRICA", "CÓD. FÁBRICA", "COD. FABRICA", "CÓD. FABRICA",
            "CODIGO DE FABRICA", "CÓDIGO DE FÁBRICA", "NOVO CODIGO DE FABRICA",
            "NOVO CÓDIGO DE FÁBRICA", "COD FABRICA", "CÓD FABRICA",
            "CODIGO FABRICA", "CÓDIGO FÁBRICA", "CODIGO_FABRICA", "codigo_fabrica",
        ])
    if not col_embalagem:
        col_embalagem = encontrar([
            "EMBALAGEM", "EMB", "QTD EMBALAGEM", "QUANTIDADE EMBALAGEM",
            "QTDE EMBALAGEM", "QTD. EMBALAGEM", "MULTIPLO", "MÚLTIPLO", "embalagem",
        ])

    if not col_codigo or not col_descricao or not col_fabrica:
        return pd.DataFrame()

    colunas = [col_codigo, col_descricao, col_fabrica]
    nomes = ["codigo", "descricao_cadastro", "codigo_fabrica_cadastro"]
    if col_embalagem:
        colunas.append(col_embalagem)
        nomes.append("embalagem")

    cadastro = df[colunas].copy()
    cadastro.columns = nomes
    cadastro["codigo"] = cadastro["codigo"].astype(str).str.extract(r"(\d+)")[0].str.zfill(5)
    cadastro["descricao_cadastro"] = cadastro["descricao_cadastro"].astype(str).str.strip()
    cadastro["codigo_fabrica_cadastro"] = cadastro["codigo_fabrica_cadastro"].astype(str).str.strip()

    if "embalagem" in cadastro.columns:
        cadastro["embalagem"] = cadastro["embalagem"].apply(br_to_float)
        cadastro["embalagem"] = pd.to_numeric(cadastro["embalagem"], errors="coerce").fillna(0).round(0).astype(int)
    else:
        cadastro["embalagem"] = 0

    cadastro = cadastro.dropna(subset=["codigo"])
    cadastro = cadastro[cadastro["codigo"].astype(str).str.lower() != "nan"]
    cadastro = cadastro.drop_duplicates(subset=["codigo"], keep="first")
    return cadastro


@st.cache_data(show_spinner="Lendo cadastro no Google Sheets...", ttl=60)
def ler_cadastro_produtos_google_public_cached(spreadsheet_id, sheet_name="Cadastro"):
    """
    Leitura simples do cadastro no Google Sheets, sem OAuth.

    Para funcionar, a planilha precisa estar compartilhada como:
    Qualquer pessoa com o link -> Leitor.

    Essa leitura é usada apenas para buscar o cadastro. Ela não cria, não edita
    e não depende de client_id, client_secret nem refresh_token.
    """
    sheet_name = str(sheet_name or "Cadastro").strip() or "Cadastro"
    urls = [
        (
            f"https://docs.google.com/spreadsheets/d/{spreadsheet_id}/gviz/tq"
            f"?tqx=out:csv&sheet={urllib.parse.quote(sheet_name)}"
        ),
        (
            f"https://docs.google.com/spreadsheets/d/{spreadsheet_id}/export"
            f"?format=csv&sheet={urllib.parse.quote(sheet_name)}"
        ),
    ]

    try:
        gid = buscar_gid_aba_google_sheets(spreadsheet_id, sheet_name)
        if gid:
            urls.insert(0, f"https://docs.google.com/spreadsheets/d/{spreadsheet_id}/export?format=csv&gid={gid}")
    except Exception:
        pass

    ultimo_erro = None
    for url in urls:
        try:
            req = urllib.request.Request(url, headers={"User-Agent": "Mozilla/5.0"})
            with urllib.request.urlopen(req, timeout=30) as resp:
                conteudo = resp.read()
            df_raw = csv_bytes_para_dataframe_raw(conteudo)
            cadastro = normalizar_cadastro_produtos_df(df_raw)
            if not cadastro.empty:
                return cadastro
            ultimo_erro = "A planilha foi lida, mas não encontrei as colunas obrigatórias do cadastro."
        except Exception as e:
            ultimo_erro = str(e)
            continue

    raise ValueError(
        "Não consegui ler a aba Cadastro do Google Sheets. "
        "Confira se a planilha está pública como 'Qualquer pessoa com o link - Leitor' "
        f"e se existe uma aba chamada Cadastro. Detalhe: {ultimo_erro}"
    )


def ler_cadastro_produtos_google():
    """
    Busca o cadastro diretamente do Google Sheets por URL pública.
    Não usa OAuth para evitar erro invalid_client no cadastro.
    """
    return ler_cadastro_produtos_google_public_cached(GOOGLE_PLANILHA_CADASTRO_ID, "Cadastro")


def aplicar_cadastro_dataframe(df_giro, cadastro):
    if cadastro is None or cadastro.empty:
        return df_giro

    df = df_giro.merge(cadastro, on="codigo", how="left")
    df["descricao"] = df["descricao_cadastro"].where(
        df["descricao_cadastro"].notna()
        & (df["descricao_cadastro"].astype(str).str.strip() != "")
        & (df["descricao_cadastro"].astype(str).str.lower() != "nan"),
        df["descricao"],
    )
    df["codigo_fabrica"] = df["codigo_fabrica_cadastro"].where(
        df["codigo_fabrica_cadastro"].notna()
        & (df["codigo_fabrica_cadastro"].astype(str).str.strip() != "")
        & (df["codigo_fabrica_cadastro"].astype(str).str.lower() != "nan"),
        df["codigo_fabrica"],
    )
    if "embalagem" in df.columns:
        df["embalagem"] = pd.to_numeric(df["embalagem"], errors="coerce").fillna(0).round(0).astype(int)
    return df.drop(columns=["descricao_cadastro", "codigo_fabrica_cadastro"], errors="ignore")


def aplicar_cadastro(df_giro, cadastro_csv):
    cadastro = ler_cadastro_produtos_csv(cadastro_csv)
    if cadastro.empty:
        return df_giro

    return aplicar_cadastro_dataframe(df_giro, cadastro)

# =========================================================
# ÚLTIMA COMPRA / PREÇO
# =========================================================

def data_alerta_icon(data_ultima_compra, meses_alerta):
    if pd.isna(data_ultima_compra):
        return ""
    hoje = pd.Timestamp.today().normalize()
    limite = hoje - pd.DateOffset(months=int(meses_alerta))
    return "⚠️" if pd.Timestamp(data_ultima_compra).normalize() <= limite else ""


def montar_info_compra(df_giro, meses_alerta_sem_compra=3):
    """
    Data de última compra: somente loja 009.
    Preço de última compra: prioriza loja 009; se não houver preço na 009, usa outra unidade com preço.
    """
    if df_giro.empty:
        return pd.DataFrame(columns=["codigo", "Data Última Compra", "Preço Última Compra"])

    df = df_giro.copy()
    df["dt_ult_compra"] = pd.to_datetime(df["dt_ult_compra"], errors="coerce", dayfirst=True)
    df["pr_ult_compra"] = pd.to_numeric(df["pr_ult_compra"], errors="coerce").fillna(0)

    resultados = []
    for codigo, grupo in df.groupby("codigo"):
        g009 = grupo[grupo["codigo_empresa"] == CODIGO_UNICA].copy()

        data_compra = pd.NaT
        data_compra_txt = ""
        if not g009.empty and g009["dt_ult_compra"].notna().any():
            idx = g009["dt_ult_compra"].idxmax()
            data_compra = g009.loc[idx, "dt_ult_compra"]
            raw = str(g009.loc[idx, "dt_ult_compra_txt"] or "").strip()
            data_compra_txt = raw if raw else format_data_br(data_compra)

        preco = 0.0
        precos_009 = g009[g009["pr_ult_compra"] > 0]["pr_ult_compra"] if not g009.empty else pd.Series(dtype=float)
        if not precos_009.empty:
            preco = float(precos_009.iloc[-1])
        else:
            precos_gerais = grupo[grupo["pr_ult_compra"] > 0]["pr_ult_compra"]
            if not precos_gerais.empty:
                preco = float(precos_gerais.iloc[-1])

        icone = data_alerta_icon(data_compra, meses_alerta_sem_compra)
        data_exibicao = f"{icone} {data_compra_txt}".strip() if data_compra_txt else icone

        resultados.append({
            "codigo": codigo,
            "Data Última Compra": data_exibicao,
            "Preço Última Compra": preco,
        })

    return pd.DataFrame(resultados)

# =========================================================
# MONTAGEM DAS TABELAS
# =========================================================

def arredondar_para_embalagem(sugestao, embalagem):
    """
    Arredonda a sugestão para cima, respeitando o múltiplo da embalagem.
    Ex.: sugestão 8 e embalagem 12 => 12; sugestão 20 e embalagem 12 => 24.
    """
    try:
        sugestao = int(math.ceil(float(sugestao or 0)))
        embalagem = int(round(float(embalagem or 0)))

        if sugestao <= 0:
            return 0
        if embalagem <= 0:
            return sugestao

        return int(math.ceil(sugestao / embalagem) * embalagem)
    except Exception:
        try:
            return int(math.ceil(float(sugestao or 0)))
        except Exception:
            return 0


@st.cache_data(show_spinner=False, ttl=3600, max_entries=6)
def montar_tabela_consolidada(
    df_giro,
    df_transito=None,
    dias_estoque_alvo=60,
    meses_alerta_sem_compra=3,
    considerar_mes_atual_media=True,
    meses_ref=None,
):
    df_giro = df_giro.copy()
    meses_ref = list(meses_ref or MESES)

    # Garantias para evitar KeyError quando algum upload não trouxer cadastro/embalagem.
    # A tabela consolidada sempre precisa dessas colunas para o groupby/agg.
    colunas_padrao = {
        "codigo_fabrica": "",
        "embalagem": 0,
        "estoque": 0,
        "codigo_empresa": "",
        "descricao": "",
        "codigo": "",
    }
    for coluna, valor_padrao in colunas_padrao.items():
        if coluna not in df_giro.columns:
            df_giro[coluna] = valor_padrao

    df_lojas = df_giro[df_giro["codigo_empresa"].isin(CODIGOS_LOJAS)].copy()
    df_unica = df_giro[df_giro["codigo_empresa"] == CODIGO_UNICA].copy()

    meses_ref = [m for m in meses_ref if m in df_giro.columns]
    if not meses_ref:
        meses_ref = [m for m in MESES_PADRAO if m in df_giro.columns]

    agg = {
        "codigo_fabrica": "first",
        "embalagem": "first",
        **{mes: "sum" for mes in meses_ref},
        "estoque": "sum",
    }

    lojas = df_lojas.groupby(["codigo", "descricao"], as_index=False).agg(agg) if not df_lojas.empty else pd.DataFrame(columns=["codigo", "descricao", *agg.keys()])
    unica = df_unica.groupby(["codigo", "descricao"], as_index=False).agg(agg) if not df_unica.empty else pd.DataFrame(columns=["codigo", "descricao", *agg.keys()])

    lojas["Média Giro Lojas"] = lojas[meses_ref].mean(axis=1).round(1) if not lojas.empty and meses_ref else []
    unica["Média Giro Única"] = unica[meses_ref].mean(axis=1).round(1) if not unica.empty and meses_ref else []

    lojas = lojas.rename(columns={
        "codigo_fabrica": "Código Fábrica",
        "embalagem": "Embalagem",
        **{mes: col_giro("Giro Lojas", mes) for mes in meses_ref},
        "estoque": "Estoque Lojas",
    })
    unica = unica.rename(columns={
        "codigo_fabrica": "Código Fábrica Única",
        "embalagem": "Embalagem Única",
        **{mes: col_giro("Giro Única", mes) for mes in meses_ref},
        "estoque": "Estoque Única",
    })

    resumo = pd.merge(lojas, unica, on=["codigo", "descricao"], how="outer").fillna(0)

    for col in ["Código Fábrica", "Código Fábrica Única"]:
        if col not in resumo.columns:
            resumo[col] = ""
        resumo[col] = resumo[col].replace(0, "")

    resumo["Código Fábrica"] = resumo.apply(
        lambda x: x["Código Fábrica"] if x["Código Fábrica"] else x["Código Fábrica Única"], axis=1
    )

    for col in ["Embalagem", "Embalagem Única"]:
        if col not in resumo.columns:
            resumo[col] = 0
        resumo[col] = pd.to_numeric(resumo[col], errors="coerce").fillna(0).round(0).astype(int)

    resumo["Embalagem"] = resumo.apply(
        lambda x: int(x["Embalagem"]) if int(x["Embalagem"]) > 0 else int(x["Embalagem Única"]),
        axis=1,
    )

    colunas_giro_geral = []
    for mes in meses_ref:
        label = label_mes_giro(mes)
        for prefixo in ["Giro Lojas", "Giro Única"]:
            col = f"{prefixo} {label}"
            if col not in resumo.columns:
                resumo[col] = 0
        col_geral = f"Giro Geral {label}"
        resumo[col_geral] = resumo[f"Giro Lojas {label}"] + resumo[f"Giro Única {label}"]
        colunas_giro_geral.append(col_geral)

    meses_media_geral = list(meses_ref)
    mes_atual = mes_atual_referencia()
    if not considerar_mes_atual_media and mes_atual in meses_media_geral and len(meses_media_geral) > 1:
        meses_media_geral = [m for m in meses_media_geral if m != mes_atual]

    colunas_media_giro_geral = [col_giro("Giro Geral", mes) for mes in meses_media_geral if col_giro("Giro Geral", mes) in resumo.columns]
    resumo["Média Giro Geral"] = resumo[colunas_media_giro_geral].mean(axis=1).round(1) if colunas_media_giro_geral else 0

    for col in ["Estoque Lojas", "Estoque Única", "Média Giro Lojas", "Média Giro Única"]:
        if col not in resumo.columns:
            resumo[col] = 0

    resumo["Estoque Atual Geral"] = resumo["Estoque Lojas"] + resumo["Estoque Única"]
    resumo["Estoque Geral"] = resumo["Estoque Atual Geral"]

    if df_transito is not None and not df_transito.empty:
        df_transito = df_transito.copy()
        df_transito["codigo"] = df_transito["codigo"].astype(str).str.extract(r"(\d+)", expand=False).fillna("").str.zfill(5)
        df_transito["Saldo em Trânsito/ABERTO"] = pd.to_numeric(df_transito.get("Saldo em Trânsito/ABERTO", 0), errors="coerce").fillna(0)
        df_transito = df_transito[df_transito["codigo"].str.strip().ne("")]
        df_transito = df_transito.groupby("codigo", as_index=False)["Saldo em Trânsito/ABERTO"].sum()
        resumo = pd.merge(resumo, df_transito.drop(columns=["descricao"], errors="ignore"), on="codigo", how="left")
    else:
        resumo["Saldo em Trânsito/ABERTO"] = 0

    resumo["Saldo em Trânsito/ABERTO"] = pd.to_numeric(resumo["Saldo em Trânsito/ABERTO"], errors="coerce").fillna(0)
    resumo["Estoque Final"] = resumo["Estoque Atual Geral"] + resumo["Saldo em Trânsito/ABERTO"]
    resumo["Alerta Estoque"] = resumo.apply(
        lambda row: "🔎 Estoque alto" if float(row.get("Média Giro Geral", 0) or 0) > 0
        and float(row.get("Estoque Final", 0) or 0) >= float(row.get("Média Giro Geral", 0) or 0) * 1.5
        else "",
        axis=1,
    )
    resumo["Estoque Alvo"] = resumo["Média Giro Geral"] * (dias_estoque_alvo / 30)
    resumo["Sugestão Sistema"] = (resumo["Estoque Alvo"] - resumo["Estoque Final"]).apply(lambda x: max(math.ceil(x), 0)).astype(int)
    resumo["Sugestão arredondada"] = resumo.apply(
        lambda row: arredondar_para_embalagem(row["Sugestão Sistema"], row.get("Embalagem", 0)),
        axis=1,
    ).astype(int)

    info_compra = montar_info_compra(df_giro, meses_alerta_sem_compra)
    resumo = pd.merge(resumo, info_compra, on="codigo", how="left")
    resumo["Preço Última Compra"] = pd.to_numeric(resumo["Preço Última Compra"], errors="coerce").fillna(0)

    resumo = resumo.drop(columns=["Código Fábrica Única", "Embalagem Única"], errors="ignore")
    return resumo.sort_values("descricao").reset_index(drop=True)


def montar_detalhe_produto(df_giro, codigo_produto):
    detalhe = df_giro[df_giro["codigo"] == codigo_produto].copy()
    if detalhe.empty:
        return pd.DataFrame()

    meses_ref = [m for m in MESES if m in detalhe.columns]
    detalhe["Média Giro"] = detalhe[meses_ref].mean(axis=1).round(1) if meses_ref else 0
    rename_meses = {mes: label_mes_giro(mes) for mes in meses_ref}
    detalhe = detalhe.rename(columns={
        "loja": "Unidade",
        "codigo_empresa": "Cód. Empresa",
        **rename_meses,
        "estoque": "Saldo em Estoque",
    })

    colunas = ["Cód. Empresa", "Unidade"] + [label_mes_giro(m) for m in meses_ref] + ["Média Giro", "Saldo em Estoque"]
    return detalhe[colunas].sort_values(["Cód. Empresa", "Unidade"])

# =========================================================
# FORMATAÇÃO / EXPORTAÇÃO
# =========================================================

def filtrar_tabela(df, campos, key):
    busca = st.text_input("Pesquisar", key=key)
    if not busca:
        return df.copy()
    busca = busca.lower()
    filtro = pd.Series(False, index=df.index)
    for campo in campos:
        if campo in df.columns:
            filtro = filtro | df[campo].astype(str).str.lower().str.contains(busca, na=False)
    return df[filtro].copy()


def colorir_colunas_consolidada(col):
    if "Lojas" in col.name:
        return ["background-color: #e8f1ff"] * len(col)
    if "Única" in col.name:
        return ["background-color: #fff1df"] * len(col)
    if "Geral" in col.name:
        return ["background-color: #eaf7ea"] * len(col)
    if "ABERTO" in col.name:
        return ["background-color: #cfe2ff; color: #084298; font-weight: 700"] * len(col)
    if "Estoque Final" in col.name:
        return ["background-color: #f3e8ff; font-weight: 600"] * len(col)
    if col.name == "Alerta Estoque":
        return ["background-color: #fef3c7; color: #92400e; font-weight: 700"] * len(col)
    if "Sistema" in col.name or "arredondada" in col.name or "Alvo" in col.name or "PEDIDO Final" in col.name:
        return ["background-color: #ffe8e8"] * len(col)
    return [""] * len(col)


def colorir_colunas_pedido(col):
    if col.name in ["Média Giro Lojas", "Estoque Lojas"]:
        return ["background-color: #e8f1ff"] * len(col)
    if col.name in ["Média Giro Única", "Estoque Única"]:
        return ["background-color: #fff1df"] * len(col)
    if col.name in ["Média Giro Geral", "Estoque Geral"]:
        return ["background-color: #eaf7ea"] * len(col)
    if col.name in ["Saldo em Trânsito/ABERTO"]:
        return ["background-color: #cfe2ff; color: #084298; font-weight: 700"] * len(col)
    if col.name in ["Estoque Final"]:
        return ["background-color: #f3e8ff; font-weight: 600"] * len(col)
    if col.name in ["Estoque Alvo", "Sugestão Sistema", "Sugestão arredondada", "PEDIDO Final", "Valor Final do Pedido"]:
        return ["background-color: #ffe8e8"] * len(col)
    return [""] * len(col)


def colunas_giro_geral_mensal(df):
    return [c for c in df.columns if str(c).startswith("Giro Geral ")]


def estilos_alerta_giro_fora_curva(row):
    estilos = pd.Series("", index=row.index)
    cols_giro = colunas_giro_geral_mensal(pd.DataFrame(columns=row.index))
    if len(cols_giro) < 2:
        return estilos

    valores = pd.to_numeric(row[cols_giro], errors="coerce").fillna(0)
    positivos = valores[valores > 0]

    colunas_alerta = set()
    if len(positivos) == 1:
        colunas_alerta.add(positivos.index[0])
    elif len(positivos) >= 2:
        for col, valor in valores.items():
            if valor <= 0:
                continue
            outros = valores.drop(index=col)
            media_outros = outros[outros > 0].mean()
            if pd.notna(media_outros) and media_outros > 0 and valor >= media_outros * 1.5:
                colunas_alerta.add(col)

    for col in colunas_alerta:
        estilos[col] = "background-color: #fed7aa; color: #9a3412; font-weight: 700"
    return estilos

def formatadores_para_tabela(df):
    fmt = {}
    dinheiro = [c for c in df.columns if "Preço" in c or "Valor" in c]
    inteiros = [c for c in df.columns if c in ["Sugestão Sistema", "Sugestão arredondada", "Sugestão de Pedido", "PEDIDO Final", "Embalagem", "Dias Estoque Pedido"]]
    for col in df.columns:
        if col in dinheiro:
            fmt[col] = format_moeda_br
        elif col in inteiros:
            fmt[col] = format_int_br
        elif pd.api.types.is_numeric_dtype(df[col]):
            fmt[col] = lambda x: format_num_br(x, 1)
    return fmt


LIMITE_CELULAS_STYLER = 18000
LIMITE_CELULAS_EDITOR = 12000


def usar_renderizacao_leve(df, limite_celulas=LIMITE_CELULAS_STYLER):
    linhas = len(df) if df is not None else 0
    colunas = len(df.columns) if df is not None else 0
    return linhas * colunas > limite_celulas


def assinatura_dataframe_colunas(df, colunas):
    if df is None or df.empty:
        return 'vazio'
    colunas_existentes = [col for col in colunas if col in df.columns]
    if not colunas_existentes:
        return f'linhas={len(df)}'
    base = df[colunas_existentes].copy()
    base = base.astype(str).fillna('')
    digest = hashlib.md5(pd.util.hash_pandas_object(base, index=False).values.tobytes()).hexdigest()
    return f'linhas={len(df)}|cols={len(colunas_existentes)}|hash={digest}'


def column_config_tabela(df, fixar_colunas=True):
    config = {}
    if fixar_colunas:
        if "codigo" in df.columns:
            config["codigo"] = st.column_config.TextColumn("Código", width="small", pinned=True)
        if "descricao" in df.columns:
            config["descricao"] = st.column_config.TextColumn("Descrição", width="large", pinned=True)
    if "Código Fábrica" in df.columns:
        config["Código Fábrica"] = st.column_config.TextColumn("Código Fábrica", width="medium")

    for col in df.columns:
        if col in config:
            continue
        if "Preço" in col or "Valor" in col:
            config[col] = st.column_config.NumberColumn(col, format="R$ %.2f")
        elif col in ["Sugestão Sistema", "Sugestão arredondada", "PEDIDO Final", "Embalagem"]:
            config[col] = st.column_config.NumberColumn(col, format="%d")
        elif pd.api.types.is_numeric_dtype(df[col]):
            config[col] = st.column_config.NumberColumn(col, format="%.1f")
    return config


def column_config_tabela(df, fixar_colunas=True):
    config = {}
    if fixar_colunas:
        if "codigo" in df.columns:
            config["codigo"] = st.column_config.TextColumn("Codigo", width="small", pinned=True)
        if "descricao" in df.columns:
            config["descricao"] = st.column_config.TextColumn("Descricao", width="large", pinned=True)

    coluna_fabrica = next((col for col in df.columns if "brica" in str(col)), None)
    if coluna_fabrica:
        config[coluna_fabrica] = st.column_config.TextColumn("Codigo Fabrica", width="medium")

    for col in df.columns:
        col_txt = str(col)
        if col in config:
            continue
        if "Pre" in col_txt or "Valor" in col_txt:
            config[col] = st.column_config.NumberColumn(col, format="R$ %.2f")
        elif "Sugest" in col_txt or col in ["PEDIDO Final", "Embalagem"]:
            config[col] = st.column_config.NumberColumn(col, format="%d")
        elif pd.api.types.is_numeric_dtype(df[col]):
            config[col] = st.column_config.NumberColumn(col, format="%.1f")
    return config


def render_tabela_interativa_colorida(df, height=650):
    if usar_renderizacao_leve(df):
        st.caption(
            "Modo rápido ativado para muitos meses/itens. "
            "A tabela fica completa, mas sem pintura célula a célula para evitar queda do Streamlit."
        )
        st.dataframe(
            df,
            use_container_width=True,
            hide_index=True,
            height=height,
            column_config=column_config_tabela(df),
        )
        return

    styled = df.style.apply(colorir_colunas_consolidada, axis=0).apply(estilos_alerta_giro_fora_curva, axis=1).format(formatadores_para_tabela(df))
    st.dataframe(
        styled,
        use_container_width=True,
        hide_index=True,
        height=height,
        column_config={
            "codigo": st.column_config.TextColumn("Código", width="small", pinned=True),
            "descricao": st.column_config.TextColumn("Descrição", width="large", pinned=True),
            "Código Fábrica": st.column_config.TextColumn("Código Fábrica", width="medium"),
        },
    )


def render_tabela_interativa_colorida(df, height=650):
    if usar_renderizacao_leve(df):
        st.caption(
            "Modo rapido ativado para muitos meses/itens. "
            "A tabela fica completa, mas sem pintura celula a celula para evitar queda do Streamlit."
        )
        st.dataframe(
            df,
            use_container_width=True,
            hide_index=True,
            height=height,
            column_config=column_config_tabela(df),
        )
        return

    styled = df.style.apply(colorir_colunas_consolidada, axis=0).apply(estilos_alerta_giro_fora_curva, axis=1).format(formatadores_para_tabela(df))
    st.dataframe(
        styled,
        use_container_width=True,
        hide_index=True,
        height=height,
        column_config=column_config_tabela(df),
    )


def gerar_csv(df):
    return df.to_csv(index=False, sep=";", decimal=",", encoding="utf-8-sig").encode("utf-8-sig")


# =========================================================
# INTEGRACAO GOOGLE DRIVE / SHEETS
# =========================================================

def google_configurado():
    if Credentials is None or Request is None or build is None or MediaIoBaseUpload is None:
        return False
    try:
        cfg = dict(st.secrets.get("google_oauth_user", {}))
        return bool(
            cfg.get("client_id")
            and cfg.get("client_secret")
            and cfg.get("refresh_token")
            and cfg.get("token_uri", "https://oauth2.googleapis.com/token")
        )
    except Exception:
        return False


def google_mensagem_configuracao():
    if Credentials is None or Request is None or build is None or MediaIoBaseUpload is None:
        return (
            "Instale as dependencias do Google no ambiente: "
            "google-api-python-client, google-auth e google-auth-httplib2."
        )
    return (
        "Configure o OAuth 2.0 da conta dona do Drive em st.secrets, na seção "
        "[google_oauth_user], com client_id, client_secret, refresh_token e token_uri. "
        "A conta autorizada deve ser gdautotintas@gmail.com. Enquanto o refresh_token "
        "não estiver configurado, o app não consegue criar planilhas no Drive."
    )


def google_oauth_user_json():
    try:
        cfg = dict(st.secrets.get("google_oauth_user", {}))
        cfg["token_uri"] = cfg.get("token_uri", "https://oauth2.googleapis.com/token")
        campos = ["client_id", "client_secret", "refresh_token", "token_uri"]
        if all(cfg.get(c) for c in campos):
            return json.dumps({c: str(cfg.get(c)) for c in campos}, sort_keys=True)
    except Exception:
        return ""
    return ""


@st.cache_resource(show_spinner=False)
def google_get_services_cached(oauth_json):
    info = json.loads(oauth_json)
    scopes = [
        "https://www.googleapis.com/auth/drive",
        "https://www.googleapis.com/auth/spreadsheets",
        "https://www.googleapis.com/auth/gmail.send",
    ]
    credentials = Credentials(
        token=None,
        refresh_token=info["refresh_token"],
        token_uri=info["token_uri"],
        client_id=info["client_id"],
        client_secret=info["client_secret"],
        scopes=scopes,
    )
    credentials.refresh(Request())
    drive_service = build("drive", "v3", credentials=credentials, cache_discovery=False)
    sheets_service = build("sheets", "v4", credentials=credentials, cache_discovery=False)
    gmail_service = build("gmail", "v1", credentials=credentials, cache_discovery=False)
    return drive_service, sheets_service, gmail_service


def google_get_services():
    oauth_json = google_oauth_user_json()
    if not oauth_json:
        raise RuntimeError(google_mensagem_configuracao())
    return google_get_services_cached(oauth_json)


def google_link_pasta(folder_id):
    return f"https://drive.google.com/drive/folders/{folder_id}"


def google_link_planilha(spreadsheet_id):
    return f"https://docs.google.com/spreadsheets/d/{spreadsheet_id}/edit"


def google_link_arquivo(file_id):
    return f"https://drive.google.com/file/d/{file_id}/view"


def google_safe_name(nome):
    nome = str(nome or "").strip()
    nome = re.sub(r"[\\/:*?\"<>|]+", "-", nome)
    nome = re.sub(r"\s+", " ", nome).strip()
    return nome or f"Pedido {datetime.now().strftime('%Y-%m-%d %H.%M')}"


def google_q_text(value):
    return str(value).replace("\\", "\\\\").replace("'", "\\'")


def google_find_file(drive_service, name, parent_id, mime_type=None):
    partes = [
        f"name = '{google_q_text(name)}'",
        f"'{google_q_text(parent_id)}' in parents",
        "trashed = false",
    ]
    if mime_type:
        partes.append(f"mimeType = '{google_q_text(mime_type)}'")
    result = drive_service.files().list(
        q=" and ".join(partes),
        spaces="drive",
        fields="files(id, name, mimeType, webViewLink)",
        pageSize=10,
        supportsAllDrives=True,
        includeItemsFromAllDrives=True,
    ).execute()
    files = result.get("files", [])
    return files[0] if files else None


def google_ensure_folder(drive_service, name, parent_id):
    mime = "application/vnd.google-apps.folder"
    existente = google_find_file(drive_service, name, parent_id, mime)
    if existente:
        return existente["id"]
    criado = drive_service.files().create(
        body={"name": name, "mimeType": mime, "parents": [parent_id]},
        fields="id",
        supportsAllDrives=True,
    ).execute()
    return criado["id"]


def google_ensure_spreadsheet(drive_service, name, parent_id):
    mime = "application/vnd.google-apps.spreadsheet"
    existente = google_find_file(drive_service, name, parent_id, mime)
    if existente:
        return existente["id"]
    criado = drive_service.files().create(
        body={"name": name, "mimeType": mime, "parents": [parent_id]},
        fields="id",
        supportsAllDrives=True,
    ).execute()
    return criado["id"]


def google_ensure_sheet_tab(sheets_service, spreadsheet_id, sheet_name):
    meta = sheets_service.spreadsheets().get(
        spreadsheetId=spreadsheet_id,
        fields="sheets(properties(title))",
    ).execute()
    existentes = {s["properties"]["title"] for s in meta.get("sheets", [])}
    if sheet_name in existentes:
        return
    sheets_service.spreadsheets().batchUpdate(
        spreadsheetId=spreadsheet_id,
        body={"requests": [{"addSheet": {"properties": {"title": sheet_name}}}]},
    ).execute()


def google_prepare_value(value):
    if pd.isna(value):
        return ""
    if isinstance(value, (datetime, date, pd.Timestamp)):
        return format_data_br(value)
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        return float(value)
    return str(value)


def google_df_to_values(df):
    df = df.copy() if df is not None else pd.DataFrame()
    header = ["" if str(col).startswith("__blank_") else col for col in df.columns]
    return [header] + [
        [google_prepare_value(row.get(col, "")) for col in df.columns]
        for _, row in df.iterrows()
    ]


def google_write_df(sheets_service, spreadsheet_id, sheet_name, df):
    google_ensure_sheet_tab(sheets_service, spreadsheet_id, sheet_name)
    sheets_service.spreadsheets().values().clear(
        spreadsheetId=spreadsheet_id,
        range=f"'{sheet_name}'!A:ZZ",
        body={},
    ).execute()
    values = google_df_to_values(df)
    sheets_service.spreadsheets().values().update(
        spreadsheetId=spreadsheet_id,
        range=f"'{sheet_name}'!A1",
        valueInputOption="USER_ENTERED",
        body={"values": values},
    ).execute()


def google_read_df(sheets_service, spreadsheet_id, sheet_name=None):
    range_name = f"'{sheet_name}'!A:ZZ" if sheet_name else "A:ZZ"
    result = sheets_service.spreadsheets().values().get(
        spreadsheetId=spreadsheet_id,
        range=range_name,
    ).execute()
    values = result.get("values", [])
    if not values:
        return pd.DataFrame()
    header = [str(c).strip() for c in values[0]]
    rows = values[1:]
    largura = len(header)
    rows = [row + [""] * (largura - len(row)) if len(row) < largura else row[:largura] for row in rows]
    return pd.DataFrame(rows, columns=header)


def google_append_rows(sheets_service, spreadsheet_id, sheet_name, rows):
    google_ensure_sheet_tab(sheets_service, spreadsheet_id, sheet_name)
    sheets_service.spreadsheets().values().append(
        spreadsheetId=spreadsheet_id,
        range=f"'{sheet_name}'!A1",
        valueInputOption="USER_ENTERED",
        insertDataOption="INSERT_ROWS",
        body={"values": rows},
    ).execute()


def google_ensure_headers(sheets_service, spreadsheet_id, sheet_name, columns):
    google_ensure_sheet_tab(sheets_service, spreadsheet_id, sheet_name)
    atual = google_read_df(sheets_service, spreadsheet_id, sheet_name)
    if atual.empty and list(atual.columns) == []:
        google_write_df(sheets_service, spreadsheet_id, sheet_name, pd.DataFrame(columns=columns))



@st.cache_data(show_spinner=False, ttl=60)
def google_get_resources_cached(_cache_key):
    """
    Recursos fixos do Drive/Sheets.

    Importante:
    - Não cria novas pastas nem planilhas centrais aqui.
    - Isso evita o erro storageQuotaExceeded da conta de serviço ao tentar criar
      arquivos nativos do Google Sheets no Meu Drive.
    - Os pedidos são gerados do zero pelo código, sem copiar planilha modelo.
    """
    google_get_services()
    return {
        "oauth_user": "gdautotintas@gmail.com",
        "pedidos_folder_id": GOOGLE_PASTA_APROVACAO_ID,
        "aprovados_folder_id": GOOGLE_PASTA_APROVADOS_ID,
        "finais_folder_id": GOOGLE_PASTA_APROVADOS_ID,
        "cadastro_id": GOOGLE_PLANILHA_CADASTRO_ID,
        "root_link": google_link_pasta(GOOGLE_PASTA_APROVACAO_ID),
        "pedidos_link": google_link_pasta(GOOGLE_PASTA_APROVACAO_ID),
        "aprovados_link": google_link_pasta(GOOGLE_PASTA_APROVADOS_ID),
        "finais_link": google_link_pasta(GOOGLE_PASTA_APROVADOS_ID),
        "cadastro_link": google_link_planilha(GOOGLE_PLANILHA_CADASTRO_ID),
        "controle_link": google_link_pasta(GOOGLE_PASTA_APROVACAO_ID),
    }


def google_get_resources():
    oauth_json = google_oauth_user_json()
    if not oauth_json:
        raise RuntimeError(google_mensagem_configuracao())
    return google_get_resources_cached(str(hash(oauth_json)))


def google_listar_planilhas_pasta(drive_service, folder_id):
    q = " and ".join([
        "mimeType = 'application/vnd.google-apps.spreadsheet'",
        f"'{google_q_text(folder_id)}' in parents",
        "trashed = false",
    ])
    arquivos = []
    page_token = None
    while True:
        result = drive_service.files().list(
            q=q,
            spaces="drive",
            fields="nextPageToken, files(id, name, webViewLink, createdTime, modifiedTime)",
            pageSize=100,
            pageToken=page_token,
            supportsAllDrives=True,
            includeItemsFromAllDrives=True,
        ).execute()
        arquivos.extend(result.get("files", []))
        page_token = result.get("nextPageToken")
        if not page_token:
            break
    return arquivos


def google_controle_padrao(status="Em edição", criado_por="", criado_em="", fornecedor="", valor="", pedido_id="", observacao=""):
    return pd.DataFrame([
        {"Campo": "Status", "Valor": status},
        {"Campo": "Criado por", "Valor": criado_por},
        {"Campo": "Criado em", "Valor": criado_em},
        {"Campo": "Última alteração", "Valor": datetime.now().strftime("%d/%m/%Y %H:%M")},
        {"Campo": "Aprovado por", "Valor": ""},
        {"Campo": "Aprovado em", "Valor": ""},
        {"Campo": "Observação", "Valor": observacao},
        {"Campo": "Fornecedor", "Valor": fornecedor},
        {"Campo": "Valor do Pedido", "Valor": valor},
        {"Campo": "ID Pedido", "Valor": pedido_id},
    ])


def google_normalizar_chave_controle(valor):
    txt = str(valor or "").strip().lower()
    txt = unicodedata.normalize("NFKD", txt).encode("ascii", "ignore").decode("ascii")
    txt = re.sub(r"[^a-z0-9]+", "_", txt).strip("_")
    return txt


def google_ler_controle_pedido(spreadsheet_id):
    """Lê a aba Controle. Aceita layout Campo/Valor ou layout antigo por cabeçalhos."""
    _, sheets_service, _ = google_get_services()
    try:
        df = google_read_df(sheets_service, spreadsheet_id, "Controle")
    except Exception:
        try:
            df = google_read_df(sheets_service, spreadsheet_id, "Aprovacao")
        except Exception:
            return {}

    if df is None or df.empty:
        return {}

    cols_norm = {google_normalizar_chave_controle(c): c for c in df.columns}
    dados = {}

    # Layout recomendado: Campo | Valor
    if "campo" in cols_norm and "valor" in cols_norm:
        col_campo = cols_norm["campo"]
        col_valor = cols_norm["valor"]
        for _, row in df.iterrows():
            chave = google_normalizar_chave_controle(row.get(col_campo, ""))
            if chave:
                dados[chave] = str(row.get(col_valor, "") or "").strip()
        return dados

    # Layout antigo: status | aprovado_por | aprovado_em | observacao
    primeira = df.iloc[0].to_dict()
    for col, val in primeira.items():
        dados[google_normalizar_chave_controle(col)] = str(val or "").strip()
    return dados


def google_escrever_controle_pedido(spreadsheet_id, **kwargs):
    _, sheets_service, _ = google_get_services()
    atual = google_ler_controle_pedido(spreadsheet_id)
    mapa = {
        "status": kwargs.get("status", atual.get("status", "Em edição")),
        "criado_por": kwargs.get("criado_por", atual.get("criado_por", "")),
        "criado_em": kwargs.get("criado_em", atual.get("criado_em", "")),
        "ultima_alteracao": datetime.now().strftime("%d/%m/%Y %H:%M"),
        "aprovado_por": kwargs.get("aprovado_por", atual.get("aprovado_por", "")),
        "aprovado_em": kwargs.get("aprovado_em", atual.get("aprovado_em", "")),
        "observacao": kwargs.get("observacao", atual.get("observacao", "")),
        "fornecedor": kwargs.get("fornecedor", atual.get("fornecedor", "")),
        "valor_do_pedido": kwargs.get("valor", atual.get("valor_do_pedido", atual.get("valor", ""))),
        "id_pedido": kwargs.get("pedido_id", atual.get("id_pedido", "")),
    }
    df = pd.DataFrame([
        {"Campo": "Status", "Valor": mapa["status"]},
        {"Campo": "Criado por", "Valor": mapa["criado_por"]},
        {"Campo": "Criado em", "Valor": mapa["criado_em"]},
        {"Campo": "Última alteração", "Valor": mapa["ultima_alteracao"]},
        {"Campo": "Aprovado por", "Valor": mapa["aprovado_por"]},
        {"Campo": "Aprovado em", "Valor": mapa["aprovado_em"]},
        {"Campo": "Observação", "Valor": mapa["observacao"]},
        {"Campo": "Fornecedor", "Valor": mapa["fornecedor"]},
        {"Campo": "Valor do Pedido", "Valor": mapa["valor_do_pedido"]},
        {"Campo": "ID Pedido", "Valor": mapa["id_pedido"]},
    ])
    google_write_df(sheets_service, spreadsheet_id, "Controle", df)


def google_enviar_email_aprovadores(nome_pedido, fornecedor, valor, link_pedido, criado_por=""):
    """Envia e-mail aos aprovadores via Gmail API usando o OAuth configurado."""
    assunto = f"Novo pedido aguardando aprovação - {fornecedor or nome_pedido}"
    corpo = f"""Olá,

Um novo pedido foi gerado e está aguardando aprovação.

Pedido: {nome_pedido}
Fornecedor: {fornecedor}
Valor estimado: {format_moeda_br(valor)}
Criado por: {criado_por}

Acesse a planilha para revisar e aprovar:
{link_pedido}

Para aprovar, altere o Status na aba Controle para Aprovado.

Atenciosamente,
Sistema de Pedidos
"""
    msg = MIMEText(corpo, "plain", "utf-8")
    msg["Subject"] = assunto
    msg["From"] = formataddr(("Sistema de Pedidos", "gdautotintas@gmail.com"))
    msg["To"] = ", ".join(GOOGLE_APROVADORES_EMAILS)

    try:
        _, _, gmail_service = google_get_services()
        raw = base64.urlsafe_b64encode(msg.as_bytes()).decode("utf-8")
        gmail_service.users().messages().send(userId="me", body={"raw": raw}).execute()
        return True, "E-mail enviado aos aprovadores."
    except Exception as e:
        return False, f"Pedido criado, mas não consegui enviar e-mail pela Gmail API: {e}"


def google_criar_spreadsheet_do_zero(drive_service, sheets_service, titulo, folder_id):
    """
    Cria uma planilha Google Sheets do zero, sem usar files().copy().

    Motivo: copiar uma planilha modelo com conta de serviço pode fazer o Google
    tentar atribuir a cópia à cota da própria conta de serviço, gerando
    storageQuotaExceeded. Aqui a planilha é criada vazia e depois estruturada
    pelo código.
    """
    ultimo_erro = None

    # Tentativa 1: criar diretamente no Drive dentro da pasta informada.
    try:
        criado = drive_service.files().create(
            body={
                "name": titulo,
                "mimeType": "application/vnd.google-apps.spreadsheet",
                "parents": [folder_id],
            },
            fields="id, webViewLink",
            supportsAllDrives=True,
        ).execute()
        return criado["id"], criado.get("webViewLink") or google_link_planilha(criado["id"])
    except Exception as e:
        ultimo_erro = e

    # Tentativa 2: criar pela Sheets API e mover para a pasta.
    # Em alguns ambientes esta rota funciona melhor do que Drive files().create().
    try:
        criado = sheets_service.spreadsheets().create(
            body={"properties": {"title": titulo}},
            fields="spreadsheetId",
        ).execute()
        spreadsheet_id = criado["spreadsheetId"]

        # Move para a pasta de aprovação. Se a planilha nasceu em uma raiz acessível,
        # remove os pais atuais e adiciona a pasta correta.
        arquivo = drive_service.files().get(
            fileId=spreadsheet_id,
            fields="parents, webViewLink",
            supportsAllDrives=True,
        ).execute()
        parents = ",".join(arquivo.get("parents", []))
        drive_service.files().update(
            fileId=spreadsheet_id,
            addParents=folder_id,
            removeParents=parents,
            fields="id, webViewLink",
            supportsAllDrives=True,
        ).execute()
        return spreadsheet_id, google_link_planilha(spreadsheet_id)
    except Exception as e2:
        raise RuntimeError(
            "Não consegui criar a planilha do pedido no Google Sheets. "
            "O código já foi ajustado para NÃO copiar modelo e gerar a planilha do zero. "
            "Se este erro continuar, a conta de serviço está impedida pelo Google de criar "
            "arquivos nativos do Google Sheets em Meu Drive. Nesse caso, use OAuth da conta "
            "dona da pasta ou um Drive Compartilhado do Google Workspace. "
            f"Erro Drive API: {ultimo_erro} | Erro Sheets API: {e2}"
        )


def google_remover_abas_padrao(sheets_service, spreadsheet_id, abas_manter):
    try:
        meta = sheets_service.spreadsheets().get(
            spreadsheetId=spreadsheet_id,
            fields="sheets(properties(sheetId,title))",
        ).execute()
        requests = []
        for sheet in meta.get("sheets", []):
            props = sheet.get("properties", {})
            title = props.get("title")
            sheet_id = props.get("sheetId")
            if title not in set(abas_manter) and sheet_id is not None:
                requests.append({"deleteSheet": {"sheetId": sheet_id}})
        if requests:
            sheets_service.spreadsheets().batchUpdate(
                spreadsheetId=spreadsheet_id,
                body={"requests": requests},
            ).execute()
    except Exception:
        # Não trava o fluxo por causa de uma aba padrão extra.
        pass


def google_formatar_planilha_pedido(sheets_service, spreadsheet_id, pedido_df):
    """Aplica formatação básica e menu de status na planilha gerada do zero."""
    meta = sheets_service.spreadsheets().get(
        spreadsheetId=spreadsheet_id,
        fields="sheets(properties(sheetId,title))",
    ).execute()
    ids = {s["properties"]["title"]: s["properties"]["sheetId"] for s in meta.get("sheets", [])}
    pedido_sheet_id = ids.get("Pedido")
    controle_sheet_id = ids.get("Controle")
    requests = []

    if pedido_sheet_id is not None:
        n_cols = max(len(pedido_df.columns), 1)
        n_rows = max(len(pedido_df) + 1, 2)
        requests.extend([
            {"repeatCell": {
                "range": {"sheetId": pedido_sheet_id, "startRowIndex": 0, "endRowIndex": 1, "startColumnIndex": 0, "endColumnIndex": n_cols},
                "cell": {"userEnteredFormat": {
                    "backgroundColor": {"red": 0.85, "green": 0.92, "blue": 0.97},
                    "horizontalAlignment": "CENTER",
                    "textFormat": {"bold": True},
                    "wrapStrategy": "WRAP",
                }},
                "fields": "userEnteredFormat(backgroundColor,textFormat,horizontalAlignment,wrapStrategy)",
            }},
            {"updateSheetProperties": {"properties": {"sheetId": pedido_sheet_id, "gridProperties": {"frozenRowCount": 1}}, "fields": "gridProperties.frozenRowCount"}},
            {"autoResizeDimensions": {
                "dimensions": {"sheetId": pedido_sheet_id, "dimension": "COLUMNS", "startIndex": 0, "endIndex": n_cols}
            }},
            {"addFilterView": {
                "filter": {
                    "title": "Filtro Pedido",
                    "range": {"sheetId": pedido_sheet_id, "startRowIndex": 0, "endRowIndex": n_rows, "startColumnIndex": 0, "endColumnIndex": n_cols},
                }
            }},
        ])

    if controle_sheet_id is not None:
        requests.extend([
            {"repeatCell": {
                "range": {"sheetId": controle_sheet_id, "startRowIndex": 0, "endRowIndex": 1, "startColumnIndex": 0, "endColumnIndex": 2},
                "cell": {"userEnteredFormat": {
                    "backgroundColor": {"red": 0.85, "green": 0.92, "blue": 0.97},
                    "horizontalAlignment": "CENTER",
                    "textFormat": {"bold": True},
                }},
                "fields": "userEnteredFormat(backgroundColor,textFormat,horizontalAlignment)",
            }},
            {"setDataValidation": {
                "range": {"sheetId": controle_sheet_id, "startRowIndex": 1, "endRowIndex": 2, "startColumnIndex": 1, "endColumnIndex": 2},
                "rule": {
                    "condition": {
                        "type": "ONE_OF_LIST",
                        "values": [
                            {"userEnteredValue": "Em edição"},
                            {"userEnteredValue": "Aguardando aprovação"},
                            {"userEnteredValue": "Aprovado"},
                            {"userEnteredValue": "Reprovado"},
                            {"userEnteredValue": "Finalizado"},
                        ],
                    },
                    "showCustomUi": True,
                    "strict": True,
                },
            }},
            {"autoResizeDimensions": {
                "dimensions": {"sheetId": controle_sheet_id, "dimension": "COLUMNS", "startIndex": 0, "endIndex": 2}
            }},
        ])

    if requests:
        try:
            sheets_service.spreadsheets().batchUpdate(
                spreadsheetId=spreadsheet_id,
                body={"requests": requests},
            ).execute()
        except Exception:
            # Formatação não deve impedir a criação do pedido.
            pass


def google_criar_planilha_pedido(nome_pedido, fornecedor, pedido_df, criado_por=""):
    drive_service, sheets_service, _ = google_get_services()
    recursos = google_get_resources()
    nome_limpo = google_safe_name(nome_pedido)
    fornecedor_limpo = google_safe_name(fornecedor)
    pedido_id = datetime.now().strftime("%Y%m%d%H%M%S")
    titulo = f"{datetime.now().strftime('%Y-%m-%d')} - {fornecedor_limpo} - {nome_limpo}"

    df_export = pedido_df.copy()
    if "zx" not in df_export.columns:
        df_export.insert(0, "zx", df_export.get("codigo", ""))

    valor = totalizar_valor_pedido(df_export)
    agora = datetime.now().strftime("%d/%m/%Y %H:%M")

    # Cria a planilha do zero. Não usa files().copy() nem planilha modelo.
    spreadsheet_id, link = google_criar_spreadsheet_do_zero(
        drive_service,
        sheets_service,
        titulo,
        GOOGLE_PASTA_APROVACAO_ID,
    )

    google_write_df(sheets_service, spreadsheet_id, "Pedido", df_export)
    google_escrever_controle_pedido(
        spreadsheet_id,
        status="Aguardando aprovação",
        criado_por=criado_por,
        criado_em=agora,
        fornecedor=fornecedor_limpo,
        valor=round(float(valor or 0), 2),
        pedido_id=pedido_id,
    )
    google_remover_abas_padrao(sheets_service, spreadsheet_id, ["Pedido", "Controle"])
    google_formatar_planilha_pedido(sheets_service, spreadsheet_id, df_export)

    email_ok, email_msg = google_enviar_email_aprovadores(nome_limpo, fornecedor_limpo, valor, link, criado_por=criado_por)

    return {
        "pedido_id": pedido_id,
        "spreadsheet_id": spreadsheet_id,
        "link": link,
        "titulo": titulo,
        "email_ok": email_ok,
        "email_msg": email_msg,
    }



def google_exportar_pedido_sheets_simples(nome_pedido, fornecedor, pedido_df, criado_por=""):
    """
    Exporta o pedido de compra diretamente para uma nova planilha Google Sheets
    na pasta de aprovação configurada, sem gerar Excel e sem depender do envio de e-mail.
    """
    drive_service, sheets_service, _ = google_get_services()

    nome_limpo = google_safe_name(nome_pedido)
    fornecedor_limpo = google_safe_name(fornecedor)
    pedido_id = datetime.now().strftime("%Y%m%d%H%M%S")
    titulo = f"{datetime.now().strftime('%Y-%m-%d')} - {fornecedor_limpo or 'Fornecedor'} - {nome_limpo}"

    df_export = pedido_df.copy()
    if "zx" not in df_export.columns:
        df_export.insert(0, "zx", df_export.get("codigo", ""))

    if "Valor Final do Pedido" not in df_export.columns:
        df_export = atualizar_valor_e_origem(df_export)

    valor = totalizar_valor_pedido(df_export)
    agora = datetime.now().strftime("%d/%m/%Y %H:%M")

    spreadsheet_id, link = google_criar_spreadsheet_do_zero(
        drive_service,
        sheets_service,
        titulo,
        GOOGLE_PASTA_APROVACAO_ID,
    )

    google_write_df(sheets_service, spreadsheet_id, "Pedido", df_export)
    google_escrever_controle_pedido(
        spreadsheet_id,
        status="Exportado em Sheets",
        criado_por=criado_por,
        criado_em=agora,
        fornecedor=fornecedor_limpo,
        valor=round(float(valor or 0), 2),
        pedido_id=pedido_id,
        observacao="Pedido exportado diretamente pelo Streamlit para Google Sheets.",
    )
    google_remover_abas_padrao(sheets_service, spreadsheet_id, ["Pedido", "Controle"])
    google_formatar_planilha_pedido(sheets_service, spreadsheet_id, df_export)

    return {
        "pedido_id": pedido_id,
        "spreadsheet_id": spreadsheet_id,
        "link": link,
        "titulo": titulo,
        "valor": valor,
        "folder_id": GOOGLE_PASTA_APROVACAO_ID,
        "folder_link": google_link_pasta(GOOGLE_PASTA_APROVACAO_ID),
    }



# =========================================================
# EXPORTAÇÃO VIA GOOGLE APPS SCRIPT (SEM OAUTH / SEM SERVICE ACCOUNT)
# =========================================================

def apps_script_configurado():
    try:
        cfg = dict(st.secrets.get("apps_script", {}))
        return bool(str(cfg.get("web_app_url", "")).strip())
    except Exception:
        return False


def apps_script_mensagem_configuracao():
    return (
        "Configure a URL do Web App do Google Apps Script em .streamlit/secrets.toml, "
        "na seção [apps_script], campo web_app_url. Use a URL terminada em /exec do deploy atual. "
        "Não precisa OAuth, refresh_token nem Service Account."
    )


def apps_script_payload_pedido(nome_pedido, fornecedor, pedido_df, criado_por=""):
    nome_limpo = google_safe_name(nome_pedido)
    fornecedor_limpo = google_safe_name(fornecedor)
    pedido_id = datetime.now().strftime("%Y%m%d%H%M%S")
    titulo = f"{datetime.now().strftime('%Y-%m-%d')} - {fornecedor_limpo or 'Fornecedor'} - {nome_limpo}"

    df_export = pedido_df.copy()

    # IMPORTANTE:
    # Para o Google Sheets, a ordem precisa ficar sem a coluna auxiliar "zx",
    # porque a fórmula solicitada considera:
    # P = Preço Última Compra
    # R = PEDIDO Final
    # T = Valor Final do Pedido = R * P
    # W = Total Geral do Pedido = soma da coluna T
    df_export = df_export.drop(columns=["zx"], errors="ignore")

    if "Valor Final do Pedido" not in df_export.columns:
        df_export = atualizar_valor_e_origem(df_export)

    ordem_oficial = [c for c in colunas_pedido_compras() if c in df_export.columns]
    extras = [c for c in df_export.columns if c not in ordem_oficial]
    df_export = df_export[ordem_oficial + extras]

    colunas_numericas = [
        c for c in df_export.columns
        if (
            c.startswith("Giro ")
            or c.startswith("Média ")
            or c.startswith("Estoque")
            or c.startswith("Sugestão")
            or c in [
                "Saldo em Trânsito/ABERTO",
                "PEDIDO Final",
                "Preço Última Compra",
                "Valor Final do Pedido",
                "Embalagem",
            ]
        )
    ]
    for col in colunas_numericas:
        if col in df_export.columns:
            if col in ["PEDIDO Final", "Sugestão Sistema", "Sugestão arredondada", "Embalagem"]:
                df_export[col] = pd.to_numeric(df_export[col], errors="coerce").fillna(0).round(0).astype(int)
            else:
                df_export[col] = df_export[col].apply(numero_planilha_para_float)

    if "PEDIDO Final" in df_export.columns and "Preço Última Compra" in df_export.columns:
        df_export["Valor Final do Pedido"] = (
            pd.to_numeric(df_export["PEDIDO Final"], errors="coerce").fillna(0)
            * pd.to_numeric(df_export["Preço Última Compra"], errors="coerce").fillna(0)
        ).round(2)

    # Garante que "Total Geral do Pedido" fique na coluna W apenas para o Apps Script
    # preencher o menu/cabecalho com a soma fixa.
    df_export = df_export.drop(columns=["Total Geral do Pedido"], errors="ignore")
    while len(df_export.columns) < 22:
        df_export[f"__blank_{len(df_export.columns) + 1}__"] = ""
    df_export["Total Geral do Pedido"] = ""

    valor = totalizar_valor_pedido(pedido_df)
    agora = datetime.now().strftime("%d/%m/%Y %H:%M")

    controle = [
        ["Campo", "Valor"],
        ["Status", "Aguardando aprovação"],
        ["Criado por", criado_por],
        ["Criado em", agora],
        ["Última alteração", agora],
        ["Aprovado por", ""],
        ["Aprovado em", ""],
        ["Observação", "Pedido criado pelo Streamlit via Google Apps Script."],
        ["Fornecedor", fornecedor_limpo],
        ["Valor do Pedido", round(float(valor or 0), 2)],
        ["ID Pedido", pedido_id],
    ]

    return {
        "action": "criar_pedido",
        "folder_id": GOOGLE_PASTA_APROVACAO_ID,
        "approved_folder_id": GOOGLE_PASTA_APROVADOS_ID,
        "approver_emails": GOOGLE_APROVADORES_EMAILS,
        "title": titulo,
        "pedido_id": pedido_id,
        "nome_pedido": nome_limpo,
        "fornecedor": fornecedor_limpo,
        "criado_por": criado_por,
        "valor": round(float(valor or 0), 2),
        "pedido": google_df_to_values(df_export),
        "controle": controle,
    }


def apps_script_post(payload):
    cfg = dict(st.secrets.get("apps_script", {}))
    url = str(cfg.get("web_app_url", "")).strip()
    if not url:
        raise RuntimeError(apps_script_mensagem_configuracao())
    if "/macros/s/" not in url or not url.rstrip("/").endswith("/exec"):
        raise RuntimeError(
            "URL do Apps Script inválida. Use a URL do Web App implantado, no formato "
            "https://script.google.com/macros/s/SEU_DEPLOY_ID/exec"
        )

    token = str(cfg.get("token", "")).strip()
    if token:
        payload["token"] = token

    data = json.dumps(payload, ensure_ascii=False).encode("utf-8")
    req = urllib.request.Request(
        url,
        data=data,
        headers={"Content-Type": "application/json"},
        method="POST",
    )
    try:
        with urllib.request.urlopen(req, timeout=60) as resp:
            body = resp.read().decode("utf-8")
    except urllib.error.HTTPError as e:
        detalhe = e.read().decode("utf-8", errors="ignore")
        if e.code == 404:
            raise RuntimeError(
                "Apps Script retornou 404. A URL do Web App está inválida, antiga ou o deploy não foi publicado. "
                "No Apps Script, clique em Implantar > Gerenciar implantações > copie a URL do app da web terminada em /exec "
                "e atualize o secret [apps_script].web_app_url no Streamlit."
            ) from e
        raise RuntimeError(f"Erro HTTP ao chamar o Apps Script: {e.code} - {detalhe}")
    except Exception as e:
        raise RuntimeError(f"Não consegui chamar o Apps Script: {e}")

    try:
        result = json.loads(body)
    except Exception:
        raise RuntimeError(f"Resposta inválida do Apps Script: {body[:500]}")

    if not result.get("ok"):
        raise RuntimeError(result.get("error") or result.get("message") or "Apps Script retornou erro desconhecido.")
    return result


def apps_script_criar_planilha_pedido(nome_pedido, fornecedor, pedido_df, criado_por=""):
    payload = apps_script_payload_pedido(nome_pedido, fornecedor, pedido_df, criado_por=criado_por)
    result = apps_script_post(payload)
    return {
        "pedido_id": payload.get("pedido_id"),
        "spreadsheet_id": result.get("spreadsheet_id"),
        "link": result.get("url") or result.get("spreadsheet_url") or result.get("edit_url"),
        "titulo": payload.get("title"),
        "valor": payload.get("valor", 0),
        "folder_id": GOOGLE_PASTA_APROVACAO_ID,
        "folder_link": google_link_pasta(GOOGLE_PASTA_APROVACAO_ID),
    }

def google_linha_pedido_por_arquivo(arquivo, status_pasta=""):
    spreadsheet_id = arquivo.get("id", "")
    controle = google_ler_controle_pedido(spreadsheet_id)
    status = controle.get("status", "") or status_pasta or ""
    valor = controle.get("valor_do_pedido", controle.get("valor", ""))
    pedido_id = controle.get("id_pedido", "") or spreadsheet_id
    fornecedor = controle.get("fornecedor", "")
    nome = arquivo.get("name", "")
    return {
        "id_pedido": pedido_id,
        "nome_pedido": nome,
        "fornecedor": fornecedor,
        "status": status,
        "valor": valor,
        "criado_em": controle.get("criado_em", arquivo.get("createdTime", "")),
        "criado_por": controle.get("criado_por", ""),
        "aprovado_em": controle.get("aprovado_em", ""),
        "aprovado_por": controle.get("aprovado_por", ""),
        "link_pedido": arquivo.get("webViewLink", google_link_planilha(spreadsheet_id)),
        "spreadsheet_id": spreadsheet_id,
        "link_autcom": "",
        "link_fornecedor": "",
        "observacao": controle.get("observacao", ""),
    }


def google_listar_pedidos():
    drive_service, _, _ = google_get_services()
    linhas = []
    for arq in google_listar_planilhas_pasta(drive_service, GOOGLE_PASTA_APROVACAO_ID):
        linhas.append(google_linha_pedido_por_arquivo(arq, "Aguardando aprovação"))
    for arq in google_listar_planilhas_pasta(drive_service, GOOGLE_PASTA_APROVADOS_ID):
        linhas.append(google_linha_pedido_por_arquivo(arq, "Aprovado"))
    df = pd.DataFrame(linhas)
    for col in GOOGLE_PEDIDOS_COLUNAS:
        if col not in df.columns:
            df[col] = ""
    if not df.empty:
        df = df.sort_values("criado_em", ascending=False)
    return df[GOOGLE_PEDIDOS_COLUNAS]


def google_salvar_pedidos_controle(df):
    # Mantido por compatibilidade. O controle agora fica dentro de cada planilha de pedido.
    return None


def google_mover_arquivo(file_id, origem_folder_id, destino_folder_id):
    drive_service, _, _ = google_get_services()
    return drive_service.files().update(
        fileId=file_id,
        addParents=destino_folder_id,
        removeParents=origem_folder_id,
        fields="id, parents, webViewLink",
        supportsAllDrives=True,
    ).execute()


def google_sincronizar_aprovacoes(usuario=""):
    drive_service, _, _ = google_get_services()
    movidos = []
    ignorados = []
    arquivos = google_listar_planilhas_pasta(drive_service, GOOGLE_PASTA_APROVACAO_ID)
    agora = datetime.now().strftime("%d/%m/%Y %H:%M")

    for arq in arquivos:
        spreadsheet_id = arq.get("id")
        controle = google_ler_controle_pedido(spreadsheet_id)
        status = str(controle.get("status", "")).strip().lower()
        status_norm = unicodedata.normalize("NFKD", status).encode("ascii", "ignore").decode("ascii")
        if status_norm == "aprovado":
            aprovado_por = controle.get("aprovado_por", "") or usuario
            aprovado_em = controle.get("aprovado_em", "") or agora
            google_escrever_controle_pedido(
                spreadsheet_id,
                status="Aprovado",
                aprovado_por=aprovado_por,
                aprovado_em=aprovado_em,
            )
            google_mover_arquivo(spreadsheet_id, GOOGLE_PASTA_APROVACAO_ID, GOOGLE_PASTA_APROVADOS_ID)
            movidos.append(arq.get("name", spreadsheet_id))
        else:
            ignorados.append({"arquivo": arq.get("name", spreadsheet_id), "status": controle.get("status", "")})

    return movidos, ignorados


def google_atualizar_status_pedido(pedido_id, status, usuario="", observacao="", link_autcom="", link_fornecedor=""):
    pedidos = google_listar_pedidos()
    mask = pedidos["id_pedido"].astype(str) == str(pedido_id)
    if not mask.any():
        raise ValueError("Pedido nao encontrado no Drive.")
    row = pedidos[mask].iloc[0].to_dict()
    spreadsheet_id = row.get("spreadsheet_id")
    agora = datetime.now().strftime("%d/%m/%Y %H:%M")
    kwargs = {"status": status, "observacao": observacao}
    if status.lower().startswith("aprov"):
        kwargs["aprovado_por"] = usuario
        kwargs["aprovado_em"] = agora
    google_escrever_controle_pedido(spreadsheet_id, **kwargs)

    if status.lower().startswith("aprov"):
        try:
            google_mover_arquivo(spreadsheet_id, GOOGLE_PASTA_APROVACAO_ID, GOOGLE_PASTA_APROVADOS_ID)
        except Exception:
            pass
    row.update(kwargs)
    return row


def google_ler_pedido_drive(spreadsheet_id):
    _, sheets_service, _ = google_get_services()
    df = google_read_df(sheets_service, spreadsheet_id, "Pedido")
    df.columns = [str(c).strip() for c in df.columns]
    return df


def google_upload_bytes(nome_arquivo, dados, mime_type, folder_id):
    drive_service, _, _ = google_get_services()
    media = MediaIoBaseUpload(BytesIO(dados), mimetype=mime_type, resumable=False)
    criado = drive_service.files().create(
        body={"name": nome_arquivo, "parents": [folder_id]},
        media_body=media,
        fields="id, webViewLink",
        supportsAllDrives=True,
    ).execute()
    return criado.get("webViewLink") or google_link_arquivo(criado["id"])


def google_registrar_acompanhamento(pedido_info):
    # Nesta versão, o acompanhamento principal é feito pela leitura das próprias
    # planilhas nas pastas de aprovação/aprovados. Mantido por compatibilidade.
    return None


def google_finalizar_pedido(pedido_id, df_tratamento, usuario=""):
    recursos = google_get_resources()
    pedidos = google_listar_pedidos()
    row = pedidos[pedidos["id_pedido"].astype(str) == str(pedido_id)]
    if row.empty:
        raise ValueError("Pedido nao encontrado.")
    row = row.iloc[0].to_dict()
    base_nome = google_safe_name(row.get("nome_pedido") or "pedido")
    autcom_bytes = gerar_excel_autcom_tratamento(df_tratamento)
    fornecedor_bytes = gerar_excel_fornecedor_tratamento(df_tratamento)
    link_autcom = google_upload_bytes(
        f"{base_nome} - importacao autcom.xlsx",
        autcom_bytes,
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        recursos["finais_folder_id"],
    )
    link_fornecedor = google_upload_bytes(
        f"{base_nome} - envio fornecedor.xlsx",
        fornecedor_bytes,
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        recursos["finais_folder_id"],
    )
    atualizado = google_atualizar_status_pedido(
        pedido_id,
        "Finalizado",
        usuario=usuario,
        link_autcom=link_autcom,
        link_fornecedor=link_fornecedor,
    )
    google_registrar_acompanhamento(atualizado)
    return link_autcom, link_fornecedor


def gerar_excel_pedido_editavel(df):
    """
    Gera uma planilha Excel editável do pedido.
    Recursos aplicados:
    - Valor Final do Pedido formulado: PEDIDO Final x Preço Última Compra.
    - Painéis congelados para facilitar navegação.
    - Coluna Total Geral do Pedido ao lado do Valor Final, pintada em amarelo.

    Observação: CSV não suporta fórmulas, congelamento de painéis nem pintura de células.
    Por isso este download é gerado em .xlsx.
    """
    if Workbook is None:
        raise RuntimeError("A biblioteca openpyxl não está instalada. Rode: python -m pip install openpyxl")

    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    df_export = df.copy()

    colunas = list(df_export.columns)
    if "Valor Final do Pedido" not in colunas:
        df_export["Valor Final do Pedido"] = 0
        colunas = list(df_export.columns)

    # Garante que a coluna Total fique exatamente ao lado de Valor Final do Pedido.
    if "Total Geral do Pedido" in df_export.columns:
        df_export = df_export.drop(columns=["Total Geral do Pedido"])

    pos_valor = list(df_export.columns).index("Valor Final do Pedido")
    colunas = list(df_export.columns)
    colunas.insert(pos_valor + 1, "Total Geral do Pedido")
    df_export["Total Geral do Pedido"] = ""
    df_export = df_export[colunas]

    wb = Workbook()
    ws = wb.active
    ws.title = "Pedido Editável"

    header_fill = PatternFill("solid", fgColor="D9EAF7")
    total_fill = PatternFill("solid", fgColor="FFF2CC")
    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col_idx, col_name in enumerate(df_export.columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = Font(bold=True)
        cell.fill = total_fill if col_name == "Total Geral do Pedido" else header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

    idx_pedido = df_export.columns.get_loc("PEDIDO Final") + 1 if "PEDIDO Final" in df_export.columns else None
    idx_preco = df_export.columns.get_loc("Preço Última Compra") + 1 if "Preço Última Compra" in df_export.columns else None
    idx_valor = df_export.columns.get_loc("Valor Final do Pedido") + 1 if "Valor Final do Pedido" in df_export.columns else None
    idx_total = df_export.columns.get_loc("Total Geral do Pedido") + 1

    for row_idx, (_, row) in enumerate(df_export.iterrows(), start=2):
        for col_idx, col_name in enumerate(df_export.columns, start=1):
            if col_name == "Valor Final do Pedido" and idx_pedido and idx_preco:
                pedido_col = get_column_letter(idx_pedido)
                preco_col = get_column_letter(idx_preco)
                value = f"={pedido_col}{row_idx}*{preco_col}{row_idx}"
            elif col_name == "Total Geral do Pedido":
                value = ""
            else:
                value = row.get(col_name, "")

            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = border

            if col_name == "Total Geral do Pedido":
                cell.fill = total_fill
            if col_name in ["Preço Última Compra", "Valor Final do Pedido", "Total Geral do Pedido"]:
                cell.number_format = 'R$ #,##0.00'
            elif col_name in ["PEDIDO Final", "Sugestão Sistema", "Sugestão arredondada", "Embalagem"]:
                cell.number_format = '0'
            elif isinstance(value, (int, float)):
                cell.number_format = '#,##0.0'

    ultima_linha = max(ws.max_row, 2)
    valor_col = get_column_letter(idx_valor) if idx_valor else None
    total_col = get_column_letter(idx_total)
    if valor_col:
        ws.cell(row=2, column=idx_total, value=f"=SUM({valor_col}2:{valor_col}{ultima_linha})")
        ws.cell(row=2, column=idx_total).fill = total_fill
        ws.cell(row=2, column=idx_total).font = Font(bold=True)
        ws.cell(row=2, column=idx_total).number_format = 'R$ #,##0.00'

    # Congela cabeçalho e as primeiras colunas de identificação.
    ws.freeze_panes = "C2"
    ws.auto_filter.ref = ws.dimensions

    for col_idx, col_name in enumerate(df_export.columns, start=1):
        letter = get_column_letter(col_idx)
        if col_name == "descricao":
            ws.column_dimensions[letter].width = 42
        elif col_name in ["Código Fábrica", "Data Última Compra", "Origem Sugestão"]:
            ws.column_dimensions[letter].width = 18
        elif col_name in ["Valor Final do Pedido", "Total Geral do Pedido", "Preço Última Compra"]:
            ws.column_dimensions[letter].width = 20
        else:
            ws.column_dimensions[letter].width = max(12, min(22, len(str(col_name)) + 2))

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


def gerar_copia_fornecedor_csv(df):
    if df is None or df.empty:
        fornecedor = pd.DataFrame(columns=["Código Fábrica", "Descrição", "Quantidade"])
    else:
        fornecedor = df.copy()
        fornecedor["PEDIDO Final"] = pd.to_numeric(fornecedor.get("PEDIDO Final", 0), errors="coerce").fillna(0).round(0).astype(int)
        fornecedor = fornecedor[fornecedor["PEDIDO Final"] > 0].copy()
        for col in ["Código Fábrica", "descricao"]:
            if col not in fornecedor.columns:
                fornecedor[col] = ""
        fornecedor = fornecedor[["Código Fábrica", "descricao", "PEDIDO Final"]].rename(columns={
            "descricao": "Descrição",
            "PEDIDO Final": "Quantidade",
        })
    return fornecedor.to_csv(index=False, sep=";", decimal=",", encoding="utf-8-sig").encode("utf-8-sig")


def gerar_excel_pedido(df_pedido):
    """
    Excel para importação no Autcom, sem cabeçalho:
    Coluna B = código
    Coluna F = quantidade
    Coluna H = valor unitário
    """
    if Workbook is None:
        raise RuntimeError("A biblioteca openpyxl não está instalada. Rode: python -m pip install openpyxl")

    wb = Workbook()
    ws = wb.active
    ws.title = "Pedido"

    linha_excel = 1
    for _, row in df_pedido.iterrows():
        qtd = int(round(float(row.get("PEDIDO Final", 0) or 0)))
        if qtd <= 0:
            continue
        ws.cell(row=linha_excel, column=2, value=str(row.get("codigo", "")).zfill(5))
        ws.cell(row=linha_excel, column=6, value=qtd)
        ws.cell(row=linha_excel, column=8, value=round(float(str(row.get("Preço Última Compra", 0)).replace(",", "." ) or 0), 2))
        ws.cell(row=linha_excel, column=8).number_format = '0.00'
        linha_excel += 1

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()



def ler_planilha_tratamento_pedido(uploaded_file):
    """
    Lê a planilha final editável enviada pelo usuário na página Tratamento de Pedido Final.
    Aceita .xlsx, .xls e .csv.
    """
    if uploaded_file is None:
        return pd.DataFrame()

    nome = str(getattr(uploaded_file, "name", "")).lower()

    try:
        uploaded_file.seek(0)
    except Exception:
        pass

    if nome.endswith((".xlsx", ".xls")):
        xls = pd.ExcelFile(uploaded_file)
        aba_pedido = next(
            (aba for aba in xls.sheet_names if normalizar_texto_simples(aba) == "PEDIDO"),
            xls.sheet_names[0] if xls.sheet_names else 0,
        )
        df = pd.read_excel(xls, sheet_name=aba_pedido, dtype=str)
        return aplicar_cabecalho_pedido_unica_sheets(df)

    tentativas = [
        {"sep": ";", "encoding": "utf-8-sig"},
        {"sep": ";", "encoding": "latin1"},
        {"sep": ",", "encoding": "utf-8-sig"},
        {"sep": ",", "encoding": "latin1"},
        {"sep": "\t", "encoding": "utf-8-sig"},
        {"sep": "\t", "encoding": "latin1"},
    ]

    ultimo_erro = None
    for tentativa in tentativas:
        try:
            uploaded_file.seek(0)
            df = pd.read_csv(
                uploaded_file,
                sep=tentativa["sep"],
                encoding=tentativa["encoding"],
                dtype=str,
                engine="python",
                on_bad_lines="skip",
            )
            return aplicar_cabecalho_pedido_unica_sheets(df)
        except Exception as e:
            ultimo_erro = str(e)
            continue

    raise RuntimeError(f"Não consegui ler a planilha enviada. Último erro: {ultimo_erro}")


def extrair_google_sheet_id_e_gid(link):
    link = str(link or "").strip()
    if not link:
        return "", "0"

    match_id = re.search(r"/spreadsheets/d/([a-zA-Z0-9-_]+)", link)
    if not match_id:
        match_id = re.search(r"[?&]id=([a-zA-Z0-9-_]+)", link)
    if not match_id:
        raise ValueError("Link do Google Sheets invalido. Cole o link completo da planilha.")

    parsed = urllib.parse.urlparse(link)
    query = urllib.parse.parse_qs(parsed.query)
    gid = "0"
    if query.get("gid"):
        gid = str(query["gid"][0] or "0")
    else:
        match_gid = re.search(r"(?:#|&)gid=([0-9]+)", link)
        if match_gid:
            gid = match_gid.group(1)

    return match_id.group(1), gid


@st.cache_data(show_spinner=False, ttl=300, max_entries=16)
def buscar_gid_aba_google_sheets(sheet_id, nome_aba="Pedido"):
    url = f"https://docs.google.com/spreadsheets/d/{sheet_id}/edit?usp=sharing"

    try:
        req = urllib.request.Request(url, headers={"User-Agent": "Mozilla/5.0"})
        with urllib.request.urlopen(req, timeout=30) as resp:
            html = resp.read().decode("utf-8", errors="ignore")
    except Exception:
        return ""

    nome_alvo = normalizar_texto_simples(nome_aba)
    padroes = [
        r'\["([^"]+)",\s*(\d+)\s*,',
        r'"name"\s*:\s*"([^"]+)"\s*,\s*"gid"\s*:\s*"?(\d+)"?',
    ]

    for padrao in padroes:
        for nome, gid in re.findall(padrao, html):
            if normalizar_texto_simples(nome) == nome_alvo:
                return str(gid).strip()

    match = re.search(r'Pedido[^0-9]{1,80}(\d{4,})', html, flags=re.IGNORECASE)
    if match:
        return match.group(1)

    return ""


def baixar_csv_google_sheets(sheet_id, gid):
    export_url = f"https://docs.google.com/spreadsheets/d/{sheet_id}/export?format=csv&gid={gid}"
    req = urllib.request.Request(export_url, headers={"User-Agent": "Mozilla/5.0"})
    with urllib.request.urlopen(req, timeout=30) as resp:
        return resp.read()


def baixar_csv_google_sheets_por_aba(sheet_id, nome_aba="Pedido"):
    sheet_name = urllib.parse.quote(str(nome_aba or "Pedido"))
    export_url = f"https://docs.google.com/spreadsheets/d/{sheet_id}/gviz/tq?tqx=out:csv&sheet={sheet_name}"
    req = urllib.request.Request(export_url, headers={"User-Agent": "Mozilla/5.0"})
    with urllib.request.urlopen(req, timeout=30) as resp:
        return resp.read()


def csv_bytes_para_dataframe_raw(conteudo):
    if not conteudo:
        return pd.DataFrame()

    for encoding in ["utf-8-sig", "latin1"]:
        try:
            texto = bytes(conteudo).decode(encoding)
            break
        except Exception:
            texto = ""
    if not texto:
        return pd.DataFrame()

    linhas = list(csv.reader(StringIO(texto)))
    if not linhas:
        return pd.DataFrame()

    max_cols = max(len(linha) for linha in linhas)
    linhas = [linha + [""] * (max_cols - len(linha)) for linha in linhas]
    return pd.DataFrame(linhas)


def detectar_linha_cabecalho_pedido(df_raw, max_linhas=50):
    if df_raw is None or df_raw.empty:
        return None

    limite = min(int(max_linhas), len(df_raw))
    melhor_idx = None
    melhor_score = -1

    termos_fortes = [
        "CODIGO", "CÓDIGO", "DESCRICAO", "DESCRIÇÃO", "PEDIDO FINAL",
        "PREÇO ÚLTIMA COMPRA", "PRECO ULTIMA COMPRA", "DATA ÚLTIMA COMPRA",
        "DATA ULTIMA COMPRA", "VALOR FINAL DO PEDIDO",
    ]

    for idx in range(limite):
        valores = [str(v or "").strip() for v in df_raw.iloc[idx].tolist()]
        norm = [normalizar_coluna(v) for v in valores if str(v or "").strip()]
        if not norm:
            continue

        tem_descricao = any(v in ["DESCRICAO", "DESCRIÇÃO"] for v in norm)
        tem_codigo = any(v in ["CODIGO", "CÓDIGO", "ZX"] for v in norm)
        tem_pedido = "PEDIDO FINAL" in norm
        tem_preco = any(v in ["PREÇO ÚLTIMA COMPRA", "PRECO ULTIMA COMPRA"] for v in norm)
        score = sum(1 for termo in termos_fortes if termo in norm)

        if tem_descricao and (tem_codigo or tem_pedido or tem_preco):
            score += 10
        if tem_pedido and tem_preco:
            score += 5

        if score > melhor_score:
            melhor_idx = idx
            melhor_score = score

    return melhor_idx if melhor_score >= 12 else None


def ler_csv_pedido_google_sheets(conteudo):
    if not conteudo:
        return pd.DataFrame()

    df_raw = csv_bytes_para_dataframe_raw(conteudo)
    idx_header = detectar_linha_cabecalho_pedido(df_raw)

    if idx_header is not None:
        headers = [str(v or "").strip() for v in df_raw.iloc[idx_header].tolist()]
        df = df_raw.iloc[idx_header + 1:].copy()
        df.columns = _deduplicar_headers_planilha(headers)
        df = df.dropna(how="all")
        df = df.loc[:, [str(c).strip() != "" for c in df.columns]]
        df = df[~df.apply(lambda row: all(str(v or "").strip() == "" for v in row), axis=1)]
        return aplicar_cabecalho_pedido_unica_sheets(df)

    df = pd.read_csv(BytesIO(conteudo), dtype=str, keep_default_na=False)
    return aplicar_cabecalho_pedido_unica_sheets(df)


@st.cache_data(show_spinner=False, ttl=300, max_entries=16)
def ler_planilha_tratamento_google_sheets_cached(link, versao_leitura="menu-pedido-v4"):
    _ = versao_leitura
    sheet_id, gid_link = extrair_google_sheet_id_e_gid(link)
    gid_pedido = buscar_gid_aba_google_sheets(sheet_id, "Pedido")
    gids_tentativa = [g for g in [gid_pedido, gid_link, "0"] if str(g or "").strip()]
    gids_tentativa = list(dict.fromkeys(gids_tentativa))

    ultimo_erro = None
    try:
        try:
            conteudo = baixar_csv_google_sheets_por_aba(sheet_id, "Pedido")
            if conteudo:
                df = ler_csv_pedido_google_sheets(conteudo)
                if planilha_tratamento_tem_colunas_obrigatorias(df):
                    return df
        except Exception as e:
            ultimo_erro = str(e)

        for gid in gids_tentativa:
            conteudo = baixar_csv_google_sheets(sheet_id, gid)
            if not conteudo:
                continue

            df = ler_csv_pedido_google_sheets(conteudo)
            if planilha_tratamento_tem_colunas_obrigatorias(df):
                return df

        return pd.DataFrame()
    except urllib.error.HTTPError as e:
        if e.code in (401, 403, 404):
            raise RuntimeError(
                "Nao consegui acessar a planilha sem credenciais. "
                "Compartilhe a planilha como 'Qualquer pessoa com o link - Leitor' e tente novamente."
            ) from e
        ultimo_erro = f"HTTP {e.code}"
    except Exception as e:
        ultimo_erro = str(e)

    raise RuntimeError(f"Erro ao baixar a aba Pedido do Google Sheets: {ultimo_erro}")


def ler_planilha_tratamento_google_sheets(link):
    return ler_planilha_tratamento_google_sheets_cached(str(link or "").strip())


@st.cache_data(show_spinner=False, ttl=300, max_entries=16)
def ler_planilha_google_sheets_aba_cached(link, nome_aba="Pedido"):
    sheet_id, gid_link = extrair_google_sheet_id_e_gid(link)
    gid_aba = buscar_gid_aba_google_sheets(sheet_id, nome_aba)
    gids_tentativa = [g for g in [gid_aba, gid_link, "0"] if str(g or "").strip()]
    gids_tentativa = list(dict.fromkeys(gids_tentativa))

    ultimo_erro = None
    try:
        try:
            conteudo = baixar_csv_google_sheets_por_aba(sheet_id, nome_aba)
            if conteudo:
                df = pd.read_csv(BytesIO(conteudo), dtype=str, keep_default_na=False)
                if not df.empty:
                    return df
        except Exception as e:
            ultimo_erro = str(e)

        for gid in gids_tentativa:
            conteudo = baixar_csv_google_sheets(sheet_id, gid)
            if not conteudo:
                continue
            df = pd.read_csv(BytesIO(conteudo), dtype=str, keep_default_na=False)
            if not df.empty:
                return df
    except urllib.error.HTTPError as e:
        if e.code in (401, 403, 404):
            raise RuntimeError(
                "Nao consegui acessar a planilha sem credenciais. "
                "Compartilhe a planilha como 'Qualquer pessoa com o link - Leitor' e tente novamente."
            ) from e
        ultimo_erro = f"HTTP {e.code}"
    except Exception as e:
        ultimo_erro = str(e)

    raise RuntimeError(f"Erro ao baixar a aba {nome_aba} do Google Sheets: {ultimo_erro}")


def ler_pedido_unica_comparativo_google_sheets(link):
    sheet_id, _gid = extrair_google_sheet_id_e_gid(str(link or "").strip())
    try:
        conteudo = baixar_csv_google_sheets_por_aba(sheet_id, "Pedido")
    except urllib.error.HTTPError as e:
        if e.code in (401, 403, 404):
            raise RuntimeError(
                "Nao consegui acessar a aba Pedido sem credenciais. "
                "Compartilhe a planilha como 'Qualquer pessoa com o link - Leitor' e tente novamente."
            ) from e
        raise RuntimeError(f"Erro ao baixar a aba Pedido do Google Sheets: HTTP {e.code}") from e
    except Exception as e:
        raise RuntimeError(f"Erro ao baixar a aba Pedido do Google Sheets: {e}") from e

    df = ler_csv_pedido_google_sheets(conteudo)
    if df is None or df.empty:
        return pd.DataFrame()
    df.columns = [str(c).strip() for c in df.columns]
    return df


def _credenciais_service_account_google():
    """Carrega a conta de serviço configurada no Streamlit Secrets."""
    if service_account is None:
        raise RuntimeError(
            "A biblioteca google-auth não está instalada. Adicione google-auth e google-api-python-client ao requirements.txt."
        )

    cfg = None
    for chave in ("gcp_service_account", "google_service_account", "service_account"):
        try:
            bloco = st.secrets.get(chave, None)
            if bloco:
                cfg = dict(bloco)
                break
        except Exception:
            continue

    if not cfg:
        raise RuntimeError(
            "Não encontrei as credenciais da conta de serviço no Streamlit Secrets. "
            "Configure o JSON em [gcp_service_account]."
        )

    # O Streamlit Secrets/TOML pode entregar a chave com \n literal, aspas extras
    # ou cabeçalho copiado com underscores. Normaliza antes de criar o PEM.
    private_key = str(cfg.get("private_key", "") or "").strip()
    if not private_key:
        raise RuntimeError(
            "A credencial gcp_service_account não possui o campo private_key. "
            "Cole a chave privada completa do arquivo JSON da conta de serviço."
        )

    # Remove aspas que eventualmente tenham sido copiadas como parte do valor.
    if len(private_key) >= 2 and private_key[0] == private_key[-1] and private_key[0] in ('\"', "'"):
        private_key = private_key[1:-1].strip()

    # Converte sequências escapadas do TOML/JSON em quebras de linha reais.
    private_key = private_key.replace("\\r\\n", "\n").replace("\\n", "\n").replace("\r\n", "\n")

    # Reconstrói o envelope PEM de maneira canônica. Isso corrige variações como:
    # ---_BEGIN_PRIVATE_KEY---, BEGIN_PRIVATE_KEY, excesso de hífens e espaços.
    # Somente o cabeçalho/rodapé são normalizados; o corpo criptográfico é preservado.
    texto_pem = private_key.strip()
    match_inicio = re.search(r"BEGIN[\s_-]*(RSA[\s_-]*)?PRIVATE[\s_-]*KEY", texto_pem, flags=re.IGNORECASE)
    match_fim = re.search(r"END[\s_-]*(RSA[\s_-]*)?PRIVATE[\s_-]*KEY", texto_pem, flags=re.IGNORECASE)

    if match_inicio and match_fim and match_fim.start() > match_inicio.end():
        tipo_rsa = bool(match_inicio.group(1))
        header = "-----BEGIN RSA PRIVATE KEY-----" if tipo_rsa else "-----BEGIN PRIVATE KEY-----"
        footer = "-----END RSA PRIVATE KEY-----" if tipo_rsa else "-----END PRIVATE KEY-----"

        corpo = texto_pem[match_inicio.end():match_fim.start()]
        # Remove apenas separadores/ruídos nas bordas do corpo. Não substitui caracteres internos.
        corpo = corpo.strip(" \t\r\n-_:;,'\"")
        corpo = corpo.replace("\r\n", "\n").replace("\r", "\n")
        linhas_corpo = [linha.strip() for linha in corpo.split("\n") if linha.strip()]

        # Se a chave veio em uma única linha, remove espaços e refaz linhas PEM de 64 caracteres.
        if len(linhas_corpo) <= 1:
            corpo_unico = re.sub(r"\s+", "", corpo)
            linhas_corpo = [corpo_unico[i:i + 64] for i in range(0, len(corpo_unico), 64)]

        private_key = header + "\n" + "\n".join(linhas_corpo) + "\n" + footer + "\n"
    else:
        # Fallback para chaves que já estejam quase corretas.
        private_key = private_key.replace("BEGIN_PRIVATE_KEY", "BEGIN PRIVATE KEY")
        private_key = private_key.replace("END_PRIVATE_KEY", "END PRIVATE KEY")
        private_key = private_key.replace("---_BEGIN", "-----BEGIN").replace("---_END", "-----END")
        if not private_key.endswith("\n"):
            private_key += "\n"

    cfg["private_key"] = private_key

    try:
        return service_account.Credentials.from_service_account_info(
            cfg,
            scopes=[
                "https://www.googleapis.com/auth/spreadsheets.readonly",
                "https://www.googleapis.com/auth/drive.readonly",
            ],
        )
    except Exception as e:
        primeira_linha = str(cfg.get("private_key", "")).splitlines()[0][:80] if cfg.get("private_key") else "vazia"
        raise RuntimeError(
            "Não foi possível carregar a chave privada da conta de serviço. "
            "A chave ainda não possui um PEM criptograficamente válido. "
            f"Primeira linha detectada após normalização: {primeira_linha!r}. "
            "O correto é '-----BEGIN PRIVATE KEY-----'. "
            f"Detalhe técnico: {e}"
        ) from e


@st.cache_data(show_spinner=False, ttl=300, max_entries=4)
def ler_tabela_precos_brasilux_google_sheets(
    spreadsheet_id=GOOGLE_PLANILHA_PRECOS_BRASILUX_ID,
    nome_aba="BRASILUX",
):
    """
    Lê especificamente a aba BRASILUX da tabela pública diretamente como CSV.

    Não utiliza conta de serviço, private_key, PEM nem Google API.
    A planilha precisa estar compartilhada como "Qualquer pessoa com o link".
    A coluna D é preservada por posição e os endpoints do Google retornam
    o valor calculado das fórmulas, não o texto da fórmula.
    """
    spreadsheet_id = str(spreadsheet_id or "").strip()
    nome_aba = str(nome_aba or "BRASILUX").strip() or "BRASILUX"
    if not spreadsheet_id:
        raise ValueError("O ID da planilha Brasilux não foi configurado.")

    # IMPORTANTE: sem o parâmetro sheet, o Google devolve a primeira aba visível,
    # que pode não ser a aba BRASILUX. Isso fazia poucos códigos serem relacionados.
    aba_url = urllib.parse.quote(nome_aba, safe="")
    urls = [
        f"https://docs.google.com/spreadsheets/d/{spreadsheet_id}/gviz/tq?tqx=out:csv&sheet={aba_url}",
        f"https://docs.google.com/spreadsheets/d/{spreadsheet_id}/export?format=csv&sheet={aba_url}",
    ]

    ultimo_erro = None
    for url in urls:
        try:
            req = urllib.request.Request(
                url,
                headers={"User-Agent": "Mozilla/5.0"},
            )
            with urllib.request.urlopen(req, timeout=30) as resposta:
                conteudo = resposta.read()

            if not conteudo:
                ultimo_erro = "O Google Sheets retornou um arquivo vazio."
                continue

            df = None
            for encoding in ("utf-8-sig", "utf-8", "latin1"):
                try:
                    df = pd.read_csv(
                        BytesIO(conteudo),
                        header=None,
                        dtype=str,
                        encoding=encoding,
                        keep_default_na=False,
                    )
                    break
                except Exception:
                    continue

            if df is None or df.empty:
                ultimo_erro = "A planilha foi acessada, mas não retornou linhas válidas."
                continue

            # Garante nomes posicionais previsíveis: COL_1=A, COL_2=B, etc.
            df.columns = [f"COL_{i + 1}" for i in range(df.shape[1])]

            # Confirma que a aba retornada possui ao menos A:D. A coluna D pode ser
            # formulada; o CSV público traz o resultado exibido na célula.
            if df.shape[1] < 4:
                ultimo_erro = (
                    f"A aba {nome_aba!r} foi acessada, mas não possui pelo menos quatro colunas (A:D)."
                )
                continue
            return df

        except urllib.error.HTTPError as e:
            if e.code in (401, 403):
                ultimo_erro = (
                    "A planilha não está acessível publicamente. "
                    "Em Compartilhar, mantenha 'Qualquer pessoa com o link' como Leitor ou Editor."
                )
            else:
                ultimo_erro = f"HTTP {e.code} ao baixar a planilha."
        except Exception as e:
            ultimo_erro = str(e)

    raise RuntimeError(
        f"Não consegui carregar a aba {nome_aba!r} da tabela Brasilux pelo link público. "
        f"Detalhe: {ultimo_erro or 'erro não identificado'}"
    )


def montar_mapa_precos_brasilux(df_tabela, codigos_referencia):
    """
    Identifica automaticamente, entre as colunas A, B e C, qual contém o código
    de fábrica com maior coincidência com o Pedido Única. O preço sempre vem da coluna D.
    """
    if df_tabela is None or df_tabela.empty or df_tabela.shape[1] < 4:
        raise ValueError("A tabela Brasilux precisa possuir pelo menos quatro colunas; o preço deve estar na coluna D.")

    referencias = {normalizar_codigo_fabrica(c) for c in (codigos_referencia or [])}
    referencias.discard("")
    if not referencias:
        raise ValueError("Não encontrei códigos de fábrica válidos no Pedido Única para relacionar com a tabela Brasilux.")

    candidatos = list(df_tabela.columns[:3])
    melhor_coluna = None
    melhor_pontuacao = -1
    for coluna in candidatos:
        codigos_coluna = df_tabela[coluna].astype(str).map(normalizar_codigo_fabrica)
        pontuacao = int(codigos_coluna.isin(referencias).sum())
        if pontuacao > melhor_pontuacao:
            melhor_pontuacao = pontuacao
            melhor_coluna = coluna

    if melhor_coluna is None or melhor_pontuacao <= 0:
        raise ValueError(
            "Não consegui relacionar os códigos do Pedido Única com as colunas A, B ou C da tabela Brasilux."
        )

    coluna_preco = df_tabela.columns[3]  # Coluna D obrigatória
    base = pd.DataFrame({
        "codigo_fabrica_norm": df_tabela[melhor_coluna].astype(str).map(normalizar_codigo_fabrica),
        "preco_brasilux": df_tabela[coluna_preco].apply(numero_planilha_para_float),
    })
    base = base[
        base["codigo_fabrica_norm"].isin(referencias)
        & (base["preco_brasilux"] > 0)
    ].copy()
    if base.empty:
        raise ValueError("Nenhum preço válido da coluna D foi encontrado para os códigos do Pedido Única.")

    # Em eventual código repetido, utiliza a última ocorrência válida da tabela.
    mapa = base.drop_duplicates("codigo_fabrica_norm", keep="last").set_index("codigo_fabrica_norm")["preco_brasilux"].to_dict()
    return mapa, melhor_coluna, melhor_pontuacao


def aplicar_precos_brasilux_no_pedido_unica(unica, mapa_precos):
    """
    No modo Brasilux, substitui exclusivamente preço e valor da base Única.
    A quantidade permanece exatamente igual à quantidade do Pedido Única.
    Itens sem preço na tabela ficam com preço zero, sem reutilizar o preço do Pedido Única.
    """
    unica = unica.copy()
    unica["preco_tabela_brasilux_encontrado"] = unica["codigo_fabrica_norm"].map(mapa_precos)
    unica["preco_unitario"] = pd.to_numeric(
        unica["preco_tabela_brasilux_encontrado"], errors="coerce"
    ).fillna(0.0)
    unica["valor_total"] = unica["quantidade"] * unica["preco_unitario"]
    return unica.drop(columns=["preco_tabela_brasilux_encontrado"], errors="ignore")


def gerar_excel_autcom_tratamento(df_tratamento):
    """
    Gera o Excel para importação no Autcom a partir da planilha de Tratamento de Pedido Final.
    Sem cabeçalho:
    - Coluna B = código da coluna zx
    - Coluna F = quantidade da coluna PEDIDO Final
    - Coluna H = preço da coluna Preço Última Compra
    """
    if Workbook is None:
        raise RuntimeError("A biblioteca openpyxl não está instalada. Rode: python -m pip install openpyxl")

    df = aplicar_cabecalho_pedido_unica_sheets(df_tratamento)
    df.columns = [str(c).strip() for c in df.columns]

    colunas_norm = {normalizar_coluna(c): c for c in df.columns}

    col_codigo = colunas_norm.get("ZX") or colunas_norm.get("CODIGO")
    col_qtd = colunas_norm.get("PEDIDO FINAL")
    col_preco = colunas_norm.get("PRECO ULTIMA COMPRA")

    faltantes = []
    if not col_codigo:
        faltantes.append("zx/codigo")
    if not col_qtd:
        faltantes.append("PEDIDO Final")
    if not col_preco:
        faltantes.append("Preço Última Compra")

    if faltantes:
        raise ValueError("A planilha enviada não possui as colunas obrigatórias: " + ", ".join(faltantes))

    wb = Workbook()
    ws = wb.active
    ws.title = "Pedido"

    linha_excel = 1
    for _, row in df.iterrows():
        codigo_raw = str(row.get(col_codigo, "")).strip()
        codigo_match = re.search(r"(\d+)", codigo_raw)
        codigo = codigo_match.group(1).zfill(5) if codigo_match else ""

        qtd = numero_planilha_para_float(row.get(col_qtd, 0))
        preco = numero_planilha_para_float(row.get(col_preco, 0))

        try:
            qtd = int(round(float(qtd)))
        except Exception:
            qtd = 0

        if not codigo or qtd <= 0:
            continue

        ws.cell(row=linha_excel, column=2, value=codigo)
        ws.cell(row=linha_excel, column=6, value=qtd)
        ws.cell(row=linha_excel, column=8, value=round(float(preco or 0), 2))
        ws.cell(row=linha_excel, column=8).number_format = '0.00'
        linha_excel += 1

    if linha_excel == 1:
        raise ValueError(
            "Nenhum item com PEDIDO Final maior que zero foi encontrado para exportar. "
            "Confira se a aba Pedido possui quantidades preenchidas na coluna PEDIDO Final."
        )

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()

def gerar_excel_fornecedor_tratamento(df_tratamento):
    """
    Gera Excel para envio ao fornecedor a partir da planilha de Tratamento Final:
    A = codigo, B = descricao, C = Codigo de fabrica, D = quantidade.
    """
    if Workbook is None:
        raise RuntimeError("A biblioteca openpyxl não está instalada. Rode: python -m pip install openpyxl")

    df = aplicar_cabecalho_pedido_unica_sheets(df_tratamento)
    df.columns = [str(c).strip() for c in df.columns]
    colunas_norm = {normalizar_coluna(c): c for c in df.columns}

    col_codigo = colunas_norm.get("ZX") or colunas_norm.get("CODIGO")
    col_descricao = colunas_norm.get("DESCRICAO") or colunas_norm.get("DESCRICAO DO ITEM")
    col_fabrica = (
        colunas_norm.get("CODIGO FABRICA") or colunas_norm.get("COD FABRICA") or
        colunas_norm.get("CODIGO DE FABRICA") or colunas_norm.get("NOVO CODIGO DE FABRICA") or
        colunas_norm.get("CODIGO_FABRICA")
    )
    col_qtd = colunas_norm.get("PEDIDO FINAL") or colunas_norm.get("QUANTIDADE") or colunas_norm.get("QTD") or colunas_norm.get("QTDE")

    faltantes = []
    if not col_codigo:
        faltantes.append("codigo/zx")
    if not col_descricao:
        faltantes.append("descricao")
    if not col_fabrica:
        faltantes.append("Código Fábrica")
    if not col_qtd:
        faltantes.append("PEDIDO Final/Quantidade")
    if faltantes:
        raise ValueError("A planilha enviada não possui as colunas obrigatórias para fornecedor: " + ", ".join(faltantes))

    wb = Workbook()
    ws = wb.active
    ws.title = "Fornecedor"
    ws.append(["Código", "Descrição", "Código de Fábrica", "Quantidade"])

    for _, row in df.iterrows():
        qtd = numero_planilha_para_float(row.get(col_qtd, 0))
        try:
            qtd = int(round(float(qtd)))
        except Exception:
            qtd = 0
        if qtd <= 0:
            continue

        codigo_raw = str(row.get(col_codigo, "")).strip()
        codigo_match = re.search(r"(\d+)", codigo_raw)
        codigo = codigo_match.group(1).zfill(5) if codigo_match else codigo_raw

        ws.append([
            codigo,
            str(row.get(col_descricao, "")).strip(),
            str(row.get(col_fabrica, "")).strip(),
            qtd,
        ])

    if ws.max_row == 1:
        raise ValueError(
            "Nenhum item com PEDIDO Final/Quantidade maior que zero foi encontrado para exportar. "
            "Confira se a aba Pedido possui quantidades preenchidas."
        )

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    larguras = {"A": 14, "B": 48, "C": 22, "D": 14}
    for col, largura in larguras.items():
        ws.column_dimensions[col].width = largura
    for cell in ws[1]:
        cell.font = cell.font.copy(bold=True)
    for cell in ws["D"]:
        cell.number_format = '0'

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()



def _coluna_por_candidatos(df, candidatos):
    colunas_norm = {normalizar_coluna(c): c for c in df.columns}
    for candidato in candidatos:
        col = colunas_norm.get(normalizar_coluna(candidato))
        if col:
            return col
    return None


def _serie_parece_data(series):
    if series is None:
        return False
    valores = [str(v or "").strip() for v in series.head(60).tolist()]
    valores = [v for v in valores if v and v.lower() not in ["nan", "none", "-"]]
    if not valores:
        return False
    qtd_datas = sum(bool(re.match(r"^\d{1,2}/\d{1,2}/\d{2,4}$", v)) for v in valores)
    return qtd_datas >= max(1, int(len(valores) * 0.55))


def _serie_parece_quantidade(series):
    if series is None:
        return False
    valores = [str(v or "").strip() for v in series.head(80).tolist()]
    valores = [v for v in valores if v and v.lower() not in ["nan", "none", "-"]]
    if not valores:
        return False
    if sum(bool(re.match(r"^\d{1,2}/\d{1,2}/\d{2,4}$", v)) for v in valores) >= max(2, int(len(valores) * 0.35)):
        return False
    numericos = [numero_planilha_para_float(v) for v in valores]
    qtd_validos = sum(1 for n in numericos if n >= 0)
    qtd_positivos = sum(1 for n in numericos if n > 0)
    qtd_inteiros = sum(1 for n in numericos if abs(n - round(n)) < 0.0001)

    # Proteção: código de fábrica/EAN/produto costuma ser uma sequência inteira longa.
    # Quantidade real geralmente é pequena/média e não deve ter maioria de valores com 6+ dígitos.
    qtd_codigos_longos = 0
    for v in valores:
        bruto = re.sub(r"\D+", "", str(v or ""))
        if len(bruto) >= 6 and abs(numero_planilha_para_float(v) - round(numero_planilha_para_float(v))) < 0.0001:
            qtd_codigos_longos += 1
    if qtd_codigos_longos >= max(2, int(len(valores) * 0.35)):
        return False

    return qtd_validos >= int(len(valores) * 0.75) and qtd_inteiros >= int(len(valores) * 0.70) and qtd_positivos > 0


def _serie_parece_origem_sugestao(series):
    if series is None:
        return False
    valores = [normalizar_texto_simples(v) for v in series.head(80).tolist()]
    valores = [v for v in valores if v]
    if not valores:
        return False
    qtd_origem = sum(("sugestao" in v or "alterado" in v or "sistema" in v) for v in valores)
    return qtd_origem >= max(3, int(len(valores) * 0.45))


def corrigir_desalinhamento_menu_pedido(df):
    if df is None or df.empty:
        return pd.DataFrame()

    df = df.copy()
    colunas = list(df.columns)
    norm_para_col = {normalizar_coluna(c): c for c in colunas}
    col_pedido = norm_para_col.get("PEDIDO FINAL")
    if not col_pedido or col_pedido not in df.columns or not _serie_parece_data(df[col_pedido]):
        return df

    idx_pedido = colunas.index(col_pedido)
    if idx_pedido + 1 >= len(colunas) or not _serie_parece_quantidade(df.iloc[:, idx_pedido + 1]):
        return df

    col_data_real = col_pedido
    col_pedido_real = colunas[idx_pedido + 1]
    col_origem_real = colunas[idx_pedido + 2] if idx_pedido + 2 < len(colunas) else None
    col_valor_real = colunas[idx_pedido + 3] if idx_pedido + 3 < len(colunas) else None
    col_embalagem_real = colunas[idx_pedido + 4] if idx_pedido + 4 < len(colunas) else None
    col_fabrica_real = colunas[idx_pedido + 5] if idx_pedido + 5 < len(colunas) else None

    df["Data Última Compra"] = df[col_data_real]
    df["PEDIDO Final"] = df[col_pedido_real]
    if col_origem_real is not None:
        df["Origem Sugestão"] = df[col_origem_real]
    if col_valor_real is not None:
        df["Valor Final do Pedido"] = df[col_valor_real]
    if col_embalagem_real is not None:
        df["Embalagem"] = df[col_embalagem_real]
    if col_fabrica_real is not None:
        df["Código Fábrica"] = df[col_fabrica_real]

    return df


CABECALHO_PEDIDO_UNICA_SHEETS = [
    "codigo",
    "descricao",
    "Giro Geral Abr/26",
    "Giro Geral Mai/26",
    "Giro Geral Jun/26",
    "Média Giro Geral",
    "Estoque Lojas",
    "Estoque Única",
    "Estoque Geral",
    "Saldo em Trânsito/ABERTO",
    "Estoque Final",
    "Estoque Alvo",
    "Sugestão Sistema",
    "Sugestão arredondada",
    "Preço Última Compra",
    "Data Última Compra",
    "PEDIDO Final",
    "Origem Sugestão",
    "Valor Final do Pedido",
    "Embalagem",
    "Código Fábrica",
]


def montar_cabecalho_pedido_dinamico(qtd_colunas, linha_menu=None):
    qtd_colunas = int(qtd_colunas or 0)
    if qtd_colunas <= 0:
        return []

    valores_menu = [str(v or "").strip() for v in (linha_menu or [])]
    origem_idx = None
    for i, valor in enumerate(valores_menu):
        if normalizar_coluna(valor) == "ORIGEM SUGESTAO":
            origem_idx = i
            break

    if origem_idx is not None and origem_idx >= 14:
        qtd_giros = max(1, origem_idx - 14)
    else:
        qtd_giros = max(1, qtd_colunas - 19)

    labels_giro = [f"Giro Geral {i + 1}" for i in range(qtd_giros)]
    headers = [
        "codigo",
        "descricao",
        *labels_giro,
        "Média Giro Geral",
        "Estoque Lojas",
        "Estoque Única",
        "Estoque Geral",
        "Saldo em Trânsito/ABERTO",
        "Estoque Final",
        "Estoque Alvo",
        "Sugestão Sistema",
        "Sugestão arredondada",
        "Preço Última Compra",
        "Data Última Compra",
        "PEDIDO Final",
        "Origem Sugestão",
        "Valor Final do Pedido",
        "Embalagem",
        "Código Fábrica",
        "Total Geral do Pedido",
    ]

    while len(headers) < qtd_colunas:
        headers.append(f"COLUNA {len(headers) + 1}")

    return headers[:qtd_colunas]


def aplicar_cabecalho_pedido_unica_sheets(df):
    if df is None or df.empty:
        return pd.DataFrame()

    df = df.copy()
    colunas = [str(c or "").strip() for c in df.columns]
    colunas_norm = [normalizar_coluna(c) for c in colunas]

    tem_codigo = any(c in ["CODIGO"] for c in colunas_norm)
    tem_pedido_final = "PEDIDO FINAL" in colunas_norm
    tem_menu_parcial_sheets = (
        any(c in ["DESCRICAO"] for c in colunas_norm)
        and "ORIGEM SUGESTAO" in colunas_norm
        and not tem_pedido_final
    )
    qtd_sem_nome = sum(
        1 for c in colunas
        if not c or c.lower().startswith("unnamed") or c.upper().startswith("COLUNA ")
    )

    if tem_codigo and tem_pedido_final and qtd_sem_nome <= 2:
        df.columns = [str(c).strip() for c in df.columns]
        return corrigir_desalinhamento_menu_pedido(df)

    if tem_menu_parcial_sheets:
        df.columns = _deduplicar_headers_planilha(montar_cabecalho_pedido_dinamico(len(colunas), colunas))
        return corrigir_desalinhamento_menu_pedido(df)

    novos = []
    for i, col in enumerate(colunas):
        headers_dinamicos = montar_cabecalho_pedido_dinamico(len(colunas), colunas)
        if i < len(headers_dinamicos):
            novos.append(headers_dinamicos[i])
        else:
            novos.append(col if col and not col.lower().startswith("unnamed") else f"COLUNA {i + 1}")

    df.columns = _deduplicar_headers_planilha(novos)
    return corrigir_desalinhamento_menu_pedido(df)


def planilha_tratamento_tem_colunas_obrigatorias(df):
    if df is None or df.empty:
        return False

    colunas_norm = {normalizar_coluna(c): c for c in df.columns}
    tem_codigo = any(col in colunas_norm for col in ["ZX", "CODIGO"])
    tem_pedido_final = "PEDIDO FINAL" in colunas_norm
    tem_preco = any(col in colunas_norm for col in ["PRECO ULTIMA COMPRA"])
    return tem_codigo and tem_pedido_final and tem_preco


def _normalizar_nome_coluna_flex(nome):
    """Normaliza nome de coluna para reconhecimento flexível."""
    txt = _texto_sem_acentos(nome).upper().strip()
    txt = re.sub(r"[^A-Z0-9]+", " ", txt)
    txt = re.sub(r"\s+", " ", txt).strip()
    return txt


def _coluna_quantidade_flexivel(df):
    """
    Encontra coluna de quantidade mesmo quando o fornecedor usa nomes variados.
    Aceita: QTD, QTDE, QUANT, QUANT., QUANTIDADE, QTD PEDIDA, QTDE SOLICITADA,
    QTD FATURADA, QTD COMPRA, QTE, QTY, QUANTITY, VOL, VOLUME etc.
    Evita confundir com preço, valor total, saldo, código, embalagem ou peso.
    """
    if df is None or df.empty:
        return None

    bloqueios = [
        "PRECO", "PREÇO", "VALOR", "VL", "VLR", "UNIT", "UNITARIO", "UNITRIO",
        "TOTAL", "IPI", "SUB", "ST", "COD", "CÓD", "CODIGO", "CÓDIGO",
        "FABRICA", "FBRICA", "REFERENCIA", "REFERÊNCIA", "SKU", "DESCR",
        "PESO", "PES", "KIL", "LIT", "LITRO", "EMB", "EMBALAGEM",
        "ESTOQUE", "SALDO", "BAIXADO", "ABERTO", "DATA", "DT",
    ]

    fortes_exatos = {
        "QTD", "QTDE", "QTE", "QTY", "QUANT", "QUANTIDADE", "QUANTITY",
        "QUANTID", "QUANTIDADE PEDIDA", "QTD PEDIDA", "QTDE PEDIDA",
        "QUANTIDADE SOLICITADA", "QTD SOLICITADA", "QTDE SOLICITADA",
        "QUANTIDADE COMPRA", "QTD COMPRA", "QTDE COMPRA",
        "QUANTIDADE DO PEDIDO", "QTD DO PEDIDO", "QTDE DO PEDIDO",
        "QUANTIDADE PEDIDO", "QTD PEDIDO", "QTDE PEDIDO",
        "PEDIDO FINAL", "QUANTIDADE FATURADA", "QTD FATURADA", "QTDE FATURADA",
        "VOLUME", "VOL",
    }

    # 1) Prioriza nomes exatos normalizados.
    for col in df.columns:
        norm = _normalizar_nome_coluna_flex(col)
        if norm in fortes_exatos:
            return col

    # 2) Depois aceita qualquer coluna que contenha uma palavra clara de quantidade,
    # desde que não contenha termos típicos de preço, código, estoque etc.
    padrao_qtd = re.compile(r"(^| )(QTD|QTDE|QTE|QTY|QUANT|QUANTIDADE|QUANTITY|VOLUME|VOL)( |$)")
    for col in df.columns:
        norm = _normalizar_nome_coluna_flex(col)
        if not norm:
            continue
        if padrao_qtd.search(norm) and not any(b in norm for b in bloqueios):
            return col

    # 3) Caso o nome seja algo como 'QTD_PED', 'QTDE-SOLIC', 'QUANT.' já estará
    # normalizado com espaço. Este fallback pega prefixos comuns.
    for col in df.columns:
        norm = _normalizar_nome_coluna_flex(col)
        compact = re.sub(r"[^A-Z0-9]+", "", norm)
        if any(compact.startswith(p) for p in ["QTD", "QTDE", "QTE", "QTY", "QUANT", "QUANTIDADE"]):
            if not any(b in norm for b in bloqueios):
                return col

    return None


def _coluna_valor_unitario_flexivel(df):
    if df is None or df.empty:
        return None

    candidatos_exatos = {
        "PRECO ULTIMA COMPRA", "PRECO", "PRECO UNITARIO", "VALOR UNITARIO",
        "VLR UNIT", "VL UNIT", "VL UNITARIO", "VR UNIT", "VR UNITARIO",
        "UNITARIO",
    }
    bloqueios = ["IPI", "ST", "QTDE", "QTD", "QUANT", "QUANTIDADE", "COD", "CODIGO", "DESCR"]

    for col in df.columns:
        norm = _normalizar_nome_coluna_flex(col)
        if norm in candidatos_exatos:
            return col

    for col in df.columns:
        norm = _normalizar_nome_coluna_flex(col)
        if not norm:
            continue
        tem_preco = ("PRECO" in norm or "VALOR UNIT" in norm or "VLR UNIT" in norm or "VL UNIT" in norm or "VR UNIT" in norm or "UNITARIO" in norm)
        if tem_preco and not any(b in norm for b in bloqueios):
            if "TOTAL" not in norm and "TOT" not in norm:
                return col

    return None


def _coluna_valor_total_flexivel(df):
    if df is None or df.empty:
        return None

    candidatos_exatos = {
        "VALOR FINAL DO PEDIDO", "VALOR TOTAL", "VL TOTAL", "VLR TOTAL",
        "TOTAL", "TOTAL GERAL", "VALOR MERCADORIA", "VL MERCADORIA",
    }

    for col in df.columns:
        norm = _normalizar_nome_coluna_flex(col)
        if norm in candidatos_exatos:
            return col

    for col in df.columns:
        norm = _normalizar_nome_coluna_flex(col)
        if not norm:
            continue
        if ("TOTAL" in norm or "VALOR MERCADORIA" in norm) and "UNIT" not in norm and "UNITARIO" not in norm:
            return col

    return None




def _texto_sem_acentos(txt):
    txt = str(txt or "")
    txt = unicodedata.normalize("NFKD", txt)
    txt = "".join(ch for ch in txt if not unicodedata.combining(ch))
    return txt


def normalizar_codigo_fabrica(valor):
    """
    Normaliza código de fábrica para comparação entre Excel e PDF.

    Correção importante para PDFs de fornecedores:
    - Alguns PDFs, como propostas Akzo/Coral, trazem o código do produto com zeros à esquerda
      (ex.: 000000000005202143), enquanto a planilha da Única pode estar como 5202143.
    - Para códigos puramente numéricos, remove zeros à esquerda para o relacionamento não falhar.
    - Para códigos alfanuméricos, mantém letras/números e remove apenas pontuação/espaços.

    Exemplos:
    - "000000000005202143" -> "5202143"
    - "5202143" -> "5202143"
    - "I :401" -> "I401"
    - "P-512" -> "P512"
    - "05.66.H0035-261" -> "0566H0035261"
    """
    raw = _texto_sem_acentos(valor).upper().strip()

    # Quando o Excel/Sheets transforma código em número, pode chegar como 5202143.0
    # ou 000000000005202143,0. Trata isso antes de remover pontuação.
    raw_sem_espaco = re.sub(r"\s+", "", raw)
    match_num_decimal_zero = re.fullmatch(r"(\d+)(?:[\.,]0+)?", raw_sem_espaco)
    if match_num_decimal_zero:
        txt = match_num_decimal_zero.group(1)
    else:
        txt = re.sub(r"[^A-Z0-9]+", "", raw)

    if txt in ["", "NAN", "NONE", "NULL", "SEM", "SNCODIGO"]:
        return ""

    # PDFs de fornecedor costumam completar código numérico com zeros à esquerda.
    # Isso quebrava o relacionamento com a planilha da Única.
    if re.fullmatch(r"\d+", txt):
        txt = txt.lstrip("0") or "0"

    return txt


def codigo_fabrica_nucleo_numerico(valor):
    """
    Retorna a parte numérica principal do código de fábrica para relacionamento seguro.

    Exemplos:
    - "AC 470200102" -> "470200102"
    - "470200102" -> "470200102"
    - "TN710041608" -> "710041608"

    O núcleo só é usado como fallback quando ele é único nos dois pedidos. Dessa forma,
    os modos de leitura anteriores continuam preservados e códigos ambíguos não são unidos.
    """
    norm = normalizar_codigo_fabrica(valor)
    if not norm:
        return ""
    m = re.search(r"(\d{4,})$", norm)
    if not m:
        return ""
    return (m.group(1).lstrip("0") or "0")


def normalizar_descricao_chave(valor):
    txt = _texto_sem_acentos(valor).upper().strip()
    txt = re.sub(r"[^A-Z0-9 ]+", " ", txt)
    txt = re.sub(r"\s+", " ", txt).strip()
    return txt


def _tokens_numericos_linha(txt):
    """Extrai números em padrão BR/US preservando a ordem visual da linha."""
    txt = str(txt or "")
    padrao = r"(?<![A-Z0-9])-?(?:R\$\s*)?\d{1,3}(?:\.\d{3})*(?:,\d+)?(?![A-Z0-9])|(?<![A-Z0-9])-?\d+(?:[\.,]\d+)?(?![A-Z0-9])"
    valores = []
    for m in re.finditer(padrao, txt, flags=re.IGNORECASE):
        bruto = m.group(0).replace("R$", "").strip()
        valores.append((m.start(), bruto, numero_planilha_para_float(bruto)))
    return valores


def _valor_parece_quantidade(valor):
    try:
        v = float(valor)
    except Exception:
        return False
    if v <= 0 or v > 100000:
        return False
    return abs(v - round(v)) < 0.0001 or v < 1000


def _inferir_qtd_preco_total_por_numeros(candidatos):
    candidatos = [float(v) for v in candidatos if pd.notna(v) and float(v or 0) > 0]
    if not candidatos:
        return 0.0, 0.0, 0.0

    melhor = None
    melhor_score = -999999.0

    for i, qtd in enumerate(candidatos[:8]):
        if not _valor_parece_quantidade(qtd):
            continue

        depois = candidatos[i + 1:]
        if not depois:
            continue

        # Em pedido de fornecedor, se aparecem unitário sem imposto e valor com imposto,
        # o unitário sem imposto costuma ser o menor valor monetário plausível depois da quantidade.
        for j, preco in enumerate(depois):
            if preco <= 0:
                continue

            totais = depois[j + 1:] or []
            total_calculado = qtd * preco
            total = 0.0
            score = 0.0

            for t in totais:
                if t <= 0:
                    continue
                if abs(total_calculado - t) <= max(0.05, abs(t) * 0.03):
                    total = t
                    score += 60
                    break
                if t >= total_calculado:
                    total = max(total, t)
                    score += 10

            if total <= 0:
                total = total_calculado

            # Prefere quantidade inteira e preços menores quando há unitário com/sem imposto.
            if abs(qtd - round(qtd)) < 0.0001:
                score += 12
            if i <= 2:
                score += 8 - i
            score -= preco * 0.0001
            if total > 0 and preco <= max(total, total_calculado):
                score += 5

            if score > melhor_score:
                melhor_score = score
                melhor = (qtd, preco, total)

    if melhor:
        qtd, preco, total = melhor

        # Segunda passada: se houver outro valor unitário menor logo após a quantidade
        # e antes do total, usa o menor. Ex.: qtd=1, unit=25, total c/imposto=25,90.
        try:
            idx_qtd = candidatos.index(qtd)
        except ValueError:
            idx_qtd = 0
        janela = [v for v in candidatos[idx_qtd + 1:] if v > 0]
        limites = []
        for v in janela:
            if total > 0 and v <= max(total, qtd * preco) and v >= 0.01:
                limites.append(v)
        unitarios = [v for v in limites if qtd <= 1 or v <= (max(total, qtd * preco) / max(qtd, 1)) * 1.35]
        if unitarios:
            preco = min(unitarios)
            if total <= 0:
                total = qtd * preco
        return qtd, preco, total

    qtd = candidatos[0] if len(candidatos) >= 1 else 0.0
    preco = min(candidatos[1:]) if len(candidatos) >= 2 else 0.0
    total = max(candidatos[1:]) if len(candidatos) >= 2 else (qtd * preco if qtd and preco else 0.0)
    return qtd, preco, total


def _inferir_qtd_preco_total_por_linha(linha, pos_codigo=0):
    """
    Heurística para PDFs de fornecedores com layouts variados.
    Após encontrar o código de fábrica na linha, procura quantidade, preço unitário e total.
    Se houver total, tenta validar pares em que qtd x preço ~= total.
    """
    linha = str(linha or "")
    numeros = [(pos, bruto, val) for pos, bruto, val in _tokens_numericos_linha(linha) if pos >= max(0, pos_codigo)]
    numeros = [(pos, bruto, val) for pos, bruto, val in numeros if pd.notna(val) and float(val) != 0]

    if not numeros:
        return 0.0, 0.0, 0.0

    return _inferir_qtd_preco_total_por_numeros([float(v) for _, _, v in numeros])


def _extrair_descricao_ao_redor_codigo(linha, codigo):
    linha = str(linha or "").strip()
    if not linha:
        return ""
    cod = str(codigo or "").strip()
    pos = linha.upper().find(cod.upper()) if cod else -1
    if pos >= 0:
        antes = linha[:pos].strip(" -|;:\t")
        depois = linha[pos + len(cod):].strip(" -|;:\t")
        depois = re.sub(r"\s+(?:R\$\s*)?\d[\d\.,]*.*$", "", depois).strip()
        desc = depois or antes
    else:
        desc = re.sub(r"\s+(?:R\$\s*)?\d[\d\.,]*.*$", "", linha).strip()
    return desc[:180]


def _referencias_codigo_fabrica(codigos_referencia=None):
    referencias = {}
    nucleos = {}
    nucleos_ambiguos = set()

    for c in (codigos_referencia or []):
        original = str(c or "").strip()
        norm = normalizar_codigo_fabrica(original)
        if norm and len(norm) >= 3:
            referencias[norm] = original

        nucleo = codigo_fabrica_nucleo_numerico(original)
        if nucleo:
            anterior = nucleos.get(nucleo)
            if anterior is None:
                nucleos[nucleo] = original
            elif normalizar_codigo_fabrica(anterior) != norm:
                nucleos_ambiguos.add(nucleo)

    # Inclui o número sem o prefixo como alias somente quando ele identifica um único
    # produto na planilha da Única. Isso cobre AC 470200102 x 470200102 sem criar
    # relacionamentos incorretos em códigos repetidos.
    for nucleo, original in nucleos.items():
        if nucleo not in nucleos_ambiguos and nucleo not in referencias:
            referencias[nucleo] = original

    return dict(sorted(referencias.items(), key=lambda kv: len(kv[0]), reverse=True))


def extrair_itens_por_codigos_em_textos(textos, codigos_referencia=None, origem_linha="Linha Fornecedor"):
    referencias = _referencias_codigo_fabrica(codigos_referencia)
    if not referencias:
        return pd.DataFrame()

    registros = []
    vistos = set()

    for linha in textos or []:
        linha = str(linha or "").strip()
        if not linha:
            continue
        linha_norm = normalizar_codigo_fabrica(linha)
        if not linha_norm:
            continue

        for cod_norm, cod_original in referencias.items():
            if cod_norm not in linha_norm:
                continue

            linha_sem_acentos = _texto_sem_acentos(linha).upper()
            cod_sem_acentos = _texto_sem_acentos(cod_original).upper()
            pos_raw = linha_sem_acentos.find(cod_sem_acentos)
            if pos_raw < 0:
                pos_raw = 0

            qtd, preco, total = _inferir_qtd_preco_total_por_linha(linha, pos_codigo=pos_raw + len(str(cod_original)))
            if qtd <= 0 and preco <= 0 and total <= 0:
                continue

            desc = _extrair_descricao_ao_redor_codigo(linha, cod_original)
            chave = (cod_norm, round(qtd, 4), round(preco, 4), round(total, 4), normalizar_descricao_chave(desc))
            if chave in vistos:
                continue
            vistos.add(chave)
            registros.append({
                "Código Fábrica": cod_original,
                "Descrição": desc,
                "Quantidade": qtd,
                "Valor Unitário": preco,
                "Valor Total": total if total > 0 else qtd * preco,
                origem_linha: linha,
                "Linha PDF": linha,
            })

    return pd.DataFrame(registros)


def _celula_parece_numero_planilha(valor):
    txt = str(valor or "").strip()
    if not txt:
        return False
    txt_limpo = txt.replace("R$", "").replace("%", "").strip()
    return bool(re.fullmatch(r"-?[\d\.\,\s]+", txt_limpo))


def _coluna_auxiliar_nao_preco_quantidade(nome_coluna):
    norm = _normalizar_nome_coluna_flex(nome_coluna)
    bloqueios = [
        "LITRO", "LITROS", "LIT", "VOLUME", "VOL", "PESO", "KG", "KIL",
        "ML", "EMB", "EMBALAGEM", "EAN", "BARRA", "BARRAS", "NCM", "IPI",
        "ST", "ICMS", "IMPOSTO", "DESCONTO",
    ]
    return any(b in norm for b in bloqueios)


def dataframe_fornecedor_tem_colunas_confiaveis(df):
    if df is None or df.empty:
        return False

    col_codigo = _coluna_por_candidatos(df, [
        "Código Fábrica", "Codigo Fabrica", "Código", "Codigo", "CÓDIGO", "Cód.",
        "Cod.", "Referência", "Referencia", "SKU", "Código Produto", "Codigo Produto",
    ])
    col_qtd = _coluna_quantidade_flexivel(df)
    col_preco = _coluna_valor_unitario_flexivel(df)

    if not col_codigo or not col_qtd or not col_preco:
        return False
    if _coluna_auxiliar_nao_preco_quantidade(col_qtd) or _coluna_auxiliar_nao_preco_quantidade(col_preco):
        return False
    return _serie_parece_quantidade(df[col_qtd])


def padronizar_dataframe_fornecedor_homologado(df):
    if df is None or df.empty:
        return pd.DataFrame()

    df_tmp = df.copy()
    df_tmp.columns = [str(c).strip() for c in df_tmp.columns]
    col_codigo = _coluna_por_candidatos(df_tmp, [
        "Código Fábrica", "Codigo Fabrica", "Código Produto", "Codigo Produto",
        "Código", "Codigo", "Cód.", "Cod.", "Produto", "SKU", "Referência",
        "Referencia", "Part Number",
    ])
    col_desc = _coluna_por_candidatos(df_tmp, [
        "Descrição", "Descricao", "Descrição do item", "Descricao do item",
        "Produto", "Item", "Nome", "Descr",
    ])
    col_qtd = _coluna_quantidade_flexivel(df_tmp)
    col_preco = _coluna_valor_unitario_flexivel(df_tmp)
    col_total = _coluna_valor_total_flexivel(df_tmp)

    if not col_codigo or not col_qtd or not col_preco:
        return pd.DataFrame()
    if _coluna_auxiliar_nao_preco_quantidade(col_qtd) or _coluna_auxiliar_nao_preco_quantidade(col_preco):
        return pd.DataFrame()
    if not _serie_parece_quantidade(df_tmp[col_qtd]):
        return pd.DataFrame()

    out = pd.DataFrame()
    out["Código Fábrica"] = df_tmp[col_codigo].astype(str).str.strip()
    out["Descrição"] = df_tmp[col_desc].astype(str).str.strip() if col_desc else out["Código Fábrica"]
    out["Quantidade"] = df_tmp[col_qtd].apply(numero_planilha_para_float)
    out["Valor Unitário"] = df_tmp[col_preco].apply(numero_planilha_para_float)
    out["Valor Total"] = df_tmp[col_total].apply(numero_planilha_para_float) if col_total else 0.0
    out["Linha Fornecedor"] = df_tmp.astype(str).agg(" | ".join, axis=1)

    out = out[out["Código Fábrica"].astype(str).str.strip().ne("")]
    out = out[out["Quantidade"] > 0].copy()
    out = out[out["Valor Unitário"] > 0].copy()
    out.loc[(out["Valor Total"] <= 0) & (out["Valor Unitário"] > 0), "Valor Total"] = out["Quantidade"] * out["Valor Unitário"]
    return out.reset_index(drop=True)


def extrair_itens_por_codigos_em_dataframe(df, codigos_referencia=None):
    if df is None or df.empty:
        return pd.DataFrame()

    referencias = _referencias_codigo_fabrica(codigos_referencia)
    if not referencias:
        return pd.DataFrame()

    registros = []
    vistos = set()
    df_tmp = df.copy()
    df_tmp.columns = [str(c).strip() for c in df_tmp.columns]

    for _, row in df_tmp.iterrows():
        valores = [str(v or "").strip() for v in row.tolist()]
        if not any(valores):
            continue

        linha = " | ".join(v for v in valores if v)
        norm_cells = [normalizar_codigo_fabrica(v) for v in valores]
        linha_norm = normalizar_codigo_fabrica(linha)

        for cod_norm, cod_original in referencias.items():
            if cod_norm not in linha_norm:
                continue

            idx_cod = None
            for i, cell_norm in enumerate(norm_cells):
                if cod_norm and cod_norm in cell_norm:
                    idx_cod = i
                    break

            if idx_cod is None:
                idx_cod = 0

            candidatos = []
            for col_nome, cell in zip(list(df_tmp.columns)[idx_cod + 1:], valores[idx_cod + 1:]):
                if _coluna_auxiliar_nao_preco_quantidade(col_nome):
                    continue
                if not _celula_parece_numero_planilha(cell):
                    continue
                nums = _tokens_numericos_linha(cell)
                candidatos.extend(float(v) for _, _, v in nums if pd.notna(v))

            if not candidatos:
                qtd, preco, total = _inferir_qtd_preco_total_por_linha(linha)
            else:
                qtd, preco, total = _inferir_qtd_preco_total_por_numeros(candidatos)

            if qtd <= 0 and preco <= 0 and total <= 0:
                continue

            desc_partes = []
            for i, cell in enumerate(valores):
                if i == idx_cod:
                    continue
                txt = str(cell or "").strip()
                if not txt:
                    continue
                if _tokens_numericos_linha(txt) and len(txt) <= 25:
                    continue
                if re.search(r"[A-Za-zÉÓÚÃÕÇáéíóúãõç]{3,}", txt):
                    desc_partes.append(txt)
            desc = " ".join(desc_partes).strip()[:180] or _extrair_descricao_ao_redor_codigo(linha, cod_original)

            chave = (cod_norm, round(qtd, 4), round(preco, 4), round(total, 4), normalizar_descricao_chave(desc))
            if chave in vistos:
                continue
            vistos.add(chave)
            registros.append({
                "Código Fábrica": cod_original,
                "Descrição": desc,
                "Quantidade": qtd,
                "Valor Unitário": preco,
                "Valor Total": total if total > 0 else qtd * preco,
                "Linha Fornecedor": linha,
                "Linha PDF": linha,
            })

    return pd.DataFrame(registros)


def extrair_itens_pdf_por_codigos(uploaded_file, codigos_referencia=None):
    """
    Lê PDF de fornecedor em vários modelos.
    Estratégia principal: usa os códigos de fábrica do pedido da Única como âncora.
    Assim, mesmo sem cabeçalho ou com layout diferente, o sistema busca o código no texto/tabela
    e tenta capturar quantidade e preço unitário na mesma linha.
    """
    referencias = _referencias_codigo_fabrica(codigos_referencia)

    linhas_texto, _linhas_tabela_pdf = extract_pdf_linhas_e_tabelas(uploaded_file)

    if referencias:
        por_ancora = extrair_itens_por_codigos_em_textos(linhas_texto, codigos_referencia, origem_linha="Linha PDF")
        if por_ancora is not None and not por_ancora.empty:
            return por_ancora

    registros = []
    vistos = set()

    for linha in linhas_texto:
        linha_norm = normalizar_codigo_fabrica(linha)
        encontrados = []

        if referencias:
            for cod_norm, cod_original in referencias.items():
                if cod_norm and cod_norm in linha_norm:
                    encontrados.append((cod_norm, cod_original))
        else:
            # Sem referência, tenta pegar candidatos parecidos com código de fábrica.
            for token in re.findall(r"\b[A-Z]{0,4}\d[A-Z0-9\-\./:]{2,}\b", _texto_sem_acentos(linha).upper()):
                cod_norm = normalizar_codigo_fabrica(token)
                if len(cod_norm) >= 3:
                    encontrados.append((cod_norm, token))

        for cod_norm, cod_original in encontrados:
            pos_raw = _texto_sem_acentos(linha).upper().find(_texto_sem_acentos(cod_original).upper())
            if pos_raw < 0:
                pos_raw = 0
            qtd, preco, total = _inferir_qtd_preco_total_por_linha(linha, pos_codigo=pos_raw + len(str(cod_original)))
            desc = _extrair_descricao_ao_redor_codigo(linha, cod_original)
            chave_visto = (cod_norm, round(qtd, 4), round(preco, 4), round(total, 4), normalizar_descricao_chave(desc))
            if chave_visto in vistos:
                continue
            vistos.add(chave_visto)
            registros.append({
                "Código Fábrica": cod_original,
                "Descrição": desc,
                "Quantidade": qtd,
                "Valor Unitário": preco,
                "Valor Total": total,
                "Linha PDF": linha,
            })

    return pd.DataFrame(registros)



def _particionar_blocos_itens_pdf(texto):
    """
    Divide texto de PDF em blocos de itens quando existe uma linha iniciando com:
    Item Abrev. Unid. Produto...
    Ex.: 1 86007 261 05.66.H003
    """
    linhas = [str(l or "").strip() for l in str(texto or "").splitlines() if str(l or "").strip()]
    blocos = []
    atual = []
    padrao_inicio = re.compile(r"^\d{1,4}\s+\d{3,6}\s+\d{3,6}\s+[A-Z0-9][A-Z0-9\.\-/]*", re.IGNORECASE)

    for linha in linhas:
        if padrao_inicio.match(linha):
            if atual:
                blocos.append(atual)
            atual = [linha]
        elif atual:
            # evita puxar rodapé/cabeçalho grande depois do fim do item
            if re.match(r"^(Página|Pedido:|Cliente:|Item\s+Abrev|Total do Pedido)", linha, flags=re.IGNORECASE):
                if re.match(r"^Total do Pedido", linha, flags=re.IGNORECASE):
                    blocos.append(atual)
                    atual = []
                continue
            atual.append(linha)
    if atual:
        blocos.append(atual)
    return blocos


def _codigo_fabrica_do_bloco_pdf(bloco):
    if not bloco:
        return "", "", ""
    primeira = str(bloco[0]).strip()
    m = re.match(r"^(\d{1,4})\s+(\d{3,6})\s+(\d{3,6})\s+([A-Z0-9][A-Z0-9\.\-/]*)", primeira, flags=re.IGNORECASE)
    if not m:
        return "", "", ""

    item, abrev, unid, produto = m.group(1), m.group(2), m.group(3), m.group(4)
    produto_completo = produto

    # Alguns PDFs quebram o final do código do produto na(s) linha(s) seguinte(s).
    # Ex.: 05.66.H003\n5 => 05.66.H0035; 05.00.M350\n0 => 05.00.M3500.
    for linha in bloco[1:4]:
        token = str(linha).strip()
        if re.fullmatch(r"[A-Z0-9]{1,3}", token, flags=re.IGNORECASE):
            produto_completo += token
            break
        # quando já começou descrição, para
        if re.search(r"[A-ZÉÓÚÃÕÇ]{3,}", token, flags=re.IGNORECASE) and not re.fullmatch(r"[A-Z0-9\.\-/]+", token, flags=re.IGNORECASE):
            break

    codigo_original = f"{produto_completo}-{unid}"
    return codigo_original, normalizar_codigo_fabrica(codigo_original), unid


def _descricao_do_bloco_pdf(bloco):
    if not bloco:
        return ""
    partes = []
    for linha in bloco[1:]:
        t = str(linha).strip()
        if not t:
            continue
        if re.fullmatch(r"[A-Z0-9]{1,3}", t, flags=re.IGNORECASE):
            continue
        if re.fullmatch(r"\d+(?:[\.,]\d+)?", t):
            continue
        # remove a cauda numérica de quantidade/preço/total da última linha
        t = re.sub(r"\s+\d+(?:[\.,]\d+)?\s+\d+(?:[\.,]\d+)?\s+\d{1,3}(?:\.\d{3})*,\d+\s*$", "", t).strip()
        if re.search(r"[A-ZÉÓÚÃÕÇ]{3,}", t, flags=re.IGNORECASE):
            partes.append(t)
    return " ".join(partes).strip()[:180]


def _qtd_preco_total_do_bloco_pdf(bloco):
    texto = " ".join(str(x or "").strip() for x in bloco)
    nums = _tokens_numericos_linha(texto)
    valores = [float(v) for _, _, v in nums if pd.notna(v)]
    if len(valores) < 3:
        return 0.0, 0.0, 0.0

    # Nos pedidos de fornecedor, normalmente os 3 últimos números do bloco são:
    # quantidade, preço unitário e preço total. Valida qtd x preço ~= total.
    for i in range(len(valores) - 3, -1, -1):
        qtd, preco, total = valores[i], valores[i + 1], valores[i + 2]
        if qtd > 0 and preco > 0 and total > 0:
            if abs(qtd * preco - total) <= max(0.10, abs(total) * 0.03):
                return qtd, preco, total
    qtd, preco, total = valores[-3], valores[-2], valores[-1]
    return qtd, preco, total


def extrair_itens_pdf_por_blocos(uploaded_file, codigos_referencia=None):
    """
    Parser complementar para PDFs em que as colunas ficam visualmente desalinhadas.
    Exemplo real Sherwin/Lazzuril:
    - Excel: 05.66.H0035-261
    - PDF: linha do produto 05.66.H003 / linha do item com UNID 261 / linha seguinte 5 0
    O parser monta Código de Fábrica como Produto-Unid e valida contra os códigos do Excel.
    """
    referencias = set()
    for c in (codigos_referencia or []):
        norm = normalizar_codigo_fabrica(c)
        if norm:
            referencias.add(norm)

    try:
        texto_pdf_completo = extract_text_from_pdf(uploaded_file)
        textos_paginas = [texto_pdf_completo]
    except Exception:
        return pd.DataFrame()

    padrao_item = re.compile(r"^(\d{1,4})\s+(\d{3,6})\s+(\d{3,6})\s+(.+)$", flags=re.IGNORECASE)
    padrao_cod_produto = re.compile(r"\b[A-Z0-9]{1,4}(?:\.[A-Z0-9]{1,8}){1,4}[A-Z0-9]*\b", flags=re.IGNORECASE)

    registros = []
    vistos = set()

    for texto in textos_paginas:
        linhas = [str(l or "").strip() for l in str(texto or "").splitlines() if str(l or "").strip()]
        for i, linha in enumerate(linhas):
            m = padrao_item.match(linha)
            if not m:
                continue

            item, abrev, unid, resto = m.group(1), m.group(2), m.group(3), m.group(4)
            nums_linha = _tokens_numericos_linha(linha)
            # item, abrev e unid também entram como número; os três últimos são qtd/preço/total.
            if len(nums_linha) < 6:
                continue
            qtd = float(nums_linha[-3][2])
            preco = float(nums_linha[-2][2])
            total = float(nums_linha[-1][2])
            if qtd <= 0 or preco <= 0:
                continue

            pos_qtd = nums_linha[-3][0]
            prefixo = linha[:pos_qtd].strip()
            prefixo_sem_inicio = re.sub(r"^\d{1,4}\s+\d{3,6}\s+\d{3,6}\s+", "", prefixo).strip()

            prev1 = linhas[i - 1] if i - 1 >= 0 else ""
            prev2 = linhas[i - 2] if i - 2 >= 0 else ""
            next1 = linhas[i + 1] if i + 1 < len(linhas) else ""

            cod_base = ""
            desc_partes = []

            cods_no_prefixo = padrao_cod_produto.findall(prefixo_sem_inicio)
            if cods_no_prefixo:
                cod_base = cods_no_prefixo[-1]
                desc_prefixo = prefixo_sem_inicio.replace(cod_base, " ").strip()
                if desc_prefixo:
                    desc_partes.append(desc_prefixo)
            else:
                # Produto pode estar na linha anterior, junto com litragem e parte da descrição.
                for prev in [prev1, prev2]:
                    cods_prev = padrao_cod_produto.findall(prev)
                    if cods_prev:
                        cod_base = cods_prev[0]
                        desc_prev = prev.replace(cod_base, " ").strip()
                        desc_prev = re.sub(r"^\d+(?:[\.,]\d+)?\s*", "", desc_prev).strip()
                        if desc_prev:
                            desc_partes.append(desc_prev)
                        break
                if prefixo_sem_inicio:
                    desc_partes.append(prefixo_sem_inicio)

            if not cod_base:
                continue

            # Gera variações com complemento quebrado na linha seguinte.
            candidatos_produto = [cod_base]
            next_tokens = str(next1 or "").split()
            if next_tokens:
                primeiro_next = re.sub(r"[^A-Z0-9]", "", _texto_sem_acentos(next_tokens[0]).upper())
                if primeiro_next and len(primeiro_next) <= 3:
                    candidatos_produto.append(cod_base + primeiro_next)

            escolhido_original = ""
            escolhido_norm = ""
            for prod in candidatos_produto:
                original = f"{prod}-{unid}"
                norm = normalizar_codigo_fabrica(original)
                if referencias and norm in referencias:
                    escolhido_original = original
                    escolhido_norm = norm
                    break
            if not escolhido_original:
                original = f"{candidatos_produto[-1]}-{unid}"
                norm = normalizar_codigo_fabrica(original)
                escolhido_original = original
                escolhido_norm = norm

            # Complementa descrição com a linha seguinte quando ela é descrição, removendo quebra de código/litragem.
            if next1:
                prox = str(next1).strip()
                prox_limpo = re.sub(r"^\d{1,3}\s+", "", prox).strip()
                prox_limpo = re.sub(r"^\d+(?:[\.,]\d+)?\s+", "", prox_limpo).strip()
                if re.search(r"[A-ZÉÓÚÃÕÇ]{3,}", prox_limpo, flags=re.IGNORECASE):
                    desc_partes.append(prox_limpo)

            descricao = " ".join([d for d in desc_partes if d]).strip()
            descricao = re.sub(r"\s+", " ", descricao)[:180]

            chave = (escolhido_norm, round(qtd, 4), round(preco, 4), round(total, 4))
            if chave in vistos:
                continue
            vistos.add(chave)
            registros.append({
                "Código Fábrica": escolhido_original,
                "Descrição": descricao,
                "Quantidade": qtd,
                "Valor Unitário": preco,
                "Valor Total": total,
                "Linha PDF": linha,
            })

    return pd.DataFrame(registros)

def _deduplicar_headers_planilha(headers):
    """Garante nomes únicos de colunas, preservando o texto original quando existir."""
    saida = []
    contagem = {}
    for i, h in enumerate(headers):
        nome = str(h or "").strip()
        if not nome or nome.lower() in ["nan", "none"]:
            nome = f"COLUNA {i + 1}"
        base = nome
        chave = normalizar_coluna(base)
        contagem[chave] = contagem.get(chave, 0) + 1
        if contagem[chave] > 1:
            nome = f"{base}.{contagem[chave]}"
        saida.append(nome)
    return saida


def _pontuar_linha_cabecalho_planilha(valores):
    """
    Pontua uma possível linha de cabeçalho em planilhas de fornecedor.
    Resolve arquivos que vêm com dados do cliente antes da tabela, como:
    CÓDIGO | DESCRIÇÃO | ... | QTDE | LITROS | VL. UNIT. | ... | VL. TOTAL
    """
    textos = [_normalizar_nome_coluna_flex(v) for v in valores]
    compactos = [re.sub(r"[^A-Z0-9]+", "", t) for t in textos]
    linha = " ".join(t for t in textos if t)

    score = 0
    tem_codigo = any(t in ["CODIGO", "COD", "SKU", "REFERENCIA", "REF"] or t.startswith("CODIGO ") or t.startswith("COD ") for t in textos) or any(c in ["CODIGO", "COD", "SKU", "REFERENCIA", "REF"] for c in compactos)
    tem_desc = any("DESCR" in t or t in ["PRODUTO", "ITEM", "NOME"] for t in textos)
    tem_qtd = any(re.search(r"(^| )(QTD|QTDE|QTE|QTY|QUANT|QUANTIDADE)( |$)", t) for t in textos) or any(c in ["QTD", "QTDE", "QTE", "QTY", "QUANT", "QUANTIDADE"] for c in compactos)
    tem_preco = any("PRECO" in t or "UNIT" in t or "VL UNIT" in t or "VR UNIT" in t or "VALOR UNIT" in t for t in textos)
    tem_total = any(("TOTAL" in t and "UNIT" not in t) or "VL TOTAL" in t or "VALOR TOTAL" in t for t in textos)

    score += 3 if tem_codigo else 0
    score += 3 if tem_desc else 0
    score += 4 if tem_qtd else 0
    score += 2 if tem_preco else 0
    score += 2 if tem_total else 0

    # Penaliza linhas de formulário/cadastro do cliente, não tabela de produtos.
    if any(p in linha for p in ["DADOS DO CLIENTE", "CNPJ", "CONDICAO PAGTO", "OBSERVACAO NF", "BAIRRO", "SUB REGIAO"]):
        score -= 4
    return score


def _ajustar_cabecalho_planilha_fornecedor(df_raw):
    """
    Quando o Excel do fornecedor possui cabeçalho fora da primeira linha,
    encontra a linha da tabela e transforma em DataFrame tabular.
    """
    if df_raw is None or df_raw.empty:
        return pd.DataFrame()

    df_scan = df_raw.copy()
    melhor_idx = None
    melhor_score = -999

    limite = min(len(df_scan), 100)
    for idx in range(limite):
        valores = list(df_scan.iloc[idx].values)
        score = _pontuar_linha_cabecalho_planilha(valores)
        if score > melhor_score:
            melhor_score = score
            melhor_idx = idx

    # Exige pelo menos código/descrição/quantidade ou pontuação equivalente.
    if melhor_idx is None or melhor_score < 7:
        return df_raw.copy()

    headers = _deduplicar_headers_planilha(list(df_scan.iloc[melhor_idx].values))
    df = df_scan.iloc[melhor_idx + 1:].copy()
    df.columns = headers

    # Remove linhas totalmente vazias e linhas de totalização.
    df = df.dropna(how="all")
    primeira_col = df.columns[0] if len(df.columns) else None
    if primeira_col:
        primeira_txt = df[primeira_col].astype(str).str.strip().str.upper()
        df = df[~primeira_txt.isin(["", "NAN", "NONE"])]
        df = df[~primeira_txt.str.contains(r"^TOTA(IS|L)?$|^TOTAL", regex=True, na=False)]

    # Remove colunas completamente vazias, mas mantém nomes reconhecidos.
    df = df.dropna(axis=1, how="all")
    return df.reset_index(drop=True)


def ler_planilha_comparativo_fornecedor(uploaded_file):
    """
    Lê Excel/CSV do fornecedor procurando automaticamente a linha real do cabeçalho.
    Importante para modelos em que a tabela começa depois de blocos como DADOS DO CLIENTE.
    """
    if uploaded_file is None:
        return pd.DataFrame()

    nome = str(getattr(uploaded_file, "name", "")).lower()
    try:
        uploaded_file.seek(0)
    except Exception:
        pass

    if nome.endswith((".xlsx", ".xls")):
        # header=None permite localizar cabeçalhos que não estão na primeira linha.
        # Para .xls, mantenha xlrd>=2.0.1 no requirements.txt.
        df_raw = pd.read_excel(uploaded_file, dtype=str, header=None)
        return _ajustar_cabecalho_planilha_fornecedor(df_raw)

    # CSV: também pode ter linhas acima do cabeçalho.
    tentativas = [
        {"sep": ";", "encoding": "utf-8-sig"},
        {"sep": ";", "encoding": "latin1"},
        {"sep": ",", "encoding": "utf-8-sig"},
        {"sep": ",", "encoding": "latin1"},
        {"sep": "\t", "encoding": "utf-8-sig"},
        {"sep": "\t", "encoding": "latin1"},
    ]
    ultimo_erro = None
    for tentativa in tentativas:
        try:
            uploaded_file.seek(0)
            df_raw = pd.read_csv(
                uploaded_file,
                sep=tentativa["sep"],
                encoding=tentativa["encoding"],
                dtype=str,
                engine="python",
                on_bad_lines="skip",
                header=None,
            )
            return _ajustar_cabecalho_planilha_fornecedor(df_raw)
        except Exception as e:
            ultimo_erro = str(e)
            continue

    raise RuntimeError(f"Não consegui ler a planilha do fornecedor. Último erro: {ultimo_erro}")



def _dataframe_de_tabela_pdf_fornecedor(linhas):
    """
    Converte uma tabela extraída do PDF do fornecedor em DataFrame padronizado.
    Corrige principalmente PDFs Akzo/Coral em que as colunas são:
    Código Produto | Descrição | Código EAN | Embalagem | Qtd. | ... | VL. Unitário | VL. Total
    """
    if not linhas:
        return pd.DataFrame()

    linhas = [[str(c or "").strip() for c in (linha or [])] for linha in linhas]
    linhas = [linha for linha in linhas if any(str(c).strip() for c in linha)]
    if len(linhas) < 2:
        return pd.DataFrame()

    max_cols = max(len(linha) for linha in linhas)
    linhas = [linha + [""] * (max_cols - len(linha)) for linha in linhas]

    header_idx = None
    melhor_score = -1
    for i, linha in enumerate(linhas[:8]):
        joined = " ".join(_normalizar_nome_coluna_flex(c) for c in linha)
        score = 0
        if re.search(r"\b(CODIGO|COD|SKU|PRODUTO)\b", joined):
            score += 2
        if re.search(r"\b(DESCRICAO|DESCR|ITEM)\b", joined):
            score += 2
        if re.search(r"\b(QTD|QTDE|QUANTIDADE|QUANT)\b", joined):
            score += 4
        if re.search(r"\b(VL|VALOR|PRECO|UNITARIO|UNIT)\b", joined):
            score += 2
        if score > melhor_score:
            melhor_score = score
            header_idx = i

    if header_idx is None or melhor_score < 4:
        return pd.DataFrame()

    headers = _deduplicar_headers_planilha([str(c).strip() or f"COLUNA {i + 1}" for i, c in enumerate(linhas[header_idx])])
    dados = linhas[header_idx + 1:]
    if not dados:
        return pd.DataFrame()

    df = pd.DataFrame(dados, columns=headers)
    df = df.dropna(how="all")
    if df.empty:
        return pd.DataFrame()

    col_codigo = _coluna_por_candidatos(df, [
        "Código Produto", "Codigo Produto", "Cód Produto", "Cod Produto", "Código", "Codigo",
        "Cód.", "Cod.", "SKU", "Referência", "Referencia", "Part Number"
    ])
    col_desc = _coluna_por_candidatos(df, [
        "Descrição", "Descricao", "Descrição do item", "Descricao do item", "Produto", "Item", "Nome", "Descr"
    ])
    col_qtd = _coluna_quantidade_flexivel(df) or _coluna_por_candidatos(df, [
        "Qtd.", "Qtd", "Qtde", "Quantidade", "Quant", "Qte", "Qty"
    ])
    col_preco = _coluna_valor_unitario_flexivel(df) or _coluna_por_candidatos(df, [
        "VL. Unitário", "VL Unitário", "VL. Unitario", "VL Unitario", "Valor Unitário", "Valor Unitario",
        "Preço Unitário", "Preco Unitario", "Preço", "Preco"
    ])
    col_total = _coluna_valor_total_flexivel(df) or _coluna_por_candidatos(df, [
        "VL. Total", "VL Total", "Valor Total", "Vlr Total", "Total"
    ])

    if not col_codigo or not col_qtd:
        return pd.DataFrame()

    # Se a coluna escolhida como quantidade parecer código/EAN, não usa esta tabela.
    if not _serie_parece_quantidade(df[col_qtd]):
        return pd.DataFrame()

    out = pd.DataFrame()
    out["Código Fábrica"] = df[col_codigo].astype(str).str.strip()
    out["Descrição"] = df[col_desc].astype(str).str.strip() if col_desc else out["Código Fábrica"]
    out["Quantidade"] = df[col_qtd].apply(numero_planilha_para_float)
    out["Valor Unitário"] = df[col_preco].apply(numero_planilha_para_float) if col_preco else 0.0
    out["Valor Total"] = df[col_total].apply(numero_planilha_para_float) if col_total else 0.0
    out["Linha PDF"] = df.astype(str).agg(" | ".join, axis=1)

    out = out[~out["Código Fábrica"].astype(str).str.upper().str.contains("TOTAL DO PEDIDO|NUMERO DE ITENS|NÚMERO DE ITENS", na=False)]
    out = out[out["Código Fábrica"].astype(str).str.extract(r"([A-Za-z0-9]{3,})", expand=False).notna()]
    out = out[out["Quantidade"] > 0].copy()
    out.loc[(out["Valor Total"] <= 0) & (out["Valor Unitário"] > 0), "Valor Total"] = out["Quantidade"] * out["Valor Unitário"]
    out.loc[(out["Valor Unitário"] <= 0) & (out["Valor Total"] > 0) & (out["Quantidade"] > 0), "Valor Unitário"] = out["Valor Total"] / out["Quantidade"]
    return out.reset_index(drop=True)



def _numero_quantidade_pdf_br(valor):
    """
    Converte quantidade impressa em PDFs brasileiros.

    Nestes relatórios a quantidade é exibida com 3 casas decimais:
    24,000 = 24 unidades; 1.500,000 = 1.500 unidades.
    Essa regra é aplicada somente aos PDFs reconhecidos como brasileiros,
    sem alterar a leitura de Excel/Sheets em que 1,000 pode significar mil.
    """
    txt = str(valor or "").strip().replace(" ", "").replace("\xa0", "")
    if not txt:
        return 0.0
    if re.fullmatch(r"-?\d{1,3}(?:\.\d{3})*,\d{3}", txt) or re.fullmatch(r"-?\d+,\d{3}", txt):
        return br_to_float(txt)
    return numero_planilha_para_float(txt)


def extrair_itens_pdf_3m(uploaded_file):
    """
    Parser específico e seguro para o layout de simulação/pedido da 3M.

    Linha esperada:
    Linha | Produto | Descrição | NCM | Agrupamento | Quantidade |
    Valor Unitário | Valor Total

    O ponto crítico deste modelo é que Quantidade usa 3 casas decimais.
    Portanto, 24,000 significa 24 e não 24.000.
    """
    if uploaded_file is None:
        return pd.DataFrame()

    try:
        texto = extract_text_from_pdf_pdfplumber(uploaded_file)
    except Exception:
        return pd.DataFrame()

    texto_up = _texto_sem_acentos(texto).upper()
    if "3M DO BRASIL" not in texto_up or "AGRUPAMENTO" not in texto_up or "VALOR UNITARIO" not in texto_up:
        return pd.DataFrame()

    padrao = re.compile(
        r"^\s*(?P<linha>\d{1,4})\s+"
        r"(?P<codigo>[A-Z]{1,4}\d[A-Z0-9]{5,})\s+"
        r"(?P<descricao>.+?)\s+"
        r"(?P<ncm>\d{8,10})\s+"
        r"(?P<agrupamento>[A-Z0-9]{2,10})\s+"
        r"(?P<qtd>-?\d{1,3}(?:\.\d{3})*,\d{3}|-?\d+,\d{3})\s+"
        r"(?P<unit>-?\d{1,3}(?:\.\d{3})*,\d{4}|-?\d+,\d{4})\s+"
        r"(?P<total>-?\d{1,3}(?:\.\d{3})*,\d{2}|-?\d+,\d{2})\s*$",
        flags=re.IGNORECASE,
    )

    registros = []
    vistos = set()
    for raw in texto.splitlines():
        linha = re.sub(r"\s+", " ", str(raw or "")).strip()
        m = padrao.match(linha)
        if not m:
            continue

        codigo = m.group("codigo").strip()
        qtd = _numero_quantidade_pdf_br(m.group("qtd"))
        preco = br_to_float(m.group("unit"))
        total = br_to_float(m.group("total"))
        if not codigo or qtd <= 0 or preco <= 0:
            continue

        # Validação contábil tolerante a IPI e demais acréscimos do documento.
        calculado = qtd * preco
        if total <= 0:
            total = calculado
        elif abs(calculado - total) > max(0.15, abs(total) * 0.25):
            # Mantém os valores impressos, mas impede captura de linha tributária aleatória.
            continue

        chave = (normalizar_codigo_fabrica(codigo), round(qtd, 6), round(preco, 6), round(total, 2))
        if chave in vistos:
            continue
        vistos.add(chave)
        registros.append({
            "Código Fábrica": codigo,
            "Descrição": re.sub(r"\s+", " ", m.group("descricao")).strip(),
            "Quantidade": qtd,
            "Valor Unitário": preco,
            "Valor Total": total,
            "Linha PDF": linha,
        })

    return pd.DataFrame(registros)


def _parse_quantidade_orcamento_venda_pdf(valor):
    """
    Converte quantidades impressas com digitos separados por espacos.
    Ex.: "2 0 00" = 2000, "6 0 0" = 600, "3 0" = 30.
    """
    txt = str(valor or "").strip()
    if not txt:
        return 0.0

    txt_limpo = re.sub(r"\s+", "", txt)
    if re.fullmatch(r"\d+", txt_limpo):
        return float(txt_limpo)

    return numero_planilha_para_float(txt)


def extrair_itens_pdf_orcamento_venda(uploaded_file):
    """
    Parser seguro para PDFs de Orcamento de Venda em que a tabela vem compactada
    em uma unica linha com quebras internas por coluna.

    Layout observado:
    ITEM | CODIGO | DESCRICAO | QUANT. | VLR.UNIT. | ICMS ST | IPI | TOTAL

    O preco usado e sempre VLR.UNIT., sem impostos. TOTAL pode conter IPI/ST.
    """
    if uploaded_file is None:
        return pd.DataFrame()

    try:
        pdf_bytes = _pdf_bytes(uploaded_file)
        texto = extract_text_from_pdf_pdfplumber_cached(pdf_bytes)
    except Exception:
        return pd.DataFrame()

    texto_norm = _texto_sem_acentos(texto).upper()
    if (
        "ORCAMENTO DE VENDA" not in texto_norm
        or "VLR.UNIT" not in texto_norm
        or "ICMS ST" not in texto_norm
        or "IPI" not in texto_norm
    ):
        return pd.DataFrame()

    registros = []
    vistos = set()

    try:
        with pdfplumber.open(BytesIO(pdf_bytes)) as pdf:
            for page in pdf.pages:
                for tabela in (page.extract_tables() or []):
                    if not tabela or len(tabela) < 2:
                        continue

                    header = [str(c or "").strip() for c in tabela[0]]
                    header_norm = [_normalizar_nome_coluna_flex(c) for c in header]
                    joined_header = " ".join(header_norm)
                    if not all(chave in joined_header for chave in ["ITEM", "CODIGO", "DESCRICAO", "QUANT"]):
                        continue
                    if not any("VLR" in h or "UNIT" in h or "VALOR" in h for h in header_norm):
                        continue

                    def idx_col(*partes):
                        for idx, nome in enumerate(header_norm):
                            if all(parte in nome for parte in partes):
                                return idx
                        return None

                    idx_codigo = idx_col("CODIGO")
                    idx_desc = idx_col("DESCRICAO")
                    idx_qtd = idx_col("QUANT")
                    idx_unit = idx_col("VLR") if idx_col("VLR") is not None else idx_col("UNIT")
                    idx_total = idx_col("TOTAL")
                    if idx_codigo is None or idx_qtd is None or idx_unit is None:
                        continue

                    for linha_tabela in tabela[1:]:
                        linha_tabela = [str(c or "").strip() for c in (linha_tabela or [])]
                        if not any(linha_tabela):
                            continue

                        colunas_split = []
                        for c in linha_tabela:
                            partes = [p.strip() for p in str(c or "").splitlines()]
                            colunas_split.append(partes)

                        qtd_linhas = max((len(p) for p in colunas_split), default=0)
                        if qtd_linhas <= 0:
                            continue

                        for i in range(qtd_linhas):
                            def valor_col(idx):
                                if idx is None or idx >= len(colunas_split):
                                    return ""
                                partes = colunas_split[idx]
                                return partes[i].strip() if i < len(partes) else ""

                            codigo = valor_col(idx_codigo)
                            descricao = valor_col(idx_desc)
                            qtd = _parse_quantidade_orcamento_venda_pdf(valor_col(idx_qtd))
                            preco = numero_planilha_para_float(valor_col(idx_unit))
                            total_impresso = numero_planilha_para_float(valor_col(idx_total))
                            if not codigo or qtd <= 0 or preco <= 0:
                                continue

                            total_sem_imposto = qtd * preco
                            total = total_sem_imposto if total_sem_imposto > 0 else total_impresso
                            chave = (normalizar_codigo_fabrica(codigo), round(qtd, 6), round(preco, 6))
                            if chave in vistos:
                                continue
                            vistos.add(chave)
                            registros.append({
                                "CÃ³digo FÃ¡brica": codigo,
                                "DescriÃ§Ã£o": re.sub(r"\s+", " ", descricao).strip(),
                                "Quantidade": qtd,
                                "Valor UnitÃ¡rio": preco,
                                "Valor Total": total,
                                "Linha PDF": " | ".join([codigo, descricao, str(qtd), str(preco), str(total_impresso)]),
                            })
    except Exception:
        return pd.DataFrame()

    return pd.DataFrame(registros)


def extrair_itens_pdf_mercanet(uploaded_file):
    """
    Parser para pedidos Mercanet/Berliner.

    Layout observado:
    Produto | Descricao | Qtde | Preco unitario liq | ... | R$ final unitario | R$ total + impostos

    Para o comparativo financeiro, usa sempre Preco unitario liq, sem impostos.
    """
    if uploaded_file is None:
        return pd.DataFrame()

    try:
        pdf_bytes = _pdf_bytes(uploaded_file)
        texto = extract_text_from_pdf_pdfplumber_cached(pdf_bytes)
    except Exception:
        return pd.DataFrame()

    texto_norm = _texto_sem_acentos(texto).upper()
    if (
        "MERCANET" not in texto_norm
        or "DADOS DO PEDIDO" not in texto_norm
        or "PRECO UNITARIO" not in texto_norm
        or "R$ TOTAL" not in texto_norm
    ):
        return pd.DataFrame()

    registros = []
    vistos = set()

    try:
        with pdfplumber.open(BytesIO(pdf_bytes)) as pdf:
            for page in pdf.pages:
                for tabela in (page.extract_tables() or []):
                    if not tabela or len(tabela) < 3:
                        continue

                    linhas = [[str(c or "").strip() for c in (linha or [])] for linha in tabela]
                    header_idx = None
                    header_norm = []
                    for idx, linha in enumerate(linhas[:8]):
                        nomes = [_normalizar_nome_coluna_flex(c) for c in linha]
                        joined = " ".join(nomes)
                        if (
                            "PRODUTO" in joined
                            and "DESCRICAO" in joined
                            and "QTDE" in joined
                            and ("PRECO" in joined or "UNITARIO" in joined)
                        ):
                            header_idx = idx
                            header_norm = nomes
                            break

                    if header_idx is None:
                        continue

                    def idx_col(*partes):
                        for col_idx, nome in enumerate(header_norm):
                            if all(parte in nome for parte in partes):
                                return col_idx
                        return None

                    def idx_col_exato(nome_exato):
                        for col_idx, nome in enumerate(header_norm):
                            if nome == nome_exato:
                                return col_idx
                        return None

                    def idx_col_ultimo(*partes):
                        encontrado = None
                        for col_idx, nome in enumerate(header_norm):
                            if all(parte in nome for parte in partes):
                                encontrado = col_idx
                        return encontrado

                    idx_produto = idx_col_exato("PRODUTO")
                    idx_desc = idx_col("DESCRICAO")
                    idx_qtd = idx_col("QTDE")
                    idx_preco = idx_col("PRECO") if idx_col("PRECO") is not None else idx_col("UNITARIO")
                    idx_total = idx_col_ultimo("TOTAL")
                    if idx_produto is None or idx_qtd is None or idx_preco is None:
                        continue

                    for row in linhas[header_idx + 1:]:
                        if not any(row):
                            continue
                        if len(row) <= max(idx_produto, idx_qtd, idx_preco):
                            continue

                        codigo = re.sub(r"\s+", " ", row[idx_produto]).strip()
                        descricao = re.sub(r"\s+", " ", row[idx_desc]).strip() if idx_desc is not None and idx_desc < len(row) else codigo
                        qtd = numero_planilha_para_float(row[idx_qtd])
                        preco = numero_planilha_para_float(row[idx_preco])
                        total_impresso = numero_planilha_para_float(row[idx_total]) if idx_total is not None and idx_total < len(row) else 0.0

                        if not codigo or qtd <= 0 or preco <= 0:
                            continue
                        if _texto_sem_acentos(codigo).upper() in {"PRODUTO", "ITENS"}:
                            continue

                        total_sem_imposto = qtd * preco
                        chave = (normalizar_codigo_fabrica(codigo), round(qtd, 6), round(preco, 6))
                        if chave in vistos:
                            continue
                        vistos.add(chave)
                        registros.append({
                            "Código Fábrica": codigo,
                            "Descrição": descricao,
                            "Quantidade": qtd,
                            "Valor Unitário": preco,
                            "Valor Total": total_sem_imposto if total_sem_imposto > 0 else total_impresso,
                            "Linha PDF": " | ".join([codigo, descricao, str(qtd), str(preco), str(total_impresso)]),
                        })
    except Exception:
        return pd.DataFrame()

    return pd.DataFrame(registros)


def extrair_itens_pdf_por_tabelas(uploaded_file):
    """Lê PDFs de fornecedor com tabela real antes das heurísticas por texto."""
    if uploaded_file is None:
        return pd.DataFrame()
    try:
        pdf_bytes = _pdf_bytes(uploaded_file)
        dfs = []
        with pdfplumber.open(BytesIO(pdf_bytes)) as pdf:
            for page in pdf.pages:
                for tabela in (page.extract_tables() or []):
                    df_tab = _dataframe_de_tabela_pdf_fornecedor(tabela)
                    if df_tab is not None and not df_tab.empty:
                        dfs.append(df_tab)
        if not dfs:
            return pd.DataFrame()
        return pd.concat(dfs, ignore_index=True)
    except Exception:
        return pd.DataFrame()


def _qtd_codigos_referencia_encontrados(df, codigos_referencia=None):
    if df is None or df.empty:
        return 0
    referencias = set(_referencias_codigo_fabrica(codigos_referencia).keys())
    if not referencias:
        return 0

    col_codigo = _coluna_por_candidatos(df, ["Código Fábrica", "Codigo Fabrica", "Código", "Codigo", "Referência", "Referencia", "SKU", "Produto"])
    if not col_codigo:
        return 0
    cods = set(df[col_codigo].astype(str).apply(normalizar_codigo_fabrica))
    return len(cods & referencias)


def extrair_texto_imagem_ocr(uploaded_file):
    if uploaded_file is None:
        return ""
    if Image is None or pytesseract is None:
        return ""
    try:
        uploaded_file.seek(0)
        img = Image.open(uploaded_file)
        return pytesseract.image_to_string(img, lang="por+eng") or ""
    except Exception:
        try:
            uploaded_file.seek(0)
            img = Image.open(uploaded_file)
            return pytesseract.image_to_string(img) or ""
        except Exception:
            return ""



def _decodificar_arquivo_texto_comparativo(uploaded_file):
    """L TXT do fornecedor aceitando os encodings mais comuns no Windows/ERP."""
    try:
        uploaded_file.seek(0)
    except Exception:
        pass
    try:
        dados = uploaded_file.read()
    except Exception:
        return ""
    if isinstance(dados, str):
        return dados
    for encoding in ("utf-8-sig", "utf-8", "cp1252", "latin1"):
        try:
            return bytes(dados).decode(encoding)
        except Exception:
            continue
    return bytes(dados).decode("latin1", errors="replace")


def extrair_itens_txt_comparativo(uploaded_file, codigos_referencia=None):
    """
    Extrai itens de pedidos TXT em formato de relatrio/pr-nota.

    Layout suportado, inclusive o arquivo MAXIRUBBER anexado:
    It | Cdigo | Descrio | TES | UM | Qt.Ped | Valor Unit. | IPI | ICM | Tab | Total

    A expresso  ancorada no incio e no fim da linha para no capturar cabealhos,
    totais ou dados cadastrais. Como fallback, usa os cdigos do Pedido nica como ncora.
    """
    texto = _decodificar_arquivo_texto_comparativo(uploaded_file)
    if not texto.strip():
        return pd.DataFrame()

    registros = []
    vistos = set()
    padrao = re.compile(
        r"^\s*(?P<item>\d{1,4})\s+"
        r"(?P<codigo>[A-Z0-9][A-Z0-9.\-/_]{2,})\s+"
        r"(?P<descricao>.+?)\s+"
        r"(?P<tes>\d{3})\s+"
        r"(?P<um>[A-Z]{1,4})\s+"
        r"(?P<qtd>-?\d+(?:[.,]\d+)?)\s+"
        r"(?P<unit>-?\d{1,3}(?:\.\d{3})*,\d{2,4}|-?\d+[.,]\d{2,4})\s+"
        r"(?P<ipi>-?\d+(?:[.,]\d+)?)\s+"
        r"(?P<icm>-?\d+(?:[.,]\d+)?)\s+"
        r"(?P<tab>[A-Z0-9]+)\s+"
        r"(?P<total>-?\d{1,3}(?:\.\d{3})*,\d{2}|-?\d+[.,]\d{2})\s*$",
        flags=re.IGNORECASE,
    )

    for raw in texto.splitlines():
        linha = str(raw or "").rstrip()
        m = padrao.match(linha)
        if not m:
            continue
        codigo = m.group("codigo").strip()
        qtd = numero_planilha_para_float(m.group("qtd"))
        preco = numero_planilha_para_float(m.group("unit"))
        total = numero_planilha_para_float(m.group("total"))
        if qtd <= 0 or not codigo:
            continue
        chave = (normalizar_codigo_fabrica(codigo), round(qtd, 6), round(preco, 6), round(total, 6))
        if chave in vistos:
            continue
        vistos.add(chave)
        registros.append({
            "Código Fábrica": codigo,
            "Descrição": re.sub(r"\s+", " ", m.group("descricao")).strip(),
            "Quantidade": qtd,
            "Valor Unitário": preco,
            "Valor Total": total if total > 0 else qtd * preco,
            "Linha Fornecedor": linha,
            "Linha TXT": linha,
        })

    df = pd.DataFrame(registros)
    if not df.empty:
        return df.reset_index(drop=True)

    # Fallback para outros TXT: procura os cdigos de fbrica do Pedido nica em cada linha.
    linhas = [l.strip() for l in texto.splitlines() if l.strip()]
    df_anchor = extrair_itens_por_codigos_em_textos(
        linhas,
        codigos_referencia,
        origem_linha="Linha TXT",
    )
    return df_anchor if df_anchor is not None else pd.DataFrame()



def extrair_itens_pdf_mastersales_brasilux(uploaded_file, codigos_referencia=None):
    """
    Parser dedicado ao RELATÓRIO DE PEDIDO MasterSales / Brasilux.

    Layout real da linha:
    Item + prefixo + código numérico + descrição + Qtde + Preço Venda + Valor Total
    + Venda c/ Imp. + IPI + ST.

    A quantidade é capturada imediatamente antes do primeiro ``R$``. Isso evita
    confundir números da descrição/embalagem, como 900 ML, 18 LT ou 25 KG, com a
    quantidade pedida.
    """
    try:
        texto = extract_text_from_pdf_pdfplumber(uploaded_file)
    except Exception:
        return pd.DataFrame()

    texto_upper = _texto_sem_acentos(texto).upper()
    if "RELATORIO DE PEDIDO" not in texto_upper or "MASTERSALES" not in texto_upper:
        return pd.DataFrame()

    referencias = _referencias_codigo_fabrica(codigos_referencia)
    registros = []
    vistos = set()

    # O trecho final exige cinco campos monetários do relatório, impedindo que a
    # regex encerre a descrição no primeiro número que encontrar.
    padrao = re.compile(
        r"^\s*(?P<item>\d+)\s+"
        r"(?:(?P<prefixo>[A-Z]{1,4})\s+)?"
        r"(?P<codigo>\d{4,18})\s+-\s+"
        r"(?P<descricao>.*?)\s+"
        r"(?P<qtd>\d+(?:[\.,]\d+)?)\s+"
        r"R\$\s*(?P<preco>[\d\.]+,\d{2})\s+"
        r"R\$\s*(?P<total>[\d\.]+,\d{2})\s+"
        r"R\$\s*(?P<preco_imp>[\d\.]+,\d{2})\s+"
        r"R\$\s*(?P<ipi>[\d\.]+,\d{2})\s+"
        r"R\$\s*(?P<st>[\d\.]+,\d{2})\s*$",
        flags=re.IGNORECASE,
    )

    for linha in str(texto or "").splitlines():
        linha = str(linha or "").strip()
        m = padrao.match(linha)
        if not m:
            continue

        prefixo = str(m.group("prefixo") or "").upper().strip()
        codigo_num = m.group("codigo").strip()
        codigo_pdf = f"{prefixo} {codigo_num}".strip()
        norm_completo = normalizar_codigo_fabrica(codigo_pdf)
        norm_numerico = normalizar_codigo_fabrica(codigo_num)

        # Quando a planilha da Única foi fornecida, devolve exatamente o código
        # existente nela. Assim o relacionamento funciona tanto para "TN 710041608"
        # quanto para "710041608".
        codigo_relacionamento = codigo_pdf
        if referencias:
            if norm_completo in referencias:
                codigo_relacionamento = referencias[norm_completo]
            elif norm_numerico in referencias:
                codigo_relacionamento = referencias[norm_numerico]
            # Não descarta o item quando ele não estiver na referência da Única.
            # O PDF MasterSales já possui colunas estruturadas e confiáveis; manter
            # todos os itens permite exibir corretamente também "Somente fornecedor".

        qtd = numero_planilha_para_float(m.group("qtd"))
        preco = numero_planilha_para_float(m.group("preco"))
        total = numero_planilha_para_float(m.group("total"))
        descricao = m.group("descricao").strip(" -")

        chave = (normalizar_codigo_fabrica(codigo_relacionamento), round(qtd, 4))
        if chave in vistos:
            continue
        vistos.add(chave)

        registros.append({
            "Código Fábrica": codigo_relacionamento,
            "Descrição": descricao,
            "Quantidade": qtd,
            "Valor Unitário": preco,
            "Valor Total": total if total > 0 else qtd * preco,
            "Linha PDF": linha,
        })

    return pd.DataFrame(registros)


MODELOS_FORNECEDOR_COMPARATIVO = {
    "Automático": "automatico",
    "Akzo Automotivo": "akzo_automotivo",
    "Maxi Rubber": "maxi_rubber",
    "Coral": "coral",
    "Sherwin Williams": "sherwin_williams",
    "Auto America": "auto_america",
    "3M": "3m",
    "Brasilux / MasterSales": "brasilux_mastersales",
}


def _modelo_fornecedor_codigo(modelo_fornecedor=None):
    modelo = str(modelo_fornecedor or "Automático").strip()
    return MODELOS_FORNECEDOR_COMPARATIVO.get(modelo, modelo.lower().strip())


def _ler_arquivo_comparativo_modelo_homologado(uploaded_file, codigos_referencia=None, modelo_fornecedor=None):
    if uploaded_file is None:
        return pd.DataFrame()

    modelo = _modelo_fornecedor_codigo(modelo_fornecedor)
    nome = str(getattr(uploaded_file, "name", "")).lower()
    try:
        uploaded_file.seek(0)
    except Exception:
        pass

    if modelo in ("automatico", ""):
        return pd.DataFrame()

    if modelo == "maxi_rubber":
        if nome.endswith(".txt"):
            return extrair_itens_txt_comparativo(uploaded_file, codigos_referencia=codigos_referencia)
        return pd.DataFrame()

    if modelo == "brasilux_mastersales":
        if nome.endswith(".pdf"):
            return extrair_itens_pdf_mastersales_brasilux(uploaded_file, codigos_referencia=codigos_referencia)
        return pd.DataFrame()

    if modelo == "3m":
        if nome.endswith(".pdf"):
            return extrair_itens_pdf_3m(uploaded_file)
        return pd.DataFrame()

    if modelo == "auto_america":
        if nome.endswith(".pdf"):
            df_pdf = extrair_itens_pdf_orcamento_venda(uploaded_file)
            if not df_pdf.empty:
                return df_pdf
            return extrair_itens_pdf_mercanet(uploaded_file)
        return pd.DataFrame()

    if modelo == "coral":
        if nome.endswith(".pdf"):
            return extrair_itens_pdf_por_tabelas(uploaded_file)
        return pd.DataFrame()

    if modelo == "sherwin_williams":
        if nome.endswith(".pdf"):
            return extrair_itens_pdf_por_blocos(uploaded_file, codigos_referencia=codigos_referencia)
        return pd.DataFrame()

    if modelo == "akzo_automotivo":
        if nome.endswith((".xlsx", ".xls", ".csv", ".txt", ".html", ".htm")):
            df_planilha = ler_planilha_comparativo_fornecedor(uploaded_file)
            df_padrao = padronizar_dataframe_fornecedor_homologado(df_planilha)
            if df_padrao is not None and not df_padrao.empty:
                return df_padrao
            if codigos_referencia and not dataframe_fornecedor_tem_colunas_confiaveis(df_planilha):
                df_anchor = extrair_itens_por_codigos_em_dataframe(df_planilha, codigos_referencia)
                if df_anchor is not None and not df_anchor.empty:
                    return df_anchor
            return df_planilha
        if nome.endswith(".pdf"):
            df_pdf = extrair_itens_pdf_por_tabelas(uploaded_file)
            if not df_pdf.empty:
                return df_pdf
            return extrair_itens_pdf_por_codigos(uploaded_file, codigos_referencia=codigos_referencia)

    return pd.DataFrame()


def ler_arquivo_comparativo(uploaded_file, codigos_referencia=None, modelo_fornecedor=None):
    if uploaded_file is None:
        return pd.DataFrame()

    nome = str(getattr(uploaded_file, "name", "")).lower()
    try:
        uploaded_file.seek(0)
    except Exception:
        pass

    df_modelo = _ler_arquivo_comparativo_modelo_homologado(
        uploaded_file,
        codigos_referencia=codigos_referencia,
        modelo_fornecedor=modelo_fornecedor,
    )
    if df_modelo is not None and not df_modelo.empty:
        return df_modelo

    try:
        uploaded_file.seek(0)
    except Exception:
        pass

    if nome.endswith(".txt"):
        return extrair_itens_txt_comparativo(uploaded_file, codigos_referencia=codigos_referencia)

    if nome.endswith(".pdf"):
        # Layout MasterSales/Brasilux: captura o Código de Fábrica e a quantidade
        # imediatamente antes do primeiro R$, sem confundir 900 ML / 18 LT / 25 KG.
        df_pdf = extrair_itens_pdf_mastersales_brasilux(
            uploaded_file, codigos_referencia=codigos_referencia
        )
        if not df_pdf.empty:
            return df_pdf

        # Layout 3M: a quantidade usa 3 casas decimais (24,000 = 24 unidades).
        # Este parser dedicado precisa vir antes das heurísticas genéricas.
        df_pdf = extrair_itens_pdf_3m(uploaded_file)
        if not df_pdf.empty:
            return df_pdf

        # Primeiro tenta tabela real do PDF. Isso evita confundir Código Produto/EAN/Código de Fábrica
        # com quantidade quando o PDF possui colunas explícitas como Qtd., VL. Unitário e VL. Total.
        df_pdf = extrair_itens_pdf_orcamento_venda(uploaded_file)
        if not df_pdf.empty:
            return df_pdf

        df_pdf = extrair_itens_pdf_mercanet(uploaded_file)
        if not df_pdf.empty:
            return df_pdf

        df_pdf = extrair_itens_pdf_por_tabelas(uploaded_file)
        if not df_pdf.empty and codigos_referencia:
            df_anchor = extrair_itens_por_codigos_em_dataframe(df_pdf, codigos_referencia)
            if df_anchor is not None and not df_anchor.empty:
                df_pdf = df_anchor
            elif _qtd_codigos_referencia_encontrados(df_pdf, codigos_referencia) <= 0:
                df_pdf = pd.DataFrame()
        if df_pdf.empty:
            df_pdf = extrair_itens_pdf_por_blocos(uploaded_file, codigos_referencia=codigos_referencia)
        if df_pdf.empty:
            df_pdf = extrair_itens_pdf_por_codigos(uploaded_file, codigos_referencia=codigos_referencia)
        if not df_pdf.empty:
            return df_pdf

        # Fallback antigo: tenta tabelas com cabeçalho quando não encontrou códigos no texto.
        _linhas_texto_pdf, linhas = extract_pdf_linhas_e_tabelas(uploaded_file)
        if not linhas:
            return pd.DataFrame()

        max_cols = max(len(linha) for linha in linhas)
        linhas = [linha + [""] * (max_cols - len(linha)) for linha in linhas]
        header_idx = 0
        for i, linha in enumerate(linhas[:15]):
            linha_norm = " ".join(normalizar_coluna(c) for c in linha)
            if any(p in linha_norm for p in ["COD", "CÓD", "DESCR", "QTD", "QTDE", "QUANT", "UNIT", "PRECO", "PREÇO"]):
                header_idx = i
                break
        headers = [str(c).strip() or f"COLUNA {i+1}" for i, c in enumerate(linhas[header_idx])]
        return pd.DataFrame(linhas[header_idx + 1:], columns=headers)

    if nome.endswith((".html", ".htm")):
        try:
            uploaded_file.seek(0)
            tabelas = pd.read_html(uploaded_file, dtype=str)
            if tabelas:
                if codigos_referencia:
                    extraidos = []
                    for tabela in tabelas:
                        if dataframe_fornecedor_tem_colunas_confiaveis(tabela):
                            continue
                        df_anchor = extrair_itens_por_codigos_em_dataframe(tabela, codigos_referencia)
                        if df_anchor is not None and not df_anchor.empty:
                            extraidos.append(df_anchor)
                    if extraidos:
                        return pd.concat(extraidos, ignore_index=True)

                # Usa a maior tabela encontrada no HTML.
                tabelas = sorted(tabelas, key=lambda d: d.shape[0] * max(d.shape[1], 1), reverse=True)
                df_html = tabelas[0].copy()
                df_html.columns = [str(c).strip() for c in df_html.columns]
                return df_html
        except Exception:
            return pd.DataFrame()

    if nome.endswith((".png", ".jpg", ".jpeg", ".webp", ".bmp", ".tif", ".tiff")):
        texto = extrair_texto_imagem_ocr(uploaded_file)
        if texto:
            linhas = [l.strip() for l in texto.splitlines() if l.strip()]
            df_img = extrair_itens_por_codigos_em_textos(linhas, codigos_referencia, origem_linha="Linha OCR")
            if df_img is not None and not df_img.empty:
                return df_img
        return pd.DataFrame()

    df_planilha = ler_planilha_comparativo_fornecedor(uploaded_file)
    if codigos_referencia:
        if not dataframe_fornecedor_tem_colunas_confiaveis(df_planilha):
            df_anchor = extrair_itens_por_codigos_em_dataframe(df_planilha, codigos_referencia)
            if df_anchor is not None and not df_anchor.empty:
                return df_anchor
    return df_planilha


def ler_multiplos_arquivos_comparativo(arquivos, codigos_referencia=None, modelo_fornecedor=None):
    """Lê e consolida um ou vários arquivos do pedido do fornecedor.

    Cada arquivo é processado pelo mesmo leitor já usado no comparativo. Os DataFrames
    são unidos antes do mapeamento e da normalização. Posteriormente,
    ``agregar_pedido_comparativo`` soma as quantidades dos códigos repetidos e calcula
    o preço unitário médio ponderado pela quantidade.

    Retorna: (dataframe_consolidado, arquivos_lidos, erros)
    """
    if arquivos is None:
        return pd.DataFrame(), [], []

    if not isinstance(arquivos, (list, tuple)):
        arquivos = [arquivos]

    frames = []
    arquivos_lidos = []
    erros = []

    for arquivo in arquivos:
        nome_arquivo = str(getattr(arquivo, "name", "arquivo fornecedor") or "arquivo fornecedor")
        try:
            try:
                arquivo.seek(0)
            except Exception:
                pass

            df_arquivo = ler_arquivo_comparativo(
                arquivo,
                codigos_referencia=codigos_referencia,
                modelo_fornecedor=modelo_fornecedor,
            )

            if df_arquivo is None or df_arquivo.empty:
                erros.append(f"{nome_arquivo}: nenhum item identificado")
                continue

            df_arquivo = df_arquivo.copy()
            df_arquivo["Arquivo de origem"] = nome_arquivo
            frames.append(df_arquivo)
            arquivos_lidos.append(nome_arquivo)
        except Exception as exc:
            erros.append(f"{nome_arquivo}: {exc}")

    if not frames:
        return pd.DataFrame(), arquivos_lidos, erros

    # sort=False mantém a ordem original e permite unir arquivos com colunas extras.
    consolidado = pd.concat(frames, ignore_index=True, sort=False)
    return consolidado, arquivos_lidos, erros


def _opcoes_colunas_mapeamento(df, incluir_vazio=True):
    cols = [str(c) for c in list(df.columns)] if df is not None else []
    return (["-- Não usar --"] + cols) if incluir_vazio else cols


def _primeira_coluna_existente(df, candidatos, permitir_vazio=True):
    if df is None or df.empty:
        return "-- Não usar --" if permitir_vazio else None
    col = _coluna_por_candidatos(df, candidatos)
    if col:
        return str(col)
    return "-- Não usar --" if permitir_vazio else (str(df.columns[0]) if len(df.columns) else None)


def _mapear_colunas_comparativo(df, prefixo, origem):
    st.markdown(f"**Mapeamento do {origem}**")
    op_obrig = _opcoes_colunas_mapeamento(df, incluir_vazio=False)
    op_opc = _opcoes_colunas_mapeamento(df, incluir_vazio=True)

    if not op_obrig:
        st.error(f"Não encontrei colunas no arquivo {origem}.")
        return None

    default_codigo = _primeira_coluna_existente(df, [
        "Código Fábrica", "Codigo Fabrica", "Cód. Fábrica", "Cod. Fabrica",
        "Código de Fábrica", "Codigo de Fabrica", "Código", "Codigo", "Cód.", "Cod.", "Cod",
        "Referência", "Referencia", "Ref", "SKU", "Produto", "Cod Produto", "Código Produto"
    ], permitir_vazio=True)
    default_desc = _primeira_coluna_existente(df, [
        "descricao", "descrição", "descrição do item", "descricao do item", "produto", "item", "nome", "descr", "Linha PDF"
    ], permitir_vazio=True)
    default_qtd = _coluna_quantidade_flexivel(df) or _primeira_coluna_existente(df, [
        "PEDIDO Final", "Quantidade", "Qtd", "Qtde", "Quant", "Qte", "Qty", "QTDE"
    ], permitir_vazio=True)
    default_preco = _coluna_valor_unitario_flexivel(df) or _primeira_coluna_existente(df, [
        "Preço Última Compra", "Preco Ultima Compra", "Preço", "Preco", "Preço Unitário", "Preco Unitario",
        "Valor Unitário", "Valor Unitario", "Vlr Unit", "Vl Unit", "VL. UNIT.", "VL UNIT", "VR.UNIT"
    ], permitir_vazio=True)
    default_total = _coluna_valor_total_flexivel(df) or _primeira_coluna_existente(df, [
        "Valor Final do Pedido", "Valor Total", "VL. TOTAL", "VL TOTAL", "Vlr Total", "Total", "Valor"
    ], permitir_vazio=True)

    def idx(opcoes, valor):
        return opcoes.index(valor) if valor in opcoes else 0

    c1, c2 = st.columns(2)
    with c1:
        col_codigo = st.selectbox("Coluna de Código de Fábrica / relacionamento", op_opc, index=idx(op_opc, default_codigo), key=f"{prefixo}_codigo")
        col_desc = st.selectbox("Coluna de descrição", op_opc, index=idx(op_opc, default_desc), key=f"{prefixo}_desc")
    with c2:
        col_qtd = st.selectbox("Coluna de quantidade", op_opc, index=idx(op_opc, default_qtd), key=f"{prefixo}_qtd")
        col_preco = st.selectbox("Coluna de valor unitário", op_opc, index=idx(op_opc, default_preco), key=f"{prefixo}_preco")
        col_total = st.selectbox("Coluna de valor total (opcional)", op_opc, index=idx(op_opc, default_total), key=f"{prefixo}_total")

    def limpar(v):
        return None if v == "-- Não usar --" else v

    return {
        "codigo": limpar(col_codigo),
        "descricao": limpar(col_desc),
        "quantidade": limpar(col_qtd),
        "preco_unitario": limpar(col_preco),
        "valor_total": limpar(col_total),
    }


def normalizar_pedido_comparativo(df, origem, mapa_colunas=None):
    df = df.copy()
    df.columns = [str(c).strip() for c in df.columns]
    mapa_colunas = mapa_colunas or {}

    # Quando houver mapeamento manual, ele manda. O reconhecimento automático fica só como fallback.
    col_fabrica = mapa_colunas.get("codigo") or _coluna_por_candidatos(df, [
        "Código Fábrica", "Codigo Fabrica", "Cód. Fábrica", "Cod. Fabrica",
        "Código de Fábrica", "Codigo de Fabrica", "Cod Fabrica", "Cód Fabrica",
        "Código", "Codigo", "Cód.", "Cod.", "Cod",
        "Referência", "Referencia", "Ref", "Código Fornecedor", "Codigo Fornecedor",
        "Cod Produto", "Código Produto", "Codigo Produto", "SKU", "Part Number", "Linha PDF",
    ])
    col_descricao = mapa_colunas.get("descricao") or _coluna_por_candidatos(df, [
        "descricao", "descrição", "descrição do item", "descricao do item", "produto", "item", "nome", "descr", "linha pdf",
    ])
    col_qtd = mapa_colunas.get("quantidade") or _coluna_quantidade_flexivel(df) or _coluna_por_candidatos(df, [
        "PEDIDO Final", "Quantidade", "Qtd", "Qtde", "Quant", "Quant.", "Qte", "Qty", "Quantity",
        "Quantidade Pedido", "Qtd Pedido", "Qtde Pedido", "Quant Pedido",
        "Quantidade Pedida", "Qtd Pedida", "Qtde Pedida", "Quant Pedida",
        "Quantidade Solicitada", "Qtd Solicitada", "Qtde Solicitada", "Quant Solicitada",
        "Quantidade Compra", "Qtd Compra", "Qtde Compra", "Qtd. Compra",
        "Quantidade Faturada", "Qtd Faturada", "Qtde Faturada",
        "Volume", "Vol",
    ])
    col_preco = mapa_colunas.get("preco_unitario") or _coluna_valor_unitario_flexivel(df) or _coluna_por_candidatos(df, [
        "Preço Última Compra", "Preco Ultima Compra", "Preço", "Preco", "Preço Unitário", "Preco Unitario",
        "Valor Unitário", "Valor Unitario", "Vlr Unit", "Vl Unit", "VL. UNIT.", "VL UNIT",
        "Vr.Unit", "VR.UNIT", "VR UNIT", "Unitário", "Unitario", "Preço Uni",
    ])
    col_total = mapa_colunas.get("valor_total") or _coluna_valor_total_flexivel(df) or _coluna_por_candidatos(df, [
        "Valor Final do Pedido", "Valor Total", "VL. TOTAL", "VL TOTAL", "Vlr Total",
        "Total", "Total Geral", "Valor", "Valor Mercadoria",
    ])

    if col_qtd and _coluna_auxiliar_nao_preco_quantidade(col_qtd):
        col_qtd_alt = _coluna_quantidade_flexivel(df)
        if col_qtd_alt and not _coluna_auxiliar_nao_preco_quantidade(col_qtd_alt):
            col_qtd = col_qtd_alt

    if col_preco and _coluna_auxiliar_nao_preco_quantidade(col_preco):
        col_preco_alt = _coluna_valor_unitario_flexivel(df)
        if col_preco_alt and not _coluna_auxiliar_nao_preco_quantidade(col_preco_alt):
            col_preco = col_preco_alt

    # Segurança contra PDF/Excel desalinhado: se a coluna mapeada como quantidade
    # parecer código de fábrica/EAN/produto, procura outra coluna ou força inferência pela linha.
    if col_qtd and col_qtd in df.columns and not _serie_parece_quantidade(df[col_qtd]):
        col_qtd_original = col_qtd
        col_qtd = None
        for candidato in df.columns:
            if candidato == col_qtd_original or candidato == col_fabrica:
                continue
            if _serie_parece_quantidade(df[candidato]) and _normalizar_nome_coluna_flex(candidato) not in ["CODIGO", "COD PRODUTO", "CODIGO PRODUTO", "CODIGO EAN"]:
                col_qtd = candidato
                break

    # Quando o arquivo veio de PDF sem cabeçalho perfeito, tenta inferir por linha completa.
    if (not col_qtd or not col_preco) and "Linha PDF" in df.columns:
        linhas_inferidas = df["Linha PDF"].astype(str).apply(_inferir_qtd_preco_total_por_linha)
        df["__qtd_inferida"] = [x[0] for x in linhas_inferidas]
        df["__preco_inferido"] = [x[1] for x in linhas_inferidas]
        df["__total_inferido"] = [x[2] for x in linhas_inferidas]
        col_qtd = col_qtd or "__qtd_inferida"
        col_preco = col_preco or "__preco_inferido"
        col_total = col_total or "__total_inferido"

    faltantes = []
    if not col_fabrica:
        faltantes.append("Código de Fábrica / relacionamento")
    if not col_qtd:
        faltantes.append("quantidade")
    if faltantes:
        raise ValueError(f"O arquivo {origem} precisa de mapeamento para: " + ", ".join(faltantes) + ".")

    out = pd.DataFrame()
    out["origem"] = origem
    out["codigo_fabrica"] = df[col_fabrica].astype(str).str.strip() if col_fabrica else ""
    out["codigo_fabrica_norm"] = out["codigo_fabrica"].apply(normalizar_codigo_fabrica)
    out["descricao"] = df[col_descricao].astype(str).str.strip() if col_descricao else out["codigo_fabrica"]
    out["quantidade"] = df[col_qtd].apply(numero_planilha_para_float)
    out["preco_unitario"] = df[col_preco].apply(numero_planilha_para_float) if col_preco else 0.0
    out["valor_total"] = df[col_total].apply(numero_planilha_para_float) if col_total else 0.0
    out.loc[(out["valor_total"] <= 0) & (out["preco_unitario"] > 0), "valor_total"] = out["quantidade"] * out["preco_unitario"]
    out.loc[(out["preco_unitario"] <= 0) & (out["valor_total"] > 0) & (out["quantidade"] > 0), "preco_unitario"] = out["valor_total"] / out["quantidade"]
    out["descricao_chave"] = out["descricao"].apply(normalizar_descricao_chave)
    out["__linha_origem"] = range(len(out))
    out = out[(out["quantidade"] > 0) | (out["preco_unitario"] > 0) | (out["valor_total"] > 0)].copy()
    return out

def agregar_pedido_comparativo(df):
    if df.empty:
        return df.copy()
    df = df.copy().reset_index(drop=True)
    df["codigo_fabrica"] = df["codigo_fabrica"].fillna("").astype(str).str.strip()
    df["codigo_fabrica_norm"] = df["codigo_fabrica_norm"].fillna("").astype(str).str.strip()

    # Regra principal do comparativo: relacionamento automático SOMENTE por Código de Fábrica.
    # Itens sem código não são aproximados por descrição; ficam separados e podem ser vinculados manualmente.
    df["chave"] = df.apply(
        lambda r: str(r["codigo_fabrica_norm"]) if str(r["codigo_fabrica_norm"]).strip() else f"SEM_CODIGO_{r.name}",
        axis=1,
    )
    # Mantém o preço unitário real lido do arquivo e calcula média ponderada pela quantidade.
    # Correção para planilhas que trazem simultaneamente VL. UNIT., UNIT.TOT e VL. TOTAL:
    # antes o agrupamento sempre fazia valor_total / quantidade, o que substituía o
    # valor unitário pela coluna de total unitário/total do item em alguns modelos.
    df["__valor_unitario_ponderado"] = pd.to_numeric(df["preco_unitario"], errors="coerce").fillna(0) * pd.to_numeric(df["quantidade"], errors="coerce").fillna(0)

    agg = df.groupby("chave", as_index=False).agg(
        codigo_fabrica=("codigo_fabrica", "first"),
        codigo_fabrica_norm=("codigo_fabrica_norm", "first"),
        descricao=("descricao", "first"),
        descricao_chave=("descricao_chave", "first"),
        quantidade=("quantidade", "sum"),
        valor_total=("valor_total", "sum"),
        __valor_unitario_ponderado=("__valor_unitario_ponderado", "sum"),
    )
    agg["preco_unitario"] = agg.apply(
        lambda r: (
            float(r["__valor_unitario_ponderado"]) / float(r["quantidade"])
            if float(r["quantidade"] or 0) > 0 and float(r["__valor_unitario_ponderado"] or 0) > 0
            else (float(r["valor_total"]) / float(r["quantidade"]) if float(r["quantidade"] or 0) > 0 and float(r["valor_total"] or 0) > 0 else 0)
        ),
        axis=1,
    )
    agg = agg.drop(columns=["__valor_unitario_ponderado"], errors="ignore")
    return agg


def codigos_referencia_comparativo(df_unica_raw, mapa_unica=None):
    try:
        unica_norm = normalizar_pedido_comparativo(df_unica_raw, "Única", mapa_unica)
        codigos = unica_norm["codigo_fabrica"].dropna().astype(str).str.strip().tolist()
        return [c for c in codigos if normalizar_codigo_fabrica(c)]
    except Exception:
        return []


def _aplicar_relacionamentos_manuais(unica, fornecedor, relacionamentos):
    relacionamentos = relacionamentos or {}
    if not relacionamentos:
        return fornecedor.copy()

    fornecedor = fornecedor.copy()
    for chave_unica, chave_fornecedor in relacionamentos.items():
        if not chave_unica or not chave_fornecedor:
            continue
        mask_f = fornecedor["chave"].astype(str) == str(chave_fornecedor)
        if mask_f.any():
            fornecedor.loc[mask_f, "chave"] = str(chave_unica)
    return fornecedor


def montar_comparativo_pedidos(df_unica_raw, df_fornecedor_raw, mapa_unica=None, mapa_fornecedor=None, relacionamentos_manuais=None, mapa_precos_brasilux=None):
    unica_normalizada = normalizar_pedido_comparativo(df_unica_raw, "Única", mapa_unica)

    # Regra do comparativo: itens com PEDIDO FINAL / quantidade da Única igual a zero
    # não entram na base de comparação. Isso evita apontar divergência de itens que
    # foram carregados na planilha, mas não foram efetivamente pedidos.
    if not unica_normalizada.empty and "quantidade" in unica_normalizada.columns:
        unica_normalizada["quantidade"] = pd.to_numeric(unica_normalizada["quantidade"], errors="coerce").fillna(0)
        unica_normalizada = unica_normalizada[unica_normalizada["quantidade"] > 0].copy()

    unica = agregar_pedido_comparativo(unica_normalizada)
    if mapa_precos_brasilux:
        unica = aplicar_precos_brasilux_no_pedido_unica(unica, mapa_precos_brasilux)
    fornecedor = agregar_pedido_comparativo(normalizar_pedido_comparativo(df_fornecedor_raw, "Fornecedor", mapa_fornecedor))
    fornecedor = _aplicar_relacionamentos_manuais(unica, fornecedor, relacionamentos_manuais)

    usados_fornecedor = set()
    linhas = []

    for _, row in unica.iterrows():
        match = None
        metodo = ""
        cod_fab = str(row.get("codigo_fabrica", "")).strip()
        cod_fab_norm = str(row.get("codigo_fabrica_norm", "")).strip()

        # 1) Relacionamento manual salvo pelo usuário.
        candidatos_manual = fornecedor[fornecedor["chave"].astype(str) == str(row.get("chave", ""))]
        candidatos_manual = candidatos_manual[~candidatos_manual.index.isin(usados_fornecedor)]
        if not candidatos_manual.empty:
            match = candidatos_manual.iloc[0]
            usados_fornecedor.add(match.name)
            metodo = "Relacionamento manual"

        # 2) Relacionamento automático pelo Código de Fábrica normalizado completo.
        if match is None and cod_fab_norm:
            candidatos = fornecedor[fornecedor["codigo_fabrica_norm"].astype(str).str.strip() == cod_fab_norm]
            candidatos = candidatos[~candidatos.index.isin(usados_fornecedor)]
            if not candidatos.empty:
                match = candidatos.iloc[0]
                usados_fornecedor.add(match.name)
                metodo = "Código de Fábrica"

        # 3) Fallback seguro pelo núcleo numérico do código.
        # Exemplo real do MasterSales: "AC 470200102" no PDF e "470200102"
        # no Pedido Única. Só relaciona quando o núcleo é único em ambos os lados.
        if match is None:
            nucleo_unica = codigo_fabrica_nucleo_numerico(cod_fab)
            if nucleo_unica:
                nucleos_unica = unica["codigo_fabrica"].apply(codigo_fabrica_nucleo_numerico)
                nucleos_forn = fornecedor["codigo_fabrica"].apply(codigo_fabrica_nucleo_numerico)
                if int((nucleos_unica == nucleo_unica).sum()) == 1 and int((nucleos_forn == nucleo_unica).sum()) == 1:
                    candidatos = fornecedor[nucleos_forn == nucleo_unica]
                    candidatos = candidatos[~candidatos.index.isin(usados_fornecedor)]
                    if not candidatos.empty:
                        match = candidatos.iloc[0]
                        usados_fornecedor.add(match.name)
                        metodo = "Código de Fábrica (número)"

        # Não usa descrição aproximada/fuzzy para evitar comparação de produtos parecidos.

        if match is None:
            linhas.append({
                "Status": "Não encontrado no fornecedor",
                "Método": "Sem correspondência",
                "Chave Única": row.get("chave", ""),
                "Chave Fornecedor": "",
                "Código Única": cod_fab,
                "Descrição Única": row.get("descricao", ""),
                "Qtd Única": row.get("quantidade", 0),
                "Preço Única": row.get("preco_unitario", 0),
                "Valor Única": row.get("valor_total", 0),
                "Código Fornecedor": "",
                "Descrição Fornecedor": "",
                "Qtd Fornecedor": 0,
                "Preço Fornecedor": 0,
                "Valor Fornecedor": 0,
                "Diferença Qtd": -float(row.get("quantidade", 0) or 0),
                "Diferença Preço": -float(row.get("preco_unitario", 0) or 0),
                "Diferença Preço %": -100.0 if float(row.get("preco_unitario", 0) or 0) > 0 else 0.0,
                "Diferença Valor": -float(row.get("valor_total", 0) or 0),
            })
            continue

        qtd_unica = float(row.get("quantidade", 0) or 0)
        preco_unica = float(row.get("preco_unitario", 0) or 0)
        valor_unica = float(row.get("valor_total", 0) or 0)
        qtd_fornecedor = float(match.get("quantidade", 0) or 0)
        preco_fornecedor = float(match.get("preco_unitario", 0) or 0)
        valor_fornecedor = float(match.get("valor_total", 0) or 0)

        dif_qtd = qtd_fornecedor - qtd_unica
        dif_preco = preco_fornecedor - preco_unica
        dif_preco_pct = (dif_preco / preco_unica * 100) if abs(preco_unica) > 0.000001 else (100.0 if abs(preco_fornecedor) > 0.000001 else 0.0)
        dif_valor = valor_fornecedor - valor_unica
        status = "OK" if abs(dif_qtd) < 0.0001 and abs(dif_preco) < 0.01 else "Divergente"
        linhas.append({
            "Status": status,
            "Método": metodo,
            "Chave Única": row.get("chave", ""),
            "Chave Fornecedor": match.get("chave", ""),
            "Código Única": cod_fab,
            "Descrição Única": row.get("descricao", ""),
            "Qtd Única": row.get("quantidade", 0),
            "Preço Única": row.get("preco_unitario", 0),
            "Valor Única": row.get("valor_total", 0),
            "Código Fornecedor": match.get("codigo_fabrica", ""),
            "Descrição Fornecedor": match.get("descricao", ""),
            "Qtd Fornecedor": match.get("quantidade", 0),
            "Preço Fornecedor": match.get("preco_unitario", 0),
            "Valor Fornecedor": match.get("valor_total", 0),
            "Diferença Qtd": dif_qtd,
            "Diferença Preço": dif_preco,
            "Diferença Preço %": dif_preco_pct,
            "Diferença Valor": dif_valor,
        })

    extras = fornecedor[~fornecedor.index.isin(usados_fornecedor)].copy()
    for _, row in extras.iterrows():
        linhas.append({
            "Status": "Somente fornecedor",
            "Método": "Sem correspondência",
            "Chave Única": "",
            "Chave Fornecedor": row.get("chave", ""),
            "Código Única": "",
            "Descrição Única": "",
            "Qtd Única": 0,
            "Preço Única": 0,
            "Valor Única": 0,
            "Código Fornecedor": row.get("codigo_fabrica", ""),
            "Descrição Fornecedor": row.get("descricao", ""),
            "Qtd Fornecedor": row.get("quantidade", 0),
            "Preço Fornecedor": row.get("preco_unitario", 0),
            "Valor Fornecedor": row.get("valor_total", 0),
            "Diferença Qtd": row.get("quantidade", 0),
            "Diferença Preço": row.get("preco_unitario", 0),
            "Diferença Preço %": 100.0 if float(row.get("preco_unitario", 0) or 0) > 0 else 0.0,
            "Diferença Valor": row.get("valor_total", 0),
        })

    return pd.DataFrame(linhas)


def gerar_relatorio_executivo_comparativo(df_comparativo, limite_itens=80):
    """
    Gera um relatório executivo em Markdown para a conferência do pedido.
    O texto é dividido por tópicos para facilitar copiar/colar no WhatsApp, e-mail ou ata.
    """
    if df_comparativo is None or df_comparativo.empty:
        return """# 📋 RELATÓRIO DE CONFERÊNCIA DO PEDIDO

## ✅ Resumo Geral

- Nenhum item foi encontrado para comparação.

## ✅ Conclusão

Não há divergências a reportar porque o comparativo está vazio.
"""

    df = df_comparativo.copy()
    for col in [
        "Qtd Única", "Qtd Fornecedor", "Preço Única", "Preço Fornecedor",
        "Diferença Qtd", "Diferença Preço", "Diferença Preço %", "Diferença Valor",
        "Valor Única", "Valor Fornecedor",
    ]:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0)

    total = len(df)
    qtd_ok = int((df["Status"] == "OK").sum()) if "Status" in df.columns else 0
    qtd_div = int((df["Status"] == "Divergente").sum()) if "Status" in df.columns else 0
    qtd_nao_fornecedor = int((df["Status"] == "Não encontrado no fornecedor").sum()) if "Status" in df.columns else 0
    qtd_somente_fornecedor = int((df["Status"] == "Somente fornecedor").sum()) if "Status" in df.columns else 0

    diverg_qtd = df[
        (df["Status"].isin(["Divergente", "Somente fornecedor", "Não encontrado no fornecedor"]))
        & (df["Diferença Qtd"].abs() > 0.0001)
    ].copy()
    diverg_preco = df[
        (df["Status"].isin(["Divergente", "Somente fornecedor", "Não encontrado no fornecedor"]))
        & (df["Diferença Preço"].abs() > 0.01)
    ].copy()

    def nome_item(row):
        codigo = str(row.get("Código Única") or row.get("Código Fornecedor") or "").strip()
        desc = str(row.get("Descrição Única") or row.get("Descrição Fornecedor") or "").strip()
        if codigo and desc:
            return f"{codigo} - {desc}"
        return codigo or desc or "item sem código"

    def sinal_num(valor, casas=1):
        try:
            valor = float(valor or 0)
            sinal = "+" if valor > 0 else ""
            return sinal + str(format_num_br(valor, casas))
        except Exception:
            return str(valor)

    def sinal_pct(valor):
        try:
            valor = float(valor or 0)
            sinal = "+" if valor > 0 else ""
            return sinal + str(format_num_br(valor, 2)) + "%"
        except Exception:
            return str(valor)

    linhas = []
    linhas.append("# 📋 RELATÓRIO DE CONFERÊNCIA DO PEDIDO")
    linhas.append("")
    linhas.append("## ✅ Resumo Geral")
    linhas.append("")
    linhas.append(f"- **Itens comparados:** {format_int_br(total)}")
    linhas.append(f"- **Itens sem divergência:** {format_int_br(qtd_ok)}")
    linhas.append(f"- **Itens com divergência:** {format_int_br(qtd_div)}")
    linhas.append(f"- **Itens no Pedido Única não encontrados no fornecedor:** {format_int_br(qtd_nao_fornecedor)}")
    linhas.append(f"- **Itens enviados pelo fornecedor que não constam no Pedido Única:** {format_int_br(qtd_somente_fornecedor)}")
    linhas.append("")

    linhas.append("## 📦 Divergências de Quantidade")
    linhas.append("")
    if diverg_qtd.empty:
        linhas.append("- Não foram identificadas divergências de quantidade.")
    else:
        linhas.append("Foram identificadas divergências nas quantidades dos seguintes itens:")
        linhas.append("")
        for _, r in diverg_qtd.head(limite_itens).iterrows():
            status = str(r.get("Status", ""))
            item = nome_item(r)
            linhas.append(f"### 🔸 {item}")
            if status == "Somente fornecedor":
                linhas.append("- **Situação:** item consta somente no pedido do fornecedor.")
                linhas.append(f"- **Quantidade no fornecedor:** {format_num_br(r.get('Qtd Fornecedor', 0), 1)}")
                linhas.append("- **Quantidade no Pedido Única:** 0")
                linhas.append("- **Ação recomendada:** verificar se o fornecedor incluiu item indevido ou se houve alteração posterior no pedido.")
            elif status == "Não encontrado no fornecedor":
                linhas.append("- **Situação:** item consta no Pedido Única, mas não foi encontrado no fornecedor.")
                linhas.append(f"- **Quantidade no Pedido Única:** {format_num_br(r.get('Qtd Única', 0), 1)}")
                linhas.append("- **Quantidade no fornecedor:** 0")
                linhas.append("- **Ação recomendada:** confirmar se houve corte, ruptura ou omissão do item pelo fornecedor.")
            else:
                linhas.append(f"- **Quantidade no Pedido Única:** {format_num_br(r.get('Qtd Única', 0), 1)}")
                linhas.append(f"- **Quantidade no fornecedor:** {format_num_br(r.get('Qtd Fornecedor', 0), 1)}")
                linhas.append(f"- **Diferença:** {sinal_num(r.get('Diferença Qtd', 0), 1)} unidade(s)")
            linhas.append("")
        if len(diverg_qtd) > limite_itens:
            linhas.append(f"- Existem mais **{format_int_br(len(diverg_qtd) - limite_itens)}** divergência(s) de quantidade não listadas neste texto.")
            linhas.append("")

    linhas.append("## 💰 Divergências de Preço")
    linhas.append("")
    if diverg_preco.empty:
        linhas.append("- Não foram identificadas divergências de preço unitário.")
    else:
        linhas.append("Foram identificadas divergências nos preços unitários dos seguintes itens:")
        linhas.append("")
        for _, r in diverg_preco.head(limite_itens).iterrows():
            status = str(r.get("Status", ""))
            item = nome_item(r)
            linhas.append(f"### 🔸 {item}")
            if status == "Somente fornecedor":
                linhas.append("- **Situação:** item consta somente no pedido do fornecedor.")
                linhas.append(f"- **Preço fornecedor:** {format_moeda_br(r.get('Preço Fornecedor', 0))}")
                linhas.append("- **Preço Pedido Única:** R$ 0,00")
                linhas.append("- **Ação recomendada:** validar se o item deve entrar no pedido antes da aprovação.")
            elif status == "Não encontrado no fornecedor":
                linhas.append("- **Situação:** item consta somente no Pedido Única.")
                linhas.append(f"- **Preço Pedido Única:** {format_moeda_br(r.get('Preço Única', 0))}")
                linhas.append("- **Preço fornecedor:** R$ 0,00")
                linhas.append("- **Ação recomendada:** confirmar se o fornecedor retirou o item ou se houve falha no arquivo recebido.")
            else:
                linhas.append(f"- **Preço Pedido Única:** {format_moeda_br(r.get('Preço Única', 0))}")
                linhas.append(f"- **Preço fornecedor:** {format_moeda_br(r.get('Preço Fornecedor', 0))}")
                linhas.append(f"- **Diferença unitária:** {format_moeda_br(r.get('Diferença Preço', 0))}")
                linhas.append(f"- **Diferença percentual:** {sinal_pct(r.get('Diferença Preço %', 0))}")
            linhas.append("")
        if len(diverg_preco) > limite_itens:
            linhas.append(f"- Existem mais **{format_int_br(len(diverg_preco) - limite_itens)}** divergência(s) de preço não listadas neste texto.")
            linhas.append("")

    somente_forn = df[df["Status"] == "Somente fornecedor"].copy()
    nao_encontrado = df[df["Status"] == "Não encontrado no fornecedor"].copy()

    linhas.append("## ⚠ Itens Apenas no Fornecedor")
    linhas.append("")
    if somente_forn.empty:
        linhas.append("- Nenhum item foi encontrado apenas no arquivo do fornecedor.")
    else:
        linhas.append("Os itens abaixo foram enviados pelo fornecedor, porém não constam no Pedido Única considerado para comparação:")
        linhas.append("")
        for _, r in somente_forn.head(limite_itens).iterrows():
            linhas.append(f"- **{nome_item(r)}** | Qtd: **{format_num_br(r.get('Qtd Fornecedor', 0), 1)}** | Preço: **{format_moeda_br(r.get('Preço Fornecedor', 0))}**")
        if len(somente_forn) > limite_itens:
            linhas.append(f"- Mais **{format_int_br(len(somente_forn) - limite_itens)}** item(ns) não listado(s).")
    linhas.append("")

    linhas.append("## ⚠ Itens do Pedido Única não encontrados no fornecedor")
    linhas.append("")
    if nao_encontrado.empty:
        linhas.append("- Nenhum item do Pedido Única ficou sem correspondência no fornecedor.")
    else:
        linhas.append("Os itens abaixo constam no Pedido Única, mas não foram encontrados no arquivo do fornecedor:")
        linhas.append("")
        for _, r in nao_encontrado.head(limite_itens).iterrows():
            linhas.append(f"- **{nome_item(r)}** | Qtd: **{format_num_br(r.get('Qtd Única', 0), 1)}** | Preço: **{format_moeda_br(r.get('Preço Única', 0))}**")
        if len(nao_encontrado) > limite_itens:
            linhas.append(f"- Mais **{format_int_br(len(nao_encontrado) - limite_itens)}** item(ns) não listado(s).")
    linhas.append("")

    percentual_ok = (qtd_ok / total * 100) if total else 0
    percentual_div = ((total - qtd_ok) / total * 100) if total else 0
    linhas.append("## 📊 Indicadores da Conferência")
    linhas.append("")
    linhas.append(f"- **Percentual sem divergência:** {format_num_br(percentual_ok, 2)}%")
    linhas.append(f"- **Percentual com algum apontamento:** {format_num_br(percentual_div, 2)}%")
    linhas.append(f"- **Total de divergências de quantidade:** {format_int_br(len(diverg_qtd))}")
    linhas.append(f"- **Total de divergências de preço:** {format_int_br(len(diverg_preco))}")
    linhas.append("")

    linhas.append("## ✅ Conclusão")
    linhas.append("")
    if qtd_div == 0 and qtd_nao_fornecedor == 0 and qtd_somente_fornecedor == 0:
        linhas.append("A conferência foi finalizada sem divergências relevantes. O pedido pode seguir para validação final.")
    else:
        linhas.append("A conferência encontrou pontos que precisam ser validados antes da aprovação final do pedido:")
        if len(diverg_qtd) > 0:
            linhas.append(f"- **{format_int_br(len(diverg_qtd))}** divergência(s) de quantidade.")
        if len(diverg_preco) > 0:
            linhas.append(f"- **{format_int_br(len(diverg_preco))}** divergência(s) de preço.")
        if qtd_somente_fornecedor > 0:
            linhas.append(f"- **{format_int_br(qtd_somente_fornecedor)}** item(ns) enviado(s) pelo fornecedor que não constam no Pedido Única.")
        if qtd_nao_fornecedor > 0:
            linhas.append(f"- **{format_int_br(qtd_nao_fornecedor)}** item(ns) do Pedido Única não encontrado(s) no fornecedor.")
        linhas.append("")
        linhas.append("Recomenda-se validar as divergências com o fornecedor antes de aprovar o pedido.")

    return "\n".join(linhas).strip()


def gerar_texto_divergencias_comparativo(df_comparativo):
    """
    Mantida por compatibilidade com o restante do app.
    Agora retorna o relatório executivo completo e um texto focado em preço.
    """
    relatorio = gerar_relatorio_executivo_comparativo(df_comparativo)
    return relatorio, relatorio

def colorir_comparativo_pedidos(row):
    status = str(row.get("Status", ""))
    if status == "OK":
        return ["background-color: #eaf7ea"] * len(row)
    if status == "Divergente":
        return ["background-color: #fff7ed"] * len(row)
    return ["background-color: #ffe8e8; font-weight: 600"] * len(row)


def gerar_excel_comparativo_pedidos(df_comparativo):
    if Workbook is None:
        raise RuntimeError("A biblioteca openpyxl não está instalada. Rode: python -m pip install openpyxl")
    wb = Workbook()
    ws = wb.active
    ws.title = "Comparativo"
    ws.append(list(df_comparativo.columns))
    for _, row in df_comparativo.iterrows():
        ws.append([row.get(col, "") for col in df_comparativo.columns])
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for cell in ws[1]:
        cell.font = cell.font.copy(bold=True)
    for idx, col in enumerate(df_comparativo.columns, start=1):
        letter = ws.cell(row=1, column=idx).column_letter
        if "Descrição" in col:
            ws.column_dimensions[letter].width = 42
        elif "%" in col:
            ws.column_dimensions[letter].width = 16
            for cell in ws[letter][1:]:
                cell.number_format = '0.00%'
                try:
                    cell.value = float(cell.value or 0) / 100
                except Exception:
                    pass
        elif "Preço" in col or "Valor" in col:
            ws.column_dimensions[letter].width = 16
            for cell in ws[letter][1:]:
                cell.number_format = 'R$ #,##0.00'
        elif "Qtd" in col:
            ws.column_dimensions[letter].width = 13
            for cell in ws[letter][1:]:
                cell.number_format = '#,##0.0'
        else:
            ws.column_dimensions[letter].width = 22
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


def render_pagina_comparativo_pedidos():
    st.markdown('<div class="page-card"><div class="page-card-title">Comparativo de Pedidos</div><div class="page-card-subtitle">Compare o pedido da Única com o fornecedor usando o Código de Fábrica como vínculo principal.</div>', unsafe_allow_html=True)

    if "relacionamentos_comparativo" not in st.session_state:
        st.session_state["relacionamentos_comparativo"] = {}

    modelo_fornecedor = st.selectbox(
        "Modelo do fornecedor",
        list(MODELOS_FORNECEDOR_COMPARATIVO.keys()),
        index=0,
        key="comparativo_modelo_fornecedor",
        help="Use Automático para manter a leitura atual. Ao selecionar um fornecedor homologado, o sistema tenta primeiro o modelo daquele fornecedor e só depois cai no automático.",
    )
    pedido_brasilux = _modelo_fornecedor_codigo(modelo_fornecedor) == "brasilux_mastersales"
    if pedido_brasilux:
        st.info(
            "Modo Brasilux ativo: itens e quantidades continuam vindo do Pedido Única; "
            "somente o preço de referência será substituído pela coluna D da tabela Brasilux."
        )
    elif _modelo_fornecedor_codigo(modelo_fornecedor) != "automatico":
        st.info(f"Modelo homologado ativo: {modelo_fornecedor}. Se esse modelo não reconhecer o arquivo, o sistema tenta a leitura automática como fallback.")

    col1, col2 = st.columns(2)
    with col1:
        link_unica_sheets = st.text_input(
            "Link do Google Sheets do pedido da Única",
            value="",
            key="link_comparativo_unica_sheets",
            placeholder="https://docs.google.com/spreadsheets/d/.../edit#gid=...",
        )
        link_unica_sheets = str(link_unica_sheets or "").strip()
        st.caption("Opcional. Se preenchido, o sistema lê a aba 'Pedido' do Sheets e ignora o upload da Única.")
        pedido_unica = st.file_uploader("Planilha do pedido da Única", type=["xlsx", "xls", "csv", "txt", "html", "htm"], key="upload_comparativo_unica")
    with col2:
        pedidos_fornecedor = st.file_uploader(
            "Pedido(s) do fornecedor",
            type=["xlsx", "xls", "csv", "txt", "pdf", "html", "htm", "png", "jpg", "jpeg", "webp", "bmp", "tif", "tiff"],
            accept_multiple_files=True,
            key="upload_comparativo_fornecedor",
            help="Você pode selecionar vários PDFs ou arquivos. O sistema consolida os itens antes de comparar com o Pedido Única.",
        )
        st.caption(
            "Quando o mesmo código aparecer em mais de um arquivo, as quantidades serão somadas "
            "e o preço do fornecedor será calculado pela média ponderada das quantidades."
        )

    st.markdown("</div>", unsafe_allow_html=True)

    if (not link_unica_sheets and not pedido_unica) or not pedidos_fornecedor:
        st.info("Cole o link ou envie o pedido da Única, e envie um ou mais arquivos do fornecedor para iniciar o comparativo.")
        return

    try:
        if link_unica_sheets:
            df_unica = ler_pedido_unica_comparativo_google_sheets(link_unica_sheets)
            origem_unica = "Google Sheets"
        else:
            df_unica = ler_arquivo_comparativo(pedido_unica)
            origem_unica = "upload"
        if df_unica.empty:
            st.error("Não consegui ler o pedido da Única.")
            return
        st.success(f"Pedido da Única lido via {origem_unica}: {len(df_unica)} linha(s).")

        st.markdown("### 1. Conferência e mapeamento das colunas")
        with st.expander("Prévia do Pedido Única", expanded=False):
            st.dataframe(df_unica.head(30), use_container_width=True, hide_index=True)

        mapa_unica = _mapear_colunas_comparativo(df_unica, "cmp_unica", "Pedido Única")
        if not mapa_unica:
            return

        codigos_ref = codigos_referencia_comparativo(df_unica, mapa_unica)

        mapa_precos_brasilux = None
        if pedido_brasilux:
            with st.spinner("Lendo a tabela de preços Brasilux..."):
                df_precos_brasilux = ler_tabela_precos_brasilux_google_sheets()
                mapa_precos_brasilux, coluna_codigo_brasilux, qtd_relacionada_brasilux = montar_mapa_precos_brasilux(
                    df_precos_brasilux, codigos_ref
                )
            st.success(
                f"Tabela Brasilux carregada: {qtd_relacionada_brasilux} código(s) relacionado(s). "
                "Preço de referência: coluna D."
            )

        df_fornecedor, arquivos_fornecedor_lidos, erros_fornecedor = ler_multiplos_arquivos_comparativo(
            pedidos_fornecedor,
            codigos_referencia=codigos_ref,
            modelo_fornecedor=modelo_fornecedor,
        )
        if df_fornecedor.empty:
            st.error(
                "Não consegui ler nenhum arquivo do fornecedor. Se algum arquivo for imagem ou PDF escaneado, "
                "o OCR/Tesseract precisa estar disponível no ambiente do deploy."
            )
            if erros_fornecedor:
                with st.expander("Detalhes dos arquivos não processados", expanded=True):
                    for erro in erros_fornecedor:
                        st.write(f"- {erro}")
            return

        st.success(
            f"Pedido do fornecedor consolidado: {len(arquivos_fornecedor_lidos)} arquivo(s) lido(s), "
            f"{len(df_fornecedor)} linha(s) antes da consolidação por código."
        )
        if erros_fornecedor:
            st.warning(
                f"{len(erros_fornecedor)} arquivo(s) não puderam ser processados. "
                "Os demais foram considerados normalmente."
            )
            with st.expander("Ver arquivos não processados", expanded=False):
                for erro in erros_fornecedor:
                    st.write(f"- {erro}")

        with st.expander("Prévia consolidada dos Pedidos do Fornecedor", expanded=False):
            colunas_preview = [c for c in ["Arquivo de origem", *list(df_fornecedor.columns)] if c in df_fornecedor.columns]
            colunas_preview = list(dict.fromkeys(colunas_preview))
            st.dataframe(df_fornecedor[colunas_preview].head(100), use_container_width=True, hide_index=True)

        # PDFs reconhecidos pelos parsers dedicados já retornam o padrão canônico.
        # Nesses casos, o comparativo segue direto, sem solicitar mapeamento manual.
        colunas_canonicas_fornecedor = {
            "Código Fábrica", "Quantidade", "Valor Unitário", "Valor Total"
        }
        if colunas_canonicas_fornecedor.issubset(set(df_fornecedor.columns)):
            mapa_fornecedor = {
                "codigo": "Código Fábrica",
                "descricao": "Descrição" if "Descrição" in df_fornecedor.columns else None,
                "quantidade": "Quantidade",
                "preco_unitario": "Valor Unitário",
                "valor_total": "Valor Total",
            }
            st.success(
                "Colunas do Pedido Fornecedor identificadas automaticamente: "
                "Código de Fábrica, Quantidade, Valor Unitário e Valor Total."
            )
        else:
            mapa_fornecedor = _mapear_colunas_comparativo(
                df_fornecedor, "cmp_fornecedor", "Pedido Fornecedor"
            )
            if not mapa_fornecedor:
                return

        st.markdown("### 2. Relacionamento manual dos itens sem identificação")
        relacionamentos = st.session_state.get("relacionamentos_comparativo", {})
        comparativo_base = montar_comparativo_pedidos(
            df_unica, df_fornecedor, mapa_unica, mapa_fornecedor, relacionamentos, mapa_precos_brasilux
        )

        nao_encontrados = comparativo_base[comparativo_base["Status"] == "Não encontrado no fornecedor"].copy()
        somente_fornecedor = comparativo_base[comparativo_base["Status"] == "Somente fornecedor"].copy()

        if not nao_encontrados.empty and not somente_fornecedor.empty:
            with st.expander("Relacionar manualmente itens não encontrados", expanded=True):
                st.caption("O vínculo automático é feito apenas por Código de Fábrica. Use esta tela somente para relacionar manualmente códigos que não bateram entre os arquivos.")
                opcoes_fornecedor = {"-- Não relacionar --": ""}
                for _, r in somente_fornecedor.iterrows():
                    label = f'{str(r.get("Código Fornecedor", "")).strip()} | {str(r.get("Descrição Fornecedor", "")).strip()} | Qtd {format_num_br(r.get("Qtd Fornecedor", 0), 1)} | R$ {format_num_br(r.get("Preço Fornecedor", 0), 2)}'
                    opcoes_fornecedor[label[:250]] = str(r.get("Chave Fornecedor", ""))

                novos_rel = dict(relacionamentos)
                limite_manual = min(len(nao_encontrados), 80)
                for i, (_, r) in enumerate(nao_encontrados.head(limite_manual).iterrows()):
                    st.markdown(f'**Única:** {str(r.get("Código Única", "")).strip()} | {str(r.get("Descrição Única", "")).strip()} | Qtd {format_num_br(r.get("Qtd Única", 0), 1)}')
                    escolha = st.selectbox(
                        "Item correspondente no fornecedor",
                        list(opcoes_fornecedor.keys()),
                        key=f"rel_manual_{i}_{str(r.get('Chave Única', ''))[:20]}",
                    )
                    chave_f = opcoes_fornecedor.get(escolha, "")
                    if chave_f:
                        novos_rel[str(r.get("Chave Única", ""))] = chave_f

                if len(nao_encontrados) > limite_manual:
                    st.info(f"Mostrando os primeiros {limite_manual} itens para relacionamento manual.")

                c_rel1, c_rel2 = st.columns(2)
                if c_rel1.button("Aplicar relacionamentos manuais", type="primary"):
                    st.session_state["relacionamentos_comparativo"] = novos_rel
                    st.rerun()
                if c_rel2.button("Limpar relacionamentos manuais"):
                    st.session_state["relacionamentos_comparativo"] = {}
                    st.rerun()
        else:
            st.info("Não há itens pendentes para relacionamento manual neste momento.")

        st.markdown("### 3. Resultado do comparativo")
        comparativo = montar_comparativo_pedidos(
            df_unica,
            df_fornecedor,
            mapa_unica,
            mapa_fornecedor,
            st.session_state.get("relacionamentos_comparativo", {}),
            mapa_precos_brasilux,
        )
        st.success(f"Comparativo gerado com {len(comparativo)} linha(s).")

        c1, c2, c3, c4 = st.columns(4)
        c1.metric("OK", int((comparativo["Status"] == "OK").sum()))
        c2.metric("Divergentes", int((comparativo["Status"] == "Divergente").sum()))
        c3.metric("Não encontrados", int((comparativo["Status"] == "Não encontrado no fornecedor").sum()))
        c4.metric("Somente fornecedor", int((comparativo["Status"] == "Somente fornecedor").sum()))

        relatorio_executivo = gerar_relatorio_executivo_comparativo(comparativo)
        with st.expander("📋 Relatório executivo da conferência", expanded=True):
            st.markdown(relatorio_executivo)
            st.text_area(
                "Texto pronto para copiar",
                relatorio_executivo,
                height=420,
                key="texto_relatorio_executivo_comparativo",
            )

        termo = st.text_input("Pesquisar no comparativo", key="busca_comparativo_pedidos").strip().lower()
        view = comparativo.copy()
        if termo:
            filtro = pd.Series(False, index=view.index)
            for col in ["Código Única", "Descrição Única", "Código Fornecedor", "Descrição Fornecedor", "Status", "Método"]:
                if col in view.columns:
                    filtro = filtro | view[col].astype(str).str.lower().str.contains(termo, na=False)
            view = view[filtro].copy()

        colunas_ocultar = ["Chave Única", "Chave Fornecedor"]
        view_exibicao = view.drop(columns=colunas_ocultar, errors="ignore")
        st.dataframe(
            view_exibicao.style.apply(colorir_comparativo_pedidos, axis=1).format(formatadores_para_tabela(view_exibicao)),
            use_container_width=True,
            hide_index=True,
            height=620,
        )

        st.download_button(
            "Baixar comparativo em Excel",
            gerar_excel_comparativo_pedidos(comparativo.drop(columns=colunas_ocultar, errors="ignore")),
            "comparativo_pedidos.xlsx",
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            type="primary",
        )
    except Exception as e:
        st.error(str(e))

def inicializar_pedido_editavel(tabela_resumo):
    colunas_base = colunas_pedido_compras(MESES)

    base = tabela_resumo.copy()
    if "Estoque Geral" not in base.columns and "Estoque Atual Geral" in base.columns:
        base["Estoque Geral"] = base["Estoque Atual Geral"]

    for col in colunas_base:
        if col not in base.columns:
            base[col] = 0 if col not in ["codigo", "descricao", "Código Fábrica", "Data Última Compra"] else ""

    # garante existência das colunas calculadas antes de ordenar
    base["PEDIDO Final"] = pd.to_numeric(base["Sugestão arredondada"], errors="coerce").fillna(0).round(0).astype(int)
    base["Origem Sugestão"] = "Sugestão do sistema"
    base["Valor Final do Pedido"] = base["PEDIDO Final"] * pd.to_numeric(base["Preço Última Compra"], errors="coerce").fillna(0)
    base = base[colunas_base].copy()
    return base

def atualizar_valor_e_origem(df):
    df = df.copy()
    df["PEDIDO Final"] = pd.to_numeric(df.get("PEDIDO Final", 0), errors="coerce").fillna(0).round(0).astype(int)
    df["Sugestão Sistema"] = pd.to_numeric(df.get("Sugestão Sistema", 0), errors="coerce").fillna(0).round(0).astype(int)
    df["Sugestão arredondada"] = pd.to_numeric(df.get("Sugestão arredondada", df["Sugestão Sistema"]), errors="coerce").fillna(0).round(0).astype(int)
    df["Preço Última Compra"] = pd.to_numeric(df.get("Preço Última Compra", 0), errors="coerce").fillna(0)
    df["Valor Final do Pedido"] = df["PEDIDO Final"] * df["Preço Última Compra"]
    df["Origem Sugestão"] = df.apply(
        lambda row: "Sugestão do sistema" if int(row["PEDIDO Final"]) == int(row["Sugestão arredondada"]) else "Alterado pelo usuário",
        axis=1,
    )
    return df




def ajustar_pedido_para_multiplo_embalagem(qtd, embalagem):
    """
    Valida o PEDIDO Final pelo múltiplo da embalagem.
    Se a quantidade não for múltipla, ajusta sempre para o próximo múltiplo acima.
    Ex.: qtd 45 e embalagem 12 => 48.
    """
    try:
        qtd = int(round(float(qtd or 0)))
    except Exception:
        qtd = 0

    try:
        embalagem = int(round(float(embalagem or 0)))
    except Exception:
        embalagem = 0

    if qtd <= 0:
        return 0
    if embalagem <= 1:
        return qtd
    if qtd % embalagem == 0:
        return qtd
    return int(math.ceil(qtd / embalagem) * embalagem)


def validar_pedidos_por_embalagem(df):
    """
    Ajusta todos os pedidos para múltiplos da embalagem e devolve mensagens de alerta.
    """
    df = df.copy()
    mensagens = []

    if "Embalagem" not in df.columns:
        df["Embalagem"] = 0

    df["PEDIDO Final"] = pd.to_numeric(df.get("PEDIDO Final", 0), errors="coerce").fillna(0).round(0).astype(int)
    df["Embalagem"] = pd.to_numeric(df.get("Embalagem", 0), errors="coerce").fillna(0).round(0).astype(int)

    for idx, row in df.iterrows():
        qtd_original = int(row.get("PEDIDO Final", 0) or 0)
        embalagem = int(row.get("Embalagem", 0) or 0)
        qtd_ajustada = ajustar_pedido_para_multiplo_embalagem(qtd_original, embalagem)

        if qtd_original > 0 and embalagem > 1 and qtd_original != qtd_ajustada:
            codigo = str(row.get("codigo", "")).zfill(5)
            descricao = str(row.get("descricao", "")).strip()
            mensagens.append(
                f"Item {codigo} - {descricao}: a embalagem é com {embalagem} unidades. "
                f"O pedido {qtd_original} foi ajustado para {qtd_ajustada}."
            )
            df.at[idx, "PEDIDO Final"] = qtd_ajustada

    return df, mensagens

def totalizar_valor_pedido(df):
    if df.empty:
        return 0.0
    qtd = pd.to_numeric(df.get("PEDIDO Final", 0), errors="coerce").fillna(0)
    preco = pd.to_numeric(df.get("Preço Última Compra", 0), errors="coerce").fillna(0)
    return float((qtd * preco).sum())



# =========================================================
# UI / EXPERIÊNCIA DO USURIO
# =========================================================

APP_NAME = "Análise de Giro e Pedido de Compra"


def aplicar_css_global():
    st.markdown(
        """
        <style>
            :root {
                --primary: #1d4ed8;
                --primary-soft: #eff6ff;
                --bg-soft: #f8fafc;
                --border: #e2e8f0;
                --text-muted: #64748b;
            }

            .main .block-container {
                padding-top: 1.25rem;
                padding-bottom: 2.5rem;
                max-width: 1500px;
            }

            [data-testid="stSidebar"] {
                background: linear-gradient(180deg, #0f172a 0%, #1e293b 100%);
            }

            [data-testid="stSidebar"] * {
                color: #f8fafc !important;
            }

            [data-testid="stSidebar"] .stRadio label,
            [data-testid="stSidebar"] .stNumberInput label {
                color: #e2e8f0 !important;
            }

            .hero-card {
                background: linear-gradient(135deg, #0f172a 0%, #1d4ed8 52%, #2563eb 100%);
                color: white;
                border-radius: 24px;
                padding: 28px 32px;
                margin-bottom: 22px;
                box-shadow: 0 18px 45px rgba(15, 23, 42, .18);
            }

            .hero-card h1 {
                margin: 0;
                font-size: 34px;
                line-height: 1.1;
                color: white;
                letter-spacing: -0.02em;
            }

            .hero-card p {
                margin: 10px 0 0 0;
                font-size: 15px;
                color: #dbeafe;
            }

            .section-title {
                font-size: 22px;
                font-weight: 800;
                margin: 20px 0 8px 0;
                color: #0f172a;
            }

            .muted {
                color: var(--text-muted);
                font-size: 14px;
            }

            .metric-card {
                background: white;
                border: 1px solid var(--border);
                border-radius: 18px;
                padding: 18px 20px;
                min-height: 118px;
                box-shadow: 0 8px 24px rgba(15, 23, 42, .06);
            }

            .metric-card .label {
                color: #64748b;
                font-size: 13px;
                font-weight: 700;
                text-transform: uppercase;
                letter-spacing: .04em;
            }

            .metric-card .value {
                margin-top: 8px;
                color: #0f172a;
                font-size: 28px;
                font-weight: 850;
                letter-spacing: -0.03em;
            }

            .metric-card .hint {
                margin-top: 5px;
                color: #64748b;
                font-size: 13px;
            }

            .upload-card {
                background: #ffffff;
                border: 1px solid var(--border);
                border-radius: 18px;
                padding: 16px 18px;
                box-shadow: 0 8px 24px rgba(15, 23, 42, .05);
                margin-bottom: 8px;
            }

            .upload-card strong {
                color: #0f172a;
                font-size: 16px;
            }

            .status-ok {
                color: #047857;
                font-weight: 800;
            }

            .status-warn {
                color: #b45309;
                font-weight: 800;
            }

            .download-card {
                background: white;
                border: 1px solid var(--border);
                border-radius: 18px;
                padding: 18px;
                box-shadow: 0 8px 24px rgba(15, 23, 42, .05);
                margin-bottom: 14px;
            }

            div[data-testid="stDataFrame"], div[data-testid="stDataEditor"] {
                border-radius: 16px;
                overflow: hidden;
                border: 1px solid #e2e8f0;
            }

            .stButton > button, .stDownloadButton > button {
                border-radius: 12px !important;
                font-weight: 800 !important;
            }


            /* Navegação lateral profissional */
            [data-testid="stSidebar"] {
                background: radial-gradient(circle at top left, #1e3a8a 0%, #0f172a 38%, #07111f 100%) !important;
                border-right: 1px solid rgba(148, 163, 184, .18);
            }
            [data-testid="stSidebar"] > div:first-child {
                padding-top: 1.15rem;
            }
            [data-testid="stSidebar"] h3 {
                font-size: 18px !important;
                letter-spacing: -0.02em;
                margin-bottom: 2px !important;
            }
            [data-testid="stSidebar"] hr {
                border-color: rgba(148, 163, 184, .18) !important;
                margin: 1.25rem 0 !important;
            }
            [data-testid="stSidebar"] [role="radiogroup"] label {
                background: rgba(15, 23, 42, .38);
                border: 1px solid transparent;
                border-radius: 14px;
                padding: 11px 12px;
                margin: 7px 0;
                transition: all .18s ease;
            }
            [data-testid="stSidebar"] [role="radiogroup"] label:hover {
                background: rgba(37, 99, 235, .16);
                border-color: rgba(96, 165, 250, .22);
            }
            [data-testid="stSidebar"] [role="radiogroup"] label:has(input:checked) {
                background: linear-gradient(135deg, rgba(37, 99, 235, .34), rgba(30, 41, 59, .7));
                border-color: rgba(96, 165, 250, .45);
                box-shadow: inset 3px 0 0 #3b82f6;
            }
            [data-testid="stSidebar"] [role="radiogroup"] label > div:first-child {
                display: none !important;
            }
            [data-testid="stSidebar"] input,
            [data-testid="stSidebar"] textarea,
            [data-testid="stSidebar"] [data-baseweb="input"] input {
                color: #0f172a !important;
                -webkit-text-fill-color: #0f172a !important;
                background: #ffffff !important;
                caret-color: #1d4ed8 !important;
                font-weight: 800 !important;
            }
            [data-testid="stSidebar"] [data-baseweb="input"] {
                background: #ffffff !important;
                border-radius: 14px !important;
                border: 1px solid rgba(96, 165, 250, .42) !important;
                box-shadow: 0 8px 26px rgba(15, 23, 42, .28);
            }
            [data-testid="stSidebar"] button {
                color: #0f172a !important;
            }
            .sidebar-brand {
                display: flex;
                gap: 12px;
                align-items: center;
                padding: 10px 8px 18px 2px;
                border-bottom: 1px solid rgba(148, 163, 184, .18);
                margin-bottom: 16px;
            }
            .sidebar-brand-icon {
                width: 38px;
                height: 38px;
                border-radius: 12px;
                display: grid;
                place-items: center;
                background: rgba(37, 99, 235, .18);
                border: 1px solid rgba(96, 165, 250, .35);
                color: #93c5fd;
            }
            .sidebar-brand-title {
                color: #ffffff;
                font-size: 18px;
                line-height: 1.1;
                font-weight: 850;
            }
            .sidebar-brand-subtitle {
                margin-top: 4px;
                color: #cbd5e1;
                font-size: 12px;
            }
            .param-card {
                background: rgba(15, 23, 42, .42);
                border: 1px solid rgba(148, 163, 184, .22);
                border-radius: 18px;
                padding: 15px 14px;
                margin-top: 8px;
                box-shadow: 0 12px 30px rgba(2, 6, 23, .22);
            }
            .param-note {
                background: rgba(37, 99, 235, .16);
                border: 1px solid rgba(96, 165, 250, .28);
                border-radius: 13px;
                padding: 10px 12px;
                color: #dbeafe;
                font-size: 12.5px;
                margin-top: 12px;
            }

            /* Cards e inputs principais */
            .page-card {
                background: #ffffff;
                border: 1px solid #dbe3ef;
                border-radius: 20px;
                padding: 20px;
                box-shadow: 0 14px 36px rgba(15, 23, 42, .06);
                margin-bottom: 16px;
            }
            .page-card-title {
                font-size: 18px;
                font-weight: 850;
                color: #0f172a;
                margin-bottom: 4px;
            }
            .page-card-subtitle {
                color: #64748b;
                font-size: 13px;
                margin-bottom: 14px;
            }
            div[data-testid="stFileUploader"] section {
                background: #f8fafc !important;
                border: 1px dashed #bfd0e6 !important;
                border-radius: 16px !important;
                padding: 14px !important;
            }
            div[data-testid="stFileUploader"] button,
            .stLinkButton > a,
            .stButton > button,
            .stDownloadButton > button,
            button[kind="formSubmit"] {
                border-radius: 12px !important;
                font-weight: 800 !important;
                min-height: 38px;
            }
            .stLinkButton > a {
                border: 1px solid #cbd5e1 !important;
                background: #ffffff !important;
                color: #0f172a !important;
            }
            div[data-baseweb="input"], div[data-baseweb="textarea"] {
                border-radius: 13px !important;
            }
            div[data-baseweb="input"] input,
            div[data-baseweb="textarea"] textarea {
                color: #0f172a !important;
                -webkit-text-fill-color: #0f172a !important;
                font-weight: 650 !important;
            }
            .sheets-badge {
                display: inline-flex;
                align-items: center;
                gap: 9px;
                background: #ecfdf5;
                border: 1px solid #bbf7d0;
                color: #047857;
                font-weight: 850;
                padding: 9px 12px;
                border-radius: 999px;
                margin: 4px 0 10px 0;
            }

        </style>
        """,
        unsafe_allow_html=True,
    )


def render_header(subtitulo="Sistema de apoio à decisão para giro, estoque e compra."):
    st.markdown(
        f"""
        <div class="hero-card">
            <h1>{APP_NAME}</h1>
            <p>{subtitulo}</p>
        </div>
        """,
        unsafe_allow_html=True,
    )


def render_metric(label, value, hint=""):
    st.markdown(
        f"""
        <div class="metric-card">
            <div class="label">{label}</div>
            <div class="value">{value}</div>
            <div class="hint">{hint}</div>
        </div>
        """,
        unsafe_allow_html=True,
    )


def render_upload_status(titulo, arquivo, obrigatorio=False):
    status = "✓ Arquivo carregado" if arquivo else ("Obrigatório" if obrigatorio else "Opcional")
    classe = "status-ok" if arquivo else "status-warn"
    nome = getattr(arquivo, "name", "") if arquivo else ""
    st.markdown(
        f"""
        <div class="upload-card">
            <strong>{titulo}</strong><br>
            <span class="{classe}">{status}</span><br>
            <span class="muted">{nome}</span>
        </div>
        """,
        unsafe_allow_html=True,
    )


def render_kpis_gerais(tabela_resumo, pedido_df=None):
    pedido_ref = pedido_df.copy() if pedido_df is not None and not pedido_df.empty else inicializar_pedido_editavel(tabela_resumo)
    pedido_ref = atualizar_valor_e_origem(pedido_ref)
    total_produtos = len(tabela_resumo)
    itens_compra = int((pd.to_numeric(pedido_ref.get("PEDIDO Final", 0), errors="coerce").fillna(0) > 0).sum())
    valor_pedido = totalizar_valor_pedido(pedido_ref)
    sem_compra = int(tabela_resumo.get("Data Última Compra", pd.Series(dtype=str)).astype(str).str.contains("⚠", na=False).sum())

    c1, c2, c3, c4 = st.columns(4)
    with c1:
        render_metric("Produtos analisados", format_int_br(total_produtos), "Itens processados no giro")
    with c2:
        render_metric("Itens com compra", format_int_br(itens_compra), "Pedido final maior que zero")
    with c3:
        render_metric("Valor do pedido", format_moeda_br(valor_pedido), "Quantidade × última compra")
    with c4:
        render_metric("Alertas sem compra", format_int_br(sem_compra), "Conforme parâmetro definido")


def render_download_card(titulo, descricao):
    st.markdown(
        f"""
        <div class="download-card">
            <strong>{titulo}</strong><br>
            <span class="muted">{descricao}</span>
        </div>
        """,
        unsafe_allow_html=True,
    )



# =========================================================
# MULTIPAGE INDEPENDENTE - RUPTURA POR MARCA
# =========================================================

def _token_numero_br_para_float(token):
    return br_to_float(token)


def _extrair_meses_cabecalho_marca(text):
    meses = []
    for raw_line in str(text or "").splitlines():
        line = str(raw_line or "")
        if "COD" in line.upper() and "DESCR" in line.upper() and "MEDIA" in line.upper():
            meses_linha = re.findall(r"\b\d{2}/\d{4}\b", line)
            if meses_linha:
                return meses_linha[:6]

    match = re.search(r"REFERENTE AOS MESES:\s*([^\n]+)", text, flags=re.IGNORECASE)
    if match:
        meses = list(reversed(re.findall(r"\d{2}/\d{4}", match.group(1))))
    if not meses:
        meses = sorted(set(re.findall(r"\b\d{2}/\d{4}\b", text)))[:4]
    return meses[:6] if meses else []


def parse_linha_giro_marca_independente(line, meses_ref):
    """
    Parser independente para o relatório de Giro Geral por Marca.
    Não interfere na lógica atual do dashboard.
    """
    line = str(line).strip()
    if not re.match(r"^\d{5}\s+", line):
        return None

    partes = line.split()
    if len(partes) < 10:
        return None

    codigo = partes[0].zfill(5)
    qtd_meses = len(meses_ref) if meses_ref else 4

    un_index = None
    for i in range(1, len(partes) - (qtd_meses + 3)):
        proximos = partes[i + 1:i + 1 + qtd_meses + 3]
        if len(proximos) >= qtd_meses + 3 and all(_eh_numero_br(t) for t in proximos[:qtd_meses + 3]):
            un_index = i
            break

    if un_index is None:
        # fallback para relatórios que usam UN como unidade principal
        for i, token in enumerate(partes):
            if token.upper() in ["UN", "LT", "GL", "CX", "PC", "MT", "KG", "DC"]:
                proximos = partes[i + 1:i + 1 + qtd_meses + 3]
                if len(proximos) >= qtd_meses + 3 and all(_eh_numero_br(t) for t in proximos[:qtd_meses + 3]):
                    un_index = i
                    break

    if un_index is None:
        return None

    descricao = " ".join(partes[1:un_index]).strip()
    unidade = partes[un_index]
    valores = partes[un_index + 1:]

    if len(valores) < qtd_meses + 3:
        return None

    registro = {
        "codigo": codigo,
        "descricao": descricao,
        "unidade": unidade,
    }

    for idx, mes in enumerate(meses_ref[:qtd_meses]):
        registro[mes] = _token_numero_br_para_float(valores[idx])

    registro["media_pdf"] = _token_numero_br_para_float(valores[qtd_meses])
    registro["previsao_30_pdf"] = _token_numero_br_para_float(valores[qtd_meses + 1])
    registro["estoque"] = _token_numero_br_para_float(valores[qtd_meses + 2])
    return registro


@st.cache_data(show_spinner="Lendo PDF de ruptura por marca...")
def parse_pdf_ruptura_por_marca(bytes_pdf):
    """
    Leitor otimizado para PDF grande.

    O relatório de Giro Geral por Marca pode ter mais de 1.500 páginas.
    pdfplumber é bom para tabelas, mas fica muito lento nesse volume.
    Aqui usamos PyMuPDF primeiro, que extrai texto página a página com muito mais velocidade,
    e só caímos para pdfplumber se PyMuPDF não estiver disponível.
    """
    registros = []
    empresa_atual = None
    relatorio_consolidado = False
    marca_cod_atual = ""
    marca_nome_atual = "SEM MARCA"
    meses_ref = []

    def processar_texto_pagina(page_text):
        nonlocal empresa_atual, relatorio_consolidado, marca_cod_atual, marca_nome_atual, meses_ref, registros

        if not meses_ref:
            meses_extraidos = _extrair_meses_cabecalho_marca(page_text or "")
            if meses_extraidos:
                meses_ref = meses_extraidos

        if not meses_ref:
            return

        if not empresa_atual:
            empresa_lista = re.search(r"EMPRESA\s*:\s*([0-9_]+)", str(page_text or ""), flags=re.IGNORECASE)
            if empresa_lista and "_" in empresa_lista.group(1):
                empresa_atual = "GERAL"
                relatorio_consolidado = True

        for raw_line in str(page_text or "").splitlines():
            line = raw_line.strip()
            if not line:
                continue

            empresa_match = re.search(r"EMPRESA\s*:\s*(\d{3})\s*-", line, flags=re.IGNORECASE)
            if empresa_match:
                empresa_atual = empresa_match.group(1)
                continue

            marca_match = re.search(r"MARCA\s*:\s*([^\n]+)", line, flags=re.IGNORECASE)
            if marca_match:
                marca_raw = marca_match.group(1).strip()
                marca_partes = re.match(r"([^\-]+)\s*-\s*(.*)", marca_raw)
                if marca_partes:
                    marca_cod_atual = marca_partes.group(1).strip()
                    marca_nome_atual = marca_partes.group(2).strip() or "SEM MARCA"
                else:
                    marca_cod_atual = ""
                    marca_nome_atual = marca_raw or "SEM MARCA"
                continue

            if not empresa_atual:
                continue
            if empresa_atual != "GERAL" and empresa_atual not in LOJAS_MAP:
                continue

            produto = parse_linha_giro_marca_independente(line, meses_ref)
            if produto:
                produto["codigo_empresa"] = empresa_atual
                produto["loja"] = "GERAL" if empresa_atual == "GERAL" else LOJAS_MAP.get(empresa_atual, empresa_atual)
                produto["tipo_unidade"] = "GERAL" if relatorio_consolidado else ("ÚNICA" if empresa_atual == CODIGO_UNICA else "LOJAS DAUTO")
                produto["marca_codigo"] = marca_cod_atual
                produto["marca"] = marca_nome_atual if marca_nome_atual else "SEM MARCA"
                registros.append(produto)

    try:
        import fitz  # PyMuPDF
    except Exception as e:
        raise RuntimeError(
            "Para esta página, instale o PyMuPDF. Rode: pip install pymupdf. "
            "O pdfplumber é lento demais para este relatório com muitas páginas."
        ) from e

    try:
        with fitz.open(stream=bytes_pdf, filetype="pdf") as doc:
            for page in doc:
                processar_texto_pagina(page.get_text("text", sort=True))
    except Exception as e:
        raise RuntimeError(f"Falha ao extrair texto do PDF com PyMuPDF: {e}") from e

    if not meses_ref:
        meses_ref = ["Mês 1", "Mês 2", "Mês 3", "Mês 4"]

    df = pd.DataFrame(registros)
    return df, meses_ref


def classificar_status_ruptura(media_mensal, estoque_geral, dias_cobertura):
    media_mensal = float(media_mensal or 0)
    estoque_geral = float(estoque_geral or 0)
    if media_mensal <= 0:
        return "SEM GIRO"
    if estoque_geral <= 0:
        return "CRÍTICO"
    if dias_cobertura <= 7:
        return "CRÍTICO"
    if dias_cobertura <= 15:
        return "ALTO"
    if dias_cobertura <= 30:
        return "ATENÇÃO"
    return "OK"


def montar_analise_ruptura_por_marca(df_ruptura, meses_ref, df_aberto_ruptura=None, dias_estoque_pedido=30):
    """
    Monta a visão gerencial de ruptura por marca.

    Regra de negócio desta tela:
    - Consolida o item por Marca + Código, somando Lojas Dauto + Única.
    - Considera o saldo de Pedidos em Aberto como estoque em trânsito.
    - O item só entra como "gera pedido" quando tem giro e o estoque considerado
      não cobre o parâmetro de dias de estoque definido na sidebar.
    - A ruptura/risco é calculada em cima do Estoque Considerado:
      Estoque Geral + Saldo em Trânsito/ABERTO.
    """
    if df_ruptura.empty:
        return pd.DataFrame(), pd.DataFrame()

    df = df_ruptura.copy()
    meses_ref = [m for m in meses_ref if m in df.columns]
    if not meses_ref:
        return pd.DataFrame(), pd.DataFrame()

    for mes in meses_ref:
        df[mes] = pd.to_numeric(df.get(mes, 0), errors="coerce").fillna(0)
    df["estoque"] = pd.to_numeric(df.get("estoque", 0), errors="coerce").fillna(0)
    df["codigo"] = df["codigo"].astype(str).str.extract(r"(\d+)")[0].fillna("").str.zfill(5)
    df["marca"] = df.get("marca", "SEM MARCA").astype(str).str.strip().replace("", "SEM MARCA")
    df["descricao"] = df.get("descricao", "").astype(str).str.strip()

    agg_dict = {mes: "sum" for mes in meses_ref}
    agg_dict.update({
        "estoque": "sum",
        "marca_codigo": "first",
        "unidade": "first",
    })

    itens = df.groupby(["marca", "codigo", "descricao"], as_index=False).agg(agg_dict)
    itens["Giro Geral"] = itens[meses_ref].sum(axis=1).round(2)
    itens["Média Giro Geral"] = itens[meses_ref].mean(axis=1).round(2)
    itens["Estoque Geral"] = pd.to_numeric(itens["estoque"], errors="coerce").fillna(0).round(2)

    if df_aberto_ruptura is not None and not df_aberto_ruptura.empty:
        aberto = df_aberto_ruptura.copy()
        aberto["codigo"] = aberto["codigo"].astype(str).str.extract(r"(\d+)")[0].fillna("").str.zfill(5)
        aberto["Saldo em Trânsito/ABERTO"] = pd.to_numeric(aberto.get("Saldo em Trânsito/ABERTO", 0), errors="coerce").fillna(0)
        aberto = aberto[aberto["codigo"].str.strip().ne("")]
        aberto = aberto.groupby("codigo", as_index=False)["Saldo em Trânsito/ABERTO"].sum()
        itens = itens.merge(aberto, on="codigo", how="left")
    else:
        itens["Saldo em Trânsito/ABERTO"] = 0

    dias_estoque_pedido = max(int(dias_estoque_pedido or 30), 1)
    itens["Saldo em Trânsito/ABERTO"] = pd.to_numeric(itens["Saldo em Trânsito/ABERTO"], errors="coerce").fillna(0).round(2)
    itens["Estoque Considerado"] = (itens["Estoque Geral"] + itens["Saldo em Trânsito/ABERTO"]).round(2)
    itens["Dias Estoque Pedido"] = dias_estoque_pedido
    itens["Estoque Alvo Pedido"] = (itens["Média Giro Geral"] * (dias_estoque_pedido / 30)).round(2)
    itens["Sugestão de Pedido"] = (itens["Estoque Alvo Pedido"] - itens["Estoque Considerado"]).apply(lambda x: max(math.ceil(float(x)), 0)).astype(int)
    itens["Necessidade 30 dias"] = itens["Sugestão de Pedido"]
    itens["Gera Pedido"] = itens.apply(
        lambda r: "SIM" if float(r.get("Média Giro Geral", 0) or 0) > 0 and int(r.get("Sugestão de Pedido", 0) or 0) > 0 else "NÃO",
        axis=1,
    )
    itens["Dias de Cobertura"] = itens.apply(
        lambda r: round((float(r["Estoque Considerado"]) / float(r["Média Giro Geral"]) * 30), 1) if float(r["Média Giro Geral"] or 0) > 0 else 9999,
        axis=1,
    )
    itens["Status"] = itens.apply(
        lambda r: classificar_status_ruptura(r["Média Giro Geral"], r["Estoque Considerado"], r["Dias de Cobertura"]),
        axis=1,
    )
    itens["Item em Ruptura"] = itens.apply(
        lambda r: "SIM" if r.get("Gera Pedido") == "SIM" and str(r.get("Status")) in ["CRÍTICO", "ALTO", "ATENÇÃO"] else "NÃO",
        axis=1,
    )
    itens["Peso Risco"] = itens["Status"].map({"CRÍTICO": 4, "ALTO": 3, "ATENÇÃO": 2, "OK": 1, "SEM GIRO": 0}).fillna(0)
    itens["Prioridade"] = itens.apply(
        lambda r: "1 - Comprar agora" if r["Status"] == "CRÍTICO" else (
            "2 - Comprar na próxima reposição" if r["Status"] == "ALTO" else (
                "3 - Acompanhar pedido" if r["Status"] == "ATENÇÃO" else "4 - Sem ação"
            )
        ),
        axis=1,
    )

    itens_risco = itens[itens["Gera Pedido"] == "SIM"].copy()

    resumo_base = itens.groupby("marca", as_index=False).agg(
        Itens_Analisados=("codigo", "count"),
        Giro_Geral_Total=("Giro Geral", "sum"),
        Media_Giro_Total=("Média Giro Geral", "sum"),
        Estoque_Geral_Total=("Estoque Geral", "sum"),
        Em_Aberto_Total=("Saldo em Trânsito/ABERTO", "sum"),
        Estoque_Considerado_Total=("Estoque Considerado", "sum"),
    )

    resumo_risco = itens_risco.groupby("marca", as_index=False).agg(
        Itens_que_Geram_Pedido=("codigo", "count"),
        Criticos=("Status", lambda s: int((s == "CRÍTICO").sum())),
        Alto=("Status", lambda s: int((s == "ALTO").sum())),
        Atencao=("Status", lambda s: int((s == "ATENÇÃO").sum())),
        Necessidade_30_dias=("Sugestão de Pedido", "sum"),
        Score_Risco=("Peso Risco", "sum"),
    )

    resumo = resumo_base.merge(resumo_risco, on="marca", how="left").fillna(0)
    for col in ["Itens_que_Geram_Pedido", "Criticos", "Alto", "Atencao", "Necessidade_30_dias", "Score_Risco"]:
        resumo[col] = pd.to_numeric(resumo[col], errors="coerce").fillna(0).astype(int)

    resumo["% Itens com Pedido"] = (
        resumo["Itens_que_Geram_Pedido"] / resumo["Itens_Analisados"].replace(0, pd.NA) * 100
    ).fillna(0).round(1)
    resumo["Dias de Cobertura"] = resumo.apply(
        lambda r: round((float(r["Estoque_Considerado_Total"]) / float(r["Media_Giro_Total"]) * 30), 1) if float(r["Media_Giro_Total"] or 0) > 0 else 9999,
        axis=1,
    )
    resumo["Ação"] = resumo.apply(
        lambda r: "Priorizar compra" if int(r["Criticos"] or 0) > 0 else (
            "Comprar na próxima reposição" if int(r["Alto"] or 0) > 0 else (
                "Acompanhar" if int(r["Atencao"] or 0) > 0 else "Sem risco relevante"
            )
        ),
        axis=1,
    )
    resumo = resumo.sort_values(
        ["Itens_que_Geram_Pedido", "Score_Risco", "Criticos", "Alto", "% Itens com Pedido"],
        ascending=[False, False, False, False, False],
    )

    resumo = resumo.rename(columns={
        "marca": "Marca",
        "Itens_Analisados": "Itens analisados",
        "Itens_que_Geram_Pedido": "Itens que geram pedido",
        "Criticos": "Críticos",
        "Atencao": "Atenção",
        "Giro_Geral_Total": "Giro Geral",
        "Media_Giro_Total": "Média Giro Geral",
        "Estoque_Geral_Total": "Estoque Geral",
        "Em_Aberto_Total": "Em Aberto",
        "Estoque_Considerado_Total": "Estoque Considerado",
        "Necessidade_30_dias": "Sugestão de Pedido",
        "Score_Risco": "Score Risco",
    })

    itens = itens.rename(columns={
        "marca": "Marca",
        "codigo": "Código",
        "descricao": "Descrição",
        "unidade": "UN",
    })
    itens = itens.sort_values(["Item em Ruptura", "Peso Risco", "Dias de Cobertura", "Média Giro Geral"], ascending=[False, False, True, False])
    return resumo.reset_index(drop=True), itens.reset_index(drop=True)

def colorir_status_ruptura(row):
    status = str(row.get("Status", ""))
    if status == "CRÍTICO":
        return ["background-color: #fee2e2; color: #7f1d1d; font-weight: 700"] * len(row)
    if status == "ALTO":
        return ["background-color: #ffedd5; color: #7c2d12; font-weight: 650"] * len(row)
    if status == "ATENÇÃO":
        return ["background-color: #fef9c3; color: #713f12"] * len(row)
    if status == "OK":
        return ["background-color: #dcfce7; color: #14532d"] * len(row)
    return ["background-color: #f1f5f9; color: #475569"] * len(row)


def colorir_resumo_marca(row):
    criticos = int(row.get("Críticos", 0) or 0)
    alto = int(row.get("Alto", 0) or 0)
    atencao = int(row.get("Atenção", 0) or 0)
    if criticos > 0:
        return ["background-color: #fee2e2; color: #7f1d1d; font-weight: 700"] * len(row)
    if alto > 0:
        return ["background-color: #ffedd5; color: #7c2d12; font-weight: 650"] * len(row)
    if atencao > 0:
        return ["background-color: #fef9c3; color: #713f12"] * len(row)
    return [""] * len(row)


def render_pagina_ruptura_por_marca():
    st.markdown('<div class="section-title"> Ruptura por Marca</div>', unsafe_allow_html=True)
    st.caption(
        "Esta página funciona separada do pedido de compra. Envie aqui o PDF específico de Giro Geral por Marca "
        "e, se houver, o PDF de Pedidos em Aberto. A análise soma lojas Dauto + Única e considera o saldo em aberto no estoque."
    )

    col_pdf_ruptura, col_pdf_aberto = st.columns(2)
    with col_pdf_ruptura:
        pdf_marca = st.file_uploader(
            "PDF - Giro Geral por Marca",
            type=["pdf"],
            key="upload_pdf_ruptura_por_marca_independente",
        )
    with col_pdf_aberto:
        pdf_pedidos_aberto_ruptura = st.file_uploader(
            "PDF - Pedidos em Aberto",
            type=["pdf"],
            key="upload_pdf_pedidos_aberto_ruptura_independente",
        )

    if not pdf_marca:
        st.info("Envie o PDF de Giro Geral por Marca para iniciar esta análise.")
        return

    st.info("Leitura otimizada ativada: o Giro por Marca usa PyMuPDF. O PDF de Pedidos em Aberto, quando enviado, entra como saldo em trânsito no cálculo da ruptura.")

    try:
        bytes_pdf = pdf_marca.getvalue()
        df_ruptura, meses_ref = parse_pdf_ruptura_por_marca(bytes_pdf)
    except Exception as e:
        st.error(f"Não consegui ler o PDF de Ruptura por Marca. Erro: {e}")
        return

    df_aberto_ruptura = pd.DataFrame(columns=["codigo", "Saldo em Trânsito/ABERTO"])
    if pdf_pedidos_aberto_ruptura:
        try:
            with st.spinner("Lendo Pedidos em Aberto para considerar no estoque..."):
                df_aberto_ruptura = parse_pedidos_compra_aberto_pdf(pdf_pedidos_aberto_ruptura)
            st.success(f"Pedidos em aberto lidos: {len(df_aberto_ruptura)} item(ns) com saldo em aberto.")
        except Exception as e:
            st.warning(f"Não consegui ler o PDF de Pedidos em Aberto. A análise seguirá apenas com o estoque atual. Erro: {e}")
            df_aberto_ruptura = pd.DataFrame(columns=["codigo", "Saldo em Trânsito/ABERTO"])

    if df_ruptura.empty:
        st.error("Não consegui extrair os itens do PDF enviado. Verifique se é o relatório de Giro Geral por Marca.")
        return

    resumo_marca, itens_marca = montar_analise_ruptura_por_marca(
        df_ruptura,
        meses_ref,
        df_aberto_ruptura,
        dias_estoque_pedido=dias_estoque_alvo,
    )
    if resumo_marca.empty:
        st.warning("O PDF foi lido, mas não houve dados suficientes para análise.")
        return

    total_itens = int(len(itens_marca))
    total_risco = int((itens_marca.get("Gera Pedido", "NÃO") == "SIM").sum())
    total_criticos = int((itens_marca["Status"] == "CRÍTICO").sum())
    total_alto = int((itens_marca["Status"] == "ALTO").sum())
    total_em_aberto = float(pd.to_numeric(itens_marca.get("Saldo em Trânsito/ABERTO", 0), errors="coerce").fillna(0).sum())
    total_sugestao = int(pd.to_numeric(itens_marca.loc[itens_marca.get("Gera Pedido", "NÃO") == "SIM", "Sugestão de Pedido"], errors="coerce").fillna(0).sum())

    c1, c2, c3, c4, c5 = st.columns(5)
    with c1:
        render_metric("Itens analisados", format_int_br(total_itens), "Lojas Dauto + Única")
    with c2:
        render_metric("Itens com pedido", format_int_br(total_risco), "Gera Pedido = SIM")
    with c3:
        render_metric("Críticos", format_int_br(total_criticos), "Sem estoque ou até 7 dias")
    with c4:
        render_metric("Sugestão total", format_int_br(total_sugestao), f"Pedido para {dias_estoque_alvo} dias")
    with c5:
        render_metric("Em aberto", format_num_br(total_em_aberto, 1), "Somado ao estoque")

    st.markdown("---")
    st.markdown('<div class="section-title">Ranking de marcas por risco</div>', unsafe_allow_html=True)
    st.caption(f"O ranking abaixo mostra, por marca, quantos itens têm giro, geram pedido para {dias_estoque_alvo} dias de estoque e continuam com risco de ruptura mesmo considerando o saldo em aberto.")

    col_busca_marca, col_zero_marca = st.columns([2, 1])
    with col_busca_marca:
        busca_marca = st.text_input("Pesquisar marca", key="busca_marca_ruptura")
    with col_zero_marca:
        mostrar_sem_risco = st.checkbox("Mostrar marcas sem risco", value=False, key="mostrar_marca_sem_risco")

    resumo_view = resumo_marca.copy()
    if not mostrar_sem_risco and "Itens que geram pedido" in resumo_view.columns:
        resumo_view = resumo_view[pd.to_numeric(resumo_view["Itens que geram pedido"], errors="coerce").fillna(0) > 0]
    if busca_marca:
        resumo_view = resumo_view[resumo_view["Marca"].astype(str).str.lower().str.contains(busca_marca.lower(), na=False)]

    st.dataframe(
        resumo_view.style.apply(colorir_resumo_marca, axis=1).format(formatadores_para_tabela(resumo_view)),
        use_container_width=True,
        hide_index=True,
        height=430,
        column_config={"Marca": st.column_config.TextColumn("Marca", pinned=True, width="large")},
    )

    st.download_button(
        "⬇ Baixar ranking de marcas em CSV",
        gerar_csv(resumo_marca),
        "ranking_ruptura_por_marca.csv",
        "text/csv",
    )

    st.markdown("---")
    st.markdown('<div class="section-title">Drill por marca</div>', unsafe_allow_html=True)
    marcas = resumo_marca["Marca"].astype(str).tolist()
    marca_selecionada = st.selectbox("Selecione a marca para abrir os produtos", marcas, key="drill_marca_ruptura")

    itens_view = itens_marca[itens_marca["Marca"].astype(str) == str(marca_selecionada)].copy()

    colf1, colf2, colf3 = st.columns([1, 1, 1])
    with colf1:
        apenas_risco = st.checkbox("Apenas itens que geram pedido", value=True, key="apenas_itens_ruptura_marca")
    with colf2:
        status_opcoes = ["Todos", "CRÍTICO", "ALTO", "ATENÇÃO", "OK", "SEM GIRO"]
        status_sel = st.selectbox("Filtrar status", status_opcoes, key="status_ruptura_marca")
    with colf3:
        busca_item = st.text_input("Pesquisar produto dentro da marca", key="busca_item_ruptura_marca")

    if apenas_risco and "Gera Pedido" in itens_view.columns:
        itens_view = itens_view[itens_view["Gera Pedido"] == "SIM"]
    if status_sel != "Todos":
        itens_view = itens_view[itens_view["Status"] == status_sel]
    if busca_item:
        termo = busca_item.lower()
        itens_view = itens_view[
            itens_view["Código"].astype(str).str.lower().str.contains(termo, na=False)
            | itens_view["Descrição"].astype(str).str.lower().str.contains(termo, na=False)
        ]

    colunas_itens = ["Código", "Descrição", "UN"] + meses_ref + [
        "Giro Geral", "Média Giro Geral", "Estoque Geral", "Saldo em Trânsito/ABERTO", "Estoque Considerado",
        "Dias de Cobertura", "Dias Estoque Pedido", "Estoque Alvo Pedido", "Sugestão de Pedido", "Gera Pedido", "Item em Ruptura", "Status", "Prioridade"
    ]
    colunas_itens = [c for c in colunas_itens if c in itens_view.columns]

    st.dataframe(
        itens_view[colunas_itens].style.apply(colorir_status_ruptura, axis=1).format(formatadores_para_tabela(itens_view[colunas_itens])),
        use_container_width=True,
        hide_index=True,
        height=560,
        column_config={
            "Código": st.column_config.TextColumn("Código", pinned=True, width="small"),
            "Descrição": st.column_config.TextColumn("Descrição", pinned=True, width="large"),
        },
    )

    st.download_button(
        "⬇ Baixar drill da marca em CSV",
        gerar_csv(itens_view[colunas_itens]),
        f"drill_ruptura_{re.sub(r'[^A-Za-z0-9]+', '_', str(marca_selecionada))}.csv",
        "text/csv",
    )

def render_pagina_pedidos_drive():
    st.markdown('<div class="section-title">Pedidos no Google Drive</div>', unsafe_allow_html=True)

    if not google_configurado():
        st.warning(google_mensagem_configuracao())
        st.code(
            """
[google_oauth_user]
client_id = "..."
client_secret = "..."
refresh_token = "..."
token_uri = "https://oauth2.googleapis.com/token"
""".strip(),
            language="toml",
        )
        st.stop()

    try:
        recursos = google_get_resources()
        st.success(f"Google Drive conectado via OAuth: {recursos.get('oauth_user', 'gdautotintas@gmail.com')}")
        c1, c2, c3 = st.columns(3)
        c1.link_button("Pedidos para aprovação", recursos["pedidos_link"])
        c2.link_button("Pedidos aprovados", recursos["aprovados_link"])
        c3.link_button("Cadastro", recursos["cadastro_link"])

        st.markdown("### Sincronizar aprovações")
        st.caption("Altere o Status para Aprovado na aba Controle da planilha. Depois clique abaixo para mover os pedidos aprovados para a pasta PEDIDOS APROVADOS.")
        usuario_sync = st.text_input("Responsável pela sincronização", value="", key="drive_usuario_sync")
        if st.button("🔄 Sincronizar aprovações", type="primary"):
            movidos, ignorados = google_sincronizar_aprovacoes(usuario=usuario_sync)
            if movidos:
                st.success(f"{len(movidos)} pedido(s) movido(s) para PEDIDOS APROVADOS.")
                for nome in movidos[:10]:
                    st.caption(f"✅ {nome}")
            else:
                st.info("Nenhum pedido com Status = Aprovado foi encontrado na pasta de aprovação.")
            st.rerun()

        pedidos = google_listar_pedidos()
        if pedidos.empty:
            st.info("Ainda nao existem pedidos registrados no Drive.")
            st.stop()

        pedidos_view = pedidos.copy()
        pedidos_view["valor"] = pd.to_numeric(pedidos_view["valor"], errors="coerce").fillna(0)
        st.markdown("### Painel de pedidos")
        st.dataframe(
            pedidos_view[[
                "id_pedido", "nome_pedido", "fornecedor", "status", "valor",
                "criado_em", "criado_por", "aprovado_em", "aprovado_por",
                "link_pedido", "link_autcom", "link_fornecedor",
            ]],
            use_container_width=True,
            hide_index=True,
            height=420,
            column_config={
                "valor": st.column_config.NumberColumn("Valor", format="R$ %.2f"),
                "link_pedido": st.column_config.LinkColumn("Pedido"),
                "link_autcom": st.column_config.LinkColumn("Autcom"),
                "link_fornecedor": st.column_config.LinkColumn("Fornecedor"),
            },
        )

        st.markdown("### Aprovar ou atualizar status")
        opcoes = {
            f"{r.get('id_pedido', '')} | {r.get('fornecedor', '')} | {r.get('nome_pedido', '')} | {r.get('status', '')}": r.get("id_pedido", "")
            for _, r in pedidos.iterrows()
        }
        pedido_label = st.selectbox("Pedido", list(opcoes.keys()), key="drive_pedido_status")
        usuario = st.text_input("Usuario responsavel", value="", key="drive_usuario_status")
        status = st.selectbox("Status", ["Aprovado", "Em edição", "Reprovado", "Finalizado"], key="drive_status")
        observacao = st.text_input("Observacao", key="drive_observacao_status")

        if st.button("Salvar status no controle", type="primary"):
            atualizado = google_atualizar_status_pedido(opcoes[pedido_label], status, usuario=usuario, observacao=observacao)
            st.success(f"Status atualizado para {atualizado.get('status')}.")
            st.rerun()
    except Exception as e:
        st.error(str(e))




# =========================================================
# MULTIPAGE - PREVISÃO FINANCEIRA DE FORNECEDORES
# =========================================================

# =========================================================
# CONFIGURAÇÕES
# =========================================================

CFOPS_BONIFICACAO = {
    "1910", "1911", "2910", "2911", "3910", "3911",
    "5910", "5911", "6910", "6911", "7910", "7911",
    "6949",  # remessa de marketing/outro enquadramento informado pelo usuário
}

# CFOPs usuais de devolução de venda/compra, incluindo operações internas,
# interestaduais e com exterior. A natureza da operação contendo "DEVOLU" também
# é usada como critério auxiliar.
CFOPS_DEVOLUCAO = {
    "1201", "1202", "1203", "1204", "1410", "1411", "1503", "1504",
    "1553", "1660", "1661", "1662",
    "2201", "2202", "2203", "2204", "2410", "2411", "2503", "2504",
    "2553", "2660", "2661", "2662",
    "3201", "3202", "3203", "3204", "3211", "3503", "3504",
    "5201", "5202", "5208", "5209", "5210", "5410", "5411", "5412",
    "5413", "5503", "5553", "5660", "5661", "5662",
    "6201", "6202", "6208", "6209", "6210", "6410", "6411", "6412",
    "6413", "6503", "6553", "6660", "6661", "6662",
    "7201", "7202", "7208", "7209", "7210", "7503", "7553",
}

COLUNAS_NOTAS = [
    "CATEGORIA",
    "NR_CHAVE_ACESSO",
    "NR_CNPJ_EMITENTE",
    "NM_EMITENTE",
    "DT_EMISSAO",
    "NR_DOCUMENTO",
    "NR_SERIE",
    "CFOPS",
    "NATUREZA_OPERACAO",
    "VL_NOTA_FISCAL",
    "QTD_ITENS",
    "ARQUIVO_ORIGEM",
]

COLUNAS_PARCELAS = [
    "CATEGORIA",
    "NR_CHAVE_ACESSO",
    "NR_CNPJ_EMITENTE",
    "NM_EMITENTE",
    "NR_DOCUMENTO",
    "NR_SERIE",
    "NR_PARCELA",
    "DT_VENCIMENTO",
    "VL_PARCELA",
    "ARQUIVO_ORIGEM",
]

COLUNAS_ITENS = [
    "CATEGORIA",
    "NR_CHAVE_ACESSO",
    "NR_CNPJ_EMITENTE",
    "NM_EMITENTE",
    "DT_EMISSAO",
    "NR_DOCUMENTO",
    "NR_SERIE",
    "CFOP",
    "NATUREZA_OPERACAO",
    "COD_PRODUTO",
    "DESCRICAO",
    "QTD",
    "VL_UNITARIO",
    "VL_TOTAL_ITEM",
    "ARQUIVO_ORIGEM",
]


# =========================================================
# FUNÇÕES XML
# =========================================================

def limpar_cnpj(cnpj: str) -> str:
    return re.sub(r"\D", "", str(cnpj or ""))


def formatar_cnpj(cnpj: str) -> str:
    c = limpar_cnpj(cnpj)
    if len(c) != 14:
        return cnpj or ""
    return f"{c[:2]}.{c[2:5]}.{c[5:8]}/{c[8:12]}-{c[12:]}"


def local_name(tag: str) -> str:
    return tag.split("}", 1)[-1] if "}" in tag else tag


def filhos_por_nome(elemento: ET.Element, nome: str) -> List[ET.Element]:
    return [e for e in elemento.iter() if local_name(e.tag) == nome]


def primeiro_texto(elemento: Optional[ET.Element], nomes: Iterable[str], default: str = "") -> str:
    if elemento is None:
        return default
    nomes_set = set(nomes)
    for e in elemento.iter():
        if local_name(e.tag) in nomes_set and e.text is not None:
            txt = str(e.text).strip()
            if txt:
                return txt
    return default


def primeiro_filho_direto(elemento: ET.Element, nome: str) -> Optional[ET.Element]:
    for child in list(elemento):
        if local_name(child.tag) == nome:
            return child
    return None


def texto_filho_direto(elemento: Optional[ET.Element], nome: str, default: str = "") -> str:
    if elemento is None:
        return default
    filho = primeiro_filho_direto(elemento, nome)
    if filho is not None and filho.text:
        return filho.text.strip()
    return default


def to_float(valor: str | None) -> float:
    if valor is None:
        return 0.0
    s = str(valor).strip()
    if not s:
        return 0.0
    # XML NF-e costuma vir com ponto decimal. Tratamento extra para vírgula.
    s = s.replace(".", "") if s.count(",") == 1 and s.count(".") > 1 else s
    s = s.replace(",", ".")
    try:
        return float(s)
    except Exception:
        return 0.0


def formatar_moeda(valor: float) -> str:
    """Formata número em padrão monetário brasileiro."""
    try:
        return f"R$ {float(valor):,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
    except Exception:
        return "R$ 0,00"


def formatar_df_moeda(df: pd.DataFrame, colunas: Iterable[str]) -> pd.DataFrame:
    """Retorna cópia do dataframe com colunas monetárias formatadas para exibição."""
    out = df.copy()
    for col in colunas:
        if col in out.columns:
            out[col] = out[col].apply(formatar_moeda)
    return out


def formatar_data(data: str) -> str:
    if not data:
        return ""
    data = data.strip()
    try:
        # dhEmi: 2026-06-01T10:20:30-03:00
        return datetime.fromisoformat(data.replace("Z", "+00:00")).strftime("%d/%m/%Y")
    except Exception:
        pass
    try:
        # dEmi: 2026-06-01
        return datetime.strptime(data[:10], "%Y-%m-%d").strftime("%d/%m/%Y")
    except Exception:
        return data


def obter_inf_nfe(root: ET.Element) -> Optional[ET.Element]:
    for e in root.iter():
        if local_name(e.tag) == "infNFe":
            return e
    return None


def obter_chave(root: ET.Element, inf_nfe: Optional[ET.Element]) -> str:
    # nfeProc/protNFe/infProt/chNFe é o caminho mais confiável quando existe.
    for e in root.iter():
        if local_name(e.tag) == "chNFe" and e.text:
            return e.text.strip()
    if inf_nfe is not None:
        id_attr = inf_nfe.attrib.get("Id", "")
        if id_attr.startswith("NFe"):
            return id_attr.replace("NFe", "").strip()
        return id_attr.strip()
    return ""


def ler_nfe_xml(conteudo: bytes, nome_arquivo: str) -> Tuple[Optional[Dict], pd.DataFrame, pd.DataFrame, Optional[str]]:
    """Retorna dict da nota, dataframe de itens, dataframe de parcelas e erro."""
    try:
        root = ET.fromstring(conteudo)
    except Exception as e:
        return None, pd.DataFrame(), pd.DataFrame(), f"XML inválido: {nome_arquivo} - {e}"

    inf_nfe = obter_inf_nfe(root)
    if inf_nfe is None:
        # XML de evento não possui itens/valores. Não entra na análise.
        return None, pd.DataFrame(), pd.DataFrame(), None

    emit = None
    ide = None
    total = None
    for child in list(inf_nfe):
        ln = local_name(child.tag)
        if ln == "emit":
            emit = child
        elif ln == "ide":
            ide = child
        elif ln == "total":
            total = child

    chave = obter_chave(root, inf_nfe)
    cnpj_emitente = limpar_cnpj(texto_filho_direto(emit, "CNPJ"))
    nome_emitente = texto_filho_direto(emit, "xNome") or texto_filho_direto(emit, "xFant")
    natureza = texto_filho_direto(ide, "natOp")
    dt_emissao = formatar_data(texto_filho_direto(ide, "dhEmi") or texto_filho_direto(ide, "dEmi"))
    nr_doc = texto_filho_direto(ide, "nNF")
    nr_serie = texto_filho_direto(ide, "serie")
    vnf = 0.0
    if total is not None:
        for e in total.iter():
            if local_name(e.tag) == "vNF":
                vnf = to_float(e.text)
                break

    linhas_itens = []
    cfops = []

    for det in [e for e in list(inf_nfe) if local_name(e.tag) == "det"]:
        prod = primeiro_filho_direto(det, "prod")
        if prod is None:
            continue
        cfop = texto_filho_direto(prod, "CFOP")
        cfops.append(cfop)
        linhas_itens.append({
            "NR_CHAVE_ACESSO": chave,
            "NR_CNPJ_EMITENTE": formatar_cnpj(cnpj_emitente),
            "NM_EMITENTE": nome_emitente,
            "DT_EMISSAO": dt_emissao,
            "NR_DOCUMENTO": nr_doc,
            "NR_SERIE": nr_serie,
            "CFOP": cfop,
            "NATUREZA_OPERACAO": natureza,
            "COD_PRODUTO": texto_filho_direto(prod, "cProd"),
            "DESCRICAO": texto_filho_direto(prod, "xProd"),
            "QTD": to_float(texto_filho_direto(prod, "qCom")),
            "VL_UNITARIO": to_float(texto_filho_direto(prod, "vUnCom")),
            "VL_TOTAL_ITEM": to_float(texto_filho_direto(prod, "vProd")),
            "ARQUIVO_ORIGEM": nome_arquivo,
        })

    cfops_unicos = sorted({c for c in cfops if c})

    # Parcelas de pagamento: cobr/dup, quando informadas pelo emissor no XML.
    linhas_parcelas = []
    for dup in [e for e in inf_nfe.iter() if local_name(e.tag) == "dup"]:
        nr_parcela = texto_filho_direto(dup, "nDup")
        dt_venc = formatar_data(texto_filho_direto(dup, "dVenc"))
        vl_parcela = to_float(texto_filho_direto(dup, "vDup"))
        linhas_parcelas.append({
            "NR_CHAVE_ACESSO": chave,
            "NR_CNPJ_EMITENTE": formatar_cnpj(cnpj_emitente),
            "NM_EMITENTE": nome_emitente,
            "NR_DOCUMENTO": nr_doc,
            "NR_SERIE": nr_serie,
            "NR_PARCELA": nr_parcela,
            "DT_VENCIMENTO": dt_venc,
            "VL_PARCELA": vl_parcela,
            "ARQUIVO_ORIGEM": nome_arquivo,
        })

    nota = {
        "NR_CHAVE_ACESSO": chave,
        "NR_CNPJ_EMITENTE": formatar_cnpj(cnpj_emitente),
        "NM_EMITENTE": nome_emitente,
        "DT_EMISSAO": dt_emissao,
        "NR_DOCUMENTO": nr_doc,
        "NR_SERIE": nr_serie,
        "CFOPS": ", ".join(cfops_unicos),
        "CFOPS_LISTA": cfops_unicos,
        "NATUREZA_OPERACAO": natureza,
        "VL_NOTA_FISCAL": vnf,
        "QTD_ITENS": len(linhas_itens),
        "ARQUIVO_ORIGEM": nome_arquivo,
    }

    return nota, pd.DataFrame(linhas_itens), pd.DataFrame(linhas_parcelas), None


# =========================================================
# LEITURA DE UPLOADS
# =========================================================

def extrair_arquivos(uploaded_files) -> List[Tuple[str, bytes]]:
    arquivos = []
    for up in uploaded_files:
        nome = up.name
        conteudo = up.read()
        if nome.lower().endswith(".zip"):
            try:
                with zipfile.ZipFile(io.BytesIO(conteudo)) as z:
                    for info in z.infolist():
                        if info.is_dir():
                            continue
                        if info.filename.lower().endswith(".xml"):
                            arquivos.append((info.filename, z.read(info)))
            except Exception as e:
                st.warning(f"Não foi possível ler o ZIP {nome}: {e}")
        elif nome.lower().endswith(".xml"):
            arquivos.append((nome, conteudo))
    return arquivos


def classificar_nota(cfops: List[str], natureza: str = "") -> str:
    cfops_validos = {str(c).strip() for c in cfops if str(c).strip()}
    natureza_norm = str(natureza or "").upper()

    # Devolução tem precedência para ficar separada das compras e bonificações.
    if any(c in CFOPS_DEVOLUCAO for c in cfops_validos) or "DEVOLU" in natureza_norm:
        return "DEVOLUCOES"

    if not cfops_validos:
        return "COMPRAS"
    tem_bonif = any(c in CFOPS_BONIFICACAO for c in cfops_validos)
    todos_bonif = all(c in CFOPS_BONIFICACAO for c in cfops_validos)
    if todos_bonif:
        return "BONIFICACOES"
    if tem_bonif:
        return "MISTAS"
    return "COMPRAS"


@st.cache_data(show_spinner=False)
def processar_uploads_cache(chaves: Tuple[Tuple[str, bytes], ...]):
    notas = []
    itens_frames = []
    parcelas_frames = []
    erros = []
    for nome, conteudo in chaves:
        nota, itens, parcelas, erro = ler_nfe_xml(conteudo, nome)
        if erro:
            erros.append(erro)
        if nota:
            categoria = classificar_nota(nota.get("CFOPS_LISTA", []), nota.get("NATUREZA_OPERACAO", ""))
            nota["CLASSIFICACAO_AUTOMATICA"] = categoria
            notas.append(nota)
            if not itens.empty:
                itens["CLASSIFICACAO_AUTOMATICA"] = categoria
                itens_frames.append(itens)
            if not parcelas.empty:
                parcelas["CLASSIFICACAO_AUTOMATICA"] = categoria
                parcelas_frames.append(parcelas)

    df_notas = pd.DataFrame(notas)
    df_itens = pd.concat(itens_frames, ignore_index=True) if itens_frames else pd.DataFrame()
    df_parcelas = pd.concat(parcelas_frames, ignore_index=True) if parcelas_frames else pd.DataFrame()
    return df_notas, df_itens, df_parcelas, erros


def processar_uploads(uploaded_files):
    arquivos = extrair_arquivos(uploaded_files)
    # Streamlit cache exige objeto hashável/serializável. Bytes em tupla funciona.
    return processar_uploads_cache(tuple(arquivos))


# =========================================================
# EXPORTAÇÃO
# =========================================================

def gerar_excel(df_resumo: pd.DataFrame, df_notas: pd.DataFrame, df_itens: pd.DataFrame, df_parcelas: pd.DataFrame) -> bytes:
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df_resumo.to_excel(writer, index=False, sheet_name="Resumo")
        df_notas.to_excel(writer, index=False, sheet_name="Notas")
        df_itens.to_excel(writer, index=False, sheet_name="Itens")
        df_parcelas.to_excel(writer, index=False, sheet_name="Parcelas")
    return output.getvalue()


def consolidar_produtos_xml(df_itens: pd.DataFrame) -> pd.DataFrame:
    if df_itens is None or df_itens.empty:
        return pd.DataFrame()

    df = df_itens.copy()
    for col in ["COD_PRODUTO", "DESCRICAO", "NM_EMITENTE", "NR_DOCUMENTO", "NR_SERIE", "CFOP"]:
        if col not in df.columns:
            df[col] = ""
        df[col] = df[col].astype(str).fillna("").str.strip()

    for col in ["QTD", "VL_UNITARIO", "VL_TOTAL_ITEM"]:
        if col not in df.columns:
            df[col] = 0
        df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0)

    df["CHAVE_PRODUTO_XML"] = df.apply(
        lambda r: normalizar_texto_simples(r.get("COD_PRODUTO", "")) + "|" + normalizar_texto_simples(r.get("DESCRICAO", "")),
        axis=1,
    )
    df["NOTA"] = df.apply(
        lambda r: (str(r.get("NR_SERIE", "")).strip() + "/" if str(r.get("NR_SERIE", "")).strip() else "") + str(r.get("NR_DOCUMENTO", "")).strip(),
        axis=1,
    )

    agrupado = df.groupby("CHAVE_PRODUTO_XML", as_index=False).agg(
        COD_PRODUTO=("COD_PRODUTO", "first"),
        DESCRICAO=("DESCRICAO", "first"),
        QTD_TOTAL=("QTD", "sum"),
        VL_TOTAL_ITEM=("VL_TOTAL_ITEM", "sum"),
        QTD_LANCAMENTOS=("QTD", "size"),
        FORNECEDORES=("NM_EMITENTE", lambda s: ", ".join(sorted({x for x in s.astype(str) if x.strip()}))),
        NOTAS=("NOTA", lambda s: ", ".join(sorted({x for x in s.astype(str) if x.strip()}))),
        CFOPS=("CFOP", lambda s: ", ".join(sorted({x for x in s.astype(str) if x.strip()}))),
    )
    agrupado["VL_UNITARIO_MEDIO"] = agrupado.apply(
        lambda r: round(float(r["VL_TOTAL_ITEM"]) / float(r["QTD_TOTAL"]), 6) if float(r["QTD_TOTAL"] or 0) > 0 else 0,
        axis=1,
    )
    agrupado = agrupado[[
        "COD_PRODUTO", "DESCRICAO", "QTD_TOTAL", "VL_UNITARIO_MEDIO", "VL_TOTAL_ITEM",
        "QTD_LANCAMENTOS", "FORNECEDORES", "NOTAS", "CFOPS",
    ]]
    return agrupado.sort_values(["DESCRICAO", "COD_PRODUTO"]).reset_index(drop=True)


def remover_notas_xml_repetidas(df_notas: pd.DataFrame, df_itens: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    if df_notas is None or df_notas.empty or "NR_CHAVE_ACESSO" not in df_notas.columns:
        return df_notas.copy() if df_notas is not None else pd.DataFrame(), df_itens.copy() if df_itens is not None else pd.DataFrame(), pd.DataFrame()

    notas = df_notas.copy()
    notas["NR_CHAVE_ACESSO"] = notas["NR_CHAVE_ACESSO"].astype(str).str.strip()
    notas_validas = notas[notas["NR_CHAVE_ACESSO"].ne("")].copy()
    notas_sem_chave = notas[notas["NR_CHAVE_ACESSO"].eq("")].copy()

    repetidas_mask = notas_validas.duplicated(subset=["NR_CHAVE_ACESSO"], keep="first")
    notas_repetidas = notas_validas[repetidas_mask].copy()
    notas_unicas = pd.concat(
        [notas_validas[~repetidas_mask].copy(), notas_sem_chave],
        ignore_index=True,
        sort=False,
    )

    itens = df_itens.copy() if df_itens is not None else pd.DataFrame()
    if not itens.empty and "NR_CHAVE_ACESSO" in itens.columns and not notas_repetidas.empty:
        chaves_repetidas = set(notas_repetidas["NR_CHAVE_ACESSO"].astype(str).str.strip())
        itens = itens[~itens["NR_CHAVE_ACESSO"].astype(str).str.strip().isin(chaves_repetidas)].copy()

    return notas_unicas.reset_index(drop=True), itens.reset_index(drop=True), notas_repetidas.reset_index(drop=True)


def gerar_excel_produtos_xml(df_notas: pd.DataFrame, df_itens: pd.DataFrame, df_produtos: pd.DataFrame) -> bytes:
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df_produtos.to_excel(writer, index=False, sheet_name="Produtos consolidados")
        df_itens.to_excel(writer, index=False, sheet_name="Itens por nota")
        df_notas.to_excel(writer, index=False, sheet_name="Notas")

        for sheet_name in writer.sheets:
            ws = writer.sheets[sheet_name]
            ws.freeze_panes = "A2"
            ws.auto_filter.ref = ws.dimensions
            for cell in ws[1]:
                cell.font = cell.font.copy(bold=True)
            for col_cells in ws.columns:
                max_len = 10
                col_letter = col_cells[0].column_letter
                header = str(col_cells[0].value or "")
                for cell in col_cells[:200]:
                    max_len = max(max_len, len(str(cell.value or "")))
                ws.column_dimensions[col_letter].width = min(max_len + 2, 48)
                if header in ["QTD_TOTAL", "QTD", "VL_UNITARIO", "VL_UNITARIO_MEDIO", "VL_TOTAL_ITEM"]:
                    for cell in col_cells[1:]:
                        cell.number_format = '#,##0.00'
    return output.getvalue()




# =========================================================
# CONTAS A PAGAR / CONCILIAÇÃO CSV x XML
# =========================================================

def normalizar_texto(valor: object) -> str:
    s = str(valor or "").upper().strip()
    s = re.sub(r"[^A-Z0-9]+", "", s)
    return s


def normalizar_numero_nf(valor: object) -> str:
    s = re.sub(r"\D", "", str(valor or ""))
    return s.lstrip("0") or ("0" if s else "")


def extrair_nf_historico(historico: object) -> str:
    texto = str(historico or "")
    padroes = [
        r"DOCUMENTO\s*(?:N[º°O\.]*|NO)?\s*0*([0-9]+)",
        r"NOTA\s*(?:FISCAL|NF)?\s*(?:N[º°O\.]*|NO)?\s*0*([0-9]+)",
        r"\bNF\s*0*([0-9]{2,})\b",
    ]
    for padrao in padroes:
        m = re.search(padrao, texto, flags=re.IGNORECASE)
        if m:
            return normalizar_numero_nf(m.group(1))
    return ""


def parse_valor_brasileiro(valor: object) -> float:
    if valor is None or (isinstance(valor, float) and pd.isna(valor)):
        return 0.0
    s = str(valor).strip().replace("R$", "").replace(" ", "")
    if not s:
        return 0.0
    if "," in s:
        s = s.replace(".", "").replace(",", ".")
    try:
        return float(s)
    except Exception:
        return 0.0


def ler_csv_financeiro(upload) -> pd.DataFrame:
    """
    Lê relatórios CSV exportados pelo sistema financeiro, inclusive quando há
    cabeçalho institucional antes da linha real de colunas.

    O arquivo de referência possui:
    - codificação ANSI/Latin-1;
    - separador ponto e vírgula;
    - várias linhas de relatório antes do cabeçalho;
    - campos HISTÓRICO eventualmente quebrados em mais de uma linha.
    """
    conteudo = upload.getvalue()
    ultimo_erro = None

    for encoding in ("utf-8-sig", "latin-1", "cp1252"):
        try:
            texto = conteudo.decode(encoding)
        except Exception as e:
            ultimo_erro = e
            continue

        linhas = texto.splitlines()
        indice_cabecalho = None

        for i, linha in enumerate(linhas):
            linha_norm = linha.upper().replace("HISTORICO", "HISTÓRICO")
            if (
                "FAVORECIDO" in linha_norm
                and "DTA.VEN" in linha_norm
                and "VAL.DUP" in linha_norm
                and "HISTÓRICO" in linha_norm
            ):
                indice_cabecalho = i
                break

        if indice_cabecalho is None:
            continue

        trecho_csv = "\n".join(linhas[indice_cabecalho:])

        try:
            df = pd.read_csv(
                io.StringIO(trecho_csv),
                sep=";",
                dtype=str,
                engine="python",
                quotechar='"',
                on_bad_lines="skip",
            )

            df.columns = [
                str(c).strip().upper().replace("HISTORICO", "HISTÓRICO")
                for c in df.columns
            ]

            # Remove linhas residuais criadas por quebras de linha do HISTÓRICO,
            # totais de relatório ou linhas completamente vazias.
            if "FAVORECIDO" in df.columns:
                df["FAVORECIDO"] = df["FAVORECIDO"].fillna("").astype(str).str.strip()
                df = df[df["FAVORECIDO"] != ""].copy()

            if "DTA.VEN" in df.columns:
                datas_teste = pd.to_datetime(
                    df["DTA.VEN"], dayfirst=True, errors="coerce"
                )
                df = df[datas_teste.notna()].copy()

            if "DTA.VEN" in df.columns and "FAVORECIDO" in df.columns:
                return df.reset_index(drop=True)

        except Exception as e:
            ultimo_erro = e

    raise ValueError(
        "Não foi possível interpretar o CSV financeiro. "
        f"Verifique se ele contém as colunas FAVORECIDO, DTA.VEN, VAL.DUP e HISTÓRICO. "
        f"Detalhe técnico: {ultimo_erro}"
    )


def preparar_contas_csv(df_csv: pd.DataFrame, hoje: pd.Timestamp) -> pd.DataFrame:
    df = df_csv.copy()
    for col in ["VAL.DUP", "VAL.PAG", "VAL.JUR", "VAL.DES"]:
        if col not in df.columns:
            df[col] = "0"
        df[col] = df[col].apply(parse_valor_brasileiro)
    if "DTA.PAG" not in df.columns:
        df["DTA.PAG"] = ""
    if "HISTÓRICO" not in df.columns:
        df["HISTÓRICO"] = ""
    if "DUPLICATA" not in df.columns:
        df["DUPLICATA"] = ""

    df["DT_VENCIMENTO"] = pd.to_datetime(df["DTA.VEN"], dayfirst=True, errors="coerce")
    df["DT_PAGAMENTO"] = pd.to_datetime(df["DTA.PAG"], dayfirst=True, errors="coerce")
    df["NR_DOCUMENTO_XML"] = df["HISTÓRICO"].apply(extrair_nf_historico)
    df["FORNECEDOR"] = df["FAVORECIDO"].fillna("").astype(str).str.strip()
    df["FORNECEDOR_CHAVE"] = df["FORNECEDOR"].apply(normalizar_texto)
    df["VALOR_A_PAGAR"] = (
        df["VAL.DUP"] + df["VAL.JUR"] - df["VAL.DES"] - df["VAL.PAG"]
    ).clip(lower=0)

    # Regra solicitada: vencimentos anteriores a hoje são tratados como pagos.
    # Também eliminamos títulos que já possuem data de pagamento ou saldo zerado.
    df = df[
        df["DT_VENCIMENTO"].notna()
        & (df["DT_VENCIMENTO"].dt.normalize() >= hoje.normalize())
        & df["DT_PAGAMENTO"].isna()
        & (df["VALOR_A_PAGAR"] > 0.005)
    ].copy()

    df["ORIGEM"] = "CSV"
    df["NR_PARCELA"] = df["DUPLICATA"].fillna("")
    return df[[
        "FORNECEDOR", "FORNECEDOR_CHAVE", "NR_DOCUMENTO_XML", "NR_PARCELA",
        "DT_VENCIMENTO", "VALOR_A_PAGAR", "ORIGEM", "HISTÓRICO"
    ]]


def preparar_parcelas_xml(df_parcelas: pd.DataFrame, hoje: pd.Timestamp) -> pd.DataFrame:
    if df_parcelas.empty:
        return pd.DataFrame(columns=[
            "FORNECEDOR", "FORNECEDOR_CHAVE", "NR_DOCUMENTO_XML", "NR_PARCELA",
            "DT_VENCIMENTO", "VALOR_A_PAGAR", "ORIGEM", "HISTÓRICO"
        ])
    df = df_parcelas.copy()
    df["DT_VENCIMENTO"] = pd.to_datetime(df["DT_VENCIMENTO"], dayfirst=True, errors="coerce")
    df["VALOR_A_PAGAR"] = pd.to_numeric(df["VL_PARCELA"], errors="coerce").fillna(0.0)
    df["FORNECEDOR"] = df["NM_EMITENTE"].fillna("").astype(str).str.strip()
    df["FORNECEDOR_CHAVE"] = df["FORNECEDOR"].apply(normalizar_texto)
    df["NR_DOCUMENTO_XML"] = df["NR_DOCUMENTO"].apply(normalizar_numero_nf)
    df["ORIGEM"] = "XML - NOTA AUSENTE NO CSV"
    df["HISTÓRICO"] = "Parcela extraída da NF-e"
    return df[
        df["DT_VENCIMENTO"].notna()
        & (df["DT_VENCIMENTO"].dt.normalize() >= hoje.normalize())
        & (df["VALOR_A_PAGAR"] > 0.005)
    ][[
        "FORNECEDOR", "FORNECEDOR_CHAVE", "NR_DOCUMENTO_XML", "NR_PARCELA",
        "DT_VENCIMENTO", "VALOR_A_PAGAR", "ORIGEM", "HISTÓRICO"
    ]].copy()


def conciliar_csv_xml(df_csv_aberto: pd.DataFrame, df_xml_aberto: pd.DataFrame) -> pd.DataFrame:
    if df_csv_aberto.empty:
        return df_xml_aberto.copy()
    if df_xml_aberto.empty:
        return df_csv_aberto.copy()

    # O histórico do CSV é a fonte indicada pelo usuário para identificar a NF.
    # Compara número da NF e fornecedor normalizado para evitar colisões entre fornecedores.
    pares_csv = set(
        zip(df_csv_aberto["FORNECEDOR_CHAVE"], df_csv_aberto["NR_DOCUMENTO_XML"])
    )
    numeros_csv = set(df_csv_aberto.loc[df_csv_aberto["NR_DOCUMENTO_XML"] != "", "NR_DOCUMENTO_XML"])

    def ja_existe(row) -> bool:
        numero = row["NR_DOCUMENTO_XML"]
        fornecedor = row["FORNECEDOR_CHAVE"]
        if not numero:
            return False
        return (fornecedor, numero) in pares_csv or numero in numeros_csv

    faltantes = df_xml_aberto[~df_xml_aberto.apply(ja_existe, axis=1)].copy()
    return pd.concat([df_csv_aberto, faltantes], ignore_index=True)


def adicionar_periodos(df: pd.DataFrame) -> pd.DataFrame:
    out = df.copy()
    data = pd.to_datetime(out["DT_VENCIMENTO"], errors="coerce")
    out["INICIO_SEMANA"] = data - pd.to_timedelta(data.dt.weekday, unit="D")
    out["FIM_SEMANA"] = out["INICIO_SEMANA"] + pd.Timedelta(days=6)
    out["SEMANA"] = out.apply(
        lambda r: f"{r['INICIO_SEMANA']:%d/%m/%Y} a {r['FIM_SEMANA']:%d/%m/%Y}", axis=1
    )
    out["MES_REF"] = data.dt.to_period("M").dt.to_timestamp()
    out["MÊS"] = out["MES_REF"].dt.strftime("%m/%Y")
    return out


def gerar_excel_contas(df_dia, df_semana, df_mes, df_fornecedor_semana, df_fornecedor_mes) -> bytes:
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df_dia.to_excel(writer, index=False, sheet_name="Contas por dia")
        df_semana.to_excel(writer, index=False, sheet_name="Consolidado semanal")
        df_mes.to_excel(writer, index=False, sheet_name="Consolidado mensal")
        df_fornecedor_semana.to_excel(writer, index=False, sheet_name="Fornecedor semana")
        df_fornecedor_mes.to_excel(writer, index=False, sheet_name="Fornecedor mês")
    return output.getvalue()

# =========================================================

# =========================================================
# PREVISÃO FINANCEIRA - FUNÇÕES COMPLEMENTARES
# =========================================================

CATEGORIAS_FINAIS = ["COMPRAS", "BONIFICAÇÕES", "DEVOLUÇÕES", "OUTRAS SITUAÇÕES"]


def categoria_padrao(row: pd.Series) -> str:
    auto = str(row.get("CLASSIFICACAO_AUTOMATICA", "COMPRAS"))
    if auto == "BONIFICACOES":
        return "BONIFICAÇÕES"
    if auto == "DEVOLUCOES":
        return "DEVOLUÇÕES"
    if auto == "MISTAS":
        return "OUTRAS SITUAÇÕES"
    return "COMPRAS"


def aplicar_classificacao_manual(df_notas: pd.DataFrame) -> pd.DataFrame:
    out = df_notas.copy()
    mapa = st.session_state.get("mapa_classificacao_xml", {})
    out["CATEGORIA_PADRAO"] = out.apply(categoria_padrao, axis=1)
    out["CATEGORIA_FINAL"] = out.apply(
        lambda r: mapa.get(str(r.get("NR_CHAVE_ACESSO", "")), r["CATEGORIA_PADRAO"]),
        axis=1,
    )
    return out


def preparar_parcelas_xml_classificadas(
    df_parcelas: pd.DataFrame,
    df_notas_classificadas: pd.DataFrame,
    hoje: pd.Timestamp,
) -> pd.DataFrame:
    colunas = [
        "FORNECEDOR", "FORNECEDOR_CHAVE", "NR_DOCUMENTO_XML", "NR_PARCELA",
        "DT_VENCIMENTO", "VALOR_A_PAGAR", "ORIGEM", "HISTÓRICO", "NR_CHAVE_ACESSO",
    ]
    if df_parcelas.empty or df_notas_classificadas.empty:
        return pd.DataFrame(columns=colunas)

    compras = df_notas_classificadas[
        df_notas_classificadas["CATEGORIA_FINAL"] == "COMPRAS"
    ][["NR_CHAVE_ACESSO", "NR_DOCUMENTO", "NM_EMITENTE", "CATEGORIA_FINAL"]].copy()

    if compras.empty:
        return pd.DataFrame(columns=colunas)

    df = df_parcelas.merge(
        compras[["NR_CHAVE_ACESSO", "CATEGORIA_FINAL"]],
        on="NR_CHAVE_ACESSO",
        how="inner",
    )
    df["DT_VENCIMENTO"] = pd.to_datetime(df["DT_VENCIMENTO"], dayfirst=True, errors="coerce")
    df["VALOR_A_PAGAR"] = pd.to_numeric(df["VL_PARCELA"], errors="coerce").fillna(0.0)
    df["FORNECEDOR"] = df["NM_EMITENTE"].fillna("").astype(str).str.strip()
    df["FORNECEDOR_CHAVE"] = df["FORNECEDOR"].apply(normalizar_texto)
    df["NR_DOCUMENTO_XML"] = df["NR_DOCUMENTO"].apply(normalizar_numero_nf)
    df["ORIGEM"] = "XML - NF AINDA NÃO LANÇADA NO AUTCOM"
    df["HISTÓRICO"] = "Parcela extraída da NF-e ainda não encontrada no CSV"

    df = df[
        df["DT_VENCIMENTO"].notna()
        & (df["DT_VENCIMENTO"].dt.normalize() >= hoje.normalize())
        & (df["VALOR_A_PAGAR"] > 0.005)
    ].copy()
    return df[colunas]


def conciliar_previsao(df_csv_aberto: pd.DataFrame, df_xml_aberto: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame]:
    """Une CSV + parcelas XML que ainda não aparecem no histórico do CSV."""
    if df_xml_aberto.empty:
        return df_csv_aberto.copy(), df_xml_aberto.copy()

    numeros_csv = set(
        df_csv_aberto.loc[df_csv_aberto["NR_DOCUMENTO_XML"] != "", "NR_DOCUMENTO_XML"].astype(str)
    ) if not df_csv_aberto.empty else set()

    pares_csv = set(
        zip(
            df_csv_aberto["FORNECEDOR_CHAVE"].astype(str),
            df_csv_aberto["NR_DOCUMENTO_XML"].astype(str),
        )
    ) if not df_csv_aberto.empty else set()

    def existe_no_csv(row: pd.Series) -> bool:
        nf = str(row.get("NR_DOCUMENTO_XML", ""))
        forn = str(row.get("FORNECEDOR_CHAVE", ""))
        if not nf:
            return False
        # Primeiro tenta fornecedor + NF; depois NF global, conforme o histórico do relatório.
        return (forn, nf) in pares_csv or nf in numeros_csv

    xml_faltantes = df_xml_aberto[~df_xml_aberto.apply(existe_no_csv, axis=1)].copy()
    combinado = pd.concat([df_csv_aberto, xml_faltantes], ignore_index=True, sort=False)
    return combinado, xml_faltantes


def preparar_periodos_previsao(df: pd.DataFrame) -> pd.DataFrame:
    out = adicionar_periodos(df)
    out["DATA"] = pd.to_datetime(out["DT_VENCIMENTO"]).dt.strftime("%d/%m/%Y")
    out["DIA_SEMANA"] = pd.to_datetime(out["DT_VENCIMENTO"]).dt.day_name(locale=None)
    return out


def tabela_periodo(df: pd.DataFrame, modo: str) -> pd.DataFrame:
    if modo == "Dia":
        agrup = (
            df.groupby(["DT_VENCIMENTO", "DATA"], as_index=False)
            .agg(QTD_TITULOS=("VALOR_A_PAGAR", "size"), VALOR_A_PAGAR=("VALOR_A_PAGAR", "sum"))
            .sort_values("DT_VENCIMENTO")
        )
        return agrup[["DATA", "QTD_TITULOS", "VALOR_A_PAGAR"]]
    if modo == "Semana":
        agrup = (
            df.groupby(["INICIO_SEMANA", "SEMANA"], as_index=False)
            .agg(QTD_TITULOS=("VALOR_A_PAGAR", "size"), VALOR_A_PAGAR=("VALOR_A_PAGAR", "sum"))
            .sort_values("INICIO_SEMANA")
        )
        return agrup[["SEMANA", "QTD_TITULOS", "VALOR_A_PAGAR"]]
    agrup = (
        df.groupby(["MES_REF", "MÊS"], as_index=False)
        .agg(QTD_TITULOS=("VALOR_A_PAGAR", "size"), VALOR_A_PAGAR=("VALOR_A_PAGAR", "sum"))
        .sort_values("MES_REF")
    )
    return agrup[["MÊS", "QTD_TITULOS", "VALOR_A_PAGAR"]]


def filtrar_periodo(df: pd.DataFrame, modo: str, periodo: str) -> pd.DataFrame:
    coluna = {"Dia": "DATA", "Semana": "SEMANA", "Mês": "MÊS"}[modo]
    return df[df[coluna] == periodo].copy()


def gerar_excel_previsao(df_base: pd.DataFrame, df_xml_faltantes: pd.DataFrame) -> bytes:
    output = io.BytesIO()
    diario = tabela_periodo(df_base, "Dia")
    semanal = tabela_periodo(df_base, "Semana")
    mensal = tabela_periodo(df_base, "Mês")
    por_fornecedor = (
        df_base.groupby("FORNECEDOR", as_index=False)
        .agg(QTD_TITULOS=("VALOR_A_PAGAR", "size"), VALOR_A_PAGAR=("VALOR_A_PAGAR", "sum"))
        .sort_values("VALOR_A_PAGAR", ascending=False)
    )
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df_base.to_excel(writer, index=False, sheet_name="Previsão completa")
        diario.to_excel(writer, index=False, sheet_name="Por dia")
        semanal.to_excel(writer, index=False, sheet_name="Por semana")
        mensal.to_excel(writer, index=False, sheet_name="Por mês")
        por_fornecedor.to_excel(writer, index=False, sheet_name="Por fornecedor")
        df_xml_faltantes.to_excel(writer, index=False, sheet_name="XML não lançado")
    return output.getvalue()




def render_pagina_previsao_financeira():
    st.markdown("""
    <style>
    [data-testid="stMetric"] {background: #ffffff; border: 1px solid #e5e7eb; padding: 14px; border-radius: 12px;}
    </style>
    """, unsafe_allow_html=True)

    st.markdown('<div class="section-title">Leitor de XML/ Previsão financeira</div>', unsafe_allow_html=True)
    st.caption("Consolidação do CSV do Autcom com parcelas de NF-e ainda não lançadas no sistema.")

    pagina = st.radio(
        "Etapa da previsão",
        ["Leitor de XMLs", "Classificação dos XMLs", "Previsão de pagamentos", "Conciliação CSV x XML"],
        horizontal=True,
        key="pagina_interna_previsao_financeira",
    )
    with st.expander("CFOPs de bonificação e remessa de marketing considerados"):
        st.code(", ".join(sorted(CFOPS_BONIFICACAO)))

    st.title("Previsão de fornecedores a pagar")
    st.caption("CSV do Autcom + parcelas de NF-e ainda não lançadas no sistema.")

    st.markdown(
        """
        <div style="background:#f8fafc;border:1px solid #dbe3ec;border-left:5px solid #2563eb;
                    border-radius:12px;padding:18px 20px;margin:12px 0 20px 0;">
            <div style="font-size:18px;font-weight:700;color:#0f172a;margin-bottom:10px;">
                Como preparar os arquivos para análise
            </div>
            <div style="color:#334155;line-height:1.65;font-size:15px;">
                <b>1.</b> No <b>Painel da Citel</b>, acesse <b>Operações</b> e gere os XMLs de entrada de fornecedores do período desejado.<br>
                <b>2.</b> No <b>Jabu, dentro do Autcom</b>, gere o CSV de fornecedores a pagar com os vencimentos do mesmo período.<br>
                <b>3.</b> Faça o upload dos dois arquivos nos campos abaixo e aguarde o processamento da análise.
            </div>
            <div style="margin-top:10px;color:#64748b;font-size:13px;">
                Para uma conciliação correta, use o mesmo período na geração do XML e do CSV.
            </div>
        </div>
        """,
        unsafe_allow_html=True,
    )

    col_up1, col_up2 = st.columns(2)
    with col_up1:
        uploads_xml = st.file_uploader(
            "XMLs ou ZIPs de NF-e",
            type=["xml", "zip"],
            accept_multiple_files=True,
            key="uploads_xml_previsao",
        )
    with col_up2:
        upload_csv = st.file_uploader(
            "CSV de contas a pagar do Autcom",
            type=["csv", "txt"],
            key="upload_csv_previsao",
        )

    # Leitura dos XMLs é opcional; o CSV sozinho já gera previsão.
    if uploads_xml:
        with st.spinner("Lendo XMLs..."):
            df_notas_raw, df_itens_raw, df_parcelas_raw, erros_xml = processar_uploads(uploads_xml)
        if not df_notas_raw.empty:
            df_notas_raw = df_notas_raw.copy()
            df_notas_raw["CLASSIFICACAO_AUTOMATICA"] = df_notas_raw.apply(
                lambda r: classificar_nota(r.get("CFOPS_LISTA", []), r.get("NATUREZA_OPERACAO", "")), axis=1
            )
            df_notas_classificadas = aplicar_classificacao_manual(df_notas_raw)
        else:
            df_notas_classificadas = pd.DataFrame()
            df_parcelas_raw = pd.DataFrame()
    else:
        df_notas_raw = pd.DataFrame()
        df_notas_classificadas = pd.DataFrame()
        df_itens_raw = pd.DataFrame()
        df_parcelas_raw = pd.DataFrame()
        erros_xml = []

    hoje = pd.Timestamp.today().normalize()

    if pagina == "Leitor de XMLs":
        st.subheader("Leitor de XMLs - produtos das notas")
        st.caption(
            "Envie XMLs ou ZIPs de NF-e para listar os produtos por nota e consolidar quantidades quando o mesmo produto aparecer mais de uma vez."
        )

        if not uploads_xml:
            st.info("Envie XMLs ou um ZIP no campo acima para ler os produtos das notas.")
            return

        if erros_xml:
            with st.expander("Arquivos XML ignorados ou com erro"):
                for erro in erros_xml:
                    st.write(erro)

        if df_itens_raw.empty:
            st.warning("Não encontrei produtos nos XMLs enviados. Confira se os arquivos são NF-e completas e não eventos/cancelamentos.")
            return

        df_notas_unicas, df_itens_unicos, df_notas_repetidas = remover_notas_xml_repetidas(df_notas_raw, df_itens_raw)
        if not df_notas_repetidas.empty:
            st.warning(
                f"Foram encontradas {len(df_notas_repetidas)} nota(s) repetida(s) pela chave de acesso. "
                "Elas foram sinalizadas e desconsideradas da consolidação."
            )
            with st.expander("Notas repetidas desconsideradas"):
                colunas_rep = [
                    "NR_CHAVE_ACESSO", "NM_EMITENTE", "NR_DOCUMENTO", "NR_SERIE",
                    "DT_EMISSAO", "VL_NOTA_FISCAL", "ARQUIVO_ORIGEM",
                ]
                colunas_rep = [c for c in colunas_rep if c in df_notas_repetidas.columns]
                st.dataframe(
                    df_notas_repetidas[colunas_rep],
                    use_container_width=True,
                    hide_index=True,
                    column_config={
                        "VL_NOTA_FISCAL": st.column_config.NumberColumn("Valor NF", format="R$ %.2f"),
                    },
                )

        produtos_xml = consolidar_produtos_xml(df_itens_unicos)
        c1, c2, c3, c4 = st.columns(4)
        c1.metric("Notas válidas", len(df_notas_unicas), f"{len(df_notas_repetidas)} repetida(s)")
        c2.metric("Produtos consolidados", len(produtos_xml))
        c3.metric("Itens considerados", len(df_itens_unicos))
        c4.metric("Valor total dos itens", formatar_moeda(float(pd.to_numeric(df_itens_unicos.get("VL_TOTAL_ITEM", 0), errors="coerce").fillna(0).sum())))

        fornecedores = sorted(df_itens_unicos["NM_EMITENTE"].dropna().astype(str).unique()) if "NM_EMITENTE" in df_itens_unicos.columns else []
        filtro_fornecedor = st.multiselect(
            "Filtrar fornecedor",
            options=fornecedores,
            default=fornecedores,
            key="filtro_fornecedor_leitor_xml",
        )

        itens_filtrados = df_itens_unicos.copy()
        if filtro_fornecedor and "NM_EMITENTE" in itens_filtrados.columns:
            itens_filtrados = itens_filtrados[itens_filtrados["NM_EMITENTE"].isin(filtro_fornecedor)].copy()
        if "NR_CHAVE_ACESSO" in itens_filtrados.columns and "NR_CHAVE_ACESSO" in df_notas_unicas.columns:
            chaves_disponiveis = sorted({str(x).strip() for x in itens_filtrados["NR_CHAVE_ACESSO"].dropna() if str(x).strip()})
            notas_disponiveis = df_notas_unicas[df_notas_unicas["NR_CHAVE_ACESSO"].astype(str).str.strip().isin(chaves_disponiveis)].copy()

            opcoes_notas = {}
            for _, nota_row in notas_disponiveis.iterrows():
                chave = str(nota_row.get("NR_CHAVE_ACESSO", "")).strip()
                label = (
                    f"NF {nota_row.get('NR_DOCUMENTO', '')} | "
                    f"{nota_row.get('NM_EMITENTE', '')} | "
                    f"{nota_row.get('DT_EMISSAO', '')} | "
                    f"{chave[-8:] if chave else ''}"
                )
                opcoes_notas[label] = chave

            notas_selecionadas_labels = st.multiselect(
                "Notas para considerar no Excel e na consolidação",
                options=list(opcoes_notas.keys()),
                default=list(opcoes_notas.keys()),
                key="notas_selecionadas_leitor_xml",
            )
            chaves_selecionadas = {opcoes_notas[label] for label in notas_selecionadas_labels}
            itens_filtrados = itens_filtrados[itens_filtrados["NR_CHAVE_ACESSO"].astype(str).str.strip().isin(chaves_selecionadas)].copy()
            df_notas_selecionadas = notas_disponiveis[notas_disponiveis["NR_CHAVE_ACESSO"].astype(str).str.strip().isin(chaves_selecionadas)].copy()
        else:
            df_notas_selecionadas = df_notas_unicas.copy()

        st.caption(f"{len(df_notas_selecionadas)} nota(s) selecionada(s) para a consolidação e para o Excel.")
        produtos_filtrados = consolidar_produtos_xml(itens_filtrados)

        busca_produto = st.text_input("Pesquisar produto/código", key="busca_produtos_xml")
        if busca_produto:
            termo = busca_produto.lower()
            produtos_filtrados = produtos_filtrados[
                produtos_filtrados["COD_PRODUTO"].astype(str).str.lower().str.contains(termo, na=False)
                | produtos_filtrados["DESCRICAO"].astype(str).str.lower().str.contains(termo, na=False)
            ].copy()

        st.markdown("### Produtos consolidados")
        st.dataframe(
            produtos_filtrados,
            use_container_width=True,
            hide_index=True,
            height=430,
            column_config={
                "QTD_TOTAL": st.column_config.NumberColumn("Quantidade total", format="%.2f"),
                "VL_UNITARIO_MEDIO": st.column_config.NumberColumn("Valor unitário médio", format="R$ %.4f"),
                "VL_TOTAL_ITEM": st.column_config.NumberColumn("Valor total", format="R$ %.2f"),
                "QTD_LANCAMENTOS": st.column_config.NumberColumn("Lançamentos", format="%d"),
            },
        )

        with st.expander("Itens por nota"):
            colunas_itens_xml = [
                "NM_EMITENTE", "NR_DOCUMENTO", "NR_SERIE", "DT_EMISSAO", "CFOP",
                "COD_PRODUTO", "DESCRICAO", "QTD", "VL_UNITARIO", "VL_TOTAL_ITEM", "ARQUIVO_ORIGEM",
            ]
            colunas_itens_xml = [c for c in colunas_itens_xml if c in itens_filtrados.columns]
            st.dataframe(
                itens_filtrados[colunas_itens_xml],
                use_container_width=True,
                hide_index=True,
                height=420,
                column_config={
                    "QTD": st.column_config.NumberColumn("Quantidade", format="%.2f"),
                    "VL_UNITARIO": st.column_config.NumberColumn("Valor unitário", format="R$ %.4f"),
                    "VL_TOTAL_ITEM": st.column_config.NumberColumn("Valor total", format="R$ %.2f"),
                },
            )

        with st.expander("Notas válidas consideradas"):
            colunas_notas_xml = [
                "NM_EMITENTE", "NR_DOCUMENTO", "NR_SERIE", "DT_EMISSAO",
                "CFOPS", "NATUREZA_OPERACAO", "VL_NOTA_FISCAL", "QTD_ITENS", "ARQUIVO_ORIGEM",
            ]
            colunas_notas_xml = [c for c in colunas_notas_xml if c in df_notas_selecionadas.columns]
            st.dataframe(
                df_notas_selecionadas[colunas_notas_xml],
                use_container_width=True,
                hide_index=True,
                height=320,
                column_config={
                    "VL_NOTA_FISCAL": st.column_config.NumberColumn("Valor NF", format="R$ %.2f"),
                    "QTD_ITENS": st.column_config.NumberColumn("Itens", format="%d"),
                },
            )

        st.download_button(
            "Baixar produtos dos XMLs em Excel",
            data=gerar_excel_produtos_xml(df_notas_selecionadas, itens_filtrados, produtos_filtrados),
            file_name="produtos_xml_consolidados.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            type="primary",
        )
        return

    if pagina == "Classificação dos XMLs":
        st.subheader("Classificação das NF-e importadas")
        st.caption(
            "Bonificações/remessas de marketing e devoluções são desconsideradas da previsão. "
            "Você pode alterar manualmente qualquer nota para Compras, Bonificações, Devoluções ou Outras Situações."
        )

        if df_notas_classificadas.empty:
            st.info("Envie XMLs ou um ZIP para classificar as notas.")
            st.stop()

        cards = st.columns(4)
        for col, categoria in zip(cards, CATEGORIAS_FINAIS):
            base_cat = df_notas_classificadas[df_notas_classificadas["CATEGORIA_FINAL"] == categoria]
            col.metric(categoria, formatar_moeda(base_cat["VL_NOTA_FISCAL"].sum()), f"{len(base_cat)} notas")

        tabela_edicao = df_notas_classificadas[[
            "NR_CHAVE_ACESSO", "NM_EMITENTE", "NR_DOCUMENTO", "DT_EMISSAO",
            "CFOPS", "NATUREZA_OPERACAO", "VL_NOTA_FISCAL",
            "CATEGORIA_PADRAO", "CATEGORIA_FINAL",
        ]].copy()
        tabela_edicao = tabela_edicao.sort_values(["CATEGORIA_FINAL", "NM_EMITENTE", "NR_DOCUMENTO"])

        editado = st.data_editor(
            tabela_edicao,
            use_container_width=True,
            hide_index=True,
            disabled=[
                "NR_CHAVE_ACESSO", "NM_EMITENTE", "NR_DOCUMENTO", "DT_EMISSAO",
                "CFOPS", "NATUREZA_OPERACAO", "VL_NOTA_FISCAL", "CATEGORIA_PADRAO",
            ],
            column_config={
                "CATEGORIA_FINAL": st.column_config.SelectboxColumn(
                    "Categoria final",
                    options=CATEGORIAS_FINAIS,
                    required=True,
                ),
                "VL_NOTA_FISCAL": st.column_config.NumberColumn("Valor da NF", format="R$ %.2f"),
            },
            key="editor_classificacao_xml",
        )

        if st.button("Salvar classificação dos XMLs", type="primary"):
            st.session_state["mapa_classificacao_xml"] = dict(
                zip(editado["NR_CHAVE_ACESSO"].astype(str), editado["CATEGORIA_FINAL"].astype(str))
            )
            st.success("Classificação salva. A previsão financeira já usará essas decisões.")
            st.rerun()

        mistas = df_notas_classificadas[
            df_notas_classificadas["CLASSIFICACAO_AUTOMATICA"] == "MISTAS"
        ]
        if not mistas.empty:
            st.warning(
                f"Existem {len(mistas)} notas mistas. Elas ficam em Outras Situações por padrão até você decidir a categoria final."
            )

        if erros_xml:
            with st.expander("Arquivos XML ignorados ou com erro"):
                for erro in erros_xml:
                    st.write(erro)

    else:
        if upload_csv is None:
            st.info("Envie o CSV de contas a pagar para montar a previsão financeira.")
            st.stop()

        try:
            df_csv_original = ler_csv_financeiro(upload_csv)
            df_csv_aberto = preparar_contas_csv(df_csv_original, hoje)
        except Exception as exc:
            st.error(str(exc))
            st.stop()

        # Reaplica decisões salvas antes de formar a previsão.
        if not df_notas_raw.empty:
            df_notas_classificadas = aplicar_classificacao_manual(df_notas_raw)
        df_xml_aberto = preparar_parcelas_xml_classificadas(df_parcelas_raw, df_notas_classificadas, hoje)
        df_previsao, df_xml_faltantes = conciliar_previsao(df_csv_aberto, df_xml_aberto)
        df_previsao = preparar_periodos_previsao(df_previsao)

        if df_previsao.empty:
            st.warning("Não há títulos futuros a pagar nos arquivos enviados.")
            st.stop()

        if pagina == "Previsão de pagamentos":
            st.subheader("Previsão consolidada")
            st.info(
                f"Data de referência: {hoje.strftime('%d/%m/%Y')}. Vencimentos anteriores são tratados como pagos e não aparecem."
            )

            fornecedores = sorted(df_previsao["FORNECEDOR"].dropna().unique())
            filtro_fornecedores = st.multiselect(
                "Fornecedores",
                options=fornecedores,
                default=fornecedores,
            )
            origem_opcoes = sorted(df_previsao["ORIGEM"].dropna().unique())
            filtro_origem = st.multiselect("Origem", options=origem_opcoes, default=origem_opcoes)
            df_view = df_previsao[
                df_previsao["FORNECEDOR"].isin(filtro_fornecedores)
                & df_previsao["ORIGEM"].isin(filtro_origem)
            ].copy()

            c1, c2, c3, c4 = st.columns(4)
            c1.metric("Total previsto", formatar_moeda(df_view["VALOR_A_PAGAR"].sum()))
            c2.metric("Fornecedores", df_view["FORNECEDOR"].nunique())
            c3.metric("Títulos/parcelas", len(df_view))
            c4.metric("Incluídos pelo XML", len(df_view[df_view["ORIGEM"].str.startswith("XML", na=False)]))

            st.markdown("### Visão do fluxo")
            modo = st.radio("Agrupar por", ["Dia", "Semana", "Mês"], horizontal=True)
            resumo_periodo = tabela_periodo(df_view, modo)
            st.dataframe(
                formatar_df_moeda(resumo_periodo, ["VALOR_A_PAGAR"]),
                use_container_width=True,
                hide_index=True,
            )

            coluna_periodo = {"Dia": "DATA", "Semana": "SEMANA", "Mês": "MÊS"}[modo]
            periodos = resumo_periodo[coluna_periodo].astype(str).tolist()
            periodo_escolhido = st.selectbox(f"Clique/escolha o {modo.lower()} para abrir o drill", periodos)
            drill_periodo = filtrar_periodo(df_view, modo, periodo_escolhido)

            st.markdown(f"### Drill do período: {periodo_escolhido}")
            resumo_fornecedor = (
                drill_periodo.groupby("FORNECEDOR", as_index=False)
                .agg(QTD_TITULOS=("VALOR_A_PAGAR", "size"), VALOR_A_PAGAR=("VALOR_A_PAGAR", "sum"))
                .sort_values("VALOR_A_PAGAR", ascending=False)
            )
            d1, d2, d3 = st.columns(3)
            d1.metric("Valor do período", formatar_moeda(drill_periodo["VALOR_A_PAGAR"].sum()))
            d2.metric("Fornecedores", drill_periodo["FORNECEDOR"].nunique())
            d3.metric("Títulos", len(drill_periodo))

            col_tab, col_graf = st.columns([1.2, 1])
            with col_tab:
                st.dataframe(
                    formatar_df_moeda(resumo_fornecedor, ["VALOR_A_PAGAR"]),
                    use_container_width=True,
                    hide_index=True,
                )
            with col_graf:
                st.bar_chart(resumo_fornecedor.set_index("FORNECEDOR")[["VALOR_A_PAGAR"]])

            fornecedor_drill = st.selectbox(
                "Fornecedor para ver os títulos do período",
                resumo_fornecedor["FORNECEDOR"].tolist(),
            )
            titulos = drill_periodo[drill_periodo["FORNECEDOR"] == fornecedor_drill].copy()
            titulos = titulos.sort_values("DT_VENCIMENTO")
            titulos_exib = titulos[[
                "DATA", "NR_DOCUMENTO_XML", "NR_PARCELA", "VALOR_A_PAGAR", "ORIGEM", "HISTÓRICO"
            ]].copy()
            st.metric("Total do fornecedor no período", formatar_moeda(titulos["VALOR_A_PAGAR"].sum()))
            st.dataframe(formatar_df_moeda(titulos_exib, ["VALOR_A_PAGAR"]), use_container_width=True, hide_index=True)

            st.markdown("### Ranking geral por fornecedor")
            geral_fornecedor = (
                df_view.groupby("FORNECEDOR", as_index=False)
                .agg(QTD_TITULOS=("VALOR_A_PAGAR", "size"), VALOR_A_PAGAR=("VALOR_A_PAGAR", "sum"))
                .sort_values("VALOR_A_PAGAR", ascending=False)
            )
            st.dataframe(formatar_df_moeda(geral_fornecedor, ["VALOR_A_PAGAR"]), use_container_width=True, hide_index=True)

            excel = gerar_excel_previsao(df_view, df_xml_faltantes)
            st.download_button(
                "Baixar previsão financeira em Excel",
                data=excel,
                file_name="previsao_fornecedores_a_pagar.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )

        else:
            st.subheader("Conciliação CSV x XML")
            st.caption(
                "O CSV é a base do Autcom. Uma NF do XML só é acrescentada quando seu número não aparece na coluna HISTÓRICO do CSV."
            )

            c1, c2, c3 = st.columns(3)
            c1.metric("Títulos futuros do CSV", len(df_csv_aberto))
            c2.metric("Parcelas XML elegíveis", len(df_xml_aberto))
            c3.metric("Parcelas XML acrescentadas", len(df_xml_faltantes))

            st.markdown("### Notas/parcelas acrescentadas pelo XML")
            if df_xml_faltantes.empty:
                st.success("Nenhuma parcela XML precisou ser acrescentada: todas as NFs elegíveis já aparecem no CSV.")
            else:
                xml_exib = df_xml_faltantes.copy()
                xml_exib["DT_VENCIMENTO"] = pd.to_datetime(xml_exib["DT_VENCIMENTO"]).dt.strftime("%d/%m/%Y")
                st.dataframe(formatar_df_moeda(xml_exib, ["VALOR_A_PAGAR"]), use_container_width=True, hide_index=True)

            st.markdown("### Títulos considerados a partir do CSV")
            csv_exib = df_csv_aberto.copy()
            csv_exib["DT_VENCIMENTO"] = pd.to_datetime(csv_exib["DT_VENCIMENTO"]).dt.strftime("%d/%m/%Y")
            st.dataframe(formatar_df_moeda(csv_exib, ["VALOR_A_PAGAR"]), use_container_width=True, hide_index=True)

            st.markdown("### XMLs desconsiderados da previsão")
            if df_notas_classificadas.empty:
                st.info("Nenhum XML enviado.")
            else:
                fora = df_notas_classificadas[df_notas_classificadas["CATEGORIA_FINAL"] != "COMPRAS"].copy()
                st.dataframe(
                    formatar_df_moeda(
                        fora[["NM_EMITENTE", "NR_DOCUMENTO", "CFOPS", "NATUREZA_OPERACAO", "VL_NOTA_FISCAL", "CATEGORIA_FINAL"]],
                        ["VL_NOTA_FISCAL"],
                    ),
                    use_container_width=True,
                    hide_index=True,
                )


# =========================================================
# APP STREAMLIT
# =========================================================

aplicar_css_global()
render_header()

st.sidebar.markdown(
    """
    <div class="sidebar-brand">
        <div class="sidebar-brand-icon">
            <svg width="23" height="23" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="1.8" stroke-linecap="round" stroke-linejoin="round">
                <path d="M4 19V5"/><path d="M4 19h16"/><rect x="7" y="10" width="3" height="6" rx="1"/><rect x="12" y="6" width="3" height="10" rx="1"/><rect x="17" y="3" width="3" height="13" rx="1"/>
            </svg>
        </div>
        <div>
            <div class="sidebar-brand-title">Análise de Giro</div>
            <div class="sidebar-brand-subtitle">Planeje melhor. Compre certo.</div>
        </div>
    </div>
    """,
    unsafe_allow_html=True,
)
pagina = st.sidebar.radio(
    "Navegação",
    ["Giro Consolidado", "Pedido de Compra", "Exportações", "Ruptura por Marca", "Comparativo de Pedidos", "Leitor de XML/ Previsão financeira", "Tratamento Final"],
    label_visibility="collapsed",
)

st.sidebar.markdown("### Parâmetros")
st.sidebar.markdown('<div class="param-card">', unsafe_allow_html=True)
dias_estoque_alvo = st.sidebar.number_input(
    "Dias de estoque alvo",
    min_value=1,
    max_value=365,
    value=60,
    step=1,
    help="Define quantos dias de cobertura de estoque o pedido deve considerar.",
)
meses_alerta_sem_compra = 3
st.sidebar.markdown(
    '<div class="param-note">Alerta sem compra fixo em <strong>03 meses</strong>.<br>Estoque Final = Estoque Atual Geral + Saldo em Trânsito/ABERTO.</div></div>',
    unsafe_allow_html=True,
)

if pagina == "Leitor de XML/ Previsão financeira":
    render_pagina_previsao_financeira()
    st.stop()

if pagina == "Ruptura por Marca":
    render_pagina_ruptura_por_marca()
    st.stop()

if pagina == "Comparativo de Pedidos":
    render_pagina_comparativo_pedidos()
    st.stop()

if pagina == "Tratamento Final":
    st.markdown('<div class="section-title">Tratamento de Pedido Final</div>', unsafe_allow_html=True)
    st.caption(
        "Envie a planilha final editável. O sistema vai gerar um Excel para importação no Autcom: "
        "coluna B = zx, coluna F = PEDIDO Final e coluna H = Preço Última Compra."
    )

    if False and google_configurado():
        st.markdown("### Usar pedido aprovado do Google Drive")
        try:
            pedidos_drive = google_listar_pedidos()
            pedidos_aprovados = pedidos_drive[pedidos_drive["status"].astype(str).str.lower().isin(["aprovado", "em edicao"])].copy()
            if pedidos_aprovados.empty:
                st.info("Não há pedidos aprovados ou em edição no controle do Drive.")
            else:
                opcoes_drive = {
                    f"{r.get('id_pedido', '')} | {r.get('fornecedor', '')} | {r.get('nome_pedido', '')} | {r.get('status', '')}": r.to_dict()
                    for _, r in pedidos_aprovados.iterrows()
                }
                pedido_drive_label = st.selectbox("Pedido do Drive", list(opcoes_drive.keys()), key="tratamento_pedido_drive")
                usuario_finalizacao = st.text_input("Finalizado por", value="", key="tratamento_usuario_drive")
                pedido_drive_info = opcoes_drive[pedido_drive_label]

                if st.button("Ler pedido do Drive e gerar arquivos finais", type="primary"):
                    df_drive = google_ler_pedido_drive(pedido_drive_info["spreadsheet_id"])
                    st.session_state["df_tratamento_drive"] = df_drive
                    st.session_state["pedido_tratamento_drive_id"] = pedido_drive_info["id_pedido"]
                    st.success(f"Pedido lido do Drive: {len(df_drive)} linha(s).")

                df_drive_preview = st.session_state.get("df_tratamento_drive")
                pedido_drive_id = st.session_state.get("pedido_tratamento_drive_id")
                if df_drive_preview is not None and pedido_drive_id:
                    colunas_preview_drive = [c for c in ["zx", "codigo", "descricao", "Código Fábrica", "PEDIDO Final", "Preço Última Compra", "Valor Final do Pedido", "Total Geral do Pedido"] if c in df_drive_preview.columns]
                    st.dataframe(
                        df_drive_preview[colunas_preview_drive].head(50) if colunas_preview_drive else df_drive_preview.head(50),
                        use_container_width=True,
                        hide_index=True,
                        height=320,
                    )
                    if st.button("Salvar arquivos finais no Drive e finalizar pedido"):
                        link_autcom, link_fornecedor = google_finalizar_pedido(
                            pedido_drive_id,
                            df_drive_preview,
                            usuario=usuario_finalizacao,
                        )
                        st.success("Pedido finalizado e arquivos salvos no Drive.")
                        c_autcom, c_forn = st.columns(2)
                        c_autcom.link_button("Abrir arquivo Autcom", link_autcom)
                        c_forn.link_button("Abrir arquivo fornecedor", link_fornecedor)
                        st.rerun()
        except Exception as e:
            st.warning(f"Não consegui usar o fluxo do Google Drive: {e}")

        st.markdown("---")
        st.markdown("### Upload manual")

    st.markdown("### Ler planilha aprovada por link")
    link_tratamento = st.text_input(
        "Link do Google Sheets aprovado",
        value="",
        key="link_tratamento_google_sheets",
        placeholder="https://docs.google.com/spreadsheets/d/.../edit#gid=0",
    )
    link_tratamento = str(link_tratamento or "").strip()
    st.caption("Para funcionar sem credenciais, a planilha precisa estar compartilhada como 'Qualquer pessoa com o link - Leitor'.")

    st.markdown("### Ou envie o arquivo manualmente")
    planilha_tratamento = st.file_uploader(
        "Planilha do Pedido Final",
        type=["xlsx", "xls", "csv"],
        key="upload_tratamento_pedido_final",
    )

    if not link_tratamento and not planilha_tratamento:
        st.info("Cole o link da planilha aprovada ou envie a planilha do pedido final para gerar o arquivo de importação Autcom.")
        st.stop()

    try:
        if link_tratamento:
            df_tratamento = ler_planilha_tratamento_google_sheets(link_tratamento)
            origem_tratamento = "Google Sheets"
        else:
            df_tratamento = ler_planilha_tratamento_pedido(planilha_tratamento)
            origem_tratamento = "upload"
        df_tratamento.columns = [str(c).strip() for c in df_tratamento.columns]

        st.success(f"Planilha lida com sucesso via {origem_tratamento}: {len(df_tratamento)} linha(s).")

        colunas_preview = [c for c in ["zx", "descricao", "Código Fábrica", "PEDIDO Final", "Preço Última Compra", "Valor Final do Pedido", "Total Geral do Pedido"] if c in df_tratamento.columns]
        if colunas_preview:
            st.dataframe(
                df_tratamento[colunas_preview].head(50),
                use_container_width=True,
                hide_index=True,
                height=360,
            )
        else:
            st.dataframe(df_tratamento.head(50), use_container_width=True, hide_index=True, height=360)

        excel_tratamento = gerar_excel_autcom_tratamento(df_tratamento)
        col_dl_autcom, col_dl_fornecedor = st.columns(2)
        with col_dl_autcom:
            st.download_button(
                "⬇ Baixar pedido tratado para importação no Autcom",
                excel_tratamento,
                "pedido_tratado_importacao_autcom.xlsx",
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                type="primary",
            )
        with col_dl_fornecedor:
            st.download_button(
                "⬇ Baixar pedido para envio ao fornecedor",
                gerar_excel_fornecedor_tratamento(df_tratamento),
                "pedido_envio_fornecedor.xlsx",
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
    except Exception as e:
        st.error(str(e))

    st.stop()

st.markdown('<div class="page-card"><div class="page-card-title">Upload dos arquivos</div><div class="page-card-subtitle">Envie o PDF de Giro para iniciar. Os demais arquivos enriquecem a análise e o pedido final.</div>', unsafe_allow_html=True)
col_upload_1, col_upload_2, col_upload_3 = st.columns(3)
cadastro_google = pd.DataFrame()

with col_upload_1:
    giro_pdf = st.file_uploader("PDF - Giro de Estoque", type=["pdf"], key="upload_giro_pdf")
    render_upload_status("Giro de Estoque", giro_pdf, obrigatorio=True)
with col_upload_2:
    pedidos_pdf = st.file_uploader("PDF - Pedidos em Aberto", type=["pdf"], key="upload_pedidos_pdf")
    render_upload_status("Pedidos em Aberto", pedidos_pdf)
with col_upload_3:
    st.markdown("**Cadastro de Produtos**")
    st.link_button("🔗 Abrir / editar cadastro no Google Sheets", google_link_planilha(GOOGLE_PLANILHA_CADASTRO_ID), use_container_width=True)
    st.caption("Após editar a planilha, volte ao app e recarregue a página para puxar a versão atualizada.")

    try:
        cadastro_google = ler_cadastro_produtos_google()
        if not cadastro_google.empty:
            st.success(f"Cadastro lido do Google Sheets: {len(cadastro_google)} item(ns).")
        else:
            st.warning("Cadastro do Google Sheets está vazio ou sem as colunas obrigatórias.")
    except Exception as e:
        st.warning(
            "Não consegui ler o cadastro do Google Sheets por leitura simples. "
            "Confira se a planilha está compartilhada como 'Qualquer pessoa com o link - Leitor' "
            f"e se existe a aba 'Cadastro'. Detalhe: {e}"
        )

    cadastro_csv = st.file_uploader("CSV - Cadastro de Produtos (fallback)", type=["csv"], key="upload_cadastro_csv")
    render_upload_status("Cadastro de Produtos", cadastro_csv)

st.markdown("</div>", unsafe_allow_html=True)

if False and pagina == "Tratamento Final":
    st.markdown('<div class="section-title">Tratamento de Pedido Final</div>', unsafe_allow_html=True)
    st.caption(
        "Envie a planilha final editável. O sistema vai gerar um Excel para importação no Autcom: "
        "coluna B = zx, coluna F = PEDIDO Final e coluna H = Preço Última Compra."
    )

    if False and google_configurado():
        st.markdown("### Usar pedido aprovado do Google Drive")
        try:
            pedidos_drive = google_listar_pedidos()
            pedidos_aprovados = pedidos_drive[pedidos_drive["status"].astype(str).str.lower().isin(["aprovado", "em edicao"])].copy()
            if pedidos_aprovados.empty:
                st.info("Não há pedidos aprovados ou em edição no controle do Drive.")
            else:
                opcoes_drive = {
                    f"{r.get('id_pedido', '')} | {r.get('fornecedor', '')} | {r.get('nome_pedido', '')} | {r.get('status', '')}": r.to_dict()
                    for _, r in pedidos_aprovados.iterrows()
                }
                pedido_drive_label = st.selectbox("Pedido do Drive", list(opcoes_drive.keys()), key="tratamento_pedido_drive")
                usuario_finalizacao = st.text_input("Finalizado por", value="", key="tratamento_usuario_drive")
                pedido_drive_info = opcoes_drive[pedido_drive_label]

                if st.button("Ler pedido do Drive e gerar arquivos finais", type="primary"):
                    df_drive = google_ler_pedido_drive(pedido_drive_info["spreadsheet_id"])
                    st.session_state["df_tratamento_drive"] = df_drive
                    st.session_state["pedido_tratamento_drive_id"] = pedido_drive_info["id_pedido"]
                    st.success(f"Pedido lido do Drive: {len(df_drive)} linha(s).")

                df_drive_preview = st.session_state.get("df_tratamento_drive")
                pedido_drive_id = st.session_state.get("pedido_tratamento_drive_id")
                if df_drive_preview is not None and pedido_drive_id:
                    colunas_preview_drive = [c for c in ["zx", "codigo", "descricao", "Código Fábrica", "PEDIDO Final", "Preço Última Compra", "Valor Final do Pedido", "Total Geral do Pedido"] if c in df_drive_preview.columns]
                    st.dataframe(
                        df_drive_preview[colunas_preview_drive].head(50) if colunas_preview_drive else df_drive_preview.head(50),
                        use_container_width=True,
                        hide_index=True,
                        height=320,
                    )
                    if st.button("Salvar arquivos finais no Drive e finalizar pedido"):
                        link_autcom, link_fornecedor = google_finalizar_pedido(
                            pedido_drive_id,
                            df_drive_preview,
                            usuario=usuario_finalizacao,
                        )
                        st.success("Pedido finalizado e arquivos salvos no Drive.")
                        c_autcom, c_forn = st.columns(2)
                        c_autcom.link_button("Abrir arquivo Autcom", link_autcom)
                        c_forn.link_button("Abrir arquivo fornecedor", link_fornecedor)
                        st.rerun()
        except Exception as e:
            st.warning(f"Não consegui usar o fluxo do Google Drive: {e}")

        st.markdown("---")
        st.markdown("### Upload manual")

    st.markdown("### Ler planilha aprovada por link")
    link_tratamento = st.text_input(
        "Link do Google Sheets aprovado",
        value="",
        key="link_tratamento_google_sheets",
        placeholder="https://docs.google.com/spreadsheets/d/.../edit#gid=0",
    )
    link_tratamento = str(link_tratamento or "").strip()
    st.caption("Para funcionar sem credenciais, a planilha precisa estar compartilhada como 'Qualquer pessoa com o link - Leitor'.")

    st.markdown("### Ou envie o arquivo manualmente")
    planilha_tratamento = st.file_uploader(
        "Planilha do Pedido Final",
        type=["xlsx", "xls", "csv"],
        key="upload_tratamento_pedido_final",
    )

    if not link_tratamento and not planilha_tratamento:
        st.info("Cole o link da planilha aprovada ou envie a planilha do pedido final para gerar o arquivo de importação Autcom.")
        st.stop()

    try:
        if link_tratamento:
            df_tratamento = ler_planilha_tratamento_google_sheets(link_tratamento)
            origem_tratamento = "Google Sheets"
        else:
            df_tratamento = ler_planilha_tratamento_pedido(planilha_tratamento)
            origem_tratamento = "upload"
        df_tratamento.columns = [str(c).strip() for c in df_tratamento.columns]

        st.success(f"Planilha lida com sucesso via {origem_tratamento}: {len(df_tratamento)} linha(s).")

        colunas_preview = [c for c in ["zx", "descricao", "Código Fábrica", "PEDIDO Final", "Preço Última Compra", "Valor Final do Pedido", "Total Geral do Pedido"] if c in df_tratamento.columns]
        if colunas_preview:
            st.dataframe(
                df_tratamento[colunas_preview].head(50),
                use_container_width=True,
                hide_index=True,
                height=360,
            )
        else:
            st.dataframe(df_tratamento.head(50), use_container_width=True, hide_index=True, height=360)

        excel_tratamento = gerar_excel_autcom_tratamento(df_tratamento)
        col_dl_autcom, col_dl_fornecedor = st.columns(2)
        with col_dl_autcom:
            st.download_button(
                "⬇ Baixar pedido tratado para importação no Autcom",
                excel_tratamento,
                "pedido_tratado_importacao_autcom.xlsx",
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                type="primary",
            )
        with col_dl_fornecedor:
            st.download_button(
                "⬇ Baixar pedido para envio ao fornecedor",
                gerar_excel_fornecedor_tratamento(df_tratamento),
                "pedido_envio_fornecedor.xlsx",
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
    except Exception as e:
        st.error(str(e))

    st.stop()

if not giro_pdf:
    st.info("Envie o PDF de Giro de Estoque para iniciar a análise.")
    st.stop()

aviso_pdf_grande(giro_pdf)
aviso_pymupdf_ausente_para_giro(giro_pdf)
with st.spinner("Lendo Giro de Estoque com cache otimizado..."):
    df_giro, MESES, texto_giro, metodo_giro = parse_giro_estoque_pdf(giro_pdf)

if df_giro.empty:
    st.error(diagnosticar_pdf_giro(texto_giro))
    with st.expander("Diagnóstico técnico do PDF de Giro"):
        st.write(f"Metodo de leitura: {metodo_giro}")
        st.write(f"Meses identificados: {MESES}")
        st.text((texto_giro or "")[:4000])
    st.stop()

meses_disponiveis_pdf = list(MESES)
MESES = meses_disponiveis_pdf

if cadastro_google is not None and not cadastro_google.empty:
    df_giro = aplicar_cadastro_dataframe(df_giro, cadastro_google)
else:
    df_giro = aplicar_cadastro(df_giro, cadastro_csv)

mes_atual_pdf = mes_atual_referencia()
if mes_atual_pdf in MESES:
    considerar_mes_atual_media = st.radio(
        f"Considerar {label_mes_giro(mes_atual_pdf)} no cálculo da Média Giro Geral?",
        ["Não", "Sim"],
        index=0,
        horizontal=True,
        key="considerar_mes_atual_media_giro",
        help="Se escolher Não, o mês atual continua aparecendo na tabela, mas não entra na Média Giro Geral nem na sugestão do pedido.",
    ) == "Sim"
else:
    considerar_mes_atual_media = True

df_transito = pd.DataFrame(columns=["codigo", "Saldo em Trânsito/ABERTO"])
if pedidos_pdf:
    aviso_pdf_grande(pedidos_pdf)
    try:
        with st.spinner("Lendo Pedidos de Compra em Aberto com cache otimizado..."):
            df_transito = parse_pedidos_compra_aberto_pdf(pedidos_pdf)
        if df_transito is None or df_transito.empty:
            st.warning(
                "Não encontrei itens em aberto nesse PDF. "
                "A análise seguirá apenas com o estoque atual."
            )
            df_transito = pd.DataFrame(columns=["codigo", "Saldo em Trânsito/ABERTO"])
        else:
            total_aberto = pd.to_numeric(df_transito.get("Saldo em Trânsito/ABERTO", 0), errors="coerce").fillna(0).sum()
            st.success(f"Pedidos em aberto lidos: {len(df_transito)} item(ns), total em aberto {format_num_br(total_aberto, 1)}.")
    except Exception as e:
        st.warning(
            "Não consegui ler o PDF de Pedidos em Aberto. "
            "A análise seguirá apenas com o estoque atual. "
            f"Detalhe: {e}"
        )
        df_transito = pd.DataFrame(columns=["codigo", "Saldo em Trânsito/ABERTO"])

tabela_resumo = montar_tabela_consolidada(
    df_giro,
    df_transito=df_transito,
    dias_estoque_alvo=dias_estoque_alvo,
    meses_alerta_sem_compra=meses_alerta_sem_compra,
    considerar_mes_atual_media=considerar_mes_atual_media,
    meses_ref=MESES,
)

assinatura_base = (
    assinatura_dataframe_colunas(
        tabela_resumo,
        ["codigo", "Código Fábrica", "Embalagem", "Saldo em Trânsito/ABERTO"],
    )
    + f"|meses={','.join(MESES)}|dias={dias_estoque_alvo}|alerta={meses_alerta_sem_compra}|mes_atual_media={int(considerar_mes_atual_media)}"
)
if st.session_state.get("assinatura_base_pedido") != assinatura_base:
    st.session_state["pedido_editado"] = inicializar_pedido_editavel(tabela_resumo)
    st.session_state["assinatura_base_pedido"] = assinatura_base

editor_pedido_key = f"editor_pedido_final_{hashlib.md5(assinatura_base.encode('utf-8')).hexdigest()[:10]}"

colunas_consolidadas = [
    "codigo", "descricao", "Código Fábrica", "Embalagem",
    *[col_giro("Giro Lojas", mes) for mes in MESES],
    "Média Giro Lojas", "Estoque Lojas",
    *[col_giro("Giro Única", mes) for mes in MESES],
    "Média Giro Única", "Estoque Única",
    *[col_giro("Giro Geral", mes) for mes in MESES],
    "Média Giro Geral", "Estoque Atual Geral", "Estoque Geral", "Saldo em Trânsito/ABERTO", "Estoque Final", "Alerta Estoque",
    "Estoque Alvo", "Sugestão Sistema", "Sugestão arredondada", "Data Última Compra", "Preço Última Compra",
]
for col in colunas_consolidadas:
    if col not in tabela_resumo.columns:
        tabela_resumo[col] = 0

render_kpis_gerais(tabela_resumo, st.session_state.get("pedido_editado"))
st.markdown("---")

if pagina == "Giro Consolidado":
    st.markdown('<div class="section-title">Giro Consolidado</div>', unsafe_allow_html=True)
    st.caption(
        "A data da última compra é puxada somente da loja 009. "
        "Quando a data ultrapassa o parâmetro de meses sem compra, aparece o ícone ⚠ ao lado da data."
    )

    tabela = tabela_resumo[colunas_consolidadas].copy()
    tabela = filtrar_tabela(tabela, ["codigo", "descricao", "Código Fábrica"], "busca_consolidada")
    render_tabela_interativa_colorida(tabela)

    st.download_button(
        "⬇ Baixar tabela consolidada em CSV",
        gerar_csv(tabela),
        "tabela_consolidada_giro_pedido.csv",
        "text/csv",
    )

    st.markdown("---")
    st.markdown('<div class="section-title">🔎 Drill por produto</div>', unsafe_allow_html=True)
    opcoes_produtos = (
        tabela_resumo["codigo"].astype(str) + " - " + tabela_resumo["descricao"].astype(str)
    ).drop_duplicates().tolist()

    produto_selecionado = st.selectbox(
        "Selecione um item para ver o giro e o saldo em estoque por unidade",
        options=[""] + opcoes_produtos,
        key="produto_drill_consolidada",
    )

    if produto_selecionado:
        codigo_produto = produto_selecionado.split(" - ")[0]
        detalhe = montar_detalhe_produto(df_giro, codigo_produto)
        st.dataframe(
            detalhe.style.format(formatadores_para_tabela(detalhe)),
            use_container_width=True,
            hide_index=True,
            height=360,
            column_config={
                "Cód. Empresa": st.column_config.TextColumn("Cód. Empresa", pinned=True),
                "Unidade": st.column_config.TextColumn("Unidade", pinned=True),
            },
        )

elif pagina == "Pedido de Compra":
    st.markdown('<div class="section-title">🛒 Pedido de Compra</div>', unsafe_allow_html=True)
    st.caption(
        "Todos os itens aparecem aqui, inclusive os com sugestão zero. "
        "A coluna PEDIDO Final é editável. A coluna Valor Final do Pedido é recalculada por quantidade × preço última compra."
    )

    pedido_base_completo = st.session_state.get("pedido_editado", inicializar_pedido_editavel(tabela_resumo)).copy()
    if "Estoque Geral" not in pedido_base_completo.columns and "Estoque Atual Geral" in pedido_base_completo.columns:
        pedido_base_completo["Estoque Geral"] = pedido_base_completo["Estoque Atual Geral"]
    pedido_base_completo = atualizar_valor_e_origem(pedido_base_completo)

    colunas_sugestao = colunas_pedido_compras(MESES)
    for col in colunas_sugestao:
        if col not in pedido_base_completo.columns:
            pedido_base_completo[col] = 0 if col not in ["codigo", "descricao", "Código Fábrica", "Data Última Compra", "Origem Sugestão"] else ""

    pedido_view = pedido_base_completo[colunas_sugestao].sort_values(["Sugestão Sistema", "descricao"], ascending=[False, True]).copy()
    pedido_view = filtrar_tabela(pedido_view, ["codigo", "descricao", "Código Fábrica"], "busca_sugestao")

    # Recalcula na hora o Valor Final do Pedido quando o usuário altera PEDIDO Final.
    estado_editor = st.session_state.get(editor_pedido_key, {})
    alteracoes_linhas = estado_editor.get("edited_rows", {}) if isinstance(estado_editor, dict) else {}
    if alteracoes_linhas:
        indices_visiveis = list(pedido_view.index)
        for posicao_linha, alteracoes in alteracoes_linhas.items():
            try:
                posicao = int(posicao_linha)
                if posicao < 0 or posicao >= len(indices_visiveis):
                    continue
                indice_real = indices_visiveis[posicao]
                if "PEDIDO Final" in alteracoes:
                    novo_pedido = pd.to_numeric(alteracoes.get("PEDIDO Final"), errors="coerce")
                    if pd.isna(novo_pedido):
                        novo_pedido = 0
                    novo_pedido = int(round(float(novo_pedido)))
                    embalagem_item = int(pd.to_numeric(pedido_view.loc[indice_real].get("Embalagem", 0), errors="coerce") or 0)
                    pedido_validado = ajustar_pedido_para_multiplo_embalagem(novo_pedido, embalagem_item)

                    if novo_pedido > 0 and embalagem_item > 1 and novo_pedido != pedido_validado:
                        descricao_item = str(pedido_view.loc[indice_real].get("descricao", "")).strip()
                        codigo_item = str(pedido_view.loc[indice_real].get("codigo", "")).zfill(5)
                        st.warning(
                            f"Item {codigo_item} - {descricao_item}: a embalagem é com {embalagem_item} unidades. "
                            f"Altere para {pedido_validado}. O sistema ajustou automaticamente para o próximo múltiplo."
                        )

                    pedido_view.loc[indice_real, "PEDIDO Final"] = pedido_validado
                    codigo_alterado = pedido_view.loc[indice_real, "codigo"]
                    mask_base = pedido_base_completo["codigo"].astype(str) == str(codigo_alterado)
                    pedido_base_completo.loc[mask_base, "PEDIDO Final"] = pedido_validado
            except Exception:
                continue

        pedido_base_completo = atualizar_valor_e_origem(pedido_base_completo)
        pedido_view = atualizar_valor_e_origem(pedido_view)
        st.session_state["pedido_editado"] = pedido_base_completo

    pedido_view = pedido_view[colunas_sugestao].copy()
    if usar_renderizacao_leve(pedido_view, LIMITE_CELULAS_EDITOR):
        st.caption(
            "Modo rápido ativado no editor para manter o Streamlit estável com muitos meses/itens."
        )
        pedido_para_editor = pedido_view
    else:
        pedido_para_editor = pedido_view.style.apply(colorir_colunas_pedido, axis=0).apply(estilos_alerta_giro_fora_curva, axis=1).format(formatadores_para_tabela(pedido_view))

    pedido_editado = st.data_editor(
        pedido_para_editor,
        use_container_width=True,
        hide_index=True,
        height=650,
        key=editor_pedido_key,
        disabled=[
            "codigo", "descricao",
            *[col_giro("Giro Geral", mes) for mes in MESES],
            "Média Giro Geral",
            "Estoque Lojas", "Estoque Única", "Estoque Geral",
            "Saldo em Trânsito/ABERTO", "Estoque Final", "Estoque Alvo",
            "Sugestão Sistema", "Sugestão arredondada",
            "Preço Última Compra", "Data Última Compra",
            "Origem Sugestão", "Valor Final do Pedido",
            "Embalagem", "Código Fábrica",
        ],
        column_config={
            "codigo": st.column_config.TextColumn("Código", pinned=True),
            "descricao": st.column_config.TextColumn("Descrição", width="large", pinned=True),
            "Código Fábrica": st.column_config.TextColumn("Código Fábrica", width="medium"),
            "Embalagem": st.column_config.NumberColumn("Embalagem", min_value=0, step=1, format="%d"),
            "Média Giro Lojas": st.column_config.NumberColumn("Média Giro Lojas", format="%.1f"),
            "Estoque Lojas": st.column_config.NumberColumn("Estoque Lojas", format="%.1f"),
            "Média Giro Única": st.column_config.NumberColumn("Média Giro Única", format="%.1f"),
            "Estoque Única": st.column_config.NumberColumn("Estoque Única", format="%.1f"),
            "Média Giro Geral": st.column_config.NumberColumn("Média Giro Geral", format="%.1f"),
            "Estoque Geral": st.column_config.NumberColumn("Estoque Geral", format="%.1f"),
            "Saldo em Trânsito/ABERTO": st.column_config.NumberColumn("Saldo em Trânsito", format="%.1f"),
            "Estoque Final": st.column_config.NumberColumn("Estoque Final", format="%.1f"),
            "Estoque Alvo": st.column_config.NumberColumn("Estoque Alvo", format="%.1f"),
            "Sugestão Sistema": st.column_config.NumberColumn("Sugestão Sistema", format="%d"),
            "Sugestão arredondada": st.column_config.NumberColumn("Sugestão arredondada", format="%d"),
            "Preço Última Compra": st.column_config.NumberColumn("Preço Última Compra", format="R$ %.2f"),
            "PEDIDO Final": st.column_config.NumberColumn("PEDIDO Final", min_value=0, step=1, format="%d"),
            "Valor Final do Pedido": st.column_config.NumberColumn("Valor Final do Pedido", format="R$ %.2f"),
        },
    )

    pedido_editado = pd.DataFrame(pedido_editado)
    pedido_editado, mensagens_validacao = validar_pedidos_por_embalagem(pedido_editado)
    if mensagens_validacao:
        st.warning("Algumas quantidades foram ajustadas para respeitar a embalagem:\n\n" + "\n".join(mensagens_validacao[:10]))
        if len(mensagens_validacao) > 10:
            st.caption(f"Mais {len(mensagens_validacao) - 10} ajuste(s) foram aplicado(s).")

        atualizacoes_validas = pedido_editado[["codigo", "PEDIDO Final"]].copy()
        mapa_validado = atualizacoes_validas.drop_duplicates("codigo", keep="last").set_index("codigo")["PEDIDO Final"]
        pedido_base_completo["PEDIDO Final"] = pedido_base_completo.apply(
            lambda row: int(mapa_validado.loc[row["codigo"]]) if row["codigo"] in mapa_validado.index else int(row["PEDIDO Final"]),
            axis=1,
        )
        st.session_state["pedido_editado"] = atualizar_valor_e_origem(pedido_base_completo)

    pedido_editado = atualizar_valor_e_origem(pedido_editado)
    pedido_editado = pedido_editado[colunas_sugestao]

    valor_editado = totalizar_valor_pedido(pedido_editado)
    st.markdown(
        f"""
        <div style="margin-top: 16px; padding: 18px; border-radius: 14px; background: #f3f6ff; border: 1px solid #d9e2ff;">
            <div style="font-size: 14px; color: #475569;">Valor final do pedido em tela</div>
            <div style="font-size: 30px; font-weight: 700; color: #0f172a;">{format_moeda_br(valor_editado)}</div>
        </div>
        """,
        unsafe_allow_html=True,
    )

    if st.button("💾 Salvar Pedido", type="primary"):
        base_completa = st.session_state["pedido_editado"].copy()
        if "Estoque Geral" not in base_completa.columns and "Estoque Atual Geral" in base_completa.columns:
            base_completa["Estoque Geral"] = base_completa["Estoque Atual Geral"]
        atualizacoes = pedido_editado[["codigo", "PEDIDO Final", "Embalagem", "descricao"]].copy()
        atualizacoes, mensagens_salvar = validar_pedidos_por_embalagem(atualizacoes)
        if mensagens_salvar:
            st.warning("Antes de salvar, o sistema ajustou quantidades para respeitar a embalagem:\n\n" + "\n".join(mensagens_salvar[:10]))
            if len(mensagens_salvar) > 10:
                st.caption(f"Mais {len(mensagens_salvar) - 10} ajuste(s) foram aplicado(s).")
        atualizacoes["PEDIDO Final"] = pd.to_numeric(atualizacoes["PEDIDO Final"], errors="coerce").fillna(0).round(0).astype(int)
        mapa_qtd = atualizacoes.drop_duplicates("codigo", keep="last").set_index("codigo")["PEDIDO Final"]
        base_completa["PEDIDO Final"] = base_completa.apply(
            lambda row: int(mapa_qtd.loc[row["codigo"]]) if row["codigo"] in mapa_qtd.index else int(row["PEDIDO Final"]),
            axis=1,
        )
        base_completa = atualizar_valor_e_origem(base_completa)
        st.session_state["pedido_editado"] = base_completa
        st.success("Pedido salvo. Vá para a página Exportar Pedido para criar o Google Sheets na pasta de aprovação.")

    st.markdown("---")
    st.markdown("### Exportar pedido em Google Sheets")
    st.markdown('<div class="sheets-badge">▦ Google Sheets</div>', unsafe_allow_html=True)
    st.caption("Nesta versão, o pedido será enviado para um Google Apps Script, que cria a planilha Google Sheets diretamente na pasta de aprovação. Não usa OAuth, refresh_token nem Service Account no Python.")
    st.link_button("Abrir pasta destino no Drive", google_link_pasta(GOOGLE_PASTA_APROVACAO_ID), use_container_width=True)

    if apps_script_configurado():
        with st.form("form_exportar_pedido_sheets"):
            nome_pedido_drive = st.text_input("Nome do pedido", value=f"Pedido {datetime.now().strftime('%d-%m-%Y')}")
            fornecedor_drive = st.text_input("Fornecedor", value="")
            usuario_drive = st.text_input("Criado por", value="")
            enviar_drive = st.form_submit_button("▦ Criar Google Sheets na pasta", type="primary")

        if enviar_drive:
            try:
                pedido_para_sheets = st.session_state.get("pedido_editado", pedido_editado).copy()
                pedido_para_sheets = atualizar_valor_e_origem(pedido_para_sheets)
                pedido_para_sheets = pedido_para_sheets[colunas_pedido_compras(MESES)]
                resultado_sheets = apps_script_criar_planilha_pedido(
                    nome_pedido_drive,
                    fornecedor_drive,
                    pedido_para_sheets,
                    criado_por=usuario_drive,
                )
                st.success("Pedido exportado em Google Sheets com sucesso via Apps Script.")
                st.write(f"Valor do pedido: **{format_moeda_br(resultado_sheets.get('valor', 0))}**")
                st.link_button("Abrir planilha criada", resultado_sheets["link"], use_container_width=True)
                st.link_button("Abrir pasta no Drive", resultado_sheets["folder_link"], use_container_width=True)
            except Exception as e:
                st.error(str(e))
    else:
        st.info(apps_script_mensagem_configuracao())
        st.code("""[apps_script]
web_app_url = "https://script.google.com/macros/s/SEU_DEPLOY_ID/exec"
# token = "opcional""", language="toml")

    with st.expander("Fallback: baixar CSV local"):
        st.download_button(
            "⬇ Baixar pedido em CSV",
            gerar_csv(pedido_editado[colunas_pedido_compras(MESES)]),
            "pedido_editavel.csv",
            "text/csv",
        )

elif pagina == "Exportações":
    st.markdown('<div class="section-title">Exportações</div>', unsafe_allow_html=True)
    st.caption("O Excel será gerado para importação no Autcom: coluna B = código, coluna F = quantidade, coluna H = valor unitário, sem cabeçalho.")

    pedido_final = st.session_state.get("pedido_editado", inicializar_pedido_editavel(tabela_resumo)).copy()
    pedido_final, mensagens_exportar = validar_pedidos_por_embalagem(pedido_final)
    if mensagens_exportar:
        st.warning("O sistema ajustou quantidades para respeitar a embalagem antes da exportação:\n\n" + "\n".join(mensagens_exportar[:10]))
        if len(mensagens_exportar) > 10:
            st.caption(f"Mais {len(mensagens_exportar) - 10} ajuste(s) foram aplicado(s).")
        st.session_state["pedido_editado"] = pedido_final.copy()
    pedido_final = atualizar_valor_e_origem(pedido_final)
    pedido_final = pedido_final[pedido_final["PEDIDO Final"] > 0].copy().sort_values("descricao")

    valor_final = totalizar_valor_pedido(pedido_final)
    st.markdown(
        f"""
        <div style="margin: 10px 0 18px 0; padding: 18px; border-radius: 14px; background: #f3f6ff; border: 1px solid #d9e2ff;">
            <div style="font-size: 14px; color: #475569;">Valor final do pedido salvo</div>
            <div style="font-size: 30px; font-weight: 700; color: #0f172a;">{format_moeda_br(valor_final)}</div>
        </div>
        """,
        unsafe_allow_html=True,
    )

    st.dataframe(
        pedido_final[[
            "codigo", "descricao", "Código Fábrica", "Sugestão Sistema", "Sugestão arredondada", "PEDIDO Final", "Preço Última Compra",
            "Valor Final do Pedido", "Data Última Compra", "Origem Sugestão",
        ]].style.format(formatadores_para_tabela(pedido_final)),
        use_container_width=True,
        hide_index=True,
        height=520,
    )

    col_dl1, col_dl2 = st.columns(2)

    with col_dl1:
        render_download_card("Excel Autcom", "Arquivo sem cabeçalho: coluna B = código, F = quantidade, H = preço.")
        try:
            excel_bytes = gerar_excel_pedido(pedido_final)
            st.download_button(
                "⬇ Baixar pedido para importação no Autcom",
                excel_bytes,
                "pedido_importacao_autcom.xlsx",
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                type="primary",
            )
        except RuntimeError as e:
            st.error(str(e))

    with col_dl2:
        render_download_card("Cópia para fornecedor", "Lista simples com código de fábrica, descrição e quantidade.")
        st.download_button(
            "⬇ Baixar cópia CSV para fornecedor",
            gerar_copia_fornecedor_csv(pedido_final),
            "copia_fornecedor.csv",
            "text/csv",
        )
